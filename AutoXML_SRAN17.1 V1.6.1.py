#modules
import xml.etree.ElementTree as ET
import datetime as dt
import time
# now = datetime.now().time() # time object
from tkinter import *
# import tkinter
# import tkinter.font as font
# import filedialog module
from tkinter import filedialog, ttk
import os
# import xml.dom.minidom
# import webbrowser
import csv

#comments:
#As of Python 3.9, ElementTree has an indent() function for pretty-printing XML trees.
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~


# class for tooltips
class CreateToolTip(object):
    '''
    create a tooltip for a given widget
    '''

    def __init__(self, widget, text='widget info'):
        self.widget = widget
        self.text = text
        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.close)

    def enter(self, event=None):
        #time.sleep(0.5)
        x = y = 0
        x, y, cx, cy = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 20
        # creates a toplevel window
        self.tw = Toplevel(self.widget)
        # Leaves only the label and removes the app window
        self.tw.wm_overrideredirect(True)
        self.tw.wm_geometry("+%d+%d" % (x, y))
        label = Label(self.tw, text=self.text, justify='left',background='white', relief='solid', borderwidth=0.5,font=("arial", "8", "normal"))
        label.pack(ipadx=1)

    def close(self, event=None):
        if self.tw:
            self.tw.destroy()


# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

NEname = ''


def auxfunct():
    print(time.time())


# Create the root window
window = Tk()
#window.overrideredirect(False) #can create a Frameless Window

#adjust the window style
#style=ttk.Style(window)
#print(style.theme_names()) #('winnative', 'clam', 'alt', 'default', 'classic', 'vista', 'xpnative')
#current_theme = style.theme_use()
#print(current_theme) #vista
#style.theme_use('clam')

#print(current_theme)
# Set window title
window.title('       AutoXML V1.6.1 BULK    \nSRAN17.1')
# window.iconbitmap("xml.ico")
# window.iconbitmap("download.ico")
try:
    window.iconbitmap("happy1.ico")
except:
    window.title('       AutoXML V1.6.1 BULK    \nSRAN17.1')
# Set window size

window.geometry('{}x{}'.format(1290, 740))
# Set window background color
window.config(background="white")

# Function for opening the
# file explorer window

label_file_explorer1 = Label(window, text="No file selected", width=130, height=1, fg="blue")
label_file_explorer1.grid(column=2, row=1)
# label_UMPT_PORT_SW = Label(window, text="", width=130, height=1, fg="blue")
# label_UMPT_PORT_SW.grid(column=2, row=2)

fo=open('LOG.txt', 'a')

def browseFiles():
    global root
    global tree
    global OUTPUT
    global fo
    fo = open('LOG.txt', 'a')
    # time.sleep(1)
    filename = filedialog.askopenfilename(initialdir="/C:/", title="Select a File",filetypes=(("XML files", "*.xml*"), ("all files", "*.*")))
    filename2 = filename
    # time.sleep(1)
    fo.write(f'\n{dt.datetime.now()} --  ' + 'Selecting file:  ' + filename2 + '\n')
    # print(filename2)
    # Change label contents
    #label_file_explorer1.configure(text="File Opened: " + filename2)
    text = Text(window, height=1, width=130, font=("arial", 10))
    text.insert(INSERT,"File Opened: " + filename2)
    text.grid(column=2, row=1)
    tree = ET.parse(filename2)
    root = tree.getroot()
    # print(root.text)
    OUTPUT = filename2

    return filename2


def board_button():
    label_UMPT_PORT_SW.configure(text=UMPT_ETHPORT_SW)


# RULES
# ET.dump(), break need to be inside most inner loop
# when removing from list, iterate list backwards!
# when removing last legacy structure format object from some class, cannot add new structure object!

# add ALMPORT with RRU
# add rxbranch tbranch for rru
# add antenna port for rru

# add rruchain
# add brd/bbp

# add cpriport in both rru and bbp functions
# add sfp in both rru and bbp functions

# change subrack number 0->1
# change subrack MO -BTS3900->BTS5900
# BTS3900-><BTS3910
# BTS3910-BTS5900

# remove RRU (+all  related MO)
# remove RRUCHAIN
# remove BBP

# change BTS3900 to BTS5900
# subrack_type_change_3900to5900()

# change BBU subrack 0 to 1
# change_bbu_subrack_id()


OUTPUT = browseFiles()

# register all  namespaces
# https://stackoverflow.com/questions/8983041/saving-xml-files-using-elementtree
ET.register_namespace('spec', "1.0.0")
ET.register_namespace('', "http://www.huawei.com/specs/bsc6000_nrm_forSyn_collapse_1.0.0")

tree = ET.parse(OUTPUT)
root = tree.getroot()
# print(dir(root))
# print(dir(tree))

# print(root.text)


# print('root tag',root.tag)
# print('root tag atributes',root.attrib)
# children = root.getchildren()
# print(children)


# xmlstr = ET.tostring(root, encoding='unicode', method='xml')
# print(xmlstr[0:10000000])
# fo=open('test.xml','w')
# fo.write(f'{dt.datetime.now()} --  {NEname}  '+xmlstr)
# fo.close()
#
# # LOAD XML AND XSL
# doc = et.parse('test.xml')
# xsl = et.parse('XSLT_Script.xsl')
#
# # CONFIGURE AND RUN TRANSFORMER
# transform = et.XSLT(xsl)
# result = transform(doc)
#
# # OUTPUT RESULT TREE TO FILE
# with open('test1.xml', 'wb') as f:
#     f.write(result)


def parse_existing_MO_name(root, MO):
    MO_list = []
    MO_dict = []  # list not dict
    MO_dict1 = []  # list not dict
    for child0 in root:  # iterate tags under root
        # print('root child tag ', child0.tag)
        # print(f'root child tag {child.tag} attributes',  child.attrib, '\n')
        # print(dir(child.tag))
        # if child.tag=='syndata':
        #     print('root child tag atributes', child.tag, child.attrib,'\n')
        # for item in child.iter('class'):
        #     print(item.attrib)

        for child1 in child0:  # iterate objects under root / syndata
            for child2 in child1:  # iterate objects under root / syndata / SECTOR
                if child2.tag == '' + str(MO):
                    MO_dict.clear()
                    #print('\n')
                    #print(MO)
                    #print(child2.tag)

                    for child3 in child2:  # iterate objects under root / syndata / SECTOR / attributes

                        # print(child3.tag) #SECTOR
                        for child4 in child3:  # iterate objects under root / syndata / SECTOR / attributes /
                            if len(child4) > 0:
                                #print({child4.tag: child4.text})  # print objects under attributes
                                MO_dict.append({child4.tag: child4.text})
                                #print('------------  WITH CHILDREN')
                            else:
                                #print({child4.tag: child4.text})
                                MO_dict.append({child4.tag: child4.text})
                            for child5 in child4:  # iterate child objects of objects under attributes (ex. under SECTORNANTENNA)
                                if len(child5) == 0:
                                    #print({child5.tag: child5.text})
                                    MO_dict.append({child5.tag: child5.text})
                                    #print('------------  NO CHILDREN')
                                else:
                                    #print({child5.tag: child5.text})
                                    MO_dict.append({child5.tag: child5.text})
                                    #print('------------  WITH CHILDREN')
                                    for child6 in child5:
                                        #print({child6.tag: child6.text})
                                        MO_dict.append({child6.tag: child6.text})
                        MO_dict1 = MO_dict.copy()
                        MO_list.append(MO_dict1)
                        # MO_dict.clear()

    return MO_list


def parse_BBP():
    BBP_list = []
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'BBP':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SN' and child4.text not in BBP_list:
                                BBP_list.append(child4.text)
                                # ET.dump(child1)
    # print('****************', BBP_list)
    return BBP_list


def LMPT_to_UMPT():
    string1 = ''
    PN_list_unique = []
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'DEVIP':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'PN' and child4.text not in PN_list_unique:
                                PN_list_unique.append(child4.text)
                                ET.dump(child1)
                                break
    flag0 = False
    for child0 in root:
        lag0 = False
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'ETHPORT':
                    for child3 in child2:
                        lag0 = False
                        for child4 in child3:
                            if child4.tag == 'PN' and child4.text == PN_list_unique[0]:
                                flag0 = True
                            if flag0 and child4.tag == 'PA':
                                port_devip_attribute = str(child4.text)  # COPPER~0, FIBER~1, AUTO~2, UNCONFIG~255
                                ET.dump(child1)
                                break
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'MPT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'TYPE':
                                MPT = child4.text
                                ET.dump(child1)
                                break

    # print('port number', PN_list_unique[0])
    # print('port attribute', port_devip_attribute)
    # print('BOARD TYPE!!!', MPT)
    if MPT == '12289':
        # if PM_list_unique[0]==1 just change in MPT
        # if PM_list_unique[0]==0 change ethport, ipsecbind, MPT
        if (PN_list_unique[0] == '1' and port_devip_attribute == '1') or (
                PN_list_unique[0] == '0' and port_devip_attribute == '0'):
            # UMPT~1, WMPT~8193, LMPT~12289
            for child0 in root:
                for child1 in child0:
                    for child2 in child1:
                        if child2.tag == '' + 'MPT':
                            for child3 in child2:
                                for child4 in child3:
                                    if child4.tag == 'TYPE' and child4.text == '12289':
                                        child4.text = '1'
                                        ET.dump(child1)
                                        break
        if PN_list_unique[0] == '0' and port_devip_attribute == '1':
            for child0 in root:
                for child1 in child0:
                    for child2 in child1:
                        if child2.tag == '' + 'MPT':
                            for child3 in child2:
                                for child4 in child3:
                                    if child4.tag == 'TYPE' and child4.text == '12289':
                                        child4.text = '1'
                                        ET.dump(child1)
                                        break
            for child0 in root:
                for child1 in child0:
                    for child2 in child1:
                        if child2.tag == '' + 'DEVIP':
                            for child3 in child2:
                                for child4 in child3:
                                    if child4.tag == 'PN' and child4.text == '0':
                                        child4.text = '1'
                                        ET.dump(child1)
                                        break
            for child0 in root:
                for child1 in child0:
                    for child2 in child1:
                        if child2.tag == '' + 'IPSECBIND':
                            for child3 in child2:
                                for child4 in child3:
                                    if child4.tag == 'PN' and child4.text == '0':
                                        child4.text = '1'
                                        ET.dump(child1)
                                        break
            # port attribute: COPPER~0, FIBER~1, AUTO~2, UNCONFIG~255
            # SPEED: 10M~0, 100M~1, 1000M~2, AUTO~3, 10G~5, 40G~6, 100G~7, 25G~8, UNCONFIG~255
            # DUPLEX MODE: FULL~1, AUTO~2, UNCONFIG~255

            flag = False
            for child0 in root:
                flag = False
                for child1 in child0:
                    for child2 in child1:
                        if child2.tag == '' + 'ETHPORT':
                            for child3 in child2:
                                flag = False
                                for child4 in child3:
                                    if child4.tag == 'PA' and child4.text == str(
                                            0):
                                        child4.text = str(1)
                                        flag = True
                                    if flag and child4.tag == 'SPEED' and child4.text == str(
                                            3):
                                        child4.text = str(2)
                                    if (flag and child4.tag == 'DUPLEX' and child4.text == str(
                                            2)) or (
                                            flag and child4.tag == 'DUPLEX' and child4.text == str(
                                            255)):
                                        child4.text = str(1)
                                        ET.dump(child1)
                                        break
    tree.write(OUTPUT)


def f1():
    for label in window.grid_slaves():
        if int(label.grid_info()["column"]) == 2 and int(label.grid_info()["row"]) > 1 and int(label.grid_info()["row"]) < 40:
            label.grid_forget()
    check_butt_list = [CheckVar1, CheckVar2, CheckVar3, CheckVar30, CheckVar300, CheckVar4, CheckVar5, CheckVar6,
                       CheckVar7, CheckVar9, CheckVar10, CheckVar11, CheckVar12, CheckVar13,
                       CheckVar14, CheckVar15, CheckVar16, CheckVar17]
    check_butt_list1 = [C1, C2, C3, C30, C300, C4, C5, C6, C7, C9, C10, C11, C12, C13, C14, C15, C16, C17]
    for item in range(0,len(check_butt_list)):
        check_butt_list[item].set(0)
        check_butt_list1[item].config(state=NORMAL)

    # def checkall1():
    #     for cb in check_butt_list1:
    #         cb.select()
    # def checkall2():
    #     for cb in check_butt_list1:
    #         cb.deselect()
    #
    # Checkbutton(window, text="select all", command=checkall1).grid(row=31, column=1, sticky=W)
    # Checkbutton(window, text="deselect all", command=checkall2).grid(row=32, column=1, sticky=W)
    #     item.set(1)
    #     item.set(0)
    # for it in [0,1,0]:
    #     CCC=CheckVar1.set(it)
    #     C1 = Checkbutton(window, text="BTS3900 to BTS5900 reconstruction", variable=CCC, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender')
    #     CCC=CheckVar2.set(it)
    #     C2 = Checkbutton(window, text="BBU subrack ID 0 to 1", variable=CCC, onvalue=1, offvalue=0, height=1,width=38, anchor="w", bg='lavender')
    #     CCC=CheckVar300.set(it)
    #     C300= Checkbutton(window, text="LMPT to UMPT", variable=CCC, onvalue=1, offvalue=0, height=1, width=38,anchor="w", bg='lavender')
    #     CCC=CheckVar30.set(it)
    #     C30 = Checkbutton(window, text="5900: L800 CD INT 80-82-84, no L700", variable=CCC, command=cb_check30,onvalue=1, offvalue=0, height=1, width=38, anchor="w", bg='lavender')
    #     CCC=CheckVar3.set(it)
    #     C3 = Checkbutton(window, text="5900: L800 AB REC/INT 60-61-62 SN 2,3 -> 4", variable=CCC, command=cb_check3,onvalue=1, offvalue=0, height=1, width=38, anchor="w", bg='lavender')
    #     CCC=CheckVar4.set(it)
    #     C4 = Checkbutton(window, text="5900: L800 CD REC + L700 AB INT 80-82-84", variable=CCC, command=cb_check4,onvalue=1, offvalue=0, height=1, width=38, anchor="w", bg='lavender')
    #     CCC=CheckVar5.set(it)
    #     C5 = Checkbutton(window, text="5900: L800 AB REC + L700 CD INT 80-82-84", variable=CCC, command=cb_check5,onvalue=1, offvalue=0, height=1, width=38, anchor="w", bg='lavender')
    #     CCC=CheckVar6.set(it)
    #     C6 = Checkbutton(window, text="5900: L700 INT 90-91-92 SN 5", variable=CCC, command=cb_check6, onvalue=1,offvalue=0, height=1, width=38, anchor="w", bg='lavender')
    #     CCC=CheckVar7.set(it)
    #     C7 = Checkbutton(window, text="5900: L1800 INT ABCD", variable=CCC, command=cb_check7, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender')
    #     CCC=CheckVar80.set(it)
    #     C80 = Checkbutton(window, text="5900: L1800 INT AC", variable=CCC, command=cb_check80, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender')
    #     CCC=CheckVar8.set(it)
    #     C8 = Checkbutton(window, text="5900: L1800 INT AB", variable=CCC, command=cb_check8, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender')
    #     CCC=CheckVar9.set(it)
    #     C9 = Checkbutton(window, text="5900: L2100 INT AB", variable=CCC, command=cb_check9, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender')
    #     CCC=CheckVar10.set(it)
    #     C10 = Checkbutton(window, text="5900: L2100 AC INT + L1800 BD REC/INT PN_3.4.5", variable=CCC,command=cb_check10, onvalue=1, offvalue=0, height=1, width=38, anchor="w", bg='lavender')
    #     CCC=CheckVar11.set(it)
    #     C11 = Checkbutton(window, text="5900: L2100 BD INT + L1800 AC REC/INT PN_3.4.5", variable=CCC,command=cb_check11, onvalue=1, offvalue=0, height=1, width=38, anchor="w", bg='lavender')
    #     CCC=CheckVar12.set(it)
    #     C12 = Checkbutton(window, text="5900: L1800 AC INT + OnAir L2100 BD _PN_3.4.5", variable=CCC,command=cb_check12, onvalue=1, offvalue=0, height=1, width=38, anchor="w", bg='lavender')
    #     CCC=CheckVar13.set(it)
    #     C13 = Checkbutton(window, text="5900: L1800 BD INT + OnAir L2100 AC _PN_3.4.5", variable=CCC,command=cb_check13, onvalue=1, offvalue=0, height=1, width=38, anchor="w", bg='lavender')
    #     CCC=CheckVar14.set(it)
    #     C14 = Checkbutton(window, text="5900: LTE 2600 AB ABCD", variable=CCC, onvalue=1, offvalue=0, height=1,width=38, anchor="w", bg='lavender')
    #     CCC=CheckVar15.set(it)
    #     C15 = Checkbutton(window, text="5900: LTE900", variable=CCC, onvalue=1, offvalue=0, height=1, width=38,anchor="w", bg='lavender')
    #     CCC=CheckVar16.set(it)
    #     C16 = Checkbutton(window, text="3900: LTE800", variable=CCC, onvalue=1, offvalue=0, height=1, width=38,anchor="w", bg='lavender')
    #     CCC=CheckVar17.set(it)
    #     C17 = Checkbutton(window, text="BTS5900 to BTS3900 reconstruction", variable=CCC, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender')
    #


def f2():
    import os
    # Mode to be set
    mode = 0o666

    # flags
    flags = os.O_RDWR | os.O_CREAT
    file = os.startfile('LOG.txt')


def UMPT_ETHPORT_SW():
    string1 = ''
    NE = ''
    #Ethport_portID_list_unique=[]
    Interface_portID_list_unique=[]
    subrack_inital_no = ''
    PN_list_unique = []
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'DEVIP':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'PN' and child4.text not in PN_list_unique:
                                PN_list_unique.append(child4.text)
                                #ET.dump(child1)
                                break
                if child2.tag == '' + 'INTERFACE':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'PORTID' and child4.text not in Interface_portID_list_unique:
                                Interface_portID_list_unique.append(child4.text)
                                #ET.dump(child1)
                                break

    if '2' in PN_list_unique:
        PN_list_unique.remove('2') # for sites where port '2' is found at DEVIP (Greyspot DT sites)

    if '71' in Interface_portID_list_unique and len(PN_list_unique)==0:
        PN_list_unique.append('1')
    if '70' in Interface_portID_list_unique and len(PN_list_unique)==0:
        PN_list_unique.append('0')
    if '72' in Interface_portID_list_unique and len(PN_list_unique)==0:
        PN_list_unique.append('2')
    if '73' in Interface_portID_list_unique and len(PN_list_unique)==0:
        PN_list_unique.append('3')



    flag0 = False
    for child0 in root:
        flag0 = False
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'ETHPORT':
                    for child3 in child2:
                        flag0 = False
                        for child4 in child3:
                            if child4.tag == 'PN' and child4.text == PN_list_unique[0]:
                                flag0 = True
                            if flag0 and child4.tag == 'PA':
                                port_devip_attribute = str(child4.text)  # COPPER~0, FIBER~1, AUTO~2, UNCONFIG~255
                                # ET.dump(child1)
                                break
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'MPT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'TYPE':
                                MPT = child4.text
                                # ET.dump(child1)
                                # break
                            if child4.tag == 'SRN':
                                subrack_inital_no = child4.text
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'NE':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'LOCATION':
                                NE = child4.text
                                # ET.dump(child1)
                                break
    for child0 in root:
        if child0.tag == '{1.0.0}fileHeader':
            # print(child0.attrib)
            header_attributes = child0.attrib
            neversion = header_attributes['neversion']
            break
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'SUBRACK':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'TYPE':
                                bbu_type=child4.text
                                break
    if bbu_type=='166':
        bbu_type='BBU5900  ' ## BBU3900~128, BBU3910~159, BBU5900~166
    elif bbu_type =='128':
        bbu_type = 'BBU3900  '
    elif bbu_type =='159':
        bbu_type = 'BBU3910  '

    flag = False
    BBP_slot_initial_unique = []
    for child0 in root:
        flag = False
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'BBP':
                    for child3 in child2:
                        flag = False
                        for child4 in child3:
                            if child4.tag == 'SN':
                                key = child4.text
                                flag = True
                            if flag and child4.tag == 'TYPE':
                                value = child4.text
                                dic = {key: value}
                                if dic not in BBP_slot_initial_unique:
                                    BBP_slot_initial_unique.append(dic)
                                    # dic.clear()
                            # ET.dump(child1)
    BBPtoslotstring = ''
    # UBBP~2, UBBP-W~181, GBBP~4098, WBBP~8194, LBBP~12290, LPMP~12546, LCOP~12802
    for item in BBP_slot_initial_unique:
        for k in item.keys():
            if item[k] == '2':
                BBPtoslotstring += ' slot ' + k + ' ' + 'UBBP'
            elif item[k] == '12290':
                BBPtoslotstring += ' slot ' + k + ' ' + 'LBBP'

    # print(BBP_slot_initial_unique)
    string1 = str(NE) + ':  '
    if MPT == '1':
        string1 = string1 + ' slot 7 UMPT'
    elif MPT == '12289':
        string1 = string1 + ' slot 7 LMPT'
    elif MPT == '8193':
        string1 = string1 + ' slot 7 WMPT'
    string1 += '__' + BBPtoslotstring + '  ~  '
    if PN_list_unique[0] == '1':
        string1 = string1 + 'Ethport 1'
    elif PN_list_unique[0] == '0':
        string1 = string1 + 'Ethport 0'
    elif PN_list_unique[0] == '2':
        string1 = string1 + 'Ethport 2'
    elif PN_list_unique[0] == '3':
        string1 = string1 + 'Ethport 3'

    if port_devip_attribute == '1':
        string1 = string1 + ' - ' + '  Fiber'
    elif port_devip_attribute == '0':
        string1 = string1 + ' - ' + '  Copper'
    elif port_devip_attribute == '2':
        string1 = string1 + ' - ' + '  Auto'
    elif port_devip_attribute == '255':
        string1 = string1 + ' - ' + ' Unconfig'
    string1 = string1 + '  ~  SRN ID  ' + subrack_inital_no + '     ' +bbu_type+'  ~  '+ neversion
    fo = open('LOG.txt', 'a')
    fo.write(f'{dt.datetime.now()} -- {NEname} Initial Hardware description: ' + string1 + '\n')
    #print(string1)
    return string1


def insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(SECTORID, CN, SRN, SN, ANTENNA_PORT_LIST):
    # xml.etree.ElementTree.SubElement(parent, tag, attrib={}, **extra)
    # ANTENNA_PORT_LIST=[]
    # SECTOR_xml_string = '''<SECTOR>
    #             <attributes>
    #                 <SECTORID>20</SECTORID>
    #                 <SECNAME>sector_20</SECNAME>
    #                 <LOCATIONNAME>Sector_20</LOCATIONNAME>
    #                 <USERLABEL>%omniFlag::false%</USERLABEL>
    #                 <ANTAZIMUTH>65535</ANTAZIMUTH>
    #                 <OLDSECTORID>65535</OLDSECTORID>
    #                 <SECTORIDFORCONVERSION>65535</SECTORIDFORCONVERSION>
    #                 <SECTORTYPEUMTS>255</SECTORTYPEUMTS><!--NULL-->
    #                 <RXANTNUM>255</RXANTNUM>
    #                 <DIVMODE>255</DIVMODE><!--NULL-->
    #                 <COVERTYPE>255</COVERTYPE><!--NULL-->
    #                 <RFCONNMODE>255</RFCONNMODE><!--NULL-->
    #                 <SECTORMODELTE>255</SECTORMODELTE><!--NULL-->
    #                 <ANTENNAMODE>255</ANTENNAMODE><!--NULL-->
    #                 <SECTORCOMBIND>255</SECTORCOMBIND><!--NULL-->
    #                 <OMNIFLAG>255</OMNIFLAG><!--NULL-->
    #                 <ORIENTOFMAJORAXIS>255</ORIENTOFMAJORAXIS>
    #                 <CONFIDENCE>255</CONFIDENCE>
    #                 <UNCERTSEMIMAJOR>2000000</UNCERTSEMIMAJOR>
    #                 <UNCERTSEMIMINOR>2000000</UNCERTSEMIMINOR>
    #                 <UNCERTALTITUDE>65535</UNCERTALTITUDE>
    #                 <SECTORANTENNA>
    #                     <element>
    #                         <CN>0</CN>
    #                         <SRN>102</SRN>
    #                         <SN>0</SN>
    #                         <ANTN>2</ANTN><!--R0C-->
    #                     </element>
    #                     <element>
    #                         <CN>0</CN>
    #                         <SRN>102</SRN>
    #                         <SN>0</SN>
    #                         <ANTN>0</ANTN><!--R0A-->
    #                     </element>
    #                 </SECTORANTENNA>
    #             </attributes>
    #         </SECTOR>'''

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'SECTOR':
                    sector_subelement0 = ET.SubElement(child1, 'SECTOR')
                    sector_subelement1 = ET.SubElement(sector_subelement0, 'attributes')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTORID')
                    sector_subelement2.text = str(SECTORID)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECNAME')
                    sector_subelement2.text = str('sector_') + str(SECTORID)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'LOCATIONNAME')
                    sector_subelement2.text = str('sector_') + str(SECTORID)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'USERLABEL')
                    sector_subelement2.text = str('% omniFlag::false %')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'ANTAZIMUTH')
                    sector_subelement2.text = str('65535')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'OLDSECTORID')
                    sector_subelement2.text = str('65535')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTORIDFORCONVERSION')
                    sector_subelement2.text = str('65535')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTORTYPEUMTS')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RXANTNUM')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'DIVMODE')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'COVERTYPE')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RFCONNMODE')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTORMODELTE')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'ANTENNAMODE')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTORCOMBIND')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'OMNIFLAG')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'ORIENTOFMAJORAXIS')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'CONFIDENCE')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'UNCERTSEMIMAJOR')
                    sector_subelement2.text = str('2000000')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'UNCERTSEMIMINOR')
                    sector_subelement2.text = str('2000000')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'UNCERTALTITUDE')
                    sector_subelement2.text = str('65535')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTORANTENNA')
                    for item in ANTENNA_PORT_LIST:
                        sector_subelement3 = ET.SubElement(sector_subelement2, 'element')
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'CN')
                        sector_subelement4.text = str(CN)
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'SRN')
                        sector_subelement4.text = str(SRN)
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'SN')
                        sector_subelement4.text = str(SN)
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'ANTN')
                        sector_subelement4.text = str(item)

                    # ET.dump(child1)
                    break
    #tree.write(OUTPUT,prettyprint=True)
    tree.write(OUTPUT)


def insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(SECTOREQMID, CN, SRN, SN, ANTENNA_PORT_LIST):
    # SECTOREQM_xml_string = '''<SECTOREQM>
	# 		<attributes>
	# 			<SECTOREQMID>20</SECTOREQMID>
	# 			<SECTORID>20</SECTORID>
	# 			<SECTOREQMANTENNA>
	# 				<element>
	# 					<CN>0</CN>
	# 					<SRN>102</SRN>
	# 					<SN>0</SN>
	# 					<ANTN>2</ANTN>
	# 					<ANTTYPE>3</ANTTYPE>
	# 					<TXBKPMODE>0</TXBKPMODE>
	# 				</element>
	# 				<element>
	# 					<CN>0</CN>
	# 					<SRN>102</SRN>
	# 					<SN>0</SN>
	# 					<ANTN>0</ANTN>
	# 					<ANTTYPE>3</ANTTYPE>
	# 					<TXBKPMODE>0</TXBKPMODE>
	# 				</element>
	# 			</SECTOREQMANTENNA>
	# 			<ANTCFGMODE>0</ANTCFGMODE>
	# 			<RRUCN>4294967295</RRUCN>
	# 			<RRUSRN>4294967295</RRUSRN>
	# 			<RRUSN>4294967295</RRUSN>
	# 			<BEAMSHAPE>255</BEAMSHAPE>
	# 			<BEAMLAYERSPLIT>255</BEAMLAYERSPLIT>
	# 			<BEAMAZIMUTHOFFSET>255</BEAMAZIMUTHOFFSET>
	# 		</attributes>
	# 	</SECTOREQM>'''

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'SECTOREQM':
                    sector_subelement0 = ET.SubElement(child1, 'SECTOREQM')
                    sector_subelement1 = ET.SubElement(sector_subelement0, 'attributes')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTOREQMID')
                    sector_subelement2.text = str(SECTOREQMID)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTORID')
                    sector_subelement2.text = str(SECTOREQMID)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTOREQMANTENNA')
                    for item in ANTENNA_PORT_LIST:
                        sector_subelement3 = ET.SubElement(sector_subelement2, 'element')
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'CN')
                        sector_subelement4.text = str(CN)
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'SRN')
                        sector_subelement4.text = str(SRN)
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'SN')
                        sector_subelement4.text = str(SN)
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'ANTN')
                        sector_subelement4.text = str(item)
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'ANTTYPE')
                        sector_subelement4.text = str('3')
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'TXBKPMODE')
                        sector_subelement4.text = str('0')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'ANTCFGMODE')
                    sector_subelement2.text = str('0')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RRUCN')
                    sector_subelement2.text = str('4294967295')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RRUSRN')
                    sector_subelement2.text = str('4294967295')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RRUSN')
                    sector_subelement2.text = str('4294967295')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'BEAMSHAPE')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'BEAMLAYERSPLIT')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'BEAMAZIMUTHOFFSET')
                    sector_subelement2.text = str('255')

                    # ET.dump(child1)
                    break
    tree.write(OUTPUT)


def insert_new_and_superior_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(SECTOREQMID, CN, SRN, SN, ANTENNA_PORT_LIST):
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'SECTOREQM':
                    sector_subelement0 = ET.SubElement(child1, 'SECTOREQM')
                    sector_subelement1 = ET.SubElement(sector_subelement0, 'attributes')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTOREQMID')
                    sector_subelement2.text = str(SECTOREQMID)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTORID')
                    sector_subelement2.text = str(SECTOREQMID)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTOREQMANTENNA')
                    for item in ANTENNA_PORT_LIST:
                        sector_subelement3 = ET.SubElement(sector_subelement2, 'element')
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'CN')
                        sector_subelement4.text = str(CN)
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'SRN')
                        sector_subelement4.text = str(SRN)
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'SN')
                        sector_subelement4.text = str(SN)
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'ANTN')
                        sector_subelement4.text = str(item[0])
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'ANTTYPE')
                        sector_subelement4.text = str(item[1])
                        if item[1] == 3:
                            sector_subelement4 = ET.SubElement(sector_subelement3, 'TXBKPMODE')
                            sector_subelement4.text = str('0')
                        else:
                            pass
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'ANTCFGMODE')
                    sector_subelement2.text = str('0')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RRUCN')
                    sector_subelement2.text = str('4294967295')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RRUSRN')
                    sector_subelement2.text = str('4294967295')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RRUSN')
                    sector_subelement2.text = str('4294967295')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'BEAMSHAPE')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'BEAMLAYERSPLIT')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'BEAMAZIMUTHOFFSET')
                    sector_subelement2.text = str('255')

                    # ET.dump(child1)
                    break
    tree.write(OUTPUT)


def insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(SECTOREQMID, SECTORID, CN, SRN, SN, ANTENNA_PORT_LIST):
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'SECTOREQM':
                    sector_subelement0 = ET.SubElement(child1, 'SECTOREQM')
                    sector_subelement1 = ET.SubElement(sector_subelement0, 'attributes')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTOREQMID')
                    sector_subelement2.text = str(SECTOREQMID)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTORID')
                    sector_subelement2.text = str(SECTORID)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTOREQMANTENNA')
                    for item in ANTENNA_PORT_LIST:
                        sector_subelement3 = ET.SubElement(sector_subelement2, 'element')
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'CN')
                        sector_subelement4.text = str(CN)
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'SRN')
                        sector_subelement4.text = str(SRN)
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'SN')
                        sector_subelement4.text = str(SN)
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'ANTN')
                        sector_subelement4.text = str(item)
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'ANTTYPE')
                        sector_subelement4.text = str('3')
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'TXBKPMODE')
                        sector_subelement4.text = str('0')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'ANTCFGMODE')
                    sector_subelement2.text = str('0')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RRUCN')
                    sector_subelement2.text = str('4294967295')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RRUSRN')
                    sector_subelement2.text = str('4294967295')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RRUSN')
                    sector_subelement2.text = str('4294967295')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'BEAMSHAPE')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'BEAMLAYERSPLIT')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'BEAMAZIMUTHOFFSET')
                    sector_subelement2.text = str('255')

                    # ET.dump(child1)
                    break
    tree.write(OUTPUT)


def insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(SECTOREQMID, SECTORID, CN, SRN, SN,ANTENNA_PORT_LIST):
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'SECTOREQM':
                    sector_subelement0 = ET.SubElement(child1, 'SECTOREQM')
                    sector_subelement1 = ET.SubElement(sector_subelement0, 'attributes')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTOREQMID')
                    sector_subelement2.text = str(SECTOREQMID)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTORID')
                    sector_subelement2.text = str(SECTORID)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTOREQMANTENNA')
                    for item in ANTENNA_PORT_LIST:
                        sector_subelement3 = ET.SubElement(sector_subelement2, 'element')
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'CN')
                        sector_subelement4.text = str(CN)
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'SRN')
                        sector_subelement4.text = str(SRN)
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'SN')
                        sector_subelement4.text = str(SN)
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'ANTN')
                        sector_subelement4.text = str(item[0])
                        sector_subelement4 = ET.SubElement(sector_subelement3, 'ANTTYPE')
                        sector_subelement4.text = str(item[1])
                        if item[1] == 3:
                            sector_subelement4 = ET.SubElement(sector_subelement3, 'TXBKPMODE')
                            sector_subelement4.text = str('0')
                        else:
                            pass
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'ANTCFGMODE')
                    sector_subelement2.text = str('0')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RRUCN')
                    sector_subelement2.text = str('4294967295')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RRUSRN')
                    sector_subelement2.text = str('4294967295')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RRUSN')
                    sector_subelement2.text = str('4294967295')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'BEAMSHAPE')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'BEAMLAYERSPLIT')
                    sector_subelement2.text = str('255')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'BEAMAZIMUTHOFFSET')
                    sector_subelement2.text = str('255')

                    # ET.dump(child1)
                    break
    tree.write(OUTPUT)


def insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(CN, SRN, SN, RRUCHAIN, Working_mode, RRU_type, RXNUM, TXNUM):
    almport_list = []
    # RRU_xml_string = '''<RRU>
	# 		<attributes>
	# 			<CN>0</CN>
	# 			<SRN>100</SRN>
	# 			<SN>0</SN>
	# 			<ADMSTATE>1</ADMSTATE><!--Unblocked-->
	# 			<IFFREQ>0</IFFREQ>
	# 			<ALMPROCSW>1</ALMPROCSW><!--OFF-->
	# 			<CUSTOMEDRFSPECSW>2048</CUSTOMEDRFSPECSW><!-- SPEC_1900M_45M_IBW_LTE:OFF;SPEC_2600M_EXPAND_LTE_TDD:OFF;SPEC_SUPPORT_TWO_LTE_CARRIERS:OFF;SPEC_1900M_60M_IBW_LTE:OFF;SPEC_1800M_45M_IBW_LTE:OFF;SPEC_1800M_55M_IBW_LTE:OFF;SPEC_2100M_60M_IBW_LTE:OFF;SPEC_2100M_45M_IBW_LTE:OFF;SPEC_1.8G_55M_2.1G_60M_IBW_LTE:OFF;SPEC_LOOSE_PSD_LIMIT:ON;SPEC_2.1G_55M_IBW_LTE:OFF -->
	# 			<ALMPROCTHRHLD>30</ALMPROCTHRHLD>
	# 			<ALMTHRHLD>20</ALMTHRHLD>
	# 			<PS>0</PS>
	# 			<RCN>17</RCN>
	# 			<TP>0</TP><!--TRUNK-->
	# 			<RS>136</RS><!--UMTS_LTE-->
	# 			<RXNUM>4</RXNUM>
	# 			<TXNUM>4</TXNUM>
	# 			<RFTXSIGNDETECTSW>0</RFTXSIGNDETECTSW><!--OFF-->
	# 			<RFTXSIGNDETECTPERIOD>0</RFTXSIGNDETECTPERIOD><!--DAY-->
	# 			<RFTXSIGNDETECTTHLD>1</RFTXSIGNDETECTTHLD><!--MIDDLE-->
	# 			<RN>0-100-0</RN> #RRU name
	# 			<RT>15</RT><!--MRRU-->
	# 			<RFDS>0</RFDS>
	# 			<FMBWH>0</FMBWH><!--5000-->
	# 			<LCPSW>0</LCPSW><!--Enable-->
	# 			<FLAG>0</FLAG>
	# 			<RUSPEC></RUSPEC>
	# 			<RFCONNCN2>4294967295</RFCONNCN2><!--NULL-->
	# 			<RFCONNSN2>4294967295</RFCONNSN2><!--NULL-->
	# 			<RFCONNSRN2>4294967295</RFCONNSRN2><!--NULL-->
	# 			<RFCONNTYPE>2</RFCONNTYPE><!--NULL-->
	# 			<TXDUPSW>255</TXDUPSW><!--DEFAULT-->
	# 			<DORMANCYSW>0</DORMANCYSW><!--OFF-->
	# 			<DORMANCYSTARTTIME>00:00:00</DORMANCYSTARTTIME>
	# 			<DORMANCYSTOPTIME>06:00:00</DORMANCYSTOPTIME>
	# 			<PAEFFSWITCH>0</PAEFFSWITCH><!--OFF-->
	# 			<SCR>255</SCR><!--Auto-->
	# 			<RXHWALMDETECTSW>0</RXHWALMDETECTSW><!--OFF-->
	# 			<RXFREQBANDMUTUALSW>0</RXFREQBANDMUTUALSW><!--OFF-->
	# 			<REMOTEFLAG>0</REMOTEFLAG><!--Undefined-->
	# 			<USERLABEL></USERLABEL>
	# 			<RFDCPWROFFALMDETECTSW>0</RFDCPWROFFALMDETECTSW><!--OFF-->
	# 			<BATTOUTPUNDERVOLTTHLD>430</BATTOUTPUNDERVOLTTHLD>
	# 			<MNTMODE>0</MNTMODE><!--NORMAL-->
	# 			<ST>2000-01-01T00:00:00</ST>
	# 			<ET>2037-12-31T23:59:59</ET>
	# 			<MMSETREMARK></MMSETREMARK>
	# 			<LEDSW>0</LEDSW><!--ON-->
	# 			<PSGID>0</PSGID>
	# 			<WIFISW>1</WIFISW><!--ON-->
	# 			<LOCATIONNAME></LOCATIONNAME>
	# 			<CIRCUITBREAKERVALUE>0</CIRCUITBREAKERVALUE><!--Default-->
	# 		</attributes>
	# 	</RRU>'''

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'RRU':
                    sector_subelement0 = ET.SubElement(child1, 'RRU')
                    sector_subelement1 = ET.SubElement(sector_subelement0, 'attributes')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'CN')
                    sector_subelement2.text = str(CN)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SRN')
                    sector_subelement2.text = str(SRN)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SN')
                    sector_subelement2.text = str(SN)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'ADMSTATE')  # BLOCKED~0, UNBLOCKED~1
                    sector_subelement2.text = str(1)
                    sector_subelement2 = ET.SubElement(sector_subelement1,'IFFREQ')  # interference freq 0~6553500, step: 100
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'ALMPROCSW')
                    sector_subelement2.text = str(1)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'CUSTOMEDRFSPECSW')
                    sector_subelement2.text = str(2048)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'ALMPROCTHRHLD')
                    sector_subelement2.text = str(30)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'ALMTHRHLD')
                    sector_subelement2.text = str(20)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'PS')  # rru position 0~20
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1,'RCN')  # radio chain number /given from already created RRUCHAIN
                    sector_subelement2.text = str(RRUCHAIN)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'TP')  # Topo Position 0/1 TRUNK, BRANCH
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RS')  # RFU working mode
                    # TDL~4, TL~6, LO~8, LFTD~12, WL~20, CL~40, GO~64, GT~68, GL~72, GLFTD~76, UO~128, UT~132, UL~136, ULFTD~140, CU~160, GU~192, GUT~196, GUL~200, GULFTD~204, MO~256, TM~260, LM~264, LFTDM~268, CM~288, CLM~296, GM~320, GTM~324, GLM~328, GLFTDM~332, UM~384, UTM~388, ULM~392, ULFTDM~396, GUM~448, GUTM~452, GULM~456, GULFTDM~460, RO~512, TR~516, LR~520, LFTDR~524, UR~640, UTR~644, ULR~648, ULFTDR~652, RM~768, TRM~772, LRM~776, LFTDRM~780, URM~896, UTRM~900, ULRM~904, ULFTDRM~908, NO~2048, TN~2052, TLN~2054, LN~2056, LFTDN~2060, CN~2080, CLN~2088, GN~2112, GTN~2116, GLN~2120, GLFTDN~2124, UN~2176, UTN~2180, ULN~2184, ULFTDN~2188, CUN~2208, GUN~2240, GUTN~2244, GULN~2248, GULFTDN~2252, MN~2304, TMN~2308, LMN~2312, LFTDMN~2316, CMN~2336, CLMN~2344, GMN~2368, GLMN~2376, GLFTDMN~2380, UMN~2432, UTMN~2436, ULMN~2440, GUMN~2496, GULMN~2504, RN~2560, TRN~2564, LRN~2568, LFTDRN~2572, URN~2688, UTRN~2692, ULRN~2696, ULFTDRN~2700, RMN~2816, TRMN~2820, LRMN~2824, LFTDRMN~2828, URMN~2944, UTRMN~2948, ULRMN~2952, ULFTDRMN~2956
                    sector_subelement2.text = str(Working_mode)  # UL=136
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RXNUM')
                    sector_subelement2.text = str(RXNUM)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'TXNUM')
                    sector_subelement2.text = str(TXNUM)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RFTXSIGNDETECTSW')
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RFTXSIGNDETECTPERIOD')
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RFTXSIGNDETECTTHLD')
                    sector_subelement2.text = str(1)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RN')
                    sector_subelement2.text = str(CN) + '-' + str(SRN) + '-' + str(SN)
                    sector_subelement2 = ET.SubElement(sector_subelement1,'RT')  ##RRU type: MRRU~15, MRXU~22, PRRU~271, MPRU~527, MPMU~783, MPRF~1039, AIRU~1295, MRIU~1551, GRRU~4111, LRRU~12303, ORRU~28687
                    sector_subelement2.text = str(RRU_type)  # MRRU
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RFDS')
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1,'FMBWH')  # frequency min bandwidth 5000~0, 4800~1, 4600~2, 4400~3, 4200~4, 3800~5, 4000~6
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'LCPSW')
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'FLAG')
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RUSPEC')
                    sector_subelement2.text = str('')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RFCONNCN2')
                    sector_subelement2.text = str(4294967295)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RFCONNSN2')
                    sector_subelement2.text = str(4294967295)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RFCONNSRN2')
                    sector_subelement2.text = str(4294967295)
                    sector_subelement2 = ET.SubElement(sector_subelement1,'RFCONNTYPE')  # INTRA_SYS_INTERCONN~0, OUTER_SYS_INTERCONN~1, NULL~2
                    sector_subelement2.text = str(2)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'TXDUPSW')
                    sector_subelement2.text = str(255)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'DORMANCYSW')
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'DORMANCYSTARTTIME')
                    sector_subelement2.text = str('00:00:00')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'DORMANCYSTOPTIME')
                    sector_subelement2.text = str('06:00:00')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'PAEFFSWITCH')
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'SCR')
                    sector_subelement2.text = str(255)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RXHWALMDETECTSW')
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RXFREQBANDMUTUALSW')
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'REMOTEFLAG')
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'USERLABEL')
                    sector_subelement2.text = str('')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'RFDCPWROFFALMDETECTSW')
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'BATTOUTPUNDERVOLTTHLD')
                    sector_subelement2.text = str(430)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'MNTMODE')
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'ST')
                    sector_subelement2.text = str('2000-01-01T00:00:00')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'ET')
                    sector_subelement2.text = str('2037-12-31T23:59:59')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'MMSETREMARK')
                    sector_subelement2.text = str('')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'LEDSW')
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'PSGID')
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'WIFISW')
                    sector_subelement2.text = str(1)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'LOCATIONNAME')
                    sector_subelement2.text = str('')
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'CIRCUITBREAKERVALUE')
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'CALPORTCONSTATUS')  # SRAN16 addition
                    sector_subelement2.text = str(0)
                    sector_subelement2 = ET.SubElement(sector_subelement1, 'PIM3CFGSW')  # SRAN16 addition
                    sector_subelement2.text = str(0)
                    # ET.dump(child1)
                    break

    # adding ALMPORT: MRRU -4 ports. LRRU - 2 ports
    # ALMPORT_xml_string = '''<ALMPORT>
	# 		<attributes>
	# 			<CN>0</CN>
	# 			<SRN>102</SRN>
	# 			<SN>0</SN>
	# 			<PN>3</PN>
	# 			<SW>0</SW><!--OFF-->
	# 			<AID>65033</AID>
	# 			<PT>0</PT><!--Digital Port-->
	# 			<AVOL>1</AVOL><!--Low Voltage-->
	# 			<DTPRD>255</DTPRD><!--System default value-->
	# 			<USERLABEL></USERLABEL>
	# 		</attributes>
	# 	</ALMPORT>'''

    if RRU_type == 15:  # MRRU
        almport_list = [0, 1, 2, 3]
    if RRU_type == 12303:  # LRRU
        almport_list = [0, 1]

    # can complete further for more RRU types

    for almportno in almport_list:
        for child0 in root:
            for child1 in child0:
                for child2 in child1:
                    if child2.tag == '' + 'ALMPORT':
                        sector_subelement0 = ET.SubElement(child1, 'ALMPORT')
                        sector_subelement1 = ET.SubElement(sector_subelement0, 'attributes')
                        sector_subelement2 = ET.SubElement(sector_subelement1, 'CN')
                        sector_subelement2.text = str(CN)
                        sector_subelement2 = ET.SubElement(sector_subelement1, 'SRN')
                        sector_subelement2.text = str(SRN)
                        sector_subelement2 = ET.SubElement(sector_subelement1, 'SN')
                        sector_subelement2.text = str(SN)
                        sector_subelement2 = ET.SubElement(sector_subelement1, 'PN')
                        sector_subelement2.text = str(almportno)
                        sector_subelement2 = ET.SubElement(sector_subelement1, 'SW')
                        sector_subelement2.text = str(0)
                        sector_subelement2 = ET.SubElement(sector_subelement1, 'AID')
                        sector_subelement2.text = str(65033)
                        sector_subelement2 = ET.SubElement(sector_subelement1, 'PT')
                        sector_subelement2.text = str(0)
                        sector_subelement2 = ET.SubElement(sector_subelement1, 'AVOL')
                        sector_subelement2.text = str(1)
                        sector_subelement2 = ET.SubElement(sector_subelement1, 'DTPRD')
                        sector_subelement2.text = str(255)
                        sector_subelement2 = ET.SubElement(sector_subelement1, 'USERLABEL')
                        sector_subelement2.text = str('')
                        # ET.dump(child1)
                        break

    # ANTENNAPORT_xml_string = '''<ANTENNAPORT>
	# 		<attributes>
	# 			<CN>0</CN>
	# 			<SRN>101</SRN>
	# 			<SN>0</SN>
	# 			<PN>3</PN><!--R0D-->
	# 			<FEEDERLENGTH>0</FEEDERLENGTH>
	# 			<DLDELAY>100</DLDELAY>
	# 			<ULDELAY>100</ULDELAY>
	# 			<PWRSWITCH>1</PWRSWITCH><!--OFF-->
	# 			<THRESHOLDTYPE>0</THRESHOLDTYPE><!--USER_DEFINED-->
	# 			<UOTHD>40</UOTHD>
	# 			<UCTHD>60</UCTHD>
	# 			<OOTHD>185</OOTHD>
	# 			<OCTHD>155</OCTHD>
	# 			<ULTRADELAYSW>0</ULTRADELAYSW><!--OFF-->
	# 		</attributes>
	# 	</ANTENNAPORT>'''

    for antennaportnumber in range(0, max(RXNUM, TXNUM)):
        for child0 in root:
            for child1 in child0:
                for child2 in child1:
                    if child2.tag == '' + 'ANTENNAPORT':
                        antennaport_subelement0 = ET.SubElement(child1, 'ANTENNAPORT')
                        antennaport_subelement1 = ET.SubElement(antennaport_subelement0, 'attributes')
                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'CN')
                        antennaport_subelement2.text = str(CN)
                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'SRN')
                        antennaport_subelement2.text = str(SRN)
                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'SN')
                        antennaport_subelement2.text = str(SN)
                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'PN')
                        antennaport_subelement2.text = str(antennaportnumber)
                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'FEEDERLENGTH')
                        antennaport_subelement2.text = str(0)
                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'DLDELAY')
                        antennaport_subelement2.text = str(100)
                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'ULDELAY')
                        antennaport_subelement2.text = str(100)
                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'PWRSWITCH')
                        antennaport_subelement2.text = str(1)
                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'THRESHOLDTYPE')
                        antennaport_subelement2.text = str(0)
                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'UOTHD')
                        antennaport_subelement2.text = str(40)
                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'UCTHD')
                        antennaport_subelement2.text = str(60)
                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'OOTHD')
                        antennaport_subelement2.text = str(185)
                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'OCTHD')
                        antennaport_subelement2.text = str(155)
                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'ULTRADELAYSW')
                        antennaport_subelement2.text = str(0)
                        # ET.dump(child1)
                        break

    # RETPORT_xml_string = '''<RETPORT>
	# 		<attributes>
	# 			<CN>0</CN>
	# 			<SRN>62</SRN>
	# 			<SN>0</SN>
	# 			<PN>2</PN><!--RET_PORT-->
	# 			<PWRSWITCH>0</PWRSWITCH><!--ON-->
	# 			<THRESHOLDTYPE>4</THRESHOLDTYPE><!--RET_ONLY_MULTICORE-->
	# 			<UOTHD>10</UOTHD>
	# 			<UCTHD>15</UCTHD>
	# 			<OOTHD>150</OOTHD>
	# 			<OCTHD>120</OCTHD>
	# 		</attributes>
	# 	</RETPORT>'''

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'RETPORT':
                    retport_subelement0 = ET.SubElement(child1, 'RETPORT')
                    retport_subelement1 = ET.SubElement(retport_subelement0, 'attributes')
                    retport_subelement2 = ET.SubElement(retport_subelement1, 'CN')
                    retport_subelement2.text = str(CN)
                    retport_subelement2 = ET.SubElement(retport_subelement1, 'SRN')
                    retport_subelement2.text = str(SRN)
                    retport_subelement2 = ET.SubElement(retport_subelement1, 'SN')
                    retport_subelement2.text = str(SN)
                    retport_subelement2 = ET.SubElement(retport_subelement1, 'PN')
                    retport_subelement2.text = str(2)  # default RETPORT enum value
                    retport_subelement2 = ET.SubElement(retport_subelement1, 'PWRSWITCH')
                    retport_subelement2.text = str(0)  # ON
                    retport_subelement2 = ET.SubElement(retport_subelement1, 'THRESHOLDTYPE')
                    retport_subelement2.text = str(4)  # RET_ONLY_MULTICORE
                    retport_subelement2 = ET.SubElement(retport_subelement1, 'UOTHD')
                    retport_subelement2.text = str(10)
                    retport_subelement2 = ET.SubElement(retport_subelement1, 'UCTHD')
                    retport_subelement2.text = str(15)
                    retport_subelement2 = ET.SubElement(retport_subelement1, 'OOTHD')
                    retport_subelement2.text = str(150)
                    retport_subelement2 = ET.SubElement(retport_subelement1, 'OCTHD')
                    retport_subelement2.text = str(120)
                    # ET.dump(child1)
                    break

    # Adding TXBRANCH
    # TXBRANCH_xml_string = '''<TXBRANCH>
	# 		<attributes>
	# 			<CN>0</CN>
	# 			<SRN>102</SRN>
	# 			<SN>0</SN>
	# 			<TXNO>3</TXNO>
	# 			<TXSW>0</TXSW><!--ON-->
	# 		</attributes>
	# 	</TXBRANCH>'''

    for txbranchno in range(0, TXNUM):
        for child0 in root:
            for child1 in child0:
                for child2 in child1:
                    if child2.tag == '' + 'TXBRANCH':
                        txbranch_subelement0 = ET.SubElement(child1, 'TXBRANCH')
                        txbranch_subelement1 = ET.SubElement(txbranch_subelement0, 'attributes')
                        txbranch_subelement2 = ET.SubElement(txbranch_subelement1, 'CN')
                        txbranch_subelement2.text = str(CN)
                        txbranch_subelement2 = ET.SubElement(txbranch_subelement1, 'SRN')
                        txbranch_subelement2.text = str(SRN)
                        txbranch_subelement2 = ET.SubElement(txbranch_subelement1, 'SN')
                        txbranch_subelement2.text = str(SN)
                        txbranch_subelement2 = ET.SubElement(txbranch_subelement1, 'TXNO')
                        txbranch_subelement2.text = str(txbranchno)
                        txbranch_subelement2 = ET.SubElement(txbranch_subelement1, 'TXSW')
                        txbranch_subelement2.text = str(0)  # ON
                        # ET.dump(child1)
                        break

    # Adding RXBRANCH
    # RXBRANCH_xml_string = '''<RXBRANCH>
	# 		<attributes>
	# 			<CN>0</CN>
	# 			<SRN>102</SRN>
	# 			<SN>0</SN>
	# 			<RXNO>1</RXNO>
	# 			<RXSW>0</RXSW><!--ON-->
	# 			<ATTEN>0</ATTEN>
	# 			<RTWPINITADJ0>0</RTWPINITADJ0>
	# 			<RTWPINITADJ1>0</RTWPINITADJ1>
	# 			<RTWPINITADJ2>0</RTWPINITADJ2>
	# 			<RTWPINITADJ3>0</RTWPINITADJ3>
	# 			<RTWPINITADJ4>0</RTWPINITADJ4>
	# 			<RTWPINITADJ5>0</RTWPINITADJ5>
	# 			<RTWPINITADJ6>0</RTWPINITADJ6>
	# 			<RTWPINITADJ7>0</RTWPINITADJ7>
	# 		</attributes>
	# 	</RXBRANCH>'''
    for rxbranchno in range(0, RXNUM):
        for child0 in root:
            for child1 in child0:
                for child2 in child1:
                    if child2.tag == '' + 'RXBRANCH':
                        rxbranch_subelement0 = ET.SubElement(child1, 'RXBRANCH')
                        rxbranch_subelement1 = ET.SubElement(rxbranch_subelement0, 'attributes')
                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'CN')
                        rxbranch_subelement2.text = str(CN)
                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'SRN')
                        rxbranch_subelement2.text = str(SRN)
                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'SN')
                        rxbranch_subelement2.text = str(SN)
                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RXNO')
                        rxbranch_subelement2.text = str(rxbranchno)
                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RXSW')
                        rxbranch_subelement2.text = str(0)  # ON
                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'ATTEN')
                        rxbranch_subelement2.text = str(0)
                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RTWPINITADJ0')
                        rxbranch_subelement2.text = str(0)
                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RTWPINITADJ1')
                        rxbranch_subelement2.text = str(0)
                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RTWPINITADJ2')
                        rxbranch_subelement2.text = str(0)
                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RTWPINITADJ3')
                        rxbranch_subelement2.text = str(0)
                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RTWPINITADJ4')
                        rxbranch_subelement2.text = str(0)
                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RTWPINITADJ5')
                        rxbranch_subelement2.text = str(0)
                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RTWPINITADJ6')
                        rxbranch_subelement2.text = str(0)
                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RTWPINITADJ7')
                        rxbranch_subelement2.text = str(0)
                        # ET.dump(child1)
                        break

    # add CPRIPORT for RRU
    # CPRIPORT_xml_string = '''<CPRIPORT>
	# 		<attributes>
	# 			<CN>0</CN>
	# 			<SRN>101</SRN>
	# 			<SN>0</SN>
	# 			<OPTN>0</OPTN>
	# 			<ADMINISTRATIVESTATE>1</ADMINISTRATIVESTATE>
	# 			<PT>8</PT>
	# 			<SPN>255</SPN>
	# 			<SFPSW>1</SFPSW>
	# 		</attributes>
	# 	</CPRIPORT>'''

    for cpriport in [0, 1]:
        for child0 in root:
            for child1 in child0:
                for child2 in child1:
                    if child2.tag == '' + 'CPRIPORT':
                        CPRIPORT_subelement0 = ET.SubElement(child1, 'CPRIPORT')
                        CPRIPORT_subelement1 = ET.SubElement(CPRIPORT_subelement0, 'attributes')
                        CPRIPORT_subelement2 = ET.SubElement(CPRIPORT_subelement1, 'CN')
                        CPRIPORT_subelement2.text = str(CN)
                        CPRIPORT_subelement2 = ET.SubElement(CPRIPORT_subelement1, 'SRN')
                        CPRIPORT_subelement2.text = str(SRN)
                        CPRIPORT_subelement2 = ET.SubElement(CPRIPORT_subelement1, 'SN')
                        CPRIPORT_subelement2.text = str(SN)
                        CPRIPORT_subelement2 = ET.SubElement(CPRIPORT_subelement1, 'OPTN')
                        CPRIPORT_subelement2.text = str(cpriport)
                        CPRIPORT_subelement2 = ET.SubElement(CPRIPORT_subelement1, 'ADMINISTRATIVESTATE')
                        CPRIPORT_subelement2.text = str(1)
                        CPRIPORT_subelement2 = ET.SubElement(CPRIPORT_subelement1, 'PT')
                        CPRIPORT_subelement2.text = str(8)  # HEI~5, CPRI~8, CPRI_E/CPRI_O~9, XCI~10
                        CPRIPORT_subelement2 = ET.SubElement(CPRIPORT_subelement1, 'SPN')
                        CPRIPORT_subelement2.text = str(255) #error previously: it was 0, changed on 6.11.2020
                        CPRIPORT_subelement2 = ET.SubElement(CPRIPORT_subelement1, 'SFPSW')
                        CPRIPORT_subelement2.text = str(1)
                        # ET.dump(child1)
                        break

    # add SFP for RRU
    # SFP_xml_string = '''<SFP>
	# 		<attributes>
	# 			<CN>0</CN>
	# 			<SRN>100</SRN>
	# 			<SN>0</SN>
	# 			<MODULEID>1</MODULEID>
	# 			<PT>8</PT>
	# 		</attributes>
	# 	</SFP>'''
    for sfp in [0, 1]:
        for child0 in root:
            for child1 in child0:
                for child2 in child1:
                    if child2.tag == '' + 'SFP':
                        SFP_subelement0 = ET.SubElement(child1, 'SFP')
                        SFP_subelement1 = ET.SubElement(SFP_subelement0, 'attributes')
                        SFP_subelement2 = ET.SubElement(SFP_subelement1, 'CN')
                        SFP_subelement2.text = str(CN)
                        SFP_subelement2 = ET.SubElement(SFP_subelement1, 'SRN')
                        SFP_subelement2.text = str(SRN)
                        SFP_subelement2 = ET.SubElement(SFP_subelement1, 'SN')
                        SFP_subelement2.text = str(SN)
                        SFP_subelement2 = ET.SubElement(SFP_subelement1, 'MODULEID')
                        SFP_subelement2.text = str(sfp)
                        SFP_subelement2 = ET.SubElement(SFP_subelement1, 'PT')
                        SFP_subelement2.text = str(8)
                        # ET.dump(child1)
                        break

    tree.write(OUTPUT)


def insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(rruchainno, HCN, HSRN, HSN, HPN):
    # RRUCHAIN_xml_string = '''<RRUCHAIN>
	# 		<attributes>
	# 			<RCN>2</RCN>
	# 			<TT>0</TT>
	# 			<BM>0</BM>
	# 			<HCN>0</HCN>
	# 			<HSRN>1</HSRN>
	# 			<HSN>4</HSN>
	# 			<HPN>2</HPN>
	# 			<BRKPOS1>255</BRKPOS1>
	# 			<BRKPOS2>255</BRKPOS2>
	# 			<AT>0</AT>
	# 			<CR>255</CR>
	# 			<LSN>255</LSN>
	# 			<PROTOCOL>254</PROTOCOL>
	# 			<SBT>6</SBT>
	# 			<CONNPORTNUM>0</CONNPORTNUM>
	# 			<USERDEFRATENEGOSW>0</USERDEFRATENEGOSW>
	# 			<BITRATESET>127</BITRATESET>
	# 			<RATECHGPERIOD>1</RATECHGPERIOD>
	# 			<RESVBW>255</RESVBW>
	# 			<LCN>255</LCN>
	# 			<LSRN>255</LSRN>
	# 			<HSPN>0</HSPN>
	# 		</attributes>
	# 	</RRUCHAIN>'''

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'RRUCHAIN':
                    rruchain_subelement0 = ET.SubElement(child1, 'RRUCHAIN')
                    rruchain_subelement1 = ET.SubElement(rruchain_subelement0, 'attributes')
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'RCN')
                    rruchain_subelement2.text = str(rruchainno)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1,'TT')  # topo type CHAIN~0, RING~1, LOADBALANCE~2, TRUNK_CHAIN~3
                    rruchain_subelement2.text = str(0)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'BM')  # backup mode /default~0
                    rruchain_subelement2.text = str(0)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'HCN')
                    rruchain_subelement2.text = str(HCN)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'HSRN')
                    rruchain_subelement2.text = str(HSRN)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'HSN')
                    rruchain_subelement2.text = str(HSN)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'HPN')
                    rruchain_subelement2.text = str(HPN)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'BRKPOS1')
                    rruchain_subelement2.text = str(255)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'BRKPOS2')
                    rruchain_subelement2.text = str(255)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'AT')
                    rruchain_subelement2.text = str(0)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'CR')
                    rruchain_subelement2.text = str(255)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'LSN')
                    rruchain_subelement2.text = str(255)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'PROTOCOL')
                    rruchain_subelement2.text = str(254)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'SBT')
                    rruchain_subelement2.text = str(6)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'CONNPORTNUM')
                    rruchain_subelement2.text = str(0)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'USERDEFRATENEGOSW')
                    rruchain_subelement2.text = str(0)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'BITRATESET')
                    rruchain_subelement2.text = str(127)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'RATECHGPERIOD')
                    rruchain_subelement2.text = str(1)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'RESVBW')
                    rruchain_subelement2.text = str(255)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'LCN')
                    rruchain_subelement2.text = str(255)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'LSRN')
                    rruchain_subelement2.text = str(255)
                    rruchain_subelement2 = ET.SubElement(rruchain_subelement1, 'HSPN')
                    rruchain_subelement2.text = str(0)
                    # ET.dump(child1)
                    break
    tree.write(OUTPUT)


def insert_only_new_BBP_CN_SRN_SN_TYPE_BBWS(CN, SRN, SN, TYPE, BBWS):
    # to be used  if needed to add  BBP board only, without SFP,CPRIPORT, CASCADEPORT, see 700(90,91,92) int case
    # BBP_xml_string = '''<BBP>
	# 		<attributes>
	# 			<CN>0</CN>
	# 			<SRN>1</SRN>
	# 			<SN>4</SN>
	# 			<TYPE>2</TYPE><!--UBBP-->
	# 			<OVERLOADALMRPTTHLD>90</OVERLOADALMRPTTHLD>
	# 			<OVERLOADALMCLRTHLD>85</OVERLOADALMCLRTHLD>
	# 			<WM>14</WM><!--Normal-->
	# 			<ADMSTATE>1</ADMSTATE><!--Unblocked-->
	# 			<TIME>0</TIME>
	# 			<BLKTP>0</BLKTP><!--Immediate-->
	# 			<HCE>255</HCE><!--FULL-->
	# 			<CPRIEX>0</CPRIEX><!--OFF-->
	# 			<BRDSPEC></BRDSPEC>
	# 			<CCNE>1</CCNE><!--ON-->
	# 			<BBWS>20</BBWS><!-- GSM:NO;UMTS:NO;LTE FDD:YES;LTE TDD:NO;NB-IoT:YES;NR:NO -->
	# 			<SRT>0</SRT><!--Default-->
	# 			<CPRIEXMODE>0</CPRIEXMODE><!--FRONT PANEL EXTENSION-->
	# 			<CPRIITFTYPE>0</CPRIITFTYPE><!--CPRI_SFP-->
	# 			<CELLDEPLOY>2</CELLDEPLOY><!--PERFORMANCE-->
	# 			<EXFUNCSW>0</EXFUNCSW><!-- COORD_OUTPUT_BACKHAUL_PT:OFF -->
	# 			<LTEFLEXSPECSW>0</LTEFLEXSPECSW><!--OFF-->
	# 		</attributes>
	# 	</BBP>'''

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'BBP':
                    bbp_subelement0 = ET.SubElement(child1, 'BBP')
                    bbp_subelement1 = ET.SubElement(bbp_subelement0, 'attributes')
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'CN')
                    bbp_subelement2.text = str(CN)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'SRN')
                    bbp_subelement2.text = str(SRN)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'SN')
                    bbp_subelement2.text = str(SN)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1,'TYPE')  # UBBP~2, UBBP-W~181, GBBP~4098, WBBP~8194, LBBP~12290, LPMP~12546, LCOP~12802
                    bbp_subelement2.text = str(TYPE)  #
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'OVERLOADALMRPTTHLD')
                    bbp_subelement2.text = str(90)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'OVERLOADALMCLRTHLD')
                    bbp_subelement2.text = str(85)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'ADMSTATE')
                    bbp_subelement2.text = str(1)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'TIME')
                    bbp_subelement2.text = str(0)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'BLKTP')
                    bbp_subelement2.text = str(0)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'HCE')
                    bbp_subelement2.text = str(255)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'CPRIEX')
                    bbp_subelement2.text = str(0)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'BRDSPEC')
                    bbp_subelement2.text = str('')
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'CCNE')
                    bbp_subelement2.text = str(1)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1,'BBWS')  ##<BBWS>20</BBWS><!-- GSM:NO;UMTS:NO;LTE FDD:YES;LTE TDD:NO;NB-IoT:YES;NR:NO --> ltefdd=4/ltefdd+iot=20
                    bbp_subelement2.text = str(BBWS)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'SRT')
                    bbp_subelement2.text = str(0)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'CPRIEXMODE')
                    bbp_subelement2.text = str(0)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'CPRIITFTYPE')
                    bbp_subelement2.text = str(0)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'CELLDEPLOY') #SRAN17.1 addition into BBP
                    bbp_subelement2.text = str(2)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'EXFUNCSW') #SRAN17.1 addition into BBP
                    bbp_subelement2.text = str(0)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'LTEFLEXSPECSW')
                    bbp_subelement2.text = str(0)
                    # ET.dump(child1)
                    break
    tree.write(OUTPUT)


def insert_new_BBP_CN_SRN_SN_TYPE_BBWS(CN, SRN, SN, TYPE, BBWS):
    # BBP_xml_string = '''<BBP>
	# 		<attributes>
	# 			<CN>0</CN>
	# 			<SRN>1</SRN>
	# 			<SN>3</SN>
	# 			<TYPE>2</TYPE><!--UBBP-->
	# 			<OVERLOADALMRPTTHLD>90</OVERLOADALMRPTTHLD>
	# 			<OVERLOADALMCLRTHLD>85</OVERLOADALMCLRTHLD>
	# 			<WM>14</WM><!--Normal-->
	# 			<ADMSTATE>1</ADMSTATE><!--Unblocked-->
	# 			<TIME>0</TIME>
	# 			<BLKTP>0</BLKTP><!--Immediate-->
	# 			<HCE>255</HCE><!--FULL-->
	# 			<CPRIEX>0</CPRIEX><!--OFF-->
	# 			<BRDSPEC></BRDSPEC>
	# 			<CCNE>1</CCNE><!--ON-->
	# 			<BBWS>20</BBWS><!-- GSM:NO;UMTS:NO;LTE FDD:YES;LTE TDD:NO;NB-IoT:YES;NR:NO -->
	# 			<SRT>0</SRT><!--Default-->
	# 			<CPRIEXMODE>0</CPRIEXMODE><!--FRONT PANEL EXTENSION-->
	# 			<CPRIITFTYPE>0</CPRIITFTYPE><!--CPRI_SFP-->
	# 			<CELLDEPLOY>2</CELLDEPLOY><!--PERFORMANCE-->
	# 			<EXFUNCSW>0</EXFUNCSW><!-- COORD_OUTPUT_BACKHAUL_PT:OFF -->
	# 			<LTEFLEXSPECSW>0</LTEFLEXSPECSW><!--OFF-->
	# 		</attributes>
	# 	</BBP>'''

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'BBP':
                    bbp_subelement0 = ET.SubElement(child1, 'BBP')
                    bbp_subelement1 = ET.SubElement(bbp_subelement0, 'attributes')
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'CN')
                    bbp_subelement2.text = str(CN)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'SRN')
                    bbp_subelement2.text = str(SRN)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'SN')
                    bbp_subelement2.text = str(SN)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1,'TYPE')  # UBBP~2, UBBP-W~181, GBBP~4098, WBBP~8194, LBBP~12290, LPMP~12546, LCOP~12802
                    bbp_subelement2.text = str(TYPE)  #
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'OVERLOADALMRPTTHLD')
                    bbp_subelement2.text = str(90)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'OVERLOADALMCLRTHLD')
                    bbp_subelement2.text = str(85)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'ADMSTATE')
                    bbp_subelement2.text = str(1)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'TIME')
                    bbp_subelement2.text = str(0)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'BLKTP')
                    bbp_subelement2.text = str(0)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'HCE')
                    bbp_subelement2.text = str(255)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'CPRIEX')
                    bbp_subelement2.text = str(0)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'BRDSPEC')
                    bbp_subelement2.text = str('')
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'CCNE')
                    bbp_subelement2.text = str(1)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1,'BBWS')  ##<BBWS>20</BBWS><!-- GSM:NO;UMTS:NO;LTE FDD:YES;LTE TDD:NO;NB-IoT:YES;NR:NO --> ltefdd=4/ltefdd+iot=20
                    bbp_subelement2.text = str(BBWS)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'SRT')
                    bbp_subelement2.text = str(0)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'CPRIEXMODE')
                    bbp_subelement2.text = str(0)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'CPRIITFTYPE')
                    bbp_subelement2.text = str(0)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'CELLDEPLOY')  # SRAN17.1 addition into BBP
                    bbp_subelement2.text = str(2)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'EXFUNCSW')  # SRAN17.1 addition into BBP
                    bbp_subelement2.text = str(0)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'LTEFLEXSPECSW')
                    bbp_subelement2.text = str(0)
                    bbp_subelement2 = ET.SubElement(bbp_subelement1, 'WM')  # SRAN16 addition
                    bbp_subelement2.text = str(14)
                    # ET.dump(child1)
                    break

    # add CPRIPORT for BBP board
    # CPRIPORT_xml_string = '''<CPRIPORT>
	# 		<attributes>
	# 			<CN>0</CN>
	# 			<SRN>1</SRN>
	# 			<SN>3</SN>
	# 			<OPTN>1</OPTN>
	# 			<ADMINISTRATIVESTATE>1</ADMINISTRATIVESTATE>
	# 			<PT>8</PT>
	# 			<SPN>0</SPN>
	# 			<SFPSW>1</SFPSW>
	# 		</attributes>
	# 	</CPRIPORT>'''

    for cpriport in [0, 1, 2, 3, 4, 5]:
        for child0 in root:
            for child1 in child0:
                for child2 in child1:
                    if child2.tag == '' + 'CPRIPORT':
                        CPRIPORT_subelement0 = ET.SubElement(child1, 'CPRIPORT')
                        CPRIPORT_subelement1 = ET.SubElement(CPRIPORT_subelement0, 'attributes')
                        CPRIPORT_subelement2 = ET.SubElement(CPRIPORT_subelement1, 'CN')
                        CPRIPORT_subelement2.text = str(CN)
                        CPRIPORT_subelement2 = ET.SubElement(CPRIPORT_subelement1, 'SRN')
                        CPRIPORT_subelement2.text = str(SRN)
                        CPRIPORT_subelement2 = ET.SubElement(CPRIPORT_subelement1, 'SN')
                        CPRIPORT_subelement2.text = str(SN)
                        CPRIPORT_subelement2 = ET.SubElement(CPRIPORT_subelement1, 'OPTN')
                        CPRIPORT_subelement2.text = str(cpriport)
                        CPRIPORT_subelement2 = ET.SubElement(CPRIPORT_subelement1, 'ADMINISTRATIVESTATE')
                        CPRIPORT_subelement2.text = str(1)
                        CPRIPORT_subelement2 = ET.SubElement(CPRIPORT_subelement1, 'PT')
                        CPRIPORT_subelement2.text = str(8)  # HEI~5, CPRI~8, CPRI_E/CPRI_O~9, XCI~10
                        CPRIPORT_subelement2 = ET.SubElement(CPRIPORT_subelement1, 'SPN')
                        CPRIPORT_subelement2.text = str(0)
                        CPRIPORT_subelement2 = ET.SubElement(CPRIPORT_subelement1, 'SFPSW')
                        CPRIPORT_subelement2.text = str(1)
                        # ET.dump(child1)
                        break

    # add SFP for BBP board
    # SFP_xml_string = '''<SFP>
	# 		<attributes>
	# 			<CN>0</CN>
	# 			<SRN>1</SRN>
	# 			<SN>4</SN>
	# 			<MODULEID>5</MODULEID>
	# 			<PT>8</PT>
	# 		</attributes>
	# 	</SFP>'''
    for sfp in [0, 1, 2, 3, 4, 5]:
        for child0 in root:
            for child1 in child0:
                for child2 in child1:
                    if child2.tag == '' + 'SFP':
                        SFP_subelement0 = ET.SubElement(child1, 'SFP')
                        SFP_subelement1 = ET.SubElement(SFP_subelement0, 'attributes')
                        SFP_subelement2 = ET.SubElement(SFP_subelement1, 'CN')
                        SFP_subelement2.text = str(CN)
                        SFP_subelement2 = ET.SubElement(SFP_subelement1, 'SRN')
                        SFP_subelement2.text = str(SRN)
                        SFP_subelement2 = ET.SubElement(SFP_subelement1, 'SN')
                        SFP_subelement2.text = str(SN)
                        SFP_subelement2 = ET.SubElement(SFP_subelement1, 'MODULEID')
                        SFP_subelement2.text = str(sfp)
                        SFP_subelement2 = ET.SubElement(SFP_subelement1, 'PT')
                        SFP_subelement2.text = str(8)
                        # ET.dump(child1)
                        break

   #  CASCADE_xml_string = '''<CASCADEPORT>
   #  <attributes>
   #   <CN>0</CN>
   #   <SRN>1</SRN>
   #   <SN>2</SN>
   #   <PN>6</PN>
   #   <PT>1</PT>
   #   <SW>0</SW>
   #   <PM>3</PM>
   #  </attributes>
   # </CASCADEPORT>'''
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'CASCADEPORT':
                    CASCADEPORT_subelement0 = ET.SubElement(child1, 'CASCADEPORT')
                    CASCADEPORT_subelement1 = ET.SubElement(CASCADEPORT_subelement0, 'attributes')
                    CASCADEPORT_subelement2 = ET.SubElement(CASCADEPORT_subelement1, 'CN')
                    CASCADEPORT_subelement2.text = str(CN)
                    CASCADEPORT_subelement2 = ET.SubElement(CASCADEPORT_subelement1, 'SRN')
                    CASCADEPORT_subelement2.text = str(SRN)
                    CASCADEPORT_subelement2 = ET.SubElement(CASCADEPORT_subelement1, 'SN')
                    CASCADEPORT_subelement2.text = str(SN)
                    CASCADEPORT_subelement2 = ET.SubElement(CASCADEPORT_subelement1, 'PN')
                    CASCADEPORT_subelement2.text = str(6)
                    CASCADEPORT_subelement2 = ET.SubElement(CASCADEPORT_subelement1, 'PT')
                    CASCADEPORT_subelement2.text = str(1)
                    CASCADEPORT_subelement2 = ET.SubElement(CASCADEPORT_subelement1, 'SW')
                    CASCADEPORT_subelement2.text = str(0)
                    CASCADEPORT_subelement2 = ET.SubElement(CASCADEPORT_subelement1, 'PM')
                    CASCADEPORT_subelement2.text = str(3)
                    # ET.dump(child1)
                    break
    tree.write(OUTPUT)


def subrack_type_change_3900to5900():
    # modify header
    for child0 in root:
        if child0.tag == '{1.0.0}fileHeader':
            # print(child0.attrib)
            header_attributes = child0.attrib
            nenrmversion = header_attributes['nenrmversion']
            nenrmversion_5900 = '5900' + str(nenrmversion[4::])
            neversion = header_attributes['neversion']
            neversion_5900 = 'BTS5900' + str(neversion[7::])
            producttype = header_attributes['producttype']
            producttype_5900 = str(8)
    for child0 in root:
        if child0.tag == '{1.0.0}fileHeader':
            child0.attrib['nenrmversion'] = nenrmversion_5900
            child0.attrib['neversion'] = neversion_5900
            child0.attrib['producttype'] = producttype_5900

    # add spec_compatibleNrmVersionList as a child of root
    spec_compatibleNrmVersionList = ET.SubElement(root, 'spec:compatibleNrmVersionList')
    spec_compatibleNrmVersion = ET.SubElement(spec_compatibleNrmVersionList, 'spec:compatibleNrmVersion')
    spec_compatibleNrmVersion.text = nenrmversion_5900

    # modify spec:syndata
    for child0 in root:
        if child0.tag == '{1.0.0}syndata':
            if 'HERTBBU' in str(child0.attrib['nermversion']):
                child0.attrib['nermversion'] = "5900" + child0.attrib['nermversion'][4::]  # "5900HERTBBUV500R009C10SPC210"
            else:
                child0.attrib['nermversion'] = nenrmversion_5900
            child0.attrib['productversion'] = neversion_5900

    # SUBRACK_xml_string = '''<SUBRACK>
	# 		<attributes>
	# 			<CN>0</CN>
	# 			<SRN>1</SRN>
	# 			<TYPE>166</TYPE>
	# 			<DESC />
	# 			<PHYSRN>255</PHYSRN>
	# 		</attributes>
	# 	</SUBRACK>'''

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'SUBRACK':
                    for child3 in child2:
                        for child4 in child3:
                            # print(child4.tag)
                            # BBU3900~128, BBU3910~159, BBU5900~166
                            if child4.tag == 'TYPE' and child4.text == str(159) or child4.tag == 'TYPE' and child4.text == str(128):
                                child4.text = str(166)
                                # print(child4.text)
                                # ET.dump(child1)
                                break

    # NODE_xml_string = '''<NODE>
	# 		<attributes>
	# 			<PRODUCTTYPE>8</PRODUCTTYPE>
	# 			<USERLABEL />
	# 			<NERMVERSION>5900HERTBBUV500R009C10SPC210</NERMVERSION>
	# 			<NODEID>1</NODEID>
	# 			<NODENAME>BXL1J2</NODENAME>
	# 			<WM>1</WM>
	# 			<SWVERSION>BTS3900_5900 V100R015C10SPC210</SWVERSION>
	# 			<HOTPATCHVERSION>BTS3900_5900 V100R015C10SPH213</HOTPATCHVERSION>
	# 			<PRODUCTVERSION>BTS5900 V100R015C10SPC210</PRODUCTVERSION>
	# 			<INTERFACEID>BTS5900 LTE V100R015C10SPC210</INTERFACEID>
	# 			<LMTVERSION>BTS5900 LTE V100R015C10SPC210</LMTVERSION>
	# 			<APPVERSION>V500R009C10SPC210</APPVERSION>
	# 			<APPHOTPATCHVERSION>V500R009C10SPH213</APPHOTPATCHVERSION>
	# 			<LTEMODE>0</LTEMODE>
	# 		</attributes>
	# 	</NODE>'''
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'NODE':
                    for child3 in child2:
                        for child4 in child3:
                            # print(child4.tag)
                            if child4.tag == 'PRODUCTTYPE' and child4.text == str(1):
                                child4.text = str(8)
                                # print(child4.text)
                            if child4.tag == 'PRODUCTVERSION':
                                child4.text = 'BTS5900' + child4.text[7::]
                            if child4.tag == 'INTERFACEID':
                                child4.text = 'BTS5900' + child4.text[7::]
                            if child4.tag == 'LMTVERSION':
                                child4.text = 'BTS5900' + child4.text[7::]
                            if child4.tag == 'NERMVERSION':
                                child4.text = '5900' + child4.text[4::]
                                # ET.dump(child1)
                                break

   #  NE_xml_string = '''<NE>
   #  <attributes>
   #   <LOCATION>OXLE21</LOCATION>
   #   <SWVERSION>BTS3900_5900 V100R015C10SPC210</SWVERSION>
   #   <NERMVERSION>3900LTEDATAV100R015C10SPC210</NERMVERSION>
   #   <INTERFACEID>BTS3900 LTE V100R015C10SPC210</INTERFACEID>
   #   <PRODUCTVERSION>BTS3900 V100R015C10SPC210</PRODUCTVERSION>
   #   <LMTVERSION>BTS3900 LTE V100R015C10SPC210</LMTVERSION>
   #   <DID>AUTODID_20150303_055829b0-c1c1-11e4-8000-0021287e6</DID>
   #   <SITENAME>O2873</SITENAME>
   #   <NENAME>OXLE21</NENAME>
   #   <HOTPATCHVERSION>BTS3900_5900 V100R015C10SPH213</HOTPATCHVERSION>
   #   <CLOUDBBID>0</CLOUDBBID>
   #  </attributes>
   # </NE>'''

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'NE':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'NERMVERSION':
                                child4.text = '5900' + str(child4.text[4::])
                            if child4.tag == 'INTERFACEID':
                                child4.text = 'BTS5900' + str(child4.text[7::])
                            if child4.tag == 'PRODUCTVERSION':
                                child4.text = 'BTS5900' + str(child4.text[7::])
                            if child4.tag == 'LMTVERSION':
                                child4.text = 'BTS5900' + str(child4.text[7::])
                                # ET.dump(child1)
                                break

   #  eNodeBFunction_xml_string = '''<eNodeBFunction>
   #  <attributes>
   #   <eNodeBFunctionName>OXLE21</eNodeBFunctionName>
   #   <objId>0</objId>
   #   <ApplicationRef>1</ApplicationRef>
   #   <eNodeBId>202942</eNodeBId>
   #   <NermVersion>3900LTEDATAV100R015C10SPC210</NermVersion>
   #   <ProductVersion>BTS3900 V100R015C10SPC210</ProductVersion>
   #  </attributes>
   # </eNodeBFunction>'''
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'eNodeBFunction':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'NermVersion':
                                child4.text = '5900' + child4.text[4::]
                            if child4.tag == 'ProductVersion':
                                child4.text = 'BTS5900' + child4.text[7::]
                                # ET.dump(child1)
                                break
    tree.write(OUTPUT)
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
def product_type_change_5900LTEto5900():
    """For changing product type from 5900 LTE to 5900 / for 5G DSS integration"""

    # modify header
    for child0 in root:
        if child0.tag == '{1.0.0}fileHeader':
            # print(child0.attrib)
            header_attributes = child0.attrib
            nenrmversion = header_attributes['nenrmversion']
            nenrmversion_5900_BTS = '5900BTS' + str(nenrmversion[7::])
            neversion = header_attributes['neversion']
            neversion_5900_BTS = 'BTS5900' + str(neversion[7::])
            producttype = header_attributes['producttype']
            producttype_5900 = str(125)
    for child0 in root:
        if child0.tag == '{1.0.0}fileHeader':
            child0.attrib['nenrmversion'] = nenrmversion_5900_BTS
            child0.attrib['neversion'] = neversion_5900_BTS
            child0.attrib['producttype'] = producttype_5900

    # add spec_compatibleNrmVersionList as a child of root
    spec_compatibleNrmVersionList = ET.SubElement(root, 'spec:compatibleNrmVersionList')
    spec_compatibleNrmVersion = ET.SubElement(spec_compatibleNrmVersionList, 'spec:compatibleNrmVersion')
    spec_compatibleNrmVersion.text = nenrmversion_5900_BTS

    # modify spec:syndata
    # for child0 in root:
    #     if child0.tag == '{1.0.0}syndata':
    #         if 'HERTBBU' in str(child0.attrib['nermversion']):
    #             #child0.attrib['nermversion'] = "5900HERTBBUV500R009" + neversion[16::]  # "5900HERTBBUV500R009C10SPC210"
    #             pass
    #         else:
    #             child0.attrib['nermversion'] = nenrmversion_5900_BTS
    #         child0.attrib['productversion'] = neversion_5900_BTS


    # NODE_xml_string = '''<NODE>
	# 		<attributes>
	# 			<PRODUCTTYPE>8</PRODUCTTYPE>
	# 			<USERLABEL />
	# 			<NERMVERSION>5900HERTBBUV500R009C10SPC210</NERMVERSION>
	# 			<NODEID>1</NODEID>
	# 			<NODENAME>BXL1J2</NODENAME>
	# 			<WM>1</WM>
	# 			<SWVERSION>BTS3900_5900 V100R015C10SPC210</SWVERSION>
	# 			<HOTPATCHVERSION>BTS3900_5900 V100R015C10SPH213</HOTPATCHVERSION>
	# 			<PRODUCTVERSION>BTS5900 V100R015C10SPC210</PRODUCTVERSION>
	# 			<INTERFACEID>BTS5900 LTE V100R015C10SPC210</INTERFACEID>
	# 			<LMTVERSION>BTS5900 LTE V100R015C10SPC210</LMTVERSION>
	# 			<APPVERSION>V500R009C10SPC210</APPVERSION>
	# 			<APPHOTPATCHVERSION>V500R009C10SPH213</APPHOTPATCHVERSION>
	# 			<LTEMODE>0</LTEMODE>
	# 		</attributes>
	# 	</NODE>'''

    """ <NODE> - LTEMode is for BTS3900 LTE, BTS5900 LTE; WMEXTENSION is for BTS5900; they need to have a value in either case, the other can be not specified"""
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'NODE':
                    for child3 in child2:
                        for child4 in child3:
                            # print(child4.tag)
                            if child4.tag == 'PRODUCTTYPE':
                                child4.text = str(125)
                                # print(child4.text)
                            if child4.tag == 'WM' and child4.text == str(1):
                                child4.text = str(0)
                            # if child4.tag == 'PRODUCTVERSION':
                            #     child4.text = 'BTS5900' + child4.text[7::]
                            if child4.tag == 'INTERFACEID':
                                child4.text = 'BTS5900' + child4.text[11::]
                            if child4.tag == 'LMTVERSION':
                                child4.text = 'BTS5900' + child4.text[11::]
                            # if child4.tag == 'NERMVERSION':
                            #     child4.text = '5900' + child4.text[4::]
                                # ET.dump(child1)
                        sector_subelement2 = ET.SubElement(child3, 'WMEXTENSION')
                        sector_subelement2.text = str(1)
                        break
    for child0 in root:
        if child0.tag == '{1.0.0}syndata':
            if str(child0.attrib['FunctionType']) == 'NODE':
                UDP_ping_detect0 = ET.SubElement(child0, 'class')
                UDP_ping_detect1 = ET.SubElement(UDP_ping_detect0, 'UDPPING')
                UDP_ping_detect2 = ET.SubElement(UDP_ping_detect1, 'attributes')
                UDP_ping_detect3 = ET.SubElement(UDP_ping_detect2, 'TIMEOUTTHD')
                UDP_ping_detect3.text = str(5000)
                UDP_ping_detect3 = ET.SubElement(UDP_ping_detect2, 'TIMEOUTCNT')
                UDP_ping_detect3.text = str(3)
                UDP_ping_detect3 = ET.SubElement(UDP_ping_detect2, 'DSCP')
                UDP_ping_detect3.text = str(48)


    for child0 in root:
        if child0.tag == '{1.0.0}syndata':
            if str(child0.attrib['FunctionType']) == 'NODE':
                SINGLEIPSWITCH0 = ET.SubElement(child0, 'class')
                SINGLEIPSWITCH1 = ET.SubElement(SINGLEIPSWITCH0, 'SINGLEIPSWITCH')
                SINGLEIPSWITCH2 = ET.SubElement(SINGLEIPSWITCH1, 'attributes')
                SINGLEIPSWITCH3 = ET.SubElement(SINGLEIPSWITCH2, 'SINGLEIPSW')
                SINGLEIPSWITCH3.text = str(0)

    for child0 in root:
        if child0.tag == '{1.0.0}syndata':
            if str(child0.attrib['FunctionType']) == 'NODE':
                CPSWITCH0 = ET.SubElement(child0, 'class')
                CPSWITCH1 = ET.SubElement(CPSWITCH0, 'CPSWITCH')
                CPSWITCH2 = ET.SubElement(CPSWITCH1, 'attributes')
                CPSWITCH3 = ET.SubElement(CPSWITCH2, 'ES')
                CPSWITCH3.text = str(0)


   #  NE_xml_string = '''<NE>
   #  <attributes>
   #   <LOCATION>OXLE21</LOCATION>
   #   <SWVERSION>BTS3900_5900 V100R015C10SPC210</SWVERSION>
   #   <NERMVERSION>3900LTEDATAV100R015C10SPC210</NERMVERSION>
   #   <INTERFACEID>BTS3900 LTE V100R015C10SPC210</INTERFACEID>
   #   <PRODUCTVERSION>BTS3900 V100R015C10SPC210</PRODUCTVERSION>
   #   <LMTVERSION>BTS3900 LTE V100R015C10SPC210</LMTVERSION>
   #   <DID>AUTODID_20150303_055829b0-c1c1-11e4-8000-0021287e6</DID>
   #   <SITENAME>O2873</SITENAME>
   #   <NENAME>OXLE21</NENAME>
   #   <HOTPATCHVERSION>BTS3900_5900 V100R015C10SPH213</HOTPATCHVERSION>
   #   <CLOUDBBID>0</CLOUDBBID>
   #  </attributes>
   # </NE>'''

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'NE':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'NERMVERSION':
                                child4.text = '5900BTS' + str(child4.text[7::])
                            if child4.tag == 'INTERFACEID':
                                child4.text = 'BTS5900' + str(child4.text[11::])
                            # if child4.tag == 'PRODUCTVERSION':
                            #     child4.text = 'BTS5900' + str(child4.text[7::])
                            if child4.tag == 'LMTVERSION':
                                child4.text = 'BTS5900' + str(child4.text[11::])
                                # ET.dump(child1)
                                break

   #  eNodeBFunction_xml_string = '''<eNodeBFunction>
   #  <attributes>
   #   <eNodeBFunctionName>OXLE21</eNodeBFunctionName>
   #   <objId>0</objId>
   #   <ApplicationRef>1</ApplicationRef>
   #   <eNodeBId>202942</eNodeBId>
   #   <NermVersion>3900LTEDATAV100R015C10SPC210</NermVersion>
   #   <ProductVersion>BTS3900 V100R015C10SPC210</ProductVersion>
   #  </attributes>
   # </eNodeBFunction>'''
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'eNodeBFunction':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'NermVersion':
                                child4.text = '5900' + child4.text[4::]
                            if child4.tag == 'ProductVersion':
                                child4.text = 'BTS5900' + child4.text[7::]
                                # ET.dump(child1)
                                break
    tree.write(OUTPUT)
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
def change_bbu_subrack_id():
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                for child3 in child2:
                    for child4 in child3:
                        if (
                                child4.tag == 'SRN' and child4.text == str(
                                0)) or (
                                child4.tag == 'HSRN' and child4.text == str(
                                0)):
                            child4.text = str(1)
                            # ET.dump(child1)
                            break
    tree.write(OUTPUT)


def change_basebandeqm_subrack_id():
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == 'BASEBANDEQM':
                    for child3 in child2:
                        for child4 in child3:
                            for child5 in child4:
                                for child6 in child5:
                                    if (child6.tag == 'SRN' and child6.text == str(0)):
                                        child6.text = str(1)
                                        # ET.dump(child1)

    tree.write(OUTPUT)


def change_bbu_subrack_id10():
    #print("has entered subrack change 1-0")
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                for child3 in child2:
                    for child4 in child3:
                        if (
                                child4.tag == 'SRN' and child4.text == str(
                                1)) or (
                                child4.tag == 'HSRN' and child4.text == str(
                                1)) or (
                                child4.tag == 'SSRN' and child4.text == str(
                                1)):
                            child4.text = str(0)
                            # ET.dump(child1)
                            break
    tree.write(OUTPUT)


def rmv_old_bbp_from_slotnumber(SN):
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == 'BBP':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SN' and child4.text == str(SN):
                                child1.remove(child2)
                                # ET.dump(child1)
                                break
    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'CPRIPORT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SN' and child4.text == str(SN):
                                child1.remove(child2)
                                # ET.dump(child1)

    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'SFP':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SN' and child4.text == str(SN):
                                child1.remove(child2)
                                ##ET.dump(child1)
    # need to solve multiple subracks with same SN BBP
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == 'CASCADEPORT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SN' and child4.text == str(SN):
                                child1.remove(child2)
                                # ET.dump(child1)
                                break
    tree.write(OUTPUT)


def rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(SRN, RET_id_RETDEVICEDATA):
    tma_list_toremove = []
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'TMA':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'DEVICENO' and child4.text not in tma_list_toremove:
                                tma_list_toremove.append(str(child4.text))
                                # ET.dump(child1)

    # print("TMAAAAAAAAAAAAAAAAAA",tma_list_toremove)
    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'RRU':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SRN' and child4.text == str(SRN):
                                child1.remove(child2)
                                # ET.dump(child1)
                                break
    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'ALMPORT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SRN' and child4.text == str(SRN):
                                child1.remove(child2)
                                # ET.dump(child1)

    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'ANTENNAPORT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SRN' and child4.text == str(SRN):
                                child1.remove(child2)
                                # ET.dump(child1)

    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'RETPORT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SRN' and child4.text == str(SRN):
                                child1.remove(child2)
                                # ET.dump(child1)

    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'TXBRANCH':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SRN' and child4.text == str(SRN):
                                child1.remove(child2)
                                # ET.dump(child1)

    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'RXBRANCH':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SRN' and child4.text == str(SRN):
                                child1.remove(child2)
                                # ET.dump(child1)

    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'CPRIPORT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SRN' and child4.text == str(SRN):
                                child1.remove(child2)
                                # ET.dump(child1)

    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'SFP':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SRN' and child4.text == str(SRN):
                                child1.remove(child2)
                                # ET.dump(child1)

    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'RET':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'CTRLSRN' and child4.text == str(SRN):
                                child1.remove(child2)
                                # ET.dump(child1)
    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'RETSUBUNIT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'CONNSRN1' and child4.text == str(SRN):
                                child1.remove(child2)
                                # ET.dump(child1)
    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'RETDEVICEDATA':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'DEVICENO' and child4.text == str(RET_id_RETDEVICEDATA):
                                child1.remove(child2)
                                # ET.dump(child1)
    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'TMA':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'DEVICENO' and child4.text in tma_list_toremove:
                                child1.remove(child2)
                                # ET.dump(child1)
    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'TMASUBUNIT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'DEVICENO' and child4.text in tma_list_toremove:
                                child1.remove(child2)
                                # ET.dump(child1)

    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'TMADEVICEDATA':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'DEVICENO' and child4.text in tma_list_toremove:
                                child1.remove(child2)
                                # ET.dump(child1)

    tree.write(OUTPUT)


def rmv_RETSUBUNIT_only(deviceno):
    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'RETSUBUNIT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'DEVICENO' and child4.text == str(deviceno):
                                child1.remove(child2)
                                # ET.dump(child1)
    tree.write(OUTPUT)


def rmv_RET_only(SRN): # complete RET, RETSUBUNIT, RETDEVICEDATA removal for 2100
    ret_devno_unique = []
    flag = False
    List_devno_srn = []
    for child0 in root:
        flag = False
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'RET':
                    for child3 in child2:
                        flag = False
                        list_devno_srn = []
                        for child4 in child3:
                            if child4.tag == 'DEVICENO':
                                list_devno_srn.append(str(child4.text))
                            if child4.tag == 'CTRLSRN':
                                list_devno_srn.append(str(child4.text))
                        List_devno_srn.append(list_devno_srn)
    #print('RETDEVICEDATA and RET CTRLSRN',List_devno_srn)

    Device_no_list_2100 = [List_devno_srn[i][0] for i in range(0, len(List_devno_srn)) if
                           List_devno_srn[i][1] in ['100', '101', '102','4']]
    #print(Device_no_list_2100)

    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'RET':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'CTRLSRN' and child4.text == str(SRN):
                                child1.remove(child2)
                                # ET.dump(child1)
    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'RETSUBUNIT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'CONNSRN1' and child4.text == str(SRN):
                                child1.remove(child2)
                                # ET.dump(child1)
    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'RETDEVICEDATA':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'DEVICENO' and child4.text in Device_no_list_2100:
                                child1.remove(child2)
                                # break
                                # ET.dump(child1)
    tree.write(OUTPUT)


def rmv_old_rruchain_ALL_from_slotnumber(HSN):
    for child0 in root:
        for child1 in child0:
            for child2 in child1[::1]:
                if child2.tag == 'RRUCHAIN':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'HSN' and child4.text == str(
                                    HSN):
                                child1.remove(child2)
                                # ET.dump(child1)
                                break

    tree.write(OUTPUT)


def rmv_old_rruchain_from_slotnumber_portnumber(HSN, HPN):
    for child0 in root:
        for child1 in child0:
            for child2 in child1[::1]:
                if child2.tag == 'RRUCHAIN':
                    for child3 in child2:
                        for item in range(0, len(child3)):
                            if (child3[
                                    item].tag == 'HSN' and
                                child3[item].text == str(HSN)) and (child3[item + 1].tag == 'HPN' and child3[item + 1].text == str(HPN)):
                                child1.remove(child2)
                                # ET.dump(child1)
                                break
    tree.write(OUTPUT)

def rmv_old_rruchain_bychainID(chainID):
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == 'RRUCHAIN':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'RCN' and child4.text == str(chainID):
                                child1.remove(child2)
                                # ET.dump(child1)
                                break
    tree.write(OUTPUT)

def rmv_old_SECTOR_ID(SECTORID):
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == 'SECTOR':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SECTORID' and child4.text == str(SECTORID):
                                child1.remove(child2)
                                # ET.dump(child1)
                                break
    tree.write(OUTPUT)
# def rmv_old_SECTOR_ID_specificSRN(SECTORID,SRN):
#     for child0 in root:
#         for child1 in child0:
#             for child2 in child1:
#                 if child2.tag == 'SECTOR':
#                     for child3 in child2:
#                         for child4 in child3:
#                             if child4.tag == 'SECTORID' and child4.text == str(SECTORID):
#                                 for child5 in child4:
#
#
#                                 child1.remove(child2)
#                                 # ET.dump(child1)
#                                 break
#    tree.write(OUTPUT)

def rmv_old_SECTOREQM_ID(SECTOREQMID):
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == 'SECTOREQM':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SECTOREQMID' and child4.text == str(
                                    SECTOREQMID):
                                child1.remove(child2)
                                # ET.dump(child1)
                                break
    tree.write(OUTPUT)


def rmv_old_UEIU():
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == 'UEIU':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SN' and child4.text == str(
                                    18):
                                child1.remove(child2)
                                # ET.dump(child1)
                                break
    tree.write(OUTPUT)


def insert_new_UPEU(CN, SRN, SN):
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'PEU':
                    upeu_subelement0 = ET.SubElement(child1, 'PEU')
                    upeu_subelement1 = ET.SubElement(upeu_subelement0, 'attributes')
                    upeu_subelement2 = ET.SubElement(upeu_subelement1, 'CN')
                    upeu_subelement2.text = str(CN)
                    upeu_subelement2 = ET.SubElement(upeu_subelement1, 'SRN')
                    upeu_subelement2.text = str(SRN)
                    upeu_subelement2 = ET.SubElement(upeu_subelement1, 'SN')
                    upeu_subelement2.text = str(SN)
                    # ET.dump(child1)
                    break
    tree.write(OUTPUT)


varx = 0


def mod_BBP_TYPE_from_slot_to_slot(SN1, SN2, Type1, Type2):
    varx = 0
    flag = False
    for child0 in root:
        flag = False
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'BBP':
                    for child3 in child2:
                        flag = False
                        for child4 in child3:
                            if child4.tag == 'SN' and child4.text == str(SN1):
                                flag = True
                            if flag and child4.tag == 'TYPE' and child4.text == str(12290):
                                varx = 1
    # print('varxvarxvarx',varx)
    i = 0
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'BBP':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SN' and child4.text == str(SN1):
                                child4.text = str(SN2)
                                if i < 1 and varx == 1:
                                    sector_subelement2 = ET.SubElement(child3, 'LTEFLEXSPECSW')
                                    sector_subelement2.text = str(0)
                                    sector_subelement2 = ET.SubElement(child3, 'CCNE')
                                    sector_subelement2.text = str(1)
                                    sector_subelement2 = ET.SubElement(child3, 'BBWS')
                                    sector_subelement2.text = str(4)
                                    sector_subelement2 = ET.SubElement(child3, 'SRT')
                                    sector_subelement2.text = str(0)
                                    sector_subelement2 = ET.SubElement(child3, 'CPRIITFTYPE')
                                    sector_subelement2.text = str(0)
                                    i = 1
                            if child4.tag == 'TYPE' and child4.text == str(Type1):
                                child4.text = str(Type2)
                            if child4.tag == 'WM':  # SRAN16.1 V1.1 bug fix
                                child4.text = str(14)
                                # print(child4.text)
                                # ET.dump(child1)
                                break
    tree.write(OUTPUT)


varx1 = 0


def mod_BBP_wbbptoubbp_TYPE_from_slot_to_slot(SN1, SN2, Type1, Type2):
    varx1 = 0
    flag = False
    for child0 in root:
        flag = False
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'BBP':
                    for child3 in child2:
                        flag = False
                        for child4 in child3:
                            if child4.tag == 'SN' and child4.text == str(SN1):
                                flag = True
                            if flag and child4.tag == 'TYPE' and child4.text == str(8194):
                                varx1 = 1
    # print('varxvarxvarx',varx1)
    i = 0
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'BBP':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SN' and child4.text == str(SN1):
                                child4.text = str(SN2)
                                if i < 1 and varx1 == 1:
                                    sector_subelement2 = ET.SubElement(child3, 'LTEFLEXSPECSW')
                                    sector_subelement2.text = str(0)
                                    sector_subelement2 = ET.SubElement(child3, 'CCNE')
                                    sector_subelement2.text = str(1)
                                    sector_subelement2 = ET.SubElement(child3, 'BBWS')
                                    sector_subelement2.text = str(2)
                                    sector_subelement2 = ET.SubElement(child3, 'SRT')
                                    sector_subelement2.text = str(0)
                                    sector_subelement2 = ET.SubElement(child3, 'CPRIITFTYPE')
                                    sector_subelement2.text = str(0)
                                    i = 1
                            if child4.tag == 'TYPE' and child4.text == str(Type1):
                                child4.text = str(Type2)
                                # print(child4.text)
                                # ET.dump(child1)
                                break
    tree.write(OUTPUT)


def mod_BBP_from_slot_to_slot(SN1, SN2, Type1, Type2, BBWS1, BBWS2):
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'BBP':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SN' and child4.text == str(SN1):
                                child4.text = str(SN2)
                            if child4.tag == 'TYPE' and child4.text == str(Type1):
                                child4.text = str(Type2)
                            if child4.tag == 'BBWS' and child4.text == str(BBWS1):
                                child4.text = str(BBWS2)
                                # print(child4.text)
                                # ET.dump(child1)
                                break

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'CPRIPORT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SN' and child4.text == str(SN1):
                                child4.text = str(SN2)
                                # ET.dump(child1)
                                break

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'SFP':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SN' and child4.text == str(SN1):
                                child4.text = str(SN2)
                                # ET.dump(child1)
                                break
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'CASCADEPORT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SN' and child4.text == str(SN1):
                                child4.text = str(SN2)
                                # ET.dump(child1)
                                break
    tree.write(OUTPUT)


def mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(SRN1, CN2, SRN2, SN2, ANTENNA_PORT_LIST2): # cannot modify number of antenna ports, from 2 to 4
    i = 0
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'SECTOR':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SECTORANTENNA':
                                for child5 in child4[::-1]:
                                    for child6 in child5:
                                        if child6.tag == 'SRN' and child6.text == str(SRN1):
                                            if i < 1:
                                                for item in ANTENNA_PORT_LIST2:
                                                    sectorantenna_subelement0 = ET.SubElement(child4, 'element')
                                                    sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0,'ANTN')
                                                    sectorantenna_subelement1.text = str(item)
                                                    sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0,'CN')
                                                    sectorantenna_subelement1.text = str(CN2)
                                                    sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0,'SN')
                                                    sectorantenna_subelement1.text = str(SN2)
                                                    sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0,'SRN')
                                                    sectorantenna_subelement1.text = str(SRN2)
                                                    i = 1
                                            child4.remove(child5)
                                    #ET.dump(child4)

    tree.write(OUTPUT)

def mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist_updated(SRN1, CN2, SRN2, SN2, ANTENNA_PORT_LIST2): # can modify number of antenna ports, from 2 to 4
    """NOT USED"""
    i = 0
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'SECTOR':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SECTORANTENNA':
                                # sectorantenna_subelement0 = ET.SubElement(child4, 'element')
                                # sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0, 'ANTN')
                                # sectorantenna_subelement1.text = str(2)
                                # sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0, 'CN')
                                # sectorantenna_subelement1.text = str(CN2)
                                # sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0, 'SN')
                                # sectorantenna_subelement1.text = str(SN2)
                                # sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0, 'SRN')
                                # sectorantenna_subelement1.text = str(SRN2)
                                # sectorantenna_subelement0 = ET.SubElement(child4, 'element')
                                # sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0, 'ANTN')
                                # sectorantenna_subelement1.text = str(3)
                                # sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0, 'CN')
                                # sectorantenna_subelement1.text = str(CN2)
                                # sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0, 'SN')
                                # sectorantenna_subelement1.text = str(SN2)
                                # sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0, 'SRN')
                                # sectorantenna_subelement1.text = str(SRN2)
                                for child5 in child4[::-1]:
                                    for child6 in child5:
                                        if child6.tag == 'SRN' and child6.text == str(SRN1):
                                            if i < 1:
                                                for item in ANTENNA_PORT_LIST2:
                                                    sectorantenna_subelement0 = ET.SubElement(child4, 'element')
                                                    sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0,'ANTN')
                                                    sectorantenna_subelement1.text = str(item)
                                                    sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0,'CN')
                                                    sectorantenna_subelement1.text = str(CN2)
                                                    sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0,'SN')
                                                    sectorantenna_subelement1.text = str(SN2)
                                                    sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0,'SRN')
                                                    sectorantenna_subelement1.text = str(SRN2)
                                                    i = 1
                                            child4.remove(child5)
                                    #ET.dump(child4)

    tree.write(OUTPUT)


def rmv_3G_all_ret_retsubunit_retdevicedata():
    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == 'RET':
                    child1.remove(child2)

        for child0 in root:
            for child1 in child0:
                for child2 in child1[::-1]:
                    if child2.tag == 'RETSUBUNIT':
                        child1.remove(child2)

            for child0 in root:
                for child1 in child0:
                    for child2 in child1[::-1]:
                        if child2.tag == 'RETDEVICEDATA':
                            child1.remove(child2)
                            # ET.dump(child1)
    tree.write(OUTPUT)


def mod_RRU_SECTOR_SECTOREQM_SRN(SRN1, SRN2):
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                for child3 in child2:
                    for child4 in child3:
                        if (child4.tag == 'SRN' and child4.text == str(SRN1)):
                            child4.text = str(SRN2)
                            # ET.dump(child1)
                            break
    tree.write(OUTPUT)


def mod_SECTOR_ID(Id1, Id2):
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'SECTOR':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SECTORID' and child4.text == str(Id1):
                                child4.text = str(Id2)
    tree.write(OUTPUT)


def mod_SECTOREQM_ID(Id1, Id2):
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'SECTOREQM':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SECTOREQMID' and child4.text == str(Id1):
                                child4.text = str(Id2)
    tree.write(OUTPUT)


def mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(SRN1, CN2, SRN2, SN2, ANTENNA_PORT_LIST2):
    i = 0
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'SECTOREQM':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SECTOREQMANTENNA':
                                for child5 in child4[::-1]:
                                    for child6 in child5:
                                        if child6.tag == 'SRN' and child6.text == str(SRN1):
                                            if i < 1:
                                                for item in ANTENNA_PORT_LIST2:
                                                    sectorantenna_subelement0 = ET.SubElement(child4, 'element')
                                                    sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0,
                                                                                              'ANTN')
                                                    sectorantenna_subelement1.text = str(item)
                                                    sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0,
                                                                                              'ANTTYPE')
                                                    sectorantenna_subelement1.text = str(3)
                                                    sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0,
                                                                                              'CN')
                                                    sectorantenna_subelement1.text = str(CN2)
                                                    sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0,
                                                                                              'SRN')
                                                    sectorantenna_subelement1.text = str(SRN2)
                                                    sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0,
                                                                                              'SN')
                                                    sectorantenna_subelement1.text = str(SN2)
                                                    sectorantenna_subelement1 = ET.SubElement(sectorantenna_subelement0,
                                                                                              'TXBKPMODE')
                                                    sectorantenna_subelement1.text = str(0)
                                                    i = 1
                                            child4.remove(child5)
                                    #ET.dump(child4)
    tree.write(OUTPUT)


def mod_RRUCHUAIN_chainno_oldHSN_newHSN(chainno, HSN1, HSN2):
    # for child0 in root:
    #     for child1 in child0:
    #         for child2 in child1:
    #             if child2.tag == '' + 'RRUCHAIN':
    #                 for child3 in child2:
    #                     for child4 in child3:
    #                         if child4.tag == 'HSN' and child4.text == str(HSN1):
    #                             child4.text = str(HSN2)
    #                             #ET.dump(child1)
    #                             break
    flag = False
    for child0 in root:
        flag = False
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'RRUCHAIN':
                    for child3 in child2:
                        flag = False
                        for child4 in child3:
                            if child4.tag == 'RCN' and child4.text == str(chainno):
                                child4.text = str(chainno)
                                flag = True
                            if flag and child4.tag == 'HSN' and child4.text == str(HSN1):
                                child4.text = str(HSN2)
                                # ET.dump(child1)
    tree.write(OUTPUT)


def mod_LRRU_to_MRRU_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(SRN1, CN2, SRN2, SN2, TXNUM2,RXNUM2, WM2, Type2):
    # function purpose is to be used on same RRU ,oldSRN=newSRn, just to chage RRU Type:e.g. LRRU to MRRU
    flag = False
    i = 0
    for child0 in root:
        flag = False
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'RRU':
                    for child3 in child2:

                        flag = False
                        for child4 in child3:
                            if child4.tag == 'SRN' and child4.text == str(
                                    SRN1):
                                child4.text = str(SRN2)
                                if i < 1:
                                    sector_subelement2 = ET.SubElement(child3, 'IFFREQ')
                                    sector_subelement2.text = str(0)
                                    sector_subelement2 = ET.SubElement(child3, 'LEDSW')
                                    sector_subelement2.text = str(0)
                                    sector_subelement2 = ET.SubElement(child3, 'LCPSW')
                                    sector_subelement2.text = str(0)
                                    sector_subelement2 = ET.SubElement(child3, 'DORMANCYSW')  # added with SRAN16.1
                                    sector_subelement2.text = str(0)
                                    i = 1
                                flag = True
                            if flag and child4.tag == 'RXNUM':
                                child4.text = str(RXNUM2)
                            if flag and child4.tag == 'TXNUM':
                                child4.text = str(TXNUM2)
                            if flag and child4.tag == 'RT':
                                child4.text = str(Type2)
                            if flag and child4.tag == 'RS':
                                child4.text = str(WM2)

                            # ET.dump(child1)

    # if child2.tag == '' + 'SECTOREQM':
    #     sector_subelement0 = ET.SubElement(child1, 'SECTOREQM')
    #     sector_subelement1 = ET.SubElement(sector_subelement0, 'attributes')
    #     sector_subelement2 = ET.SubElement(sector_subelement1, 'SECTOREQMID')
    #     sector_subelement2.text = str(SECTOREQMID)
    tree.write(OUTPUT)


def ULOCELLSECTOREQM_change():
    flag11 = False
    for child0 in root:
        flag11 = False
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'ULOCELLSECTOREQM':

                    for child3 in child2:
                        flag11 = False
                        for child4 in child3:
                            if child4.tag == 'ULOCELLID' and str(child4.text[-1]) == '1':
                                flag11 = True
                            if flag11 and child4.tag == 'SECTOREQMID':
                                child4.text = '0'
    flag12 = False
    for child0 in root:
        for child1 in child0:
            flag12 = False
            for child2 in child1:
                if child2.tag == '' + 'ULOCELLSECTOREQM':

                    for child3 in child2:
                        flag12 = False
                        for child4 in child3:
                            if child4.tag == 'ULOCELLID' and str(child4.text[-1]) == '2':
                                flag12 = True
                            if flag12 and child4.tag == 'SECTOREQMID':
                                child4.text = '2'
    flag13 = False
    for child0 in root:
        for child1 in child0:
            flag13 = False
            for child2 in child1:
                if child2.tag == '' + 'ULOCELLSECTOREQM':

                    for child3 in child2:
                        flag13 = False
                        for child4 in child3:
                            if child4.tag == 'ULOCELLID' and str(child4.text[-1]) == '3':
                                flag13 = True
                            if flag13 and child4.tag == 'SECTOREQMID':
                                child4.text = '4'
    flag14 = False
    for child0 in root:
        for child1 in child0:
            flag14 = False
            for child2 in child1:
                if child2.tag == '' + 'ULOCELLSECTOREQM':

                    for child3 in child2:
                        flag14 = False
                        for child4 in child3:
                            if child4.tag == 'ULOCELLID' and str(
                                    child4.text[-1]) == '5':
                                flag14 = True
                            if flag14 and child4.tag == 'SECTOREQMID':
                                child4.text = '1'

    flag16 = False
    for child0 in root:
        for child1 in child0:
            flag16 = False
            for child2 in child1:
                if child2.tag == '' + 'ULOCELLSECTOREQM':

                    for child3 in child2:
                        flag16 = False
                        for child4 in child3:
                            if child4.tag == 'ULOCELLID' and str(
                                    child4.text[-1]) == '6':
                                flag16 = True
                            if flag16 and child4.tag == 'SECTOREQMID':
                                child4.text = '3'

    flag17 = False
    for child0 in root:
        for child1 in child0:
            flag17 = False
            for child2 in child1:
                if child2.tag == '' + 'ULOCELLSECTOREQM':

                    for child3 in child2:
                        flag17 = False
                        for child4 in child3:
                            if child4.tag == 'ULOCELLID' and str(
                                    child4.text[-1]) == '7':
                                flag17 = True
                            if flag17 and child4.tag == 'SECTOREQMID':
                                child4.text = '5'

    flag18 = False
    for child0 in root:
        for child1 in child0:
            flag18 = False
            for child2 in child1:
                if child2.tag == '' + 'ULOCELLSECTOREQM':

                    for child3 in child2:
                        flag18 = False
                        for child4 in child3:
                            if child4.tag == 'ULOCELLID' and str(
                                    child4.text[-1]) == '8':
                                flag18 = True
                            if flag18 and child4.tag == 'SECTOREQMID':
                                child4.text = '1'

    flag19 = False
    for child0 in root:
        for child1 in child0:
            flag19 = False
            for child2 in child1:
                if child2.tag == '' + 'ULOCELLSECTOREQM':

                    for child3 in child2:
                        flag19 = False
                        for child4 in child3:
                            if child4.tag == 'ULOCELLID' and str(
                                    child4.text[-1]) == '9':
                                flag19 = True
                            if flag19 and child4.tag == 'SECTOREQMID':
                                child4.text = '3'

    flag20 = False
    for child0 in root:
        for child1 in child0:
            flag20 = False
            for child2 in child1:
                if child2.tag == '' + 'ULOCELLSECTOREQM':

                    for child3 in child2:
                        flag20 = False
                        for child4 in child3:
                            if child4.tag == 'ULOCELLID' and str(
                                    child4.text[-1]) == '0':
                                flag20 = True
                            if flag20 and child4.tag == 'SECTOREQMID':
                                child4.text = '5'
    tree.write(OUTPUT)

def mod_RRU_onlyRRUCHAINno(SRN,RCN1,RCN2):
    flag = False
    for child0 in root:
        flag = False
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'RRU':
                    for child3 in child2:
                        flag = False
                        for child4 in child3:
                            if child4.tag == 'SRN' and child4.text == str(SRN):
                                flag = True
                            if flag and child4.tag == 'RCN' and child4.text==str(RCN1):
                                child4.text = str(RCN2)
                                # ET.dump(child1)

def mod_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(SRN1, CN2, SRN2, SN2, TXNUM2, RXNUM2, WM2, Type2):
    flag = False
    for child0 in root:
        flag = False
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'RRU':
                    for child3 in child2:
                        flag = False
                        for child4 in child3:
                            if child4.tag == 'SRN' and child4.text == str(SRN1):
                                child4.text = str(SRN2)
                                flag = True
                            if flag and child4.tag == 'RXNUM':
                                child4.text = str(RXNUM2)
                            if flag and child4.tag == 'TXNUM':
                                child4.text = str(TXNUM2)
                            if flag and child4.tag == 'RT':
                                child4.text = str(Type2)
                            if flag and child4.tag == 'RS':
                                child4.text = str(WM2)
                                # ET.dump(child1)

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'ALMPORT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SRN' and child4.text == str(SRN1):
                                child4.text = str(SRN2)
                                # ET.dump(child1)
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'RETPORT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SRN' and child4.text == str(SRN1):
                                child4.text = str(SRN2)
                                # ET.dump(child1)
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'CPRIPORT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SRN' and child4.text == str(SRN1):
                                child4.text = str(SRN2)
                                # ET.dump(child1)
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'SFP':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SRN' and child4.text == str(SRN1):
                                child4.text = str(SRN2)
                                # ET.dump(child1)
    i = 0
    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == '' + 'TXBRANCH':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SRN' and child4.text == str(SRN1):
                                if i < 1:
                                    for item in range(0, TXNUM2):
                                        txbranch_subelement0 = ET.SubElement(child1, 'TXBRANCH')
                                        txbranch_subelement1 = ET.SubElement(txbranch_subelement0, 'attributes')
                                        txbranch_subelement2 = ET.SubElement(txbranch_subelement1, 'CN')
                                        txbranch_subelement2.text = str(CN2)
                                        txbranch_subelement2 = ET.SubElement(txbranch_subelement1, 'SRN')
                                        txbranch_subelement2.text = str(SRN2)
                                        txbranch_subelement2 = ET.SubElement(txbranch_subelement1, 'SN')
                                        txbranch_subelement2.text = str(SN2)
                                        txbranch_subelement2 = ET.SubElement(txbranch_subelement1, 'TXNO')
                                        txbranch_subelement2.text = str(item)
                                        txbranch_subelement2 = ET.SubElement(txbranch_subelement1, 'TXSW')
                                        txbranch_subelement2.text = str(0)  # ON
                                        i = 1
                                child1.remove(child2)
                                # ET.dump(child1)
    i = 0
    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == '' + 'RXBRANCH':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SRN' and child4.text == str(
                                    SRN1):
                                if i < 1:
                                    for item in range(0, RXNUM2):
                                        rxbranch_subelement0 = ET.SubElement(child1, 'RXBRANCH')
                                        rxbranch_subelement1 = ET.SubElement(rxbranch_subelement0, 'attributes')
                                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'CN')
                                        rxbranch_subelement2.text = str(CN2)
                                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'SRN')
                                        rxbranch_subelement2.text = str(SRN2)
                                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'SN')
                                        rxbranch_subelement2.text = str(SN2)
                                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RXNO')
                                        rxbranch_subelement2.text = str(item)
                                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RXSW')
                                        rxbranch_subelement2.text = str(0)  # ON
                                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'ATTEN')
                                        rxbranch_subelement2.text = str(0)
                                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RTWPINITADJ0')
                                        rxbranch_subelement2.text = str(0)
                                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RTWPINITADJ1')
                                        rxbranch_subelement2.text = str(0)
                                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RTWPINITADJ2')
                                        rxbranch_subelement2.text = str(0)
                                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RTWPINITADJ3')
                                        rxbranch_subelement2.text = str(0)
                                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RTWPINITADJ4')
                                        rxbranch_subelement2.text = str(0)
                                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RTWPINITADJ5')
                                        rxbranch_subelement2.text = str(0)
                                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RTWPINITADJ6')
                                        rxbranch_subelement2.text = str(0)
                                        rxbranch_subelement2 = ET.SubElement(rxbranch_subelement1, 'RTWPINITADJ7')
                                        rxbranch_subelement2.text = str(0)
                                        i = 1
                                child1.remove(child2)
                                # ET.dump(child1)

    i = 0
    for child0 in root:
        for child1 in child0:
            for child2 in child1[::-1]:
                if child2.tag == '' + 'ANTENNAPORT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SRN' and child4.text == str(SRN1):
                                if i < 1:
                                    for item in range(0, max(TXNUM2, RXNUM2)):
                                        antennaport_subelement0 = ET.SubElement(child1, 'ANTENNAPORT')
                                        antennaport_subelement1 = ET.SubElement(antennaport_subelement0, 'attributes')
                                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'CN')
                                        antennaport_subelement2.text = str(CN2)
                                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'SRN')
                                        antennaport_subelement2.text = str(SRN2)
                                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'SN')
                                        antennaport_subelement2.text = str(SN2)
                                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'PN')
                                        antennaport_subelement2.text = str(item)
                                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'FEEDERLENGTH')
                                        antennaport_subelement2.text = str(0)
                                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'DLDELAY')
                                        antennaport_subelement2.text = str(100)
                                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'ULDELAY')
                                        antennaport_subelement2.text = str(100)
                                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'PWRSWITCH')
                                        antennaport_subelement2.text = str(1)
                                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1,'THRESHOLDTYPE')
                                        antennaport_subelement2.text = str(0)
                                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'UOTHD')
                                        antennaport_subelement2.text = str(40)
                                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'UCTHD')
                                        antennaport_subelement2.text = str(60)
                                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'OOTHD')
                                        antennaport_subelement2.text = str(185)
                                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'OCTHD')
                                        antennaport_subelement2.text = str(155)
                                        antennaport_subelement2 = ET.SubElement(antennaport_subelement1, 'ULTRADELAYSW')
                                        antennaport_subelement2.text = str(0)
                                        i = 1
                                child1.remove(child2)
                                # ET.dump(child1)

   #  RET_xml_string = '''<RET>
   #  <attributes>
   #   <DEVICENO>8</DEVICENO>
   #   <RETTYPE>1</RETTYPE>
   #   <POLARTYPE>2</POLARTYPE>
   #   <SCENARIO>1</SCENARIO>
   #   <SUBUNITNUM>1</SUBUNITNUM>
   #   <ANTENNAFORM>0</ANTENNAFORM>
   #   <DEVICENAME>OXLE21S8</DEVICENAME>
   #   <CTRLCN>0</CTRLCN>
   #   <CTRLSRN>68</CTRLSRN>
   #   <CTRLSN>0</CTRLSN>
   #   <VENDORCODE>KA</VENDORCODE>
   #   <SERIALNO>E4L2755862C-Y1</SERIALNO>
   #  </attributes>
   # </RET>'''
   #  RETSUBUNIT_xml_string = '''<RETSUBUNIT>
   #  <attributes>
   #   <DEVICENO>8</DEVICENO>
   #   <SUBUNITNO>1</SUBUNITNO>
   #   <CONNCN1>0</CONNCN1>
   #   <CONNSRN1>68</CONNSRN1>
   #   <CONNSN1>0</CONNSN1>
   #   <CONNPN1>0</CONNPN1>
   #   <CONNCN2>255</CONNCN2>
   #   <CONNPN2>1</CONNPN2>
   #   <TILT>60</TILT>
   #   <AER>5</AER>
   #   <SUBNAME>OXLE21I</SUBNAME>
   #  </attributes>
   # </RETSUBUNIT>'''

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'RET':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'CTRLSRN' and child4.text == str(
                                    SRN1):
                                child4.text = str(SRN2)
                                # ET.dump(child1)

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'RETSUBUNIT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'CONNSRN1' and child4.text == str(
                                    SRN1):
                                child4.text = str(SRN2)
                                # ET.dump(child1)
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'TMA':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'CTRLSRN' and child4.text == str(
                                    SRN1):
                                child4.text = str(SRN2)
                                # ET.dump(child1)
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'TMASUBUNIT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'CTRLSRN' and child4.text == str(
                                    SRN1):
                                child4.text = str(SRN2)
                                # ET.dump(child1)
    tree.write(OUTPUT)


def ETHPORT_adjust_thd():
    ethport_flag_ocr = False
    ethport_flag_clr = False
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                ethport_flag_ocr = False
                ethport_flag_clr = False
                if child2.tag == '' + 'ETHPORT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'RXBCPKTALMOCRTHD':
                                ethport_flag_ocr = True
                            if child4.tag == 'RXBCPKTALMCLRTHD':
                                ethport_flag_clr = True
                        if ethport_flag_ocr == False and ethport_flag_clr == False:
                            sector_subelement2 = ET.SubElement(child3, 'RXBCPKTALMOCRTHD')
                            sector_subelement2.text = str(320)
                            sector_subelement2 = ET.SubElement(child3, 'RXBCPKTALMCLRTHD')
                            sector_subelement2.text = str(288)
    tree.write(OUTPUT)


def CELL_SECTOREQM_mapping():
    list_cell_unique = []
    list_sectoreqm_unique = []
    string_out = 'Existing Cell:  '
    X1 = parse_existing_MO_name(root, 'Cell')
    X2 = parse_existing_MO_name(root, 'SECTOREQM')
    # print(X1)
    # print(X2)
    for i in X1:
        for j in i:
            for j in range(0, len(X1)):
                for item in X1[j]:
                    for key in item.keys():
                        if key == 'CellName':
                            if item[key][1] + '.' + item[key][6] not in list_cell_unique:
                                list_cell_unique.append(item[key][1] + '.' + item[key][6])
    for i in X2:
        for j in i:
            for j in range(0, len(X2)):
                for item in X2[j]:
                    for key in item.keys():
                        if key == 'SECTOREQMID':
                            if item[key] not in list_sectoreqm_unique:
                                list_sectoreqm_unique.append(item[key])

    # print(list_cell_unique)
    # if len(list_cell_unique)!=len(list_sector_unique):
    #     for k in range(0,len(list_cell_unique)-len(list_sector_unique)):
    #         list_sector_unique.append('')
    list_sectoreqm_unique.sort(key=int)
    list_cell_unique.sort()
    for i in range(0, len(list_cell_unique)):
        string_out += list_cell_unique[i] + ',  '
    string_out += '\nExisting SectorEqm: '
    for i in range(0, len(list_sectoreqm_unique)):
        string_out += list_sectoreqm_unique[i] + ', '
    # print(string_out)
    fo = open('LOG.txt', 'a')
    fo.write(f'{dt.datetime.now()} -- {NEname} Smart Agent lookup: ' + string_out + '\n')
    return string_out,list_cell_unique,list_sectoreqm_unique

#Future versions

# def change_ip():
#     window1 = Tk()
#     window1.title('Change IP addresses')
#     window1.geometry('{}x{}'.format(300, 100))
#     window1.config(background="white")

# def change_specific_ip():
#     window2 = Tk()
#     window2.title('Change specific 4G addresses')
#     window2.geometry('{}x{}'.format(500, 280))
#     window2.config(background="white")
#     OaM_IP = StringVar(window2)
#     OaM_SNM, OaM_DR, CUP_IP, CUP_SNM, CUP_DR, Sync_IP, Sync_SNM, Sync_DR, Sync_Source, Sync_Source_SNM, CUP_INNER_IP, CUP_INNER_SNM, SecGW_IP = \
#     StringVar(window2), StringVar(window2), StringVar(window2), StringVar(window2), StringVar(window2), StringVar(window2), StringVar(window2), StringVar(window2), StringVar(window2), StringVar(window2), StringVar(window2), StringVar(window2), StringVar(window2)
#
#     #Collect old IP addresses:
#     ACLRULE_var1 = parse_existing_MO_name(root, 'ACLRULE')
#     ACLRULE_4G_inner_unique = []
#     for i in ACLRULE_var1:
#         for j in i:
#             for j in range(0, len(ACLRULE_var1)):
#                 flag_4G_inner = False
#                 for item in ACLRULE_var1[j]:
#                     for key in item.keys():
#                         if key == 'ACLID' and item[key] == '3300':
#                             flag_4G_inner = True
#                         if flag_4G_inner and key == 'SIP':
#                             if item[key] not in ACLRULE_4G_inner_unique:
#                                 ACLRULE_4G_inner_unique.append(item[key])
#     try:
#         S1_4G_inner_old = ACLRULE_4G_inner_unique[0]
#         print(S1_4G_inner_old)
#     except:
#         pass
#
#     ACLRULE_5G_inner_unique = []
#     for i in ACLRULE_var1:
#         for j in i:
#             for j in range(0, len(ACLRULE_var1)):
#                 flag_5G_inner = False
#                 for item in ACLRULE_var1[j]:
#                     for key in item.keys():
#                         if key == 'ACLID' and item[key] == '3301':
#                             flag_5G_inner = True
#                         if flag_5G_inner and key == 'SIP':
#                             if item[key] not in ACLRULE_5G_inner_unique:
#                                 ACLRULE_5G_inner_unique.append(item[key])
#     try:
#         S1_5G_inner_old = ACLRULE_5G_inner_unique[0]
#         print(S1_5G_inner_old)
#     except:
#         pass
#
#     OMCH_var1 = parse_existing_MO_name(root, 'OMCH')
#     OMCH_localip_unique = []
#     for i in OMCH_var1:
#         for j in i:
#             for j in range(0, len(OMCH_var1)):
#                 for item in OMCH_var1[j]:
#                     for key in item.keys():
#                         if key == 'IP':
#                             if item[key] not in OMCH_localip_unique:
#                                 OMCH_localip_unique.append(item[key])
#     try:
#         OMCH_ip_old = OMCH_localip_unique[0]
#         print(OMCH_ip_old)
#     except:
#         pass
#
#     Sync_var1 = parse_existing_MO_name(root, 'IPCLKLNK')
#     Sync_localip_unique = []
#     Sync_source_unique=[]
#     for i in Sync_var1:
#         for j in i:
#             for j in range(0, len(Sync_var1)):
#                 for item in Sync_var1[j]:
#                     for key in item.keys():
#                         if key == 'CIP':
#                             if item[key] not in Sync_localip_unique:
#                                 Sync_localip_unique.append(item[key])
#                         if key == 'SIP':
#                             if item[key] not in Sync_source_unique:
#                                 Sync_source_unique.append(item[key])
#     try:
#         Sync_ip_old = Sync_localip_unique[0]
#         print(Sync_ip_old)
#         Sync_source_old=Sync_source_unique[0]
#         print(Sync_source_old)
#     except:pass
#
#     Certreq_var1=parse_existing_MO_name(root, 'CERTREQ')
#     S1_4G_outer_unique=[]
#     for i in Certreq_var1:
#         for j in i:
#             for j in range(0, len(Certreq_var1)):
#                 for item in Certreq_var1[j]:
#                     for key in item.keys():
#                         if key == 'LOCALIP':
#                             if item[key] not in S1_4G_outer_unique:
#                                 S1_4G_outer_unique.append(item[key])
#     try:
#         S1_4G_outer_old = S1_4G_outer_unique[0]
#         print(S1_4G_outer_old)
#     except:pass
#
#     Ikepeer_var1=parse_existing_MO_name(root, 'IKEPEER')
#     Ikepeer_4G_unique=[]
#     for i in Ikepeer_var1:
#         for j in i:
#             for j in range(0, len(Ikepeer_var1)):
#                 flag_4G_ikepeer = False
#                 for item in Ikepeer_var1[j]:
#                     for key in item.keys():
#                         if key == 'PEERNAME' and item[key] == 's1x2peer':
#                             flag_4G_ikepeer = True
#                         if flag_4G_ikepeer and key == 'REMOTEIP':
#                             if item[key] not in Ikepeer_4G_unique:
#                                 Ikepeer_4G_unique.append(item[key])
#     try:
#         Ikepeer_4G_old = Ikepeer_4G_unique[0]
#         print(Ikepeer_4G_old)
#     except:pass
#
#     # Vlanmap_legacy_var1=parse_existing_MO_name(root, 'VLANMAP')
#     # Nexthop_4g_legacy_unique=[]
#     # Nexthop_sync_legacy_unique=[]
#     # Nexthop_oam_legacy_unique=[]
#     # for i in Vlanmap_legacy_var1:
#     #     for j in i:
#     #         for j in range(0, len(Vlanmap_legacy_var1)):
#     #             flag_4G_ikepeer = False
#     #             for item in Vlanmap_legacy_var1[j]:
#     #                 for key in item.keys():
#     #                     if key == 'VLANID' and item[key] in []:
#     #                         flag_4G_ikepeer = True
#     #                     if flag_4G_ikepeer and key == 'REMOTEIP':
#     #                         if item[key] not in Ikepeer_4G_unique:
#     #                             Ikepeer_4G_unique.append(item[key])
#
#     Iprt_legacy_var1=parse_existing_MO_name(root, 'IPRT')
#     Nexthop_4G_legacy_unique=[]
#     Nexthop_sync_legacy_unique=[]
#     Nexthop_oam_legacy_unique=[]
#
#
#     def submit():
#         CUP_IP_value = CUP_IP.get()
#         CUP_SNM_value = CUP_SNM.get()
#         CUP_DR_value = CUP_DR.get()
#         CUP_INNER_IP_value = CUP_INNER_IP.get()
#         CUP_INNER_SNM_value = CUP_INNER_SNM.get()
#         SecGW_IP_value=SecGW_IP.get()
#         OaM_IP_value = OaM_IP.get()
#         OaM_SNM_value = OaM_SNM.get()
#         OaM_DR_value = OaM_DR.get()
#         Sync_IP_value = Sync_IP.get()
#         Sync_SNM_value = Sync_SNM.get()
#         Sync_DR_value = Sync_DR.get()
#         Sync_Source_value = Sync_Source.get()
#         Sync_Source_SNM_value = Sync_Source_SNM.get()
#
#         print("CUP_IP : " + CUP_IP_value)
#         print("CUP_SNM : " + CUP_SNM_value)
#         print("CUP_DR : " + CUP_DR_value)
#         print("CUP_INNER : " + CUP_INNER_IP_value)
#         print("CUP_INNER_SNM : " + CUP_INNER_SNM_value)
#         print("SecGW_IP : " + SecGW_IP_value)
#         print("OaM_IP : " + OaM_IP_value)
#         print("OaM_SNM : " + OaM_SNM_value)
#         print("OaM_DR : " + OaM_DR_value)
#         print("Sync_IP : " + Sync_IP_value)
#         print("Sync_SNM : " + Sync_SNM_value)
#         print("Sync_DR : " + Sync_DR_value)
#         print("Sync_Source : " + Sync_Source_value)
#         print("CUP_IP : " + CUP_IP_value)
#         print("Sync_Source_SNM : " + Sync_Source_SNM_value)
#
#
#
#
#
#         # for child0 in root:
#         #     for child1 in child0:
#         #         for child2 in child1:
#         #             if child2.tag == '' + 'OMCH':
#         #                 for child3 in child2:
#         #                     for child4 in child3:
#         #                         if child4.tag == 'IP' and OaM_IP_value.count('.')==3 and ' ' not in OaM_IP_value:
#         #                             child4.text =OaM_IP_value
#         #                             # ET.dump(child1)
#         #                             break
#         # tree.write(OUTPUT)
#         # for child0 in root:
#         #     for child1 in child0:
#         #         for child2 in child1:
#         #             if child2.tag == '' + 'OMCH':
#         #                 for child3 in child2:
#         #                     for child4 in child3:
#         #                         if child4.tag == 'MASK' and OaM_SNM_value.count('.')==3 and ' ' not in OaM_SNM_value:
#         #                             child4.text =OaM_SNM_value
#         #                             # ET.dump(child1)
#         #                             break
#         # tree.write(OUTPUT)
#         # time.sleep(10)
#         # window2.destroy()
#
#     def label_text_grid(txt,col,row_label,row_text,ip_var):
#         """creates label and text widgets for this window"""
#
#         #name_var.set("")
#         label=Label(window2,text=txt,bg='white').grid(column=col, row=row_label, sticky="W")
#         entry=Entry(window2, textvariable = ip_var, font=("arial", 10))
#         entry.grid(column=col, row=row_text, sticky="W")
#         #return label, entry
#
#     sub_btn = Button(window2, text='Submit', command=submit)
#     label_text_grid('CUP_IP',1,1,2,CUP_IP)
#     # Label(window2, text='CUP_IP').grid(column=1, row=1, sticky="W")
#     # entry = Entry(window2, textvariable=CUP_IP, font=("arial", 10))
#     # entry.grid(row=2,column=1, sticky="W")
#     label_text_grid('CUP_SNM', 2, 1, 2,CUP_SNM)
#     label_text_grid('CUP_DR',3,1,2,CUP_DR)
#     label_text_grid('CUP_INNER_IP', 1, 3, 4,CUP_INNER_IP)
#     label_text_grid('CUP_INNER_SNM', 2, 3, 4,CUP_INNER_SNM)
#     label_text_grid('SecGW_IP', 1, 5, 6,SecGW_IP)
#     label_text_grid('Sync_IP', 1, 7, 8,Sync_IP)
#     label_text_grid('Sync_SNM', 2, 7, 8,Sync_SNM)
#     label_text_grid('Sync_DR', 3, 7, 8,Sync_DR)
#     label_text_grid('Sync_Source', 1, 9, 10,Sync_Source)
#     label_text_grid('Sync_Source_SNM', 2, 9, 10,Sync_Source_SNM)
#     label_text_grid('OaM_IP', 1, 11, 12,OaM_IP)
#     label_text_grid('OaM_SNM', 2, 11, 12,OaM_SNM)
#     label_text_grid('OaM_DR', 3, 11, 12,OaM_DR)
#
#
#     sub_btn.grid(row=13, column=1)
#     window2.mainloop()

    # def change_csv_all_ip():
    #     filename3 = filedialog.askopenfilename(initialdir="/C:/", title="Select a File",filetypes=(("CSV files", "*.csv*"), ("all files", "*.*")))
    #     fo_csv=open(filename3,'r')
    #     csvreader = csv.reader(fo_csv, delimiter=';', quotechar='|')
    #     for row in csvreader:
    #         #print(', '.join(row))
    #         print(row)
    #         print(len(row))
    #
    #     OaM_IP = ''
    #     OaM_SNM = ''
    #     OaM_DR = ''
    #     CUP_IP = ''
    #     CUP_SNM = ''
    #     CUP_DR = ''
    #     Sync_IP = ''
    #     Sync_SNM = ''
    #     Sync_DR = ''
    #     Sync_Source = ''
    #     Sync_Source_SNM = ''
    #     CUP_INNER_IP = ''
    #     CUP_INNER_SNM = ''
    # # filename3 = filedialog.askopenfilename(initialdir="/C:/", title="Select a File",filetypes=(("CSV files", "*.csv*"), ("all files", "*.*")))
    # # fo_csv=open(filename3,'r')
    # # csvreader = csv.reader(fo_csv, delimiter=';', quotechar='|')
    # # for row in csvreader:
    # #     #print(', '.join(row))
    # #     print(row)
    # #     print(len(row))
    # #
    #Button_ip_1=Button(window2, text="Change specific 4G IP", command=change_specific_ip, width=30)
    #Button_ip_1.grid(column=1, row=1, sticky="W")
    # Button_ip_2=Button(window1, text="Change CSV IP", command=change_csv_all_ip, width=30)
    # Button_ip_2.grid(column=1, row=2, sticky="W")

def insert_IPADDR4(ITFID,IP,MASK,VRFIDX,USERLABEL):
    child2_tag_list_unique=[]
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag not in child2_tag_list_unique:
                    child2_tag_list_unique.append(child2.tag)

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'VRF':    #will create a new class under same parent as 'VRF' object
                    if 'IPADDR4' not in child2_tag_list_unique:
                        new_class_IPADDR4_subelement0 = ET.SubElement(child0, 'class')
                        new_class_IPADDR4_subelement1 = ET.SubElement(new_class_IPADDR4_subelement0, 'IPADDR4')
                        new_class_IPADDR4_subelement2 = ET.SubElement(new_class_IPADDR4_subelement1, 'attributes')
                        new_class_IPADDR4_subelement3 =  ET.SubElement(new_class_IPADDR4_subelement2, 'ITFD')
                        new_class_IPADDR4_subelement3.text=str(ITFID)
                        new_class_IPADDR4_subelement3 = ET.SubElement(new_class_IPADDR4_subelement2, 'IP')
                        new_class_IPADDR4_subelement3.text = str(IP)
                        new_class_IPADDR4_subelement3 = ET.SubElement(new_class_IPADDR4_subelement2, 'MASK')
                        new_class_IPADDR4_subelement3.text = str(MASK)
                        new_class_IPADDR4_subelement3 = ET.SubElement(new_class_IPADDR4_subelement2, 'VRFIDX')
                        new_class_IPADDR4_subelement3.text = str(VRFIDX)
                        new_class_IPADDR4_subelement3 = ET.SubElement(new_class_IPADDR4_subelement2, 'USERLABEL')
                        new_class_IPADDR4_subelement3.text = str(USERLABEL)
                        break
                    else:
                        for child0 in root:
                            for child1 in child0:
                                for child2 in child1:
                                    if child2.tag == '' + 'IPADDR4':
                                        new_class_IPADDR4_subelement0 = ET.SubElement(child1, 'IPADDR4')
                                        new_class_IPADDR4_subelement1 = ET.SubElement(new_class_IPADDR4_subelement0, 'attributes')
                                        new_class_IPADDR4_subelement2 = ET.SubElement(new_class_IPADDR4_subelement1,'ITFD')
                                        new_class_IPADDR4_subelement2.text = str(ITFID)
                                        new_class_IPADDR4_subelement2 = ET.SubElement(new_class_IPADDR4_subelement1,'IP')
                                        new_class_IPADDR4_subelement2.text = str(IP)
                                        new_class_IPADDR4_subelement2 = ET.SubElement(new_class_IPADDR4_subelement1,'MASK')
                                        new_class_IPADDR4_subelement2.text = str(MASK)
                                        new_class_IPADDR4_subelement2 = ET.SubElement(new_class_IPADDR4_subelement1,'VRFIDX')
                                        new_class_IPADDR4_subelement2.text = str(VRFIDX)
                                        new_class_IPADDR4_subelement2 = ET.SubElement(new_class_IPADDR4_subelement1,'USERLABEL')
                                        new_class_IPADDR4_subelement2.text = str(USERLABEL)
                                        break
    tree.write(OUTPUT)
def insert_IPROUTE4(RTIDX,VRFIDX,DSTIP,DSTMASK,NEXTHOP,USERLABEL):
    child2_tag_list_unique = []
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag not in child2_tag_list_unique:
                    child2_tag_list_unique.append(child2.tag)

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'VRF':  # will create a new class under same parent as 'VRF' object ,select correctly the syndata object
                    if 'IPROUTE4' not in child2_tag_list_unique:
                        new_class_IPADDR4_subelement0 = ET.SubElement(child0, 'class')
                        new_class_IPADDR4_subelement1 = ET.SubElement(new_class_IPADDR4_subelement0, 'IPROUTE4')
                        new_class_IPADDR4_subelement2 = ET.SubElement(new_class_IPADDR4_subelement1, 'attributes')
                        new_class_IPADDR4_subelement3 = ET.SubElement(new_class_IPADDR4_subelement2, 'RTIDX')
                        new_class_IPADDR4_subelement3.text = str(RTIDX)
                        new_class_IPADDR4_subelement3 = ET.SubElement(new_class_IPADDR4_subelement2, 'VRFIDX')
                        new_class_IPADDR4_subelement3.text = str(VRFIDX)
                        new_class_IPADDR4_subelement3 = ET.SubElement(new_class_IPADDR4_subelement2, 'DSTIP')
                        new_class_IPADDR4_subelement3.text = str(DSTIP)
                        new_class_IPADDR4_subelement3 = ET.SubElement(new_class_IPADDR4_subelement2, 'DSTMASK')
                        new_class_IPADDR4_subelement3.text = str(DSTMASK)
                        new_class_IPADDR4_subelement3 = ET.SubElement(new_class_IPADDR4_subelement2, 'RTTYPE')
                        new_class_IPADDR4_subelement3.text = str(0)
                        new_class_IPADDR4_subelement3 = ET.SubElement(new_class_IPADDR4_subelement2, 'NEXTHOP')
                        new_class_IPADDR4_subelement3.text = str(NEXTHOP)
                        new_class_IPADDR4_subelement3 = ET.SubElement(new_class_IPADDR4_subelement2, 'MTUSWITCH')
                        new_class_IPADDR4_subelement3.text = str(0)
                        new_class_IPADDR4_subelement3 = ET.SubElement(new_class_IPADDR4_subelement2, 'PREF')
                        new_class_IPADDR4_subelement3.text = str(60)
                        new_class_IPADDR4_subelement3 = ET.SubElement(new_class_IPADDR4_subelement2, 'USERLABEL')
                        new_class_IPADDR4_subelement3.text = str(USERLABEL)
                        break
                    else:
                        for child0 in root:
                            for child1 in child0:
                                for child2 in child1:
                                    if child2.tag == '' + 'IPROUTE4':
                                        new_class_IPADDR4_subelement0 = ET.SubElement(child1, 'IPROUTE4')
                                        new_class_IPADDR4_subelement1 = ET.SubElement(new_class_IPADDR4_subelement0,'attributes')
                                        new_class_IPADDR4_subelement2 = ET.SubElement(new_class_IPADDR4_subelement1,'RTIDX')
                                        new_class_IPADDR4_subelement2.text = str(RTIDX)
                                        new_class_IPADDR4_subelement2 = ET.SubElement(new_class_IPADDR4_subelement1,'VRFIDX')
                                        new_class_IPADDR4_subelement2.text = str(VRFIDX)
                                        new_class_IPADDR4_subelement2 = ET.SubElement(new_class_IPADDR4_subelement1,'DSTIP')
                                        new_class_IPADDR4_subelement2.text = str(DSTIP)
                                        new_class_IPADDR4_subelement2 = ET.SubElement(new_class_IPADDR4_subelement1,'DSTMASK')
                                        new_class_IPADDR4_subelement2.text = str(DSTMASK)
                                        new_class_IPADDR4_subelement2 = ET.SubElement(new_class_IPADDR4_subelement1,'RTTYPE')
                                        new_class_IPADDR4_subelement2.text = str(0)
                                        new_class_IPADDR4_subelement2 = ET.SubElement(new_class_IPADDR4_subelement1,'NEXTHOP')
                                        new_class_IPADDR4_subelement2.text = str(NEXTHOP)
                                        new_class_IPADDR4_subelement2 = ET.SubElement(new_class_IPADDR4_subelement1,'MTUSWITCH')
                                        new_class_IPADDR4_subelement2.text = str(0)
                                        new_class_IPADDR4_subelement2 = ET.SubElement(new_class_IPADDR4_subelement1,'PREF')
                                        new_class_IPADDR4_subelement2.text = str(60)
                                        new_class_IPADDR4_subelement2 = ET.SubElement(new_class_IPADDR4_subelement1,'USERLABEL')
                                        new_class_IPADDR4_subelement2.text = str(USERLABEL)
                                        break
    tree.write(OUTPUT)
def old2new_basic_transmission_model_convertor():
    # GTRANSPARA mod
    # ETHPORT mod
    # IPADDR4 add
    # IPROUTE4 add
    # INTERFACE add
    # IPSECBINDITF add
    # DEVIP rmv
    # IPRT rmv
    # VLANMAP rmv
    # IPSECBIND rmv

    devip_ip_list=[]
    devip_mask_list=[]
    devip_userlabel_list=[]
    iprt_rtidx_list=[]
    iprt_destip_list=[]
    iprt_destmask_list=[]
    iprt_nexthop_list=[]
    iprt_descri_list=[]
    PN_list_unique = []
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'DEVIP':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'PN' and child4.text not in PN_list_unique:
                                PN_list_unique.append(child4.text)
                                # ET.dump(child1)
                                break
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'DEVIP':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'IP':
                                devip_ip_list.append(child4.text)
                            if child4.tag == 'MASK':
                                devip_mask_list.append(child4.text)
                            if child4.tag == 'USERLABEL':
                                devip_userlabel_list.append(child4.text)

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'IPRT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'RTIDX':
                                iprt_rtidx_list.append(child4.text)
                            if child4.tag == 'DSTIP':
                                iprt_destip_list.append(child4.text)
                            if child4.tag == 'DSTMASK':
                                iprt_destmask_list.append(child4.text)
                            if child4.tag == 'NEXTHOP':
                                iprt_nexthop_list.append(child4.text)
                            if child4.tag == 'DESCRI':
                                iprt_descri_list.append(child4.text)


    # create a list with the interfaces, create the interfaces, replace cnt from 1st arg  in function call insert_IPADDR4

    for cnt in range(0,len(devip_ip_list)):
        insert_IPADDR4(cnt,devip_ip_list[cnt],devip_mask_list[cnt],'0', devip_userlabel_list[cnt])

    for cnt in range(0,len(iprt_destip_list)):
        insert_IPROUTE4(iprt_rtidx_list[cnt],0,iprt_destip_list[cnt],iprt_rtidx_list[cnt],iprt_nexthop_list[cnt],iprt_descri_list[cnt])
    # insert_IPADDR4(0, devip_ip_list[0], devip_mask_list[0], '0', devip_userlabel_list[0])
    # insert_IPADDR4(1, devip_ip_list[1], devip_mask_list[1], '0', devip_userlabel_list[1])
    # insert_IPADDR4(2, devip_ip_list[2], devip_mask_list[2], '0', devip_userlabel_list[2])
    # insert_IPADDR4(3, devip_ip_list[3], devip_mask_list[3], '0', devip_userlabel_list[3])



    """<class>
   <IPADDR4>
    <attributes>
     <ITFID>0</ITFID>
     <IP>172.26.100.238</IP>
     <MASK>255.255.254.0</MASK>
     <VRFIDX>0</VRFIDX>
     <USERLABEL>S1_X2 Inner IP</USERLABEL>
    </attributes>
   </IPADDR4>
   <IPADDR4>
    <attributes>
     <ITFID>1</ITFID>
     <IP>10.125.245.163</IP>
     <MASK>255.255.255.224</MASK>
     <VRFIDX>0</VRFIDX>
     <USERLABEL>OM local IP</USERLABEL>
    </attributes>
   </IPADDR4>
   <IPADDR4>
    <attributes>
     <ITFID>2</ITFID>
     <IP>192.170.3.218</IP>
     <MASK>255.255.255.252</MASK>
     <VRFIDX>0</VRFIDX>
     <USERLABEL>IPCLK local</USERLABEL>
    </attributes>
   </IPADDR4>
   <IPADDR4>
    <attributes>
     <ITFID>3</ITFID>
     <IP>192.168.50.19</IP>
     <MASK>255.255.254.0</MASK>
     <VRFIDX>0</VRFIDX>
     <USERLABEL>S1_X2 Outer IP</USERLABEL>
    </attributes>
   </IPADDR4>
  </class>"""

    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'GTRANSPARA':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'TRANSCFGMODE' and child4.text == str(0):
                                child4.text = str(1)
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'ETHPORT':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'PORTID':
                                if '0' in PN_list_unique:
                                    child4.text = str(70)
                                elif '1' in PN_list_unique:
                                    child4.text = str(71)

    tree.write(OUTPUT)


#############################################################~~~~MAIN~~~~##################################################################################
#############################################################~~~~MAIN~~~~##################################################################################
#############################################################~~~~MAIN~~~~##################################################################################
#############################################################~~~~MAIN~~~~##################################################################################
#############################################################~~~~MAIN~~~~##################################################################################


# print("MAIN\n")
# parse_existing_MO(root,'SUBRACK')
# C = parse_existing_MO_name(root, 'SFP')

# print(C)
# parse_existing_MO(root,'Cell')
# parse_existing_MO(root,'SCTPPEER')
# print("VALUES SAVED\n")


# for item in C:
#    print(item)

# insert_new_BBP(CN,SRN,SN,TYPE,BBWS)
##type: UBBP~2, UBBP-W~181, GBBP~4098, WBBP~8194, LBBP~12290, LPMP~12546, LCOP~12802
# <BBWS>20</BBWS><!-- GSM:NO;UMTS:NO;LTE FDD:YES;LTE TDD:NO;NB-IoT:YES;NR:NO --> ltefdd=4/ltefdd+iot=20
# insert_new_BBP(0, Subrack_global, 5, 2, 20)

# insert_new_RRUCHAIN(rruchainno,HCN,HSRN,HSN,HPN):
# insert_new_RRUCHAIN(25, 0, 1, 5, 5)

# working mode of RRU: #TDL~4, TL~6, LO~8, LFTD~12, WL~20, CL~40, GO~64, GT~68, GL~72, GLFTD~76, UO~128, UT~132, UL~136, ULFTD~140, CU~160, GU~192, GUT~196, GUL~200, GULFTD~204, MO~256, TM~260, LM~264, LFTDM~268, CM~288, CLM~296, GM~320, GTM~324, GLM~328, GLFTDM~332, UM~384, UTM~388, ULM~392, ULFTDM~396, GUM~448, GUTM~452, GULM~456, GULFTDM~460, RO~512, TR~516, LR~520, LFTDR~524, UR~640, UTR~644, ULR~648, ULFTDR~652, RM~768, TRM~772, LRM~776, LFTDRM~780, URM~896, UTRM~900, ULRM~904, ULFTDRM~908, NO~2048, TN~2052, TLN~2054, LN~2056, LFTDN~2060, CN~2080, CLN~2088, GN~2112, GTN~2116, GLN~2120, GLFTDN~2124, UN~2176, UTN~2180, ULN~2184, ULFTDN~2188, CUN~2208, GUN~2240, GUTN~2244, GULN~2248, GULFTDN~2252, MN~2304, TMN~2308, LMN~2312, LFTDMN~2316, CMN~2336, CLMN~2344, GMN~2368, GLMN~2376, GLFTDMN~2380, UMN~2432, UTMN~2436, ULMN~2440, GUMN~2496, GULMN~2504, RN~2560, TRN~2564, LRN~2568, LFTDRN~2572, URN~2688, UTRN~2692, ULRN~2696, ULFTDRN~2700, RMN~2816, TRMN~2820, LRMN~2824, LFTDRMN~2828, URMN~2944, UTRMN~2948, ULRMN~2952, ULFTDRMN~2956
# insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(CN,SRN,SN,RRUCHAIN,Working_mode,RRU_type,RXNUM,TXNUM):
# insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 105, 0, 25, 136, 15, 4, 4)

# insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(SECTORID,CN,SRN,SN,ANTENNA_PORT_LIST):
# insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(90, 0, 105, 0, [0, 1, 2, 3])  # create sector with ABCD

# insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(SECTOREQMID,CN,SRN,SN,ANTENNA_PORT_LIST)
# insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(90, 0, 105, 0, [0, 3])


# rmv_old_bbp_from_slotnumber(2)
# rmv_old_rruchain_ALL_from_slotnumber(2)
# rmv_old_rruchain_from_slotnumber_portnumber(3,2)
# rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(60)
# rmv_old_SECTOREQM_ID(19)
# rmv_old_SECTOR_ID(19)

# scenarios
# 5900 / 2100 TUV single rru
# 5900/ 2100 TUV shared with 1800 and UMTS
# 5900 / 1800 single
# 5900 / 1800 shared with 2100 and UMTS
# 5900 /800 +700+900
# 5900 / 2600
# 5900/ 800 single
# 5900/ 700 single
#
# 3900 / 2100 TUV ??
# 3900 / 1800 HIK ??
# 3900 / 800 ABC ??
# 3900 / 2600 EFG ??


def SCENARIO1_3900to5900_newLTE2100_AB_oldLTE800_AB_notsharedlte1800():
    # SCENARIO1 [HXL253] BTS3900->BTS5900, add 2100 4T4R BD with existing 800 2T2R AB,no LTE1800  or it is not shared with LTE2100
    # LTE1800  does not exist or LTE1800 is single RRU 67,68,69 slot 3
    change_bbu_subrack_id()
    subrack_type_change_3900to5900()
    insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, 1, 4, 2, 20)
    insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, 1, 3, 2, 4)
    rmv_old_bbp_from_slotnumber(2)
    # <BBWS>20</BBWS><!-- GSM:NO;UMTS:NO;LTE FDD:YES;LTE TDD:NO;NB-IoT:YES;NR:NO --> ltefdd=4/ltefdd+iot=20

    insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(0, 0, 1, 4, 0)
    insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(1, 0, 1, 4, 1)
    insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(2, 0, 1, 4, 2)
    insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(18, 0, 1, 3, 3)
    insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(19, 0, 1, 3, 4)
    insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(20, 0, 1, 3, 5)
    rmv_old_rruchain_ALL_from_slotnumber(2)
    # TDL~4, TL~6, LO~8, LFTD~12, WL~20, CL~40, GO~64, GT~68, GL~72, GLFTD~76, UO~128, UT~132, UL~136, ULFTD~140, CU~160, GU~192, GUT~196, GUL~200, GULFTD~204, MO~256, TM~260, LM~264, LFTDM~268, CM~288, CLM~296, GM~320, GTM~324, GLM~328, GLFTDM~332, UM~384, UTM~388, ULM~392, ULFTDM~396, GUM~448, GUTM~452, GULM~456, GULFTDM~460, RO~512, TR~516, LR~520, LFTDR~524, UR~640, UTR~644, ULR~648, ULFTDR~652, RM~768, TRM~772, LRM~776, LFTDRM~780, URM~896, UTRM~900, ULRM~904, ULFTDRM~908, NO~2048, TN~2052, TLN~2054, LN~2056, LFTDN~2060, CN~2080, CLN~2088, GN~2112, GTN~2116, GLN~2120, GLFTDN~2124, UN~2176, UTN~2180, ULN~2184, ULFTDN~2188, CUN~2208, GUN~2240, GUTN~2244, GULN~2248, GULFTDN~2252, MN~2304, TMN~2308, LMN~2312, LFTDMN~2316, CMN~2336, CLMN~2344, GMN~2368, GLMN~2376, GLFTDMN~2380, UMN~2432, UTMN~2436, ULMN~2440, GUMN~2496, GULMN~2504, RN~2560, TRN~2564, LRN~2568, LFTDRN~2572, URN~2688, UTRN~2692, ULRN~2696, ULFTDRN~2700, RMN~2816, TRMN~2820, LRMN~2824, LFTDRMN~2828, URMN~2944, UTRMN~2948, ULRMN~2952, ULFTDRMN~2956

    insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 18, 264, 15, 2, 2)
    insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 19, 264, 15, 2, 2)
    insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 20, 264, 15, 2, 2)

    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(18, 0, 100, 0, [0, 1])
    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 0, 101, 0, [0, 1])
    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 0, 102, 0, [0, 1])

    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(18, 0, 100, 0, [0, 1])
    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 0, 101, 0, [0, 1])
    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 0, 102, 0, [0, 1])


# mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(60,0,80,0,[0,2,3])
# mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(61,0,82,0,[0,2,3])
# mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(62,0,84,0,[0,2,3])

# mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(60,0,80,0,[0,2,3])

# mod_BBP_from_slot_to_slot(2,4,2,2,20,20)
# mod_RRUCHUAIN_oldHSN_newHSN(2,4)
#
# mod_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(60,0,80,0,4,4,264,15)
# mod_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(61,0,82,0,4,4,264,15)
# mod_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(62,0,84,0,4,4,264,15)


# what do i want in the end:
#  BBU5900
# 0. 3900to5900
# subrack1
# 1. LTE800 AB
# 2. LTE800 CD +LTE700 AB
# 2..LTE800 AB +LTE700 CD
# 3. LTE700 single slot 5
# 3..LTE 700 single slot 4
# 4. LTE1800 AB ABCD
# 5. LTE2100 AB
# 6. LTE2100 AC +LTE1800 BD p012
# 7. LTE2100 BD +LTE1800 AC p012
# 6. LTE2100 AC +LTE1800 BD p345
# 7. LTE2100 BD +LTE1800 AC p345
# 8. LTE 2600 AB ABCD
# 9. LTE900 ??
# BBU3900

# CELL_SECTOREQM_mapping()


# Create a File Explorer label
# textBox=Label(window, height=10, width=30,text="CELL-SECTOR")
# textBox.grid(column=2, row=3)
def label_f():
    textBox = Label(window, text=CELL_SECTOREQM_mapping()[0], font='arial 11 bold').grid(column=2, row=3)


def label_g():
    #textBox = Label(window, text=UMPT_ETHPORT_SW(), font='arial 11 bold').grid(column=2, row=2)
    text_label_g=Text(window,height=1,width=130,font =("arial", 10))
    text_label_g.insert(INSERT, UMPT_ETHPORT_SW())
    text_label_g.grid(column=2, row=2)


label_file_explorer = Label(window, text="BRD BUTTON", width=40, height=1, fg="blue")
# label_Execution_logs = Label(window,text="Board,sofware, subrack type output",width=100, height=1,fg="red").grid(column=2,row=2)
label_space1 = Label(window, text="", width=80, height=4, fg="white")
button_explore = Button(window, text="Open .XML File", command=browseFiles, width=30)
button_explore_tip = CreateToolTip(button_explore, "Select a new XML file to restart process.")
button_UMPT_ETHPORT_SW = Button(window, text="Initial Hardware/Software", command=label_g, width=40)
button_UMPT_ETHPORT_SW_tip = CreateToolTip(button_UMPT_ETHPORT_SW, "Press Me!\nShow info on initial BBU configuration, ethport, subrackID, software.")
button_show_cells = Button(window, text="Initial Cell & SECTOREQM", command=label_f, width=40)


button_show_cells_tip = CreateToolTip(button_show_cells, "Press Me!\nList of configured Cells and configured SECTOREQM.\n Missing SECTOREQM for defined Cells can show what needs to be integrated.")

def close_window():
    window.destroy()


button_exit = Button(window, text="EXIT", command=close_window, width=40)

link = Label(window, text="Help", fg="blue",bg='white', cursor="hand2")

link.bind("<Button-1>", lambda event: os.startfile('help.docx'))
#link.grid(column=1, row=39,sticky='W')
link.place(x=50, y=685)
# Grid method is chosen for placing
# the widgets at respective positions
# in a table like structure by
# specifying rows and columns
label_file_explorer.grid(column=1, row=1)

button_explore.grid(column=1, row=1,sticky="W")
button_UMPT_ETHPORT_SW.grid(column=1, row=2)
button_show_cells.grid(column=1, row=3)
button_exit.grid(column=1, row=35)

#button_ip_change=Button(window, text="IP CHANGE", command=change_specific_ip, width=40)
#button_ip_change.grid(column=1, row=36)

def cb_check2():
    if CheckVar2.get() == 1:
        C16.config(state=DISABLED)
    else:
        C16.config(state=NORMAL)

def cb_check300():
    if CheckVar300.get() == 1:
        C3.config(state=DISABLED)
        C4.config(state=DISABLED)
        C5.config(state=DISABLED)
        C6.config(state=DISABLED)
        C30.config(state=DISABLED)
    else:
        C3.config(state=NORMAL)
        C4.config(state=NORMAL)
        C5.config(state=NORMAL)
        C6.config(state=NORMAL)
        C30.config(state=NORMAL)

def cb_check30():
    if CheckVar30.get() == 1:
        C3.config(state=DISABLED)
        C4.config(state=DISABLED)
        C5.config(state=DISABLED)
        C6.config(state=DISABLED)
        C300.config(state=DISABLED)
    else:
        C3.config(state=NORMAL)
        C4.config(state=NORMAL)
        C5.config(state=NORMAL)
        C6.config(state=NORMAL)
        C300.config(state=NORMAL)


def cb_check3():
    if CheckVar3.get() == 1:
        C30.config(state=DISABLED)
        C4.config(state=DISABLED)
        C5.config(state=DISABLED)
        C300.config(state=DISABLED)
    else:
        C30.config(state=NORMAL)
        C4.config(state=NORMAL)
        C5.config(state=NORMAL)
        C300.config(state=NORMAL)


def cb_check4():
    if CheckVar4.get() == 1:
        C30.config(state=DISABLED)
        C3.config(state=DISABLED)
        C5.config(state=DISABLED)
        C6.config(state=DISABLED)
        C300.config(state=DISABLED)

    else:
        C30.config(state=NORMAL)
        C3.config(state=NORMAL)
        C5.config(state=NORMAL)
        C6.config(state=NORMAL)
        C300.config(state=NORMAL)


def cb_check5():
    if CheckVar5.get() == 1:
        C30.config(state=DISABLED)
        C3.config(state=DISABLED)
        C4.config(state=DISABLED)
        C6.config(state=DISABLED)
        C300.config(state=DISABLED)
    else:
        C30.config(state=NORMAL)
        C3.config(state=NORMAL)
        C4.config(state=NORMAL)
        C6.config(state=NORMAL)
        C300.config(state=NORMAL)


def cb_check6():
    if CheckVar6.get() == 1:
        C30.config(state=DISABLED)
        C4.config(state=DISABLED)
        C5.config(state=DISABLED)
        C300.config(state=DISABLED)
    else:
        C30.config(state=NORMAL)
        C4.config(state=NORMAL)
        C5.config(state=NORMAL)
        C300.config(state=NORMAL)


def cb_check7():
    if CheckVar7.get() == 1:
        #C80.config(state=DISABLED)
        #C8.config(state=DISABLED)
        # C9.config(state=DISABLED)
        C10.config(state=DISABLED)
        C11.config(state=DISABLED)
        C12.config(state=DISABLED)
        C13.config(state=DISABLED)

    else:
        #C80.config(state=NORMAL)
        #C8.config(state=NORMAL)
        # C9.config(state=NORMAL)
        C10.config(state=NORMAL)
        C11.config(state=NORMAL)
        C12.config(state=NORMAL)
        C13.config(state=NORMAL)


# def cb_check80():
#     if CheckVar80.get() == 1:
#         C7.config(state=DISABLED)
#         C8.config(state=DISABLED)
#         # C9.config(state=DISABLED)
#         C10.config(state=DISABLED)
#         C11.config(state=DISABLED)
#         C12.config(state=DISABLED)
#         C13.config(state=DISABLED)
#
#     else:
#         C7.config(state=NORMAL)
#         C8.config(state=NORMAL)
#         # C9.config(state=NORMAL)
#         C10.config(state=NORMAL)
#         C11.config(state=NORMAL)
#         C12.config(state=NORMAL)
#         C13.config(state=NORMAL)


# def cb_check8():
#     if CheckVar8.get() == 1:
#         C7.config(state=DISABLED)
#         C80.config(state=DISABLED)
#         # C9.config(state=DISABLED)
#         C10.config(state=DISABLED)
#         C11.config(state=DISABLED)
#         C12.config(state=DISABLED)
#         C13.config(state=DISABLED)
#
#     else:
#         C7.config(state=NORMAL)
#         C80.config(state=NORMAL)
#         # C9.config(state=NORMAL)
#         C10.config(state=NORMAL)
#         C11.config(state=NORMAL)
#         C12.config(state=NORMAL)
#         C13.config(state=NORMAL)


def cb_check9():
    if CheckVar9.get() == 1:
        C10.config(state=DISABLED)
        C11.config(state=DISABLED)
        C12.config(state=DISABLED)
        C13.config(state=DISABLED)


    else:
        C10.config(state=NORMAL)
        C11.config(state=NORMAL)
        C12.config(state=NORMAL)
        C13.config(state=NORMAL)


def cb_check10():
    if CheckVar10.get() == 1:
        C7.config(state=DISABLED)
        #C80.config(state=DISABLED)
        #C8.config(state=DISABLED)
        C9.config(state=DISABLED)
        C11.config(state=DISABLED)
        C12.config(state=DISABLED)
        C13.config(state=DISABLED)

    else:
        C7.config(state=NORMAL)
        #C80.config(state=NORMAL)
        #C8.config(state=NORMAL)
        C9.config(state=NORMAL)
        C11.config(state=NORMAL)
        C12.config(state=NORMAL)
        C13.config(state=NORMAL)


def cb_check11():
    if CheckVar11.get() == 1:
        C7.config(state=DISABLED)
        #C80.config(state=DISABLED)
        #C8.config(state=DISABLED)
        C9.config(state=DISABLED)
        C10.config(state=DISABLED)
        C12.config(state=DISABLED)
        C13.config(state=DISABLED)

    else:
        C7.config(state=NORMAL)
        #C80.config(state=NORMAL)
        #C8.config(state=NORMAL)
        C9.config(state=NORMAL)
        C10.config(state=NORMAL)
        C12.config(state=NORMAL)
        C13.config(state=NORMAL)


def cb_check12():
    if CheckVar12.get() == 1:
        C7.config(state=DISABLED)
        #C80.config(state=DISABLED)
        #C8.config(state=DISABLED)
        C9.config(state=DISABLED)
        C11.config(state=DISABLED)
        C10.config(state=DISABLED)
        C13.config(state=DISABLED)

    else:
        C7.config(state=NORMAL)
        #C80.config(state=NORMAL)
        #C8.config(state=NORMAL)
        C9.config(state=NORMAL)
        C11.config(state=NORMAL)
        C10.config(state=NORMAL)
        C13.config(state=NORMAL)


def cb_check13():
    if CheckVar13.get() == 1:
        C7.config(state=DISABLED)
        #C80.config(state=DISABLED)
        #C8.config(state=DISABLED)
        C9.config(state=DISABLED)
        C11.config(state=DISABLED)
        C12.config(state=DISABLED)
        C10.config(state=DISABLED)

    else:
        C7.config(state=NORMAL)
        #C80.config(state=NORMAL)
        #C8.config(state=NORMAL)
        C9.config(state=NORMAL)
        C11.config(state=NORMAL)
        C12.config(state=NORMAL)
        C10.config(state=NORMAL)


def cb_check16():
    if CheckVar16.get() == 1:
        C2.config(state=DISABLED)
    else:
        C2.config(state=NORMAL)


def cb_check16_17():
    C16.config(state=DISABLED)
    C17.config(state=DISABLED)

def cb_check_17():
    if CheckVar17.get() == 1:
        for item in [C1, C3, C30, C300, C4, C5, C6, C7, C9, C10, C11, C12, C13, C14, C15, C18,C19, C20, C21, C22, C23]:
            item.config(state=DISABLED)
    else:
        for item in [C1, C3, C30, C300, C4, C5, C6, C7, C9, C10, C11, C12, C13, C14, C15, C18,C19, C20, C21, C22, C23]:
            item.config(state=NORMAL)


def cb_check18():
    if CheckVar18.get() == 1:
        for item in [C1, C3, C30, C300, C4, C5, C6, C7, C9, C10, C11, C12, C13, C14, C15, C19, C20, C21, C22,C23]:
            item.config(state=DISABLED)
    else:
        for item in [C1, C3, C30, C300, C4, C5, C6, C7, C9, C10, C11, C12, C13, C14, C15, C19, C20, C21, C22,C23]:
            item.config(state=NORMAL)


def cb_check19():
    if CheckVar19.get() == 1:
        for item in [C1, C3, C30, C300, C4, C5, C6, C7, C9, C10, C11, C12, C13, C14, C15, C18, C20, C21, C22,C23]:
            item.config(state=DISABLED)
    else:
        for item in [C1, C3, C30, C300, C4, C5, C6, C7, C9, C10, C11, C12, C13, C14, C15, C18, C20, C21, C22,C23]:
            item.config(state=NORMAL)


def cb_check20():
    if CheckVar20.get() == 1:
        for item in [C1, C3, C30, C300, C4, C5, C6, C7, C9, C10, C11, C12, C13, C14, C15, C18, C19, C21, C22,C23]:
            item.config(state=DISABLED)
    else:
        for item in [C1, C3, C30, C300, C4, C5, C6, C7, C9, C10, C11, C12, C13, C14, C15, C18, C19, C21, C22,C23]:
            item.config(state=NORMAL)


def cb_check21():
    if CheckVar21.get() == 1:
        for item in [C1, C3, C30, C300, C4, C5, C6, C7, C9, C10, C11, C12, C13, C14, C15, C18, C19, C20, C22,C23]:
            item.config(state=DISABLED)
    else:
        for item in [C1, C3, C30, C300, C4, C5, C6, C7, C9, C10, C11, C12, C13, C14, C15, C18, C19, C20, C22,C23]:
            item.config(state=NORMAL)


def cb_check22():
    if CheckVar22.get() == 1:
        for item in [C1, C3, C30, C300, C4, C5, C6, C7, C9, C10, C11, C12, C13, C14, C15, C18, C19, C20, C21,C23]:
            item.config(state=DISABLED)
    else:
        for item in [C1, C3, C30, C300, C4, C5, C6, C7, C9, C10, C11, C12, C13, C14, C15, C18, C19, C20, C21,C23]:
            item.config(state=NORMAL)


def cb_check23():
    if CheckVar23.get() == 1:
        for item in [C1, C3, C30, C300, C4, C5, C6, C7, C9, C10, C11, C12, C13, C14, C15, C18, C19, C20, C21,C22]:
            item.config(state=DISABLED)
    else:
        for item in [C1, C3, C30, C300, C4, C5, C6, C7, C9, C10, C11, C12, C13, C14, C15, C18, C19, C20, C21,C22]:
            item.config(state=NORMAL)


# Let the window wait for any events
CheckVar1 = IntVar()
CheckVar2 = IntVar()
CheckVar3 = IntVar()
CheckVar30 = IntVar()
CheckVar300 = IntVar()
CheckVar4 = IntVar()
CheckVar5 = IntVar()
CheckVar6 = IntVar()
CheckVar7 = IntVar()
# CheckVar80 = IntVar()
# CheckVar8 = IntVar()
CheckVar9 = IntVar()
CheckVar10 = IntVar()
CheckVar11 = IntVar()
CheckVar12 = IntVar()
CheckVar13 = IntVar()
CheckVar14 = IntVar()
CheckVar15 = IntVar()
CheckVar16 = IntVar()
CheckVar17 = IntVar()
CheckVar18 = IntVar()
CheckVar19 = IntVar()
CheckVar20 = IntVar()
CheckVar21 = IntVar()
CheckVar22 = IntVar()
CheckVar23 = IntVar()
CheckVar24 = IntVar()

C1 = Checkbutton(window, text="NE Type Reconst. BTS3900 LTE to BTS5900 LTE", variable=CheckVar1, onvalue=1, offvalue=0, height=1,width=38, anchor="w", bg='lavender',fg='navy')
C1_tip = CreateToolTip(C1, "BBU reconstruction: BBU3900/BBU3910 to BBU 5900 and NE type reconstruction to BTS5900 LTE.\nLTE 800 reconfiguration not included within this option.")
C2 = Checkbutton(window, text="BBU subrack ID 0 -> 1", variable=CheckVar2, command=cb_check2, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender',fg='navy')
C2_tip=CreateToolTip(C2, "Change the BBU Subrack ID from 0 -> 1")
C300 = Checkbutton(window, text="5900: L8 AB INT 80-82-84, no L7 cells  [5519et]", variable=CheckVar300, command=cb_check300, onvalue=1, offvalue=0, height=1, width=38,anchor="w", bg='lavender',fg='navy')
C300_tip=CreateToolTip(C300, "L800 integration with RRU5519et, antenna ports AB, without L700")
C30 = Checkbutton(window, text="5900: L8 CD INT 80-82-84, no L7 cells  [5509t]", variable=CheckVar30, command=cb_check30,onvalue=1, offvalue=0, height=1, width=38, anchor="w", bg='lavender',fg='navy')
C30_tip=CreateToolTip(C30, "L800 integration with RRU5509t, antenna ports CD, without L700")
C3 = Checkbutton(window, text="5900: L8 AB REC/INT 60-61-62 SN 2,3 -> 4", variable=CheckVar3, command=cb_check3,onvalue=1, offvalue=0, height=1, width=38, anchor="w", bg='lavender',fg='navy')
C3_tip=CreateToolTip(C3, "L800 integration or L800 reconfiguration with single band RRU,on UBBP slot 4/0.1.2 SRM 60-61-62")
C4 = Checkbutton(window, text="5900: L8 CD REC/INT + L7 AB INT 80-82-84  [5509t]", variable=CheckVar4, command=cb_check4,onvalue=1, offvalue=0, height=1, width=38, anchor="w", bg='lavender',fg='navy')
C4_tip=CreateToolTip(C4, "[L800 reconfiguration or integration, antenna ports CD]  and/or  [L700 integration, antenna ports AB] with multi band RRU5509t,on UBBP slot 4/0.1.2 SRN 80-82-84.]\n In case one of the  bands is missing, it will execute only for the other.")
C5 = Checkbutton(window, text="5900: L8 AB REC/INT + L7 CD INT 80-82-84  [5519et]", variable=CheckVar5, command=cb_check5,onvalue=1, offvalue=0, height=1, width=38, anchor="w", bg='lavender',fg='navy')
C5_tip=CreateToolTip(C5, "[L800 reconfiguration or integration,antenna ports AB]  and/or  [L700 integration,antenna ports CD] with multi band 5519et,on UBBP slot 4/0.1.2 SRN 80-82-84.\n In case one of the  bands is missing, it will execute only for the other.")
C6 = Checkbutton(window, text="5900: L7 AB INT 90-91-92 SN 5", variable=CheckVar6, command=cb_check6, onvalue=1,offvalue=0, height=1, width=38, anchor="w", bg='lavender',fg='navy')
C6_tip=CreateToolTip(C6, "L700 integration with singleband RRUs on slot 5/0.1.2 SRN 90-91-92")
C7 = Checkbutton(window, text="5900: L18 AB/ABCD INT 67-68-69 singleband RRU", variable=CheckVar7, command=cb_check7, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender',fg='navy')
#C80 = Checkbutton(window, text="5900: L1800 INT AC", variable=CheckVar80, command=cb_check80, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender')
#C8 = Checkbutton(window, text="5900: L1800 INT AB", variable=CheckVar8, command=cb_check8, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender')
C7_tip=CreateToolTip(C7, "L1800 integration with singleband RRUs on slot 3/0.1.2 SRN 67-68-69, 2T2R or 4T4R")
C9 = Checkbutton(window, text="5900: L21 AB INT singleband RRU", variable=CheckVar9, command=cb_check9, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender',fg='navy')
C9_tip=CreateToolTip(C9, "L2100 integration with singleband RRUs on slot 3/3.4.5 SRN 100-101-102, 2T2R")
C10 = Checkbutton(window, text="5900: L21 AC INT + L18 BD/ABCD  REC/INT   [5501]", variable=CheckVar10,command=cb_check10, onvalue=1, offvalue=0, height=1, width=38, anchor="w", bg='lavender',fg='navy')
C10_tip=CreateToolTip(C10,'[LTE2100 integration, antenna ports AC]  and/or  [LTE1800 reconfiguration or integration, antenna ports BD/ABCD 2t2r/4t4r], with multiband RRU5501.\nIn case one of the  bands is missing, it will execute only for the other.')
C11 = Checkbutton(window, text="5900: L21 BD INT + L18 AC/ABCD  REC/INT   [5501]", variable=CheckVar11,command=cb_check11, onvalue=1, offvalue=0, height=1, width=38, anchor="w", bg='lavender',fg='navy')
C11_tip=CreateToolTip(C11,'[LTE2100 integration, antenna ports BD]  and/or  [LTE1800 reconfiguration or integration, antenna ports AC/ABCD 2t2r/4t4r], with multiband RRU5501.\nIn case one of the  bands is missing, it will execute only for the other.')
C12 = Checkbutton(window, text="5900: L18 AC/ABCD INT + L21 AB->BD  REC  [5501]", variable=CheckVar12,command=cb_check12, onvalue=1, offvalue=0, height=1, width=38, anchor="w", bg='lavender',fg='navy')
C12_tip=CreateToolTip(C12,'[LTE1800 integration, antenna ports AC/ABCD 2t2t/4t4r]  and/or  [LTE2100 reconfiguration, antenna ports AB to BD], with multiband RRU5501.\nIn case one of the  bands is missing, it will execute only for the other.')
C13 = Checkbutton(window, text="5900: L18 BD/ABCD INT + L21 AB->AC  REC  [5501]", variable=CheckVar13,command=cb_check13, onvalue=1, offvalue=0, height=1, width=38, anchor="w", bg='lavender',fg='navy')
C13_tip=CreateToolTip(C13,'[LTE1800 integration, antenna ports BD/ABCD 2t2t/4t4r]  and/or  [LTE2100 reconfiguration, antenna ports AB to AC], with multiband RRU5501.\nIn case one of the  bands is missing, it will execute only for the other.')
C14 = Checkbutton(window, text="5900: L26 AB/ABCD INT 64-65-66 SN 5", variable=CheckVar14, onvalue=1, offvalue=0, height=1, width=38,anchor="w", bg='lavender',fg='navy')
C14_tip=CreateToolTip(C14, "L2600 integration on slot 5/3.4.5 SRN 64-65-66, 2T2R or 4T4R")
C15 = Checkbutton(window, text="5900: L9 AB INT 80-82-84 ", variable=CheckVar15, onvalue=1, offvalue=0, height=1, width=38,anchor="w", bg='lavender',fg='navy')
C15_tip=CreateToolTip(C15,"L900 integration, .")
C16 = Checkbutton(window, text="BBU subrack ID 1 -> 0", variable=CheckVar16, command=cb_check16, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender',fg='navy')
#C17 = Checkbutton(window, text="                                                    ", variable=CheckVar17,state=DISABLED, onvalue=1, offvalue=0, height=1, width=38, anchor="w", bg='lavender')
C16_tip=CreateToolTip(C16, "Change the BBU Subrack ID from 1 -> 0")
C17 = Checkbutton(window, text="4G/5G: NE Type Reconst.BTS5900 LTE -> BTS5900", variable=CheckVar17,command=cb_check_17, onvalue=1, offvalue=0, height=1, width=38, anchor="w", bg='SlateGray1',fg='navy')
C17_tip=CreateToolTip(C17, "Convert BTS5900 LTE to BTS5900 as a prerequisite for 4G5G CoMPT")
#legacy options 3G/ not used in GUI
C18 = Checkbutton(window, text="3G Refarming UBBP ", variable=CheckVar18, command=cb_check18, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender')
C19 = Checkbutton(window, text="3G Refarming WBBP", variable=CheckVar19, command=cb_check19, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender')
C20 = Checkbutton(window, text="3G Refarming UBBP AC", variable=CheckVar20, command=cb_check20, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender')
C21 = Checkbutton(window, text="3G Refarming UBBP BD", variable=CheckVar21, command=cb_check21, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender')
C22 = Checkbutton(window, text="3G Refarming WBBP AC", variable=CheckVar22, command=cb_check22, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender')
C23 = Checkbutton(window, text="3G Refarming WBBP BD", variable=CheckVar23, command=cb_check23, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender')
C24 = Checkbutton(window, text="4G/5G: Old2New Transmission model convert(TBD)", variable=CheckVar24, command=cb_check23, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='SlateGray1',fg='navy')
# Ctest=Checkbutton(window, text="TEST", variable=CheckVar23,command=cb_check23, onvalue=1, offvalue=0,height=1, width=38, anchor="w", bg='lavender')

C1.grid(column=1, row=6)
C2.grid(column=1, row=7)
C300.grid(column=1, row=9)
C30.grid(column=1, row=10)
C3.grid(column=1, row=11)
C4.grid(column=1, row=12)
C5.grid(column=1, row=13)
C6.grid(column=1, row=14)
C7.grid(column=1, row=15)
#C80.grid(column=1, row=16)
#C8.grid(column=1, row=17)
C9.grid(column=1, row=18)
C10.grid(column=1, row=19)
C11.grid(column=1, row=20)
C12.grid(column=1, row=21)
C13.grid(column=1, row=22)
C14.grid(column=1, row=23)
C15.grid(column=1, row=24)
C16.grid(column=1, row=8)

C17.grid(column=1, row=25)
# C18.grid(column=1, row=26)
# C19.grid(column=1, row=27)
# C20.grid(column=1, row=28)
# C21.grid(column=1, row=29)
# C22.grid(column=1, row=30)
# C23.grid(column=1, row=31)
C24.grid(column=1, row=26)
def smart_agent():
    #old2new_basic_transmission_model_convertor()
    #window1=Tk()
    text=Text(window,height=3,width=130,font =("arial", 10),fg='green')

    Intelligent_message0='Existing Cells:  '
    Intelligent_message1='\nToDo: Integration  '
    Intelligent_message2=''
    intelligent_flag_1=0
    Label_h_result=CELL_SECTOREQM_mapping()
    Label_h_result_cells=Label_h_result[1]
    Label_h_result_sectoreqms = Label_h_result[2]

    if 'X.A,' in Label_h_result[0] and 'X.B,' in Label_h_result[0] and 'X.C,' in Label_h_result[0]:
        Intelligent_message0+= 'LTE800:ABC, '
    elif ('X.A,' in Label_h_result[0] and 'X.B,' in Label_h_result[0]) and ('X.C,' not in Label_h_result[0]):
        Intelligent_message0 += 'LTE800:AB, '
    elif ('X.A,' in Label_h_result[0]) and ('X.B,' not in Label_h_result[0] and 'X.C,' not in Label_h_result[0]):
        Intelligent_message0 += 'LTE800:A, '
    if 'X.D' in Label_h_result[0]:
        Intelligent_message0 += 'D, '

    if 'X.E,' in Label_h_result[0] and 'X.F,' in Label_h_result[0] and 'X.G,' in Label_h_result[0]:
        Intelligent_message0 += 'LTE2600:EFG, '
    elif ('X.E,' in Label_h_result[0] and 'X.F,' in Label_h_result[0]) and ('X.G,' not in Label_h_result[0]):
        Intelligent_message0 += 'LTE2600:EF, '
    elif ('X.E,' in Label_h_result[0]) and ('X.F,' not in Label_h_result[0] and 'X.G,' not in Label_h_result[0]):
        Intelligent_message0 += 'LTE2600:E, '

    if 'X.N,' in Label_h_result[0] and 'X.O,' in Label_h_result[0] and 'X.P,' in Label_h_result[0]:
        Intelligent_message0 += 'LTE2600:NOP, '
    elif ('X.N,' in Label_h_result[0] and 'X.O,' in Label_h_result[0]) and ('X.P,' not in Label_h_result[0]):
        Intelligent_message0 += 'LTE2600:NO, '
    elif ('X.N,' in Label_h_result[0]) and ('X.O,' not in Label_h_result[0] and 'X.P,' not in Label_h_result[0]):
        Intelligent_message0 += 'LTE2600:N, '

    if 'X.H,' in Label_h_result[0] and 'X.I,' in Label_h_result[0] and 'X.K,' in Label_h_result[0]:
        Intelligent_message0 += 'LTE1800:HIK, '
    elif ('X.H,' in Label_h_result[0] and 'X.I,' in Label_h_result[0]) and ('X.K,' not in Label_h_result[0]):
        Intelligent_message0 += 'LTE1800:HI, '
    elif ('X.H,' in Label_h_result[0]) and ('X.I,' not in Label_h_result[0] and 'X.K,' not in Label_h_result[0]):
        Intelligent_message0 += 'LTE1800:H, '

    if 'X.T,' in Label_h_result[0] and 'X.U,' in Label_h_result[0] and 'X.V,' in Label_h_result[0]:
        Intelligent_message0 += 'LTE2100:TUV, '
    elif ('X.T,' in Label_h_result[0] and 'X.U,' in Label_h_result[0]) and ('X.V,' not in Label_h_result[0]):
        Intelligent_message0 += 'LTE2100:TU, '
    elif ('X.T,' in Label_h_result[0]) and ('X.U,' not in Label_h_result[0] and 'X.V,' not in Label_h_result[0]):
        Intelligent_message0 += 'LTE2100:T, '

    if 'G.A,' in Label_h_result[0] and 'G.B,' in Label_h_result[0] and 'G.C,' in Label_h_result[0]:
        Intelligent_message0 += 'LTE900:ABC  '
    elif ('G.A,' in Label_h_result[0] and 'G.B,' in Label_h_result[0]) and ('G.C,' not in Label_h_result[0]):
        Intelligent_message0 += 'LTE900:AB, '
    elif ('G.A,' in Label_h_result[0]) and ('G.B,' not in Label_h_result[0] and 'G.C,' not in Label_h_result[0]):
        Intelligent_message0 += 'LTE900:A, '

    if 'P.A,' in Label_h_result[0] and 'P.B,' in Label_h_result[0] and 'P.C,' in Label_h_result[0]:
        Intelligent_message0 += 'LTE700:ABC '
    elif ('P.A,' in Label_h_result[0] and 'P.B,' in Label_h_result[0]) and ('P.C,' not in Label_h_result[0]):
        Intelligent_message0 += 'LTE700:AB, '
    elif ('P.A,' in Label_h_result[0]) and ('P.B,' not in Label_h_result[0] and 'P.C,' not in Label_h_result[0]):
        Intelligent_message0 += 'LTE700:A, '

    Umpt_ethport_sw_function_result=UMPT_ETHPORT_SW()
    if 'BBU5900' in Umpt_ethport_sw_function_result:
        C1.config(state=DISABLED)
    if 'SRN ID  1' in Umpt_ethport_sw_function_result:
        C2.config(state=DISABLED)
    if 'SRN ID  0' in Umpt_ethport_sw_function_result:
        C16.config(state=DISABLED)
    if '0' in Label_h_result_sectoreqms or '1' in Label_h_result_sectoreqms or '2' in Label_h_result_sectoreqms:
        C300.config(state=DISABLED)
        C30.config(state=DISABLED)
        #C3.config(state=DISABLED)
    if '7' in Label_h_result_sectoreqms or '8' in Label_h_result_sectoreqms or '9' in Label_h_result_sectoreqms:
        C7.config(state=DISABLED)
    if '18' in Label_h_result_sectoreqms or '19' in Label_h_result_sectoreqms or '20' in Label_h_result_sectoreqms:
        C9.config(state=DISABLED)
    if ('18' in Label_h_result_sectoreqms or '19' in Label_h_result_sectoreqms or '20' in Label_h_result_sectoreqms) and ('7' in Label_h_result_sectoreqms or '8' in Label_h_result_sectoreqms or '9' in Label_h_result_sectoreqms):
        C10.config(state=DISABLED)
        C11.config(state=DISABLED)
    if ('18' not in Label_h_result_sectoreqms and '19' not in Label_h_result_sectoreqms and '20' not in Label_h_result_sectoreqms) and ('7' in Label_h_result_sectoreqms or '8' in Label_h_result_sectoreqms or '9' in Label_h_result_sectoreqms):
        C12.config(state=DISABLED)
        C13.config(state=DISABLED)
    if '30' in Label_h_result_sectoreqms or '31' in Label_h_result_sectoreqms or '32' in Label_h_result_sectoreqms:
        C6.config(state=DISABLED)
    if 'P.A' in Label_h_result_cells or 'P.B' in Label_h_result_cells or 'P.C' in Label_h_result_cells:
        C300.config(state=DISABLED)
        C30.config(state=DISABLED)
    else:
        C6.config(state=DISABLED)

    if ('30' in Label_h_result_sectoreqms or '31' in Label_h_result_sectoreqms or '32' in Label_h_result_sectoreqms) and ('0' in Label_h_result_sectoreqms or '1' in Label_h_result_sectoreqms or '2' in Label_h_result_sectoreqms):
        C6.config(state=DISABLED)
        C4.config(state=DISABLED)
        C5.config(state=DISABLED)
    if '24' in Label_h_result_sectoreqms or '25' in Label_h_result_sectoreqms or '26' in Label_h_result_sectoreqms:
        C15.config(state=DISABLED)
    if '4' in Label_h_result_sectoreqms or '5' in Label_h_result_sectoreqms or '6' in Label_h_result_sectoreqms:
        C14.config(state=DISABLED)


    #disable GUI only by missing cells
    if 'X.A' not in Label_h_result_cells and 'X.B' not in Label_h_result_cells and 'X.C' not in Label_h_result_cells and 'X.D' not in Label_h_result_cells:
        C300.config(state=DISABLED)
        C30.config(state=DISABLED)
    if 'X.A' not in Label_h_result_cells and 'X.B' not in Label_h_result_cells and 'X.C' not in Label_h_result_cells and 'X.D' not in Label_h_result_cells and \
            'P.A' not in Label_h_result_cells and 'P.B' not in Label_h_result_cells and 'P.C' not in Label_h_result_cells:
        C4.config(state=DISABLED)
        C5.config(state=DISABLED)
    if 'G.A' not in Label_h_result_cells and 'G.B' not in Label_h_result_cells and 'G.C' not in Label_h_result_cells:
        C15.config(state=DISABLED)
    if 'P.A' not in Label_h_result_cells and 'P.B' not in Label_h_result_cells and 'P.C' not in Label_h_result_cells:
        C6.config(state=DISABLED)
    if 'X.E' not in Label_h_result_cells and 'X.F' not in Label_h_result_cells and 'X.G' not in Label_h_result_cells:
        C14.config(state=DISABLED)
    if 'X.T' not in Label_h_result_cells and 'X.U' not in Label_h_result_cells and 'X.V' not in Label_h_result_cells:
        C9.config(state=DISABLED)
    if 'X.H' not in Label_h_result_cells and 'X.I' not in Label_h_result_cells and 'X.K' not in Label_h_result_cells:
        C7.config(state=DISABLED)
    if ('X.H' not in Label_h_result_cells and 'X.I' not in Label_h_result_cells and 'X.K' not in Label_h_result_cells) and ('X.T' not in Label_h_result_cells and 'X.U' not in Label_h_result_cells and 'X.V' not in Label_h_result_cells):
        C9.config(state=DISABLED)
        C10.config(state=DISABLED)
        C11.config(state=DISABLED)
        C12.config(state=DISABLED)
        C13.config(state=DISABLED)
    #disable GUI by already existing Cells and SectorEqm
    if ('X.T' in Label_h_result_cells or 'X.U' in Label_h_result_cells or 'X.V' in Label_h_result_cells) and ('18' in Label_h_result_sectoreqms or '19' in Label_h_result_sectoreqms or '20' in Label_h_result_sectoreqms):
        C10.config(state=DISABLED)
        C11.config(state=DISABLED)
        # C12.config(state=DISABLED)
        # C13.config(state=DISABLED)

    if ('X.H' in Label_h_result_cells or 'X.I' in Label_h_result_cells or 'X.K' in Label_h_result_cells) and ('7' in Label_h_result_sectoreqms or '8' in Label_h_result_sectoreqms or '9' in Label_h_result_sectoreqms):
        # C10.config(state=DISABLED)
        # C11.config(state=DISABLED)
        C12.config(state=DISABLED)
        C13.config(state=DISABLED)
    if ('X.A' in Label_h_result_cells or 'X.B' in Label_h_result_cells or 'X.C' in Label_h_result_cells) and ('0' in Label_h_result_sectoreqms or '1' in Label_h_result_sectoreqms or '2' in Label_h_result_sectoreqms) and 'BBU5900' in Umpt_ethport_sw_function_result:
        C3.config(state=DISABLED)
    if ('X.A' in Label_h_result_cells or 'X.B' in Label_h_result_cells or 'X.C' in Label_h_result_cells) and ('0' in Label_h_result_sectoreqms or '1' in Label_h_result_sectoreqms or '2' in Label_h_result_sectoreqms) and 'BBU5900' not in Umpt_ethport_sw_function_result:
        Intelligent_message2+=' ❗ L800 reconfiguration needed.'

    #disable GUI by missing SectorEqm
    if ('X.A' in Label_h_result_cells or 'X.B' in Label_h_result_cells or 'X.C' in Label_h_result_cells) and '0' not in Label_h_result_sectoreqms:
        Intelligent_message1+='L800, '
        intelligent_flag_1=1
    if ('P.A' in Label_h_result_cells or 'P.B' in Label_h_result_cells or 'P.C' in Label_h_result_cells) and '30' not in Label_h_result_sectoreqms:
        Intelligent_message1+='L700, '
        intelligent_flag_1 = 1
    if ('G.A' in Label_h_result_cells or 'G.B' in Label_h_result_cells or 'G.C' in Label_h_result_cells) and '24' not in Label_h_result_sectoreqms:
        Intelligent_message1+='L900, '
        intelligent_flag_1 = 1
    if ('X.H' in Label_h_result_cells or 'X.I' in Label_h_result_cells or 'X.K' in Label_h_result_cells) and '7' not in Label_h_result_sectoreqms:
        Intelligent_message1+='L1800, '
        intelligent_flag_1 = 1
    if ('X.E' in Label_h_result_cells or 'X.F' in Label_h_result_cells or 'X.G' in Label_h_result_cells) and '4' not in Label_h_result_sectoreqms:
        Intelligent_message1+='L2600, '
        intelligent_flag_1 = 1
    if ('X.T' in Label_h_result_cells or 'X.U' in Label_h_result_cells or 'X.V' in Label_h_result_cells) and '18' not in Label_h_result_sectoreqms:
        Intelligent_message1+='L2100, '
        intelligent_flag_1 = 1
    if 'BBU5900' not in Umpt_ethport_sw_function_result:
        Intelligent_message1+=' ❗ BBU5900 reconstruction needed.'
    if intelligent_flag_1==0:
        Intelligent_message1='\nToDo: No new integration scenario possible! Verify remaining options, or CLEAR screen and change file.'
        if ('X.A' in Label_h_result_cells or 'X.B' in Label_h_result_cells or 'X.C' in Label_h_result_cells) and ('BBU3900' in Umpt_ethport_sw_function_result or 'BBU3910' in Umpt_ethport_sw_function_result):
            Intelligent_message1+=' ❗ BBU5900 reconstruction + L800 Reconfiguration possible.'

    text.insert(INSERT,Intelligent_message0+'\n'+Label_h_result[0].split('\n')[1]+Intelligent_message1+Intelligent_message2,)
    text.grid(column=2, row=3)
    #textBox = Label(window, text=Intelligent_message0+'\n'+Label_h_result[0].split('\n')[1]+Intelligent_message1, font='arial 11 bold').grid(column=2, row=3)


#########################################################################################################################
enodeB_name = []
BTS3900_to_BTS5900_reconstruction = []
BBU_subrack_ID_0_to_1 = []
BBU_subrack_ID_1_to_0 = []
LMPT_to_UMPT = []
L800_CD_INT_NO_700 = []
L800_AB_REC_INT_60_61_62_SN_2_3_to_4 = []
L800_CD_REC_L700_AB_INT_80_82_84 = []
L800_AB_REC_L700_CD_INT_80_82_84 = []
L700_INT_90_91_92_SN_5 = []
L1800_INT_ABCD = []
L1800_INT_AC = []
L1800_INT_AB = []
L2100_INT_AB = []
L2100_AC_INT_L1800_BD_REC_INT = []
L2100_BD_INT_L1800_AC_REC_INT = []
L1800_AC_INT_L2100_AB_BD_REC = []
L1800_BD_INT_L2100_AB_AC_REC = []
LTE_2600_AB_INT = []
LTE900_INT = []
Refarming_3G_UBBP = []
Refarming_3G_WBBP = []
Refarming_3G_UBBP_AC = []
Refarming_3G_UBBP_BD = []
Refarming_3G_WBBP_AC = []
Refarming_3G_WBBP_BD = []

def bulk_step1():
    from tkinter import filedialog
    import openpyxl
    from glob import glob
    import xml.etree.ElementTree as ET

    folder_selected = filedialog.askdirectory()
    # print(folder_selected)

    xml_folder_file_list = glob(folder_selected + "/*.xml")
    print(xml_folder_file_list)
    for file in xml_folder_file_list:
        print(file)

    wb1 = openpyxl.load_workbook('NE.xlsx', data_only=True)
    sheet1 = wb1['Sheet1']
    for i in range(2, len(sheet1['1']) + 1):
        enodeB_name.append(sheet1.cell(row=1, column=i).value)
        BTS3900_to_BTS5900_reconstruction.append(sheet1.cell(row=2, column=i).value)
        BBU_subrack_ID_0_to_1.append(sheet1.cell(row=3, column=i).value)
        BBU_subrack_ID_1_to_0.append(sheet1.cell(row=4, column=i).value)
        LMPT_to_UMPT.append(sheet1.cell(row=5, column=i).value)
        L800_CD_INT_NO_700.append(sheet1.cell(row=6, column=i).value)
        L800_AB_REC_INT_60_61_62_SN_2_3_to_4.append(sheet1.cell(row=7, column=i).value)
        L800_CD_REC_L700_AB_INT_80_82_84.append(sheet1.cell(row=8, column=i).value)
        L800_AB_REC_L700_CD_INT_80_82_84.append(sheet1.cell(row=9, column=i).value)
        L700_INT_90_91_92_SN_5.append(sheet1.cell(row=10, column=i).value)
        L1800_INT_ABCD.append(sheet1.cell(row=11, column=i).value)
        # L1800_INT_AC.append(sheet1.cell(row=12, column=i).value)
        # L1800_INT_AB.append(sheet1.cell(row=13, column=i).value)
        L2100_INT_AB.append(sheet1.cell(row=12, column=i).value)
        L2100_AC_INT_L1800_BD_REC_INT.append(sheet1.cell(row=13, column=i).value)
        L2100_BD_INT_L1800_AC_REC_INT.append(sheet1.cell(row=14, column=i).value)
        L1800_AC_INT_L2100_AB_BD_REC.append(sheet1.cell(row=15, column=i).value)
        L1800_BD_INT_L2100_AB_AC_REC.append(sheet1.cell(row=16, column=i).value)
        LTE_2600_AB_INT.append(sheet1.cell(row=17, column=i).value)
        LTE900_INT.append(sheet1.cell(row=18, column=i).value)
        Refarming_3G_UBBP.append(sheet1.cell(row=19, column=i).value)
        Refarming_3G_WBBP.append(sheet1.cell(row=20, column=i).value)
        Refarming_3G_UBBP_AC.append(sheet1.cell(row=21, column=i).value)
        Refarming_3G_UBBP_BD.append(sheet1.cell(row=22, column=i).value)
        Refarming_3G_WBBP_AC.append(sheet1.cell(row=23, column=i).value)
        Refarming_3G_WBBP_BD.append(sheet1.cell(row=24, column=i).value)
    dict_BTS3900_to_BTS5900_reconstruction = dict(zip(enodeB_name, BTS3900_to_BTS5900_reconstruction))
    dict_BBU_subrack_ID_0_to_1 = dict(zip(enodeB_name, BBU_subrack_ID_0_to_1))
    dict_BBU_subrack_ID_1_to_0 = dict(zip(enodeB_name, BBU_subrack_ID_1_to_0))
    dict_LMPT_to_UMPT = dict(zip(enodeB_name, LMPT_to_UMPT))
    dict_L800_CD_INT_NO_700 = dict(zip(enodeB_name, L800_CD_INT_NO_700))
    dict_L800_AB_REC_INT_60_61_62_SN_2_3_to_4 = dict(zip(enodeB_name, L800_AB_REC_INT_60_61_62_SN_2_3_to_4))
    dict_L800_CD_REC_L700_AB_INT_80_82_84 = dict(zip(enodeB_name, L800_CD_REC_L700_AB_INT_80_82_84))
    dict_L800_AB_REC_L700_CD_INT_80_82_84 = dict(zip(enodeB_name, L800_AB_REC_L700_CD_INT_80_82_84))
    dict_L700_INT_90_91_92_SN_5 = dict(zip(enodeB_name, L700_INT_90_91_92_SN_5))
    dict_L1800_INT_ABCD = dict(zip(enodeB_name, L1800_INT_ABCD))
    # dict_L1800_INT_AC = dict(zip(enodeB_name, L1800_INT_AC))
    # dict_L1800_INT_AB = dict(zip(enodeB_name, L1800_INT_AB))
    dict_L2100_INT_AB = dict(zip(enodeB_name, L2100_INT_AB))
    dict_L2100_AC_INT_L1800_BD_REC_INT = dict(zip(enodeB_name, L2100_AC_INT_L1800_BD_REC_INT))
    dict_L2100_BD_INT_L1800_AC_REC_INT = dict(zip(enodeB_name, L2100_BD_INT_L1800_AC_REC_INT))
    dict_L1800_AC_INT_L2100_AB_BD_REC = dict(zip(enodeB_name, L1800_AC_INT_L2100_AB_BD_REC))
    dict_L1800_BD_INT_L2100_AB_AC_REC = dict(zip(enodeB_name, L1800_BD_INT_L2100_AB_AC_REC))
    dict_LTE_2600_AB_INT = dict(zip(enodeB_name, LTE_2600_AB_INT))
    dict_LTE900_INT = dict(zip(enodeB_name, LTE900_INT))
    dict_Refarming_3G_UBBP = dict(zip(enodeB_name, Refarming_3G_UBBP))
    dict_Refarming_3G_WBBP = dict(zip(enodeB_name, Refarming_3G_WBBP))
    dict_Refarming_3G_UBBP_AC = dict(zip(enodeB_name, Refarming_3G_UBBP_AC))
    dict_Refarming_3G_UBBP_BD = dict(zip(enodeB_name, Refarming_3G_UBBP_BD))
    dict_Refarming_3G_WBBP_AC = dict(zip(enodeB_name, Refarming_3G_WBBP_AC))
    dict_Refarming_3G_WBBP_BD = dict(zip(enodeB_name, Refarming_3G_WBBP_BD))

    # print(enodeB_name)
    # print(dict_BTS3900_to_BTS5900_reconstruction)
    # print(dict_BBU_subrack_ID_0_to_1)
    # print(dict_BBU_subrack_ID_1_to_0)
    # print(dict_LMPT_to_UMPT)
    # print(dict_L800_CD_INT_NO_700)
    # print(dict_L800_AB_REC_INT_60_61_62_SN_2_3_to_4)
    # print(dict_L800_CD_REC_L700_AB_INT_80_82_84)
    # print(dict_L800_AB_REC_L700_CD_INT_80_82_84)
    # print(dict_L700_INT_90_91_92_SN_5)
    # print(dict_L1800_INT_ABCD)
    # print(dict_L1800_INT_AC)
    # print(dict_L1800_INT_AB)
    # print(dict_L2100_INT_AB)
    # print(dict_L2100_AC_INT_L1800_BD_REC_INT)
    # print(dict_L2100_BD_INT_L1800_AC_REC_INT)
    # print(dict_L1800_AC_INT_L2100_AB_BD_REC)
    # print(dict_L1800_BD_INT_L2100_AB_AC_REC)
    # print(dict_LTE_2600_AB_INT)
    # print(dict_LTE900_INT)
    print(dict_Refarming_3G_UBBP)
    print(dict_Refarming_3G_WBBP)
    print(dict_Refarming_3G_UBBP_AC)
    print(dict_Refarming_3G_UBBP_BD)
    print(dict_Refarming_3G_WBBP_AC)
    print(dict_Refarming_3G_WBBP_BD)

    wb1.close()

    return [xml_folder_file_list,enodeB_name,dict_BTS3900_to_BTS5900_reconstruction,dict_BBU_subrack_ID_0_to_1,\
            dict_BBU_subrack_ID_1_to_0,dict_LMPT_to_UMPT,dict_L800_CD_INT_NO_700,dict_L800_AB_REC_INT_60_61_62_SN_2_3_to_4,dict_L800_CD_REC_L700_AB_INT_80_82_84,\
            dict_L800_AB_REC_L700_CD_INT_80_82_84,dict_L700_INT_90_91_92_SN_5,dict_L1800_INT_ABCD,dict_L2100_INT_AB,\
            dict_L2100_AC_INT_L1800_BD_REC_INT,dict_L2100_BD_INT_L1800_AC_REC_INT,dict_L1800_AC_INT_L2100_AB_BD_REC,dict_L1800_BD_INT_L2100_AB_AC_REC,\
            dict_LTE_2600_AB_INT,dict_LTE900_INT,dict_Refarming_3G_UBBP,dict_Refarming_3G_WBBP,dict_Refarming_3G_UBBP_AC,dict_Refarming_3G_UBBP_BD,\
            dict_Refarming_3G_WBBP_AC,dict_Refarming_3G_WBBP_BD]
#########################################################################################################################

#define selection variables for Bulk configuration mode; initialize with 0 value/
CheckkVar1, CheckkVar2, CheckkVar300, CheckkVar30, CheckkVar3, CheckkVar4, CheckkVar5,\
CheckkVar6, CheckkVar7, CheckkVar9,\
CheckkVar10, CheckkVar11, CheckkVar12, CheckkVar13, CheckkVar14, CheckkVar15, CheckkVar16,\
CheckkVar17, CheckkVar18, CheckkVar19, CheckkVar20, CheckkVar21, CheckkVar22, CheckkVar23 =\
0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0

#define execution function for single and bulk configuration modes
def executor_function():
    # CheckkVar1, CheckkVar2, CheckkVar300, CheckkVar30, CheckkVar3, CheckkVar4, CheckkVar5, CheckkVar6, CheckkVar7, CheckkVar80, CheckkVar8, CheckkVar9, CheckkVar10, CheckkVar11, CheckkVar12, CheckkVar13, CheckkVar14, CheckkVar15, CheckkVar16, CheckkVar17, CheckkVar18, CheckkVar19, CheckkVar20, CheckkVar21, CheckkVar22, CheckkVar23 =\
    # 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0
    start = time.time()
    global NEname
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'NE':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'LOCATION':
                                NEname = child4.text
                                # ET.dump(child1)
                                break
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'SUBRACK':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'SRN':
                                Subrack_BBU_initial = child4.text

    if CheckVar2.get() == 1 or CheckkVar2==1:
        Subrack_global = str(1)
    if CheckVar16.get() == 1 or CheckkVar16==1:
        Subrack_global = str(0)
    if CheckVar2.get() == 0 and CheckVar16.get() == 0 and CheckkVar2==0 and CheckkVar16==0:
        Subrack_global = str(Subrack_BBU_initial)

    #########################################################################################################################################################################################
    #########################################################################################################################################################################################
    # BTS3900 to BTS5900
    if CheckVar1.get() == 1 or CheckkVar1==1:
        #print('has entered in reconstruction!')
        PEU_var1 = parse_existing_MO_name(root, 'PEU')
        PEU_list_unique = []
        #Label(window, text='BBU type change executed..', width=80, fg='green').grid(column=2, row=6)
        subrack_type_change_3900to5900()
        rmv_old_UEIU()
        for i in PEU_var1:
            for j in i:
                for j in range(0, len(PEU_var1)):
                    for item in PEU_var1[j]:
                        for key in item.keys():
                            if key == 'SN':
                                if item[key] not in PEU_list_unique:
                                    PEU_list_unique.append(item[key])
        fo = open('LOG.txt', 'a')
        fo.write(f'{dt.datetime.now()} --  {NEname}  {NEname}' + 'BBU type change executed..\n')
        if '18' not in PEU_list_unique:
            insert_new_UPEU(0, Subrack_global, 18)
        text = Text(window, height=1, width=130, font=("arial", 10))
        text.insert(INSERT, 'BBU type change executed..')
        text.grid(column=2, row=6)
    #########################################################################################################################################################################################
    # BBU subrack no 0 to 1
    if CheckVar2.get() or CheckkVar2 == 1:
        #Label(window, text='BBU subrack ID change 0->1 executed..', width=80, fg='green').grid(column=2, row=7)
        change_bbu_subrack_id()
        change_basebandeqm_subrack_id()  # only for 3G
        fo = open('LOG.txt', 'a')
        fo.write(f'{dt.datetime.now()} --  {NEname}  ' + 'BBU subrack ID change 0->1 executed..\n')
        text = Text(window, height=1, width=130, font=("arial", 10))
        text.insert(INSERT, 'BBU subrack ID change 0->1 executed..')
        text.grid(column=2, row=7)
    #########################################################################################################################################################################################
    # BTS5900: LTE800 80-82-84 AB INTEGRATED , no LTE700  .. [5519et]
    if CheckVar300.get() == 1 or CheckkVar300 == 1:
        #Label(window, text='BTS5900: LTE800 80-82-84 AB INTEGRATED , no LTE700  .. [5519et] ', width=80, fg='green').grid(column=2, row=9)
        if CheckVar300.get() == 1 or CheckkVar300 == 1:
            BBP300 = parse_existing_MO_name(root, 'BBP')
            SECTOR_var300 = parse_existing_MO_name(root, 'SECTOR')
            Chain_var300 = parse_existing_MO_name(root, 'RRUCHAIN')
            Cell_var300 = parse_existing_MO_name(root, 'Cell')
            # print(BBP1)
            BBP_list300_unique = []
            SECTOR_list300_unique = []
            Chain_list300_unique = []
            Cell_band_800_list300 = []
            Cell_list_list300 = []

            for i in BBP300:
                for j in i:
                    for j in range(0, len(BBP300)):
                        for item in BBP300[j]:
                            for key in item.keys():
                                if key == 'SN':
                                    if item[key] not in BBP_list300_unique:
                                        BBP_list300_unique.append(item[key])
            for i in Cell_var300:
                for j in i:
                    for j in range(0, len(Cell_var300)):
                        for item in Cell_var300[j]:
                            for key in item.keys():
                                if key == 'LocalCellId':
                                    if item[key] not in Cell_list_list300:
                                        Cell_list_list300.append(item[key])

            if CheckVar14.get() == 1 or CheckkVar14 == 1:
                if '4' in Cell_list_list300:
                    insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 64, 0, 4, 8, 15, 4, 4)
                if '5' in Cell_list_list300:
                    insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 65, 0, 5, 8, 15, 4, 4)
                if '6' in Cell_list_list300:
                    insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 66, 0, 6, 8, 15, 4, 4)

            if '4' not in BBP_list300_unique:
                insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, Subrack_global, 4, 2, 20)
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(0, 0, Subrack_global, 4, 0)
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(1, 0, Subrack_global, 4, 1)
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(2, 0, Subrack_global, 4, 2)
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 80, 0, 0, 328, 15, 4, 4)
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 82, 0, 1, 328, 15, 4, 4)
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 84, 0, 2, 328, 15, 4, 4)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 80, 0, [0, 1,2,3])
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 82, 0, [0, 1,2,3])
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 0, 84, 0, [0, 1,2,3])
            insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 80, 0, [0, 1])
            insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 82, 0, [0, 1])
            insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 0, 84, 0, [0, 1])
            #Label(window, text='BTS5900: LTE800 80-82-84 AB INTEGRATED , no LTE700  .. [5519et]', width=80,fg='green').grid(column=2, row=9)
            fo = open('LOG.txt', 'a')
            fo.write(f'{dt.datetime.now()} --  {NEname}  ' + 'BTS5900: LTE800 80-82-84 AB INTEGRATED , no LTE700  .. [5519et]\n')
            text = Text(window, height=1, width=130, font=("arial", 10))
            text.insert(INSERT, 'BTS5900: LTE800 80-82-84 AB INTEGRATED , no LTE700  .. [5519et] ')
            text.grid(column=2, row=9)
            #########################################################################################################################################################################################
    # BTS5900: LTE800 80-82-84 CD INTEGRATED , no LTE700  .. [5509t]
    if CheckVar30.get() == 1 or CheckkVar30 == 1:
        BBP30 = parse_existing_MO_name(root, 'BBP')
        SECTOR_var30 = parse_existing_MO_name(root, 'SECTOR')
        Chain_var30 = parse_existing_MO_name(root, 'RRUCHAIN')
        Cell_var30 = parse_existing_MO_name(root, 'Cell')
        # print(BBP1)
        BBP_list30_unique = []
        SECTOR_list30_unique = []
        Chain_list30_unique = []
        Cell_band_800_list30 = []
        Cell_list_list30=[]

        for i in BBP30:
            for j in i:
                for j in range(0, len(BBP30)):
                    for item in BBP30[j]:
                        for key in item.keys():
                            if key == 'SN':
                                if item[key] not in BBP_list30_unique:
                                    BBP_list30_unique.append(item[key])
        for i in Cell_var30:
            for j in i:
                for j in range(0, len(Cell_var30)):
                    for item in Cell_var30[j]:
                        for key in item.keys():
                            if key == 'LocalCellId':
                                if item[key] not in Cell_list_list30:
                                    Cell_list_list30.append(item[key])

        if CheckVar14.get() == 1 or CheckkVar14==1:
            if '4' in Cell_list_list30:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 64, 0, 4, 8, 15, 4, 4)
            if '5' in Cell_list_list30:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 65, 0, 5, 8, 15, 4, 4)
            if '6' in Cell_list_list30:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 66, 0, 6, 8, 15, 4, 4)

        if '4' not in BBP_list30_unique:
            insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, Subrack_global, 4, 2, 20)
        insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(0, 0, Subrack_global, 4, 0)
        insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(1, 0, Subrack_global, 4, 1)
        insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(2, 0, Subrack_global, 4, 2)
        insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 80, 0, 0, 328, 15, 4, 4)
        insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 82, 0, 1, 328, 15, 4, 4)
        insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 84, 0, 2, 328, 15, 4, 4)
        insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 80, 0, [0,1,2, 3])
        insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 82, 0, [0,1,2, 3])
        insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 0, 84, 0, [0,1,2, 3])
        insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 80, 0, [2, 3])
        insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 82, 0, [2, 3])
        insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 0, 84, 0, [2, 3])
        #Label(window, text='BTS5900: LTE800 80-82-84 CD INTEGRATED , no LTE700  .. [5509t]', width=80, fg='green').grid(column=2,row=10)
        fo = open('LOG.txt', 'a')
        fo.write(f'{dt.datetime.now()} --  {NEname}  ' + 'BTS5900: LTE800 80-82-84 INTEGRATED, no LTE700  ..\n')
        text = Text(window, height=1, width=130, font=("arial", 10))
        text.insert(INSERT, 'BTS5900: LTE800 80-82-84 CD INTEGRATED , no LTE700  .. [5509t]')
        text.grid(column=2, row=10)
    #########################################################################################################################################################################################
    # BTS5900: integrate/reconfigure  LTE800 60-61-62 AB 2,3->4
    if CheckVar3.get() == 1 or CheckkVar3 == 1:
        BBP3 = parse_existing_MO_name(root, 'BBP')
        SECTOR_var3 = parse_existing_MO_name(root, 'SECTOR')
        SECTOREQM_var3=parse_existing_MO_name(root, 'SECTOREQM')
        Chain_var3 = parse_existing_MO_name(root, 'RRUCHAIN')
        Cell_var3 = parse_existing_MO_name(root, 'Cell')
        # print(BBP1)
        BBP_list3_unique = []
        SECTOR_list3_unique = []
        SECTOREQM_list3_unique=[]
        Chain_list3_unique = []
        Cell_band_800_list3 = []
        Cell_band_700_list3 = []
        Cell_band_900_list3 = []
        Cell_band_1800_2100_list3 = []
        Cell_band_2600_list3 = []

        for i in BBP3:
            for j in i:
                for j in range(0, len(BBP3)):
                    for item in BBP3[j]:
                        for key in item.keys():
                            if key == 'SN':
                                if item[key] not in BBP_list3_unique:
                                    BBP_list3_unique.append(item[key])
        for i in SECTOR_var3:
            for j in i:
                for j in range(0, len(SECTOR_var3)):
                    for item in SECTOR_var3[j]:
                        for key in item.keys():
                            if key == 'SECTORID':
                                if item[key] not in SECTOR_list3_unique:
                                    SECTOR_list3_unique.append(item[key])
        for i in SECTOREQM_var3:
            for j in i:
                for j in range(0, len(SECTOREQM_var3)):
                    for item in SECTOREQM_var3[j]:
                        for key in item.keys():
                            if key == 'SECTOREQMID':
                                if item[key] not in SECTOREQM_list3_unique:
                                    SECTOREQM_list3_unique.append(item[key])
        for i in Chain_var3:
            for j in i:
                for j in range(0, len(Chain_var3)):
                    for item in Chain_var3[j]:
                        for key in item.keys():
                            if key == 'RCN':
                                if item[key] not in Chain_list3_unique:
                                    Chain_list3_unique.append(item[key])
        cnt = 0
        for i in Cell_var3:
            for j in i:
                for j in range(0, len(Cell_var3)):
                    for item in Cell_var3[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['0', '1', '2']:
                                if item[key] not in Cell_band_800_list3:
                                    Cell_band_800_list3.append(item[key])
        for i in Cell_var3:
            for j in i:
                for j in range(0, len(Cell_var3)):
                    for item in Cell_var3[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['24', '25', '26']:
                                if item[key] not in Cell_band_900_list3:
                                    Cell_band_900_list3.append(item[key])
        for i in Cell_var3:
            for j in i:
                for j in range(0, len(Cell_var3)):
                    for item in Cell_var3[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['7', '8', '9', '18', '19', '20','21','22','23']:
                                if item[key] not in Cell_band_1800_2100_list3:
                                    Cell_band_1800_2100_list3.append(item[key])
        for i in Cell_var3:
            for j in i:
                for j in range(0, len(Cell_var3)):
                    for item in Cell_var3[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['30', '31', '32']:
                                if item[key] not in Cell_band_700_list3:
                                    Cell_band_700_list3.append(item[key])
        for i in Cell_var3:
            for j in i:
                for j in range(0, len(Cell_var3)):
                    for item in Cell_var3[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['4', '5', '6']:
                                if item[key] not in Cell_band_2600_list3:
                                    Cell_band_2600_list3.append(item[key])
        # print(BBP_list3_unique)
        # print(SECTOR_list3_unique)
        # print(Chain_list3_unique)
        # print(Cell_band_800_list3)
        if '4' not in BBP_list3_unique:
            insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, Subrack_global, 4, 2, 20)

        # SN 5 addition is the case when lte800 is 60-61-62, we move  to slot 4, but there is also lte700 in slot 5, we need to keep slot 5
        # in corelation with 700 integration option to be selected in future lines , for which adding slot 5 BBP might fail!!!
        if ('30' in Cell_band_700_list3 or '31' in Cell_band_700_list3 or '32' in Cell_band_700_list3) and '5' not in BBP_list3_unique:
            # insert_only_new_BBP_CN_SRN_SN_TYPE_BBWS(0, Subrack_global, 5, 2, 4)
            insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, Subrack_global, 5, 2, 4)
        if '2' in BBP_list3_unique and '3' not in BBP_list3_unique:
            if '4' in Cell_band_2600_list3 or '5' in Cell_band_2600_list3 or '6' in Cell_band_2600_list3:
                pass
                # there is LTE2600 on site
            # we leave board 2 if 1800/2100 on site and no board 3 exists ( in order to pe possible to add later board 3)
            else:
                if '7' in Cell_band_1800_2100_list3 or '8' in Cell_band_1800_2100_list3 or '9' in Cell_band_1800_2100_list3 or '18' in Cell_band_1800_2100_list3 or '19' in Cell_band_1800_2100_list3 or '20' in Cell_band_1800_2100_list3:

                    pass
                else:
                    rmv_old_bbp_from_slotnumber(2)

        if '3' in BBP_list3_unique:
            if '7' in Cell_band_1800_2100_list3 or '8' in Cell_band_1800_2100_list3 or '9' in Cell_band_1800_2100_list3 or '18' in Cell_band_1800_2100_list3 or '19' in Cell_band_1800_2100_list3 or '20' in Cell_band_1800_2100_list3:
                mod_BBP_TYPE_from_slot_to_slot(3, 3, 12290, 2)
                mod_BBP_TYPE_from_slot_to_slot(2, 2, 12290, 2)

                # there is LTE2100/Lte1800 on brd 3 on site
            else:
                rmv_old_bbp_from_slotnumber(3)

        #if CheckVar7.get() == 1 or CheckVar80.get() == 1:
        if CheckVar7.get() == 1 or CheckkVar7==1:
            if '7' in Cell_band_1800_2100_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 67, 0, 7, 8, 15, 4, 4)
            if '8' in Cell_band_1800_2100_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 68, 0, 8, 8, 15, 4, 4)
            if '9' in Cell_band_1800_2100_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 69, 0, 9, 8, 15, 4, 4)

        # if CheckVar8.get() == 1:
        #     if '7' in Cell_band_1800_2100_list3:
        #         insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 67, 0, 7, 8, 15, 2, 2)
        #     if '8' in Cell_band_1800_2100_list3:
        #         insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 68, 0, 8, 8, 15, 2, 2)
        #     if '9' in Cell_band_1800_2100_list3:
        #         insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 69, 0, 9, 8, 15, 2, 2)

        if CheckVar10.get() == 1 or CheckVar11.get() == 1 or CheckkVar10==1 or CheckkVar11==1:
            if '18' in Cell_band_1800_2100_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 18, 136, 15, 4, 4)
            if '19' in Cell_band_1800_2100_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 19, 136, 15, 4, 4)
            if '20' in Cell_band_1800_2100_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 20, 136, 15, 4, 4)

        if CheckVar9.get() == 1 or CheckkVar9==1:
            if '18' in Cell_band_1800_2100_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 18, 136, 15, 2, 2)
            if '19' in Cell_band_1800_2100_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 19, 136, 15, 2, 2)
            if '20' in Cell_band_1800_2100_list3 and '18' in Cell_band_1800_2100_list3 and '19' in Cell_band_1800_2100_list3 :
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 20, 136, 15, 2, 2)
            if '20' in Cell_band_1800_2100_list3 and '18' not in Cell_band_1800_2100_list3 and '19' not in Cell_band_1800_2100_list3 :
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 20, 136, 15, 2, 2)
            # if '21' in Cell_band_1800_2100_list3 and '18' in Cell_band_1800_2100_list3 and '19' in Cell_band_1800_2100_list3 :
            #     insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 103, 0, 21, 136, 15, 2, 2)
            if '21' in Cell_band_1800_2100_list3 and '18' not in Cell_band_1800_2100_list3 and '19' not in Cell_band_1800_2100_list3 and '20' in Cell_band_1800_2100_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 21, 136, 15, 2, 2)
            if '21' in Cell_band_1800_2100_list3 and '18' not in Cell_band_1800_2100_list3 and '19' not in Cell_band_1800_2100_list3 and '20' not in Cell_band_1800_2100_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 21, 136, 15, 2, 2)
            # if '22' in Cell_band_1800_2100_list3 and '18' in Cell_band_1800_2100_list3 and '19' in Cell_band_1800_2100_list3 and  '20' in Cell_band_1800_2100_list3 and '21' in Cell_band_1800_2100_list3:
            #     insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 104, 0, 22, 136, 15, 2, 2)
            if '22' in Cell_band_1800_2100_list3 and '18' not in Cell_band_1800_2100_list3 and '19' not in Cell_band_1800_2100_list3 and  '20' not in Cell_band_1800_2100_list3 and '21' not in Cell_band_1800_2100_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 22, 136, 15, 2, 2)
            if '22' in Cell_band_1800_2100_list3 and '18' not in Cell_band_1800_2100_list3 and '19' not in Cell_band_1800_2100_list3 and  '20' not in Cell_band_1800_2100_list3 and '21' in Cell_band_1800_2100_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 22, 136, 15, 2, 2)
            if '23' in Cell_band_1800_2100_list3 and '18' not in Cell_band_1800_2100_list3 and '19' not in Cell_band_1800_2100_list3 and  '20' not in Cell_band_1800_2100_list3 and '21' not in Cell_band_1800_2100_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 23, 136, 15, 2, 2)
            if '23' in Cell_band_1800_2100_list3 and '18' not in Cell_band_1800_2100_list3 and '19' not in Cell_band_1800_2100_list3 and  '20' not in Cell_band_1800_2100_list3 and '21'  in Cell_band_1800_2100_list3 and '22'  in Cell_band_1800_2100_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 23, 136, 15, 2, 2)


        if CheckVar14.get() == 1 or CheckkVar14==1:
            if '4' in Cell_band_2600_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 64, 0, 4, 8, 15, 4, 4)
            if '5' in Cell_band_2600_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 65, 0, 5, 8, 15, 4, 4)
            if '6' in Cell_band_2600_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 66, 0, 6, 8, 15, 4, 4)

        if CheckVar15.get() == 1 or CheckkVar15==1:
            if '24' in Cell_band_900_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 80, 0, 24, 8, 15, 2, 2)
            if '25' in Cell_band_900_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 82, 0, 25, 8, 15, 2, 2)
            if '26' in Cell_band_900_list3:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 84, 0, 26, 8, 15, 2, 2)

        if '0' in SECTOREQM_list3_unique and ( (('0' in Chain_list3_unique and '2' in BBP_list3_unique) or ('1' in Chain_list3_unique and '2' in BBP_list3_unique) or ('2' in Chain_list3_unique and '2' in BBP_list3_unique))):
            mod_RRUCHUAIN_chainno_oldHSN_newHSN(0, 2, 4)
            mod_RRUCHUAIN_chainno_oldHSN_newHSN(1, 2, 4)
            mod_RRUCHUAIN_chainno_oldHSN_newHSN(2, 2, 4)
            mod_LRRU_to_MRRU_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(60, 0, 60, 0, 2, 2, 264,15)
            mod_LRRU_to_MRRU_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(61, 0, 61, 0, 2, 2, 264,15)
            mod_LRRU_to_MRRU_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(62, 0, 62, 0, 2, 2, 264,15)
            #Label(window, text='BTS5900: LTE800 60-61-62 AB RECONFIGURED, moved to SN 4 ..', width=80, fg='green').grid(column=2, row=11)
            fo = open('LOG.txt', 'a')
            fo.write(f'{dt.datetime.now()} --  {NEname}  ' + 'BTS5900: LTE800 60-61-62 AB RECONFIGURED, moved to SN 4 ..\n')
            text = Text(window, height=1, width=130, font=("arial", 10))
            text.insert(INSERT, 'BTS5900: LTE800 60-61-62 AB RECONFIGURED, moved to SN 4 ..')
            text.grid(column=2, row=11)
        elif '0' in SECTOREQM_list3_unique and ((('0' in Chain_list3_unique and '3' in BBP_list3_unique) or ('1' in Chain_list3_unique and '3' in BBP_list3_unique) or ('2' in Chain_list3_unique and '3' in BBP_list3_unique)) and ('4' not in Chain_list3_unique)):
            mod_RRUCHUAIN_chainno_oldHSN_newHSN(0, 3, 4)
            mod_RRUCHUAIN_chainno_oldHSN_newHSN(1, 3, 4)
            mod_RRUCHUAIN_chainno_oldHSN_newHSN(2, 3, 4)
            mod_LRRU_to_MRRU_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(60, 0, 60, 0, 2, 2, 264,15)
            mod_LRRU_to_MRRU_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(61, 0, 61, 0, 2, 2, 264,15)
            mod_LRRU_to_MRRU_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(62, 0, 62, 0, 2, 2, 264,15)

            #Label(window, text='BTS5900: LTE800 60-61-62 AB RECONFIGURED, moved to SN 4 ..', width=80, fg='green').grid(column=2, row=11)
            fo = open('LOG.txt', 'a')
            fo.write(f'{dt.datetime.now()} --  {NEname}  ' + 'BTS5900: LTE800 60-61-62 AB RECONFIGURED, moved to SN 4 ..\n')
            text = Text(window, height=1, width=130, font=("arial", 10))
            text.insert(INSERT, 'BTS5900: LTE800 60-61-62 AB RECONFIGURED, moved to SN 4 ..')
            text.grid(column=2, row=11)
        else:
            if ('0' not in Chain_list3_unique and '1' not in Chain_list3_unique and '2' not in Chain_list3_unique):
                for i in range(0, len(Cell_band_800_list3)):  # i in (0,1,2) as cell local id same as bbp 4 port id
                    insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(i, 0, Subrack_global, 4, i)
                    insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 60 + i, 0, i, 264, 15, 2, 2)
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(i, 0, 60 + i, 0, [0, 1])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(i, 0, 60 + i, 0, [0, 1])
            else:
                for i in range(10, 10+len(Cell_band_800_list3)):  # i in (0,1,2) as cell local id same as bbp 4 port id
                    insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(i, 0, Subrack_global, 4, i-10)
                    insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 60 + i-10, 0, i, 264, 15, 2, 2)
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(i-10, 0, 60 + i-10, 0, [0, 1])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(i-10, 0, 60 + i-10, 0, [0, 1])

            #Label(window, text='BTS5900: LTE800 60-61-62 AB INTEGRATED ..', width=80, fg='green').grid(column=2, row=11)
            fo = open('LOG.txt', 'a')
            fo.write(f'{dt.datetime.now()} --  {NEname}  ' + 'BTS5900: LTE800 60-61-62 AB INTEGRATED .. PLEASE MANUALLY REMOVE SLOT 2 IF NEEDED!!!\n')
            text = Text(window, height=1, width=130, font=("arial", 10))
            text.insert(INSERT, 'BTS5900: LTE800 60-61-62 AB INTEGRATED ..')
            text.grid(column=2, row=11)
    #########################################################################################################################################################################################
    # BTS5900: LTE800 CD REC/INT + INT LTE700 AB
    if CheckVar4.get() == 1 or CheckkVar4 == 1:
        BBP4 = parse_existing_MO_name(root, 'BBP')
        SECTOR_var4 = parse_existing_MO_name(root, 'SECTOR')
        SECTOREQM_var4 = parse_existing_MO_name(root, 'SECTOREQM')
        Chain_var4 = parse_existing_MO_name(root, 'RRUCHAIN')
        Cell_var4 = parse_existing_MO_name(root, 'Cell')
        RRU_var4 = parse_existing_MO_name(root, 'RRU')
        # print(BBP1)
        BBP_list4_unique = []
        SECTOR_list4_unique = []
        SECTOREQM_list4_unique = []
        Chain_list4_unique = []
        Cell_band_800_list4_unique = []
        Cell_band_700_list4_unique = []
        RRU_list4_unique = []
        Cell_band_1800_2100_list4 = []
        Cell_band_2600_list4 = []
        for i in BBP4:
            for j in i:
                for j in range(0, len(BBP4)):
                    for item in BBP4[j]:
                        for key in item.keys():
                            if key == 'SN':
                                if item[key] not in BBP_list4_unique:
                                    BBP_list4_unique.append(item[key])
        for i in Cell_var4:
            for j in i:
                for j in range(0, len(Cell_var4)):
                    for item in Cell_var4[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['7', '8', '9', '18', '19', '20','21','22','23']:
                                if item[key] not in Cell_band_1800_2100_list4:
                                    Cell_band_1800_2100_list4.append(item[key])
        for i in Cell_var4:
            for j in i:
                for j in range(0, len(Cell_var4)):
                    for item in Cell_var4[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['30', '31', '32']:
                                if item[key] not in Cell_band_700_list4_unique:
                                    Cell_band_700_list4_unique.append(item[key])
        for i in Cell_var4:
            for j in i:
                for j in range(0, len(Cell_var4)):
                    for item in Cell_var4[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['0', '1', '2']:
                                if item[key] not in Cell_band_800_list4_unique:
                                    Cell_band_800_list4_unique.append(item[key])
        for i in Cell_var4:
            for j in i:
                for j in range(0, len(Cell_var4)):
                    for item in Cell_var4[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['4', '5', '6']:
                                if item[key] not in Cell_band_2600_list4:
                                    Cell_band_2600_list4.append(item[key])
        for i in SECTOR_var4:
            for j in i:
                for j in range(0, len(SECTOR_var4)):
                    for item in SECTOR_var4[j]:
                        for key in item.keys():
                            if key == 'SECTORID':
                                if item[key] not in SECTOR_list4_unique:
                                    SECTOR_list4_unique.append(item[key])
        for i in SECTOREQM_var4:
            for j in i:
                for j in range(0, len(SECTOREQM_var4)):
                    for item in SECTOREQM_var4[j]:
                        for key in item.keys():
                            if key == 'SECTOREQMID':
                                if item[key] not in SECTOREQM_list4_unique:
                                    SECTOREQM_list4_unique.append(item[key])
        for i in Chain_var4:
            for j in i:
                for j in range(0, len(Chain_var4)):
                    for item in Chain_var4[j]:
                        for key in item.keys():
                            if key == 'RCN':
                                if item[key] not in Chain_list4_unique:
                                    Chain_list4_unique.append(item[key])
        for i in RRU_var4:
            for j in i:
                for j in range(0, len(RRU_var4)):
                    for item in RRU_var4[j]:
                        for key in item.keys():
                            if key == 'SRN':
                                if item[key] not in RRU_list4_unique:
                                    RRU_list4_unique.append(item[key])

        if '4' not in BBP_list4_unique:
            insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, Subrack_global, 4, 2, 20)
        if '5' not in BBP_list4_unique:
            insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, Subrack_global, 5, 2, 4)
        if '2' in BBP_list4_unique and '3' not in BBP_list4_unique:
            if '4' in Cell_band_2600_list4 or '5' in Cell_band_2600_list4 or '6' in Cell_band_2600_list4:
                pass
                # there is LTE2600 on site
            # we leave board 2 if 1800/2100 on site and no board 3 exists ( in order to pe possible to add later board 3)
            else:
                if '7' in Cell_band_1800_2100_list4 or '8' in Cell_band_1800_2100_list4 or '9' in Cell_band_1800_2100_list4 or '18' in Cell_band_1800_2100_list4 or '19' in Cell_band_1800_2100_list4 or '20' in Cell_band_1800_2100_list4:

                    pass
                else:
                    rmv_old_bbp_from_slotnumber(2)

        if '3' in BBP_list4_unique:
            if '7' in Cell_band_1800_2100_list4 or '8' in Cell_band_1800_2100_list4 or '9' in Cell_band_1800_2100_list4 or '18' in Cell_band_1800_2100_list4 or '19' in Cell_band_1800_2100_list4 or '20' in Cell_band_1800_2100_list4:
                mod_BBP_TYPE_from_slot_to_slot(3, 3, 12290, 2)
                mod_BBP_TYPE_from_slot_to_slot(2, 2, 12290, 2)

                # there is LTE2100/Lte1800 on brd 3 on site
            else:
                rmv_old_bbp_from_slotnumber(3)
        if ('0' in Chain_list4_unique and '2' in BBP_list4_unique) or ('1' in Chain_list4_unique and '2' in BBP_list4_unique) or ('2' in Chain_list4_unique and '2' in BBP_list4_unique):
            mod_RRUCHUAIN_chainno_oldHSN_newHSN(0, 2, 4)
            mod_RRUCHUAIN_chainno_oldHSN_newHSN(1, 2, 4)
            mod_RRUCHUAIN_chainno_oldHSN_newHSN(2, 2, 4)
        if (('0' in Chain_list4_unique and '3' in BBP_list4_unique and '2' not in BBP_list4_unique) or ('1' in Chain_list4_unique and '3' in BBP_list4_unique and '2' not in BBP_list4_unique) or ('2' in Chain_list4_unique and '3' in BBP_list4_unique and '2' not in BBP_list4_unique)) and ('7' not in SECTOR_list4_unique or '18' not in SECTOR_list4_unique):
            mod_RRUCHUAIN_chainno_oldHSN_newHSN(0, 3, 4)
            mod_RRUCHUAIN_chainno_oldHSN_newHSN(1, 3, 4)
            mod_RRUCHUAIN_chainno_oldHSN_newHSN(2, 3, 4)

        #if CheckVar7.get() == 1 or CheckVar80.get() == 1:
        if CheckVar7.get() == 1 or CheckkVar7==1:
            if '7' in Cell_band_1800_2100_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 67, 0, 7, 8, 15, 4, 4)
            if '8' in Cell_band_1800_2100_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 68, 0, 8, 8, 15, 4, 4)
            if '9' in Cell_band_1800_2100_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 69, 0, 9, 8, 15, 4, 4)

        # if CheckVar8.get() == 1:
        #     if '7' in Cell_band_1800_2100_list4:
        #         insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 67, 0, 7, 8, 15, 2, 2)
        #     if '8' in Cell_band_1800_2100_list4:
        #         insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 68, 0, 8, 8, 15, 2, 2)
        #     if '9' in Cell_band_1800_2100_list4:
        #         insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 69, 0, 9, 8, 15, 2, 2)

        if CheckVar10.get() == 1 or CheckVar11.get() == 1 or CheckkVar10 == 1 or CheckkVar11 == 1:
            if '18' in Cell_band_1800_2100_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 18, 136, 15, 4, 4)
            if '19' in Cell_band_1800_2100_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 19, 136, 15, 4, 4)
            if '20' in Cell_band_1800_2100_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 20, 136, 15, 4, 4)

        if CheckVar9.get() == 1 or CheckkVar9==1:
            if '18' in Cell_band_1800_2100_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 18, 136, 15, 2, 2)
            if '19' in Cell_band_1800_2100_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 19, 136, 15, 2, 2)
            if '20' in Cell_band_1800_2100_list4 and '18' in Cell_band_1800_2100_list4 and '19' in Cell_band_1800_2100_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 20, 136, 15, 2, 2)
            if '20' in Cell_band_1800_2100_list4 and '18' not in Cell_band_1800_2100_list4 and '19' not in Cell_band_1800_2100_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 20, 136, 15, 2, 2)
            if '21' in Cell_band_1800_2100_list4 and '18' not in Cell_band_1800_2100_list4 and '19' not in Cell_band_1800_2100_list4 and '20' in Cell_band_1800_2100_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 21, 136, 15, 2, 2)
            if '21' in Cell_band_1800_2100_list4 and '18' not in Cell_band_1800_2100_list4 and '19' not in Cell_band_1800_2100_list4 and '20' not in Cell_band_1800_2100_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 21, 136, 15, 2, 2)
            if '22' in Cell_band_1800_2100_list4 and '18' not in Cell_band_1800_2100_list4 and '19' not in Cell_band_1800_2100_list4 and '20' not in Cell_band_1800_2100_list4 and '21' not in Cell_band_1800_2100_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 22, 136, 15, 2, 2)
            if '22' in Cell_band_1800_2100_list4 and '18' not in Cell_band_1800_2100_list4 and '19' not in Cell_band_1800_2100_list4 and '20' not in Cell_band_1800_2100_list4 and '21' in Cell_band_1800_2100_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 22, 136, 15, 2, 2)
            if '23' in Cell_band_1800_2100_list4 and '18' not in Cell_band_1800_2100_list4 and '19' not in Cell_band_1800_2100_list4 and '20' not in Cell_band_1800_2100_list4 and '21' not in Cell_band_1800_2100_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 23, 136, 15, 2, 2)
            if '23' in Cell_band_1800_2100_list4 and '18' not in Cell_band_1800_2100_list4 and '19' not in Cell_band_1800_2100_list4 and '20' not in Cell_band_1800_2100_list4 and '21' in Cell_band_1800_2100_list4 and '22' in Cell_band_1800_2100_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 23, 136, 15, 2, 2)

        if CheckVar14.get() == 1 or CheckkVar14==1:
            if '4' in Cell_band_2600_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 64, 0, 4, 8, 15, 4, 4)
            if '5' in Cell_band_2600_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 65, 0, 5, 8, 15, 4, 4)
            if '6' in Cell_band_2600_list4:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 66, 0, 6, 8, 15, 4, 4)

        # 800 REC CD
        if '60' in RRU_list4_unique and '0' in Cell_band_800_list4_unique:
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 80, 0, 0, 328, 15, 4, 4)
            rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(60, 0)
            rmv_RETSUBUNIT_only(0) # in case above function does not remove retsubunit
            #mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(60, 0, 80, 0, [0,1,2, 3]) # another option is to rmv+ add
            rmv_old_SECTOR_ID(0)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 80, 0, [0, 1, 2, 3])
            mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(60, 0, 80, 0, [2, 3]) # another option is to rmv+ add
        if '61' in RRU_list4_unique and '1' in Cell_band_800_list4_unique:
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 82, 0, 1, 328, 15, 4, 4)
            rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(61, 1)
            rmv_RETSUBUNIT_only(1) # in case above function does not remove retsubunit
            #mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(61, 0, 82, 0, [0,1,2, 3]) # another option is to rmv+ add
            rmv_old_SECTOR_ID(1)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 82, 0, [0, 1, 2, 3])
            mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(61, 0, 82, 0, [2, 3]) # another option is to rmv+ add
        if '62' in RRU_list4_unique  and '2' in Cell_band_800_list4_unique:
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 84, 0, 2, 328, 15, 4, 4)
            rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(62, 2)
            rmv_RETSUBUNIT_only(2) # in case above function does not remove retsubunit
            #mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(62, 0, 84, 0, [0,1,2, 3]) # another option is to rmv+ add
            rmv_old_SECTOR_ID(2)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 0, 84, 0, [0, 1, 2, 3])
            mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(62, 0, 84, 0, [2, 3]) # another option is to rmv+ add

        # insert700 AB
        if '30' not in SECTOREQM_list4_unique and '30' in Cell_band_700_list4_unique:
            #insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            #insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(30,0,0,80,0,[0,1])
            if '0' in SECTOR_list4_unique and '0' in SECTOREQM_list4_unique and '60' not in RRU_list4_unique:
                rmv_old_SECTOR_ID(0)
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 80, 0, [0, 1, 2, 3])
            if '0' not in SECTOR_list4_unique and '0' not in Cell_band_800_list4_unique:
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 80, 0, [0, 1, 2, 3])
            if '60' not in RRU_list4_unique and '0' not in Chain_list4_unique:
                #mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(80,0,80,0,[0,1,2,3])
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(0, 0, Subrack_global, 4, 0)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 80, 0, 0, 328, 15, 4, 4)

        # insert800 CD
        if '0' not in SECTOREQM_list4_unique and '0' in Cell_band_800_list4_unique and '60' not in RRU_list4_unique:
            #insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            #insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(0,0,0,80,0,[2,3])
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 80, 0, [0, 1, 2, 3])
            if '30' not in Cell_band_700_list4_unique:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(0, 0, Subrack_global, 4, 0)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 80, 0, 0, 328, 15, 4, 4)

        # insert700 AB
        if '31' not in SECTOREQM_list4_unique and '31' in Cell_band_700_list4_unique:
            #insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            #insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(31,1,0,82,0,[0,1])
            if '1' in SECTOR_list4_unique and '1' in SECTOREQM_list4_unique and '61' not in RRU_list4_unique:
                rmv_old_SECTOR_ID(1)
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 82, 0, [0, 1, 2, 3])
            if '1' not in SECTOR_list4_unique and '1' not in Cell_band_800_list4_unique:
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 82, 0, [0, 1, 2, 3])
            if '61' not in RRU_list4_unique and '1' not in Chain_list4_unique:
                #mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(82, 0, 82, 0, [0, 1, 2, 3])
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(1, 0, Subrack_global, 4, 1)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 82, 0, 1, 328, 15, 4, 4)

        # insert800 CD
        if '1' not in SECTOREQM_list4_unique and '1' in Cell_band_800_list4_unique and '61' not in RRU_list4_unique:
            #insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            #insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(1,1,0,82,0,[2,3])
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 82, 0, [0, 1, 2, 3])
            if '31' not in Cell_band_700_list4_unique:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(1, 0, Subrack_global, 4, 1)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 82, 0, 1, 328, 15, 4, 4)


        # insert700 AB
        if '32' not in SECTOREQM_list4_unique and '32' in Cell_band_700_list4_unique:
            #insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            #insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(32,2,0,84,0,[0,1])
            if '2' in SECTOR_list4_unique and '2' in SECTOREQM_list4_unique and '62' not in RRU_list4_unique:
                rmv_old_SECTOR_ID(2)
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 0, 84, 0, [0, 1, 2, 3])
            if '2' not in SECTOR_list4_unique and '2' not in Cell_band_800_list4_unique:
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 0, 84, 0, [0, 1, 2, 3])
            if '62' not in RRU_list4_unique and '2' not in Chain_list4_unique:
                #mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(84, 0, 84, 0, [0, 1, 2, 3])
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(2, 0, Subrack_global, 4, 2)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 84, 0, 2, 328, 15, 4, 4)
                # if '2' not in SECTOR_list4_unique and '2' not in Cell_band_800_list4_unique:
                #     insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 0, 84, 0, [0, 1,2,3])
        # insert800 CD
        if '2' not in SECTOREQM_list4_unique and '2' in Cell_band_800_list4_unique and '62' not in RRU_list4_unique:
            #insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            #insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(2,2,0,84,0,[2,3])
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 0, 84, 0, [0, 1, 2, 3])
            if '32' not in Cell_band_700_list4_unique:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(2, 0, Subrack_global, 4, 2)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 84, 0, 2, 328, 15, 4, 4)


        # V1.2 introduced:
        if '24' in SECTOR_list4_unique:
            rmv_old_SECTOR_ID(24)
            rmv_old_SECTOREQM_ID(24)
            insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(24, 0, 0, 80, 0, [0, 1])
        if '25' in SECTOR_list4_unique:
            rmv_old_SECTOR_ID(25)
            rmv_old_SECTOREQM_ID(25)
            insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(25, 1, 0, 82, 0, [0, 1])
        if '26' in SECTOR_list4_unique:
            rmv_old_SECTOR_ID(26)
            rmv_old_SECTOREQM_ID(26)
            insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(26, 2, 0, 84, 0, [0, 1])

        #Label(window, text='LTE800 CD reconfigured  + LTE700 AB INTEGRATED ..', width=80, fg='green').grid(column=2,row=12)
        fo = open('LOG.txt', 'a')
        fo.write(f'{dt.datetime.now()} --  {NEname}  ' + 'LTE800 CD reconfigured  + LTE700 AB INTEGRATED ..\n')
        text = Text(window, height=1, width=130, font=("arial", 10))
        text.insert(INSERT, 'LTE800 CD reconfigured  + LTE700 AB INTEGRATED ..')
        text.grid(column=2, row=12)
    #########################################################################################################################################################################################
    # BTS5900: LTE800 AB REC/INT + INT LTE700 CD
    if CheckVar5.get() == 1 or CheckkVar5 == 1:
        BBP5 = parse_existing_MO_name(root, 'BBP')
        SECTOR_var5 = parse_existing_MO_name(root, 'SECTOR')
        SECTOREQM_var5 = parse_existing_MO_name(root, 'SECTOREQM')
        Chain_var5 = parse_existing_MO_name(root, 'RRUCHAIN')
        Cell_var5 = parse_existing_MO_name(root, 'Cell')
        RRU_var5 = parse_existing_MO_name(root, 'RRU')
        # print(BBP1)
        BBP_list5_unique = []
        SECTOR_list5_unique = []
        SECTOREQM_list5_unique=[]
        Chain_list5_unique = []
        Cell_band_800_list5_unique = []
        Cell_band_700_list5_unique = []
        RRU_list5_unique = []
        Cell_band_1800_2100_list5 = []
        Cell_band_2600_list5 = []
        for i in BBP5:
            for j in i:
                for j in range(0, len(BBP5)):
                    for item in BBP5[j]:
                        for key in item.keys():
                            if key == 'SN':
                                if item[key] not in BBP_list5_unique:
                                    BBP_list5_unique.append(item[key])
        for i in Cell_var5:
            for j in i:
                for j in range(0, len(Cell_var5)):
                    for item in Cell_var5[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['30', '31', '32']:
                                if item[key] not in Cell_band_700_list5_unique:
                                    Cell_band_700_list5_unique.append(item[key])
        for i in Cell_var5:
            for j in i:
                for j in range(0, len(Cell_var5)):
                    for item in Cell_var5[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['0', '1', '2']:
                                if item[key] not in Cell_band_800_list5_unique:
                                    Cell_band_800_list5_unique.append(item[key])
        for i in Cell_var5:
            for j in i:
                for j in range(0, len(Cell_var5)):
                    for item in Cell_var5[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['4', '5', '6']:
                                if item[key] not in Cell_band_2600_list5:
                                    Cell_band_2600_list5.append(item[key])
        for i in Cell_var5:
            for j in i:
                for j in range(0, len(Cell_var5)):
                    for item in Cell_var5[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['7', '8', '9', '18', '19', '20','21','22','23']:
                                if item[key] not in Cell_band_1800_2100_list5:
                                    Cell_band_1800_2100_list5.append(item[key])
        for i in SECTOR_var5:
            for j in i:
                for j in range(0, len(SECTOR_var5)):
                    for item in SECTOR_var5[j]:
                        for key in item.keys():
                            if key == 'SECTORID':
                                if item[key] not in SECTOR_list5_unique:
                                    SECTOR_list5_unique.append(item[key])
        for i in SECTOREQM_var5:
            for j in i:
                for j in range(0, len(SECTOREQM_var5)):
                    for item in SECTOREQM_var5[j]:
                        for key in item.keys():
                            if key == 'SECTOREQMID':
                                if item[key] not in SECTOREQM_list5_unique:
                                    SECTOREQM_list5_unique.append(item[key])
        for i in Chain_var5:
            for j in i:
                for j in range(0, len(Chain_var5)):
                    for item in Chain_var5[j]:
                        for key in item.keys():
                            if key == 'RCN':
                                if item[key] not in Chain_list5_unique:
                                    Chain_list5_unique.append(item[key])
        for i in RRU_var5:
            for j in i:
                for j in range(0, len(RRU_var5)):
                    for item in RRU_var5[j]:
                        for key in item.keys():
                            if key == 'SRN':
                                if item[key] not in RRU_list5_unique:
                                    RRU_list5_unique.append(item[key])

        if '4' not in BBP_list5_unique:
            insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, Subrack_global, 4, 2, 20)
        if '5' not in BBP_list5_unique:
            insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, Subrack_global, 5, 2, 4)
        if '2' in BBP_list5_unique and '3' not in BBP_list5_unique:
            if '4' in Cell_band_2600_list5 or '5' in Cell_band_2600_list5 or '6' in Cell_band_2600_list5:
                pass
                # there is LTE2600 on site
            # we leave board 2 if 1800/2100 on site and no board 3 exists ( in order to pe possible to add later board 3)
            else:
                if '7' in Cell_band_1800_2100_list5 or '8' in Cell_band_1800_2100_list5 or '9' in Cell_band_1800_2100_list5 or '18' in Cell_band_1800_2100_list5 or '19' in Cell_band_1800_2100_list5 or '20' in Cell_band_1800_2100_list5:

                    pass
                else:
                    rmv_old_bbp_from_slotnumber(2)
        if '3' in BBP_list5_unique:
            if '7' in Cell_band_1800_2100_list5 or '8' in Cell_band_1800_2100_list5 or '9' in Cell_band_1800_2100_list5 or '18' in Cell_band_1800_2100_list5 or '19' in Cell_band_1800_2100_list5 or '20' in Cell_band_1800_2100_list5:
                mod_BBP_TYPE_from_slot_to_slot(3, 3, 12290, 2)
                mod_BBP_TYPE_from_slot_to_slot(2, 2, 12290, 2)

                # there is LTE2100/Lte1800 on brd 3 on site
            else:
                rmv_old_bbp_from_slotnumber(3)
        if ('0' in Chain_list5_unique and '2' in BBP_list5_unique) or ('1' in Chain_list5_unique and '2' in BBP_list5_unique) or ('2' in Chain_list5_unique and '2' in BBP_list5_unique):
            mod_RRUCHUAIN_chainno_oldHSN_newHSN(0, 2, 4)
            mod_RRUCHUAIN_chainno_oldHSN_newHSN(1, 2, 4)
            mod_RRUCHUAIN_chainno_oldHSN_newHSN(2, 2, 4)

        if (('0' in Chain_list5_unique and '3' in BBP_list5_unique and '2' not in BBP_list5_unique) or ('1' in Chain_list5_unique and '3' in BBP_list5_unique and '2' not in BBP_list5_unique) or ('2' in Chain_list5_unique and '3' in BBP_list5_unique and '2' not in BBP_list5_unique)) and ('7' not in SECTOR_list5_unique or '18' not in SECTOR_list5_unique):
            mod_RRUCHUAIN_chainno_oldHSN_newHSN(0, 3, 4)
            mod_RRUCHUAIN_chainno_oldHSN_newHSN(1, 3, 4)
            mod_RRUCHUAIN_chainno_oldHSN_newHSN(2, 3, 4)

        #if CheckVar7.get() == 1 or CheckVar80.get() == 1:
        if CheckVar7.get() == 1 or CheckkVar7==1:
            if '7' in Cell_band_1800_2100_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 67, 0, 7, 8, 15, 4, 4)
            if '8' in Cell_band_1800_2100_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 68, 0, 8, 8, 15, 4, 4)
            if '9' in Cell_band_1800_2100_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 69, 0, 9, 8, 15, 4, 4)

        # if CheckVar8.get() == 1:
        #     if '7' in Cell_band_1800_2100_list5:
        #         insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 67, 0, 7, 8, 15, 2, 2)
        #     if '8' in Cell_band_1800_2100_list5:
        #         insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 68, 0, 8, 8, 15, 2, 2)
        #     if '9' in Cell_band_1800_2100_list5:
        #         insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 69, 0, 9, 8, 15, 2, 2)

        if CheckVar10.get() == 1 or CheckVar11.get() == 1 or CheckkVar10 == 1 or CheckkVar11==1:
            if '18' in Cell_band_1800_2100_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 18, 136, 15, 4, 4)
            if '19' in Cell_band_1800_2100_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 19, 136, 15, 4, 4)
            if '20' in Cell_band_1800_2100_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 20, 136, 15, 4, 4)

        if CheckVar9.get() == 1 or CheckkVar9==1:
            if '18' in Cell_band_1800_2100_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 18, 136, 15, 2, 2)
            if '19' in Cell_band_1800_2100_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 19, 136, 15, 2, 2)
            if '20' in Cell_band_1800_2100_list5 and '18' in Cell_band_1800_2100_list5 and '19' in Cell_band_1800_2100_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 20, 136, 15, 2, 2)
            if '20' in Cell_band_1800_2100_list5 and '18' not in Cell_band_1800_2100_list5 and '19' not in Cell_band_1800_2100_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 20, 136, 15, 2, 2)
            # if '21' in Cell_band_1800_2100_list3 and '18' in Cell_band_1800_2100_list3 and '19' in Cell_band_1800_2100_list3 :
            #     insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 103, 0, 21, 136, 15, 2, 2)
            if '21' in Cell_band_1800_2100_list5 and '18' not in Cell_band_1800_2100_list5 and '19' not in Cell_band_1800_2100_list5 and '20' in Cell_band_1800_2100_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 21, 136, 15, 2, 2)
            if '21' in Cell_band_1800_2100_list5 and '18' not in Cell_band_1800_2100_list5 and '19' not in Cell_band_1800_2100_list5 and '20' not in Cell_band_1800_2100_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 21, 136, 15, 2, 2)
            # if '22' in Cell_band_1800_2100_list3 and '18' in Cell_band_1800_2100_list3 and '19' in Cell_band_1800_2100_list3 and  '20' in Cell_band_1800_2100_list3 and '21' in Cell_band_1800_2100_list3:
            #     insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 104, 0, 22, 136, 15, 2, 2)
            if '22' in Cell_band_1800_2100_list5 and '18' not in Cell_band_1800_2100_list5 and '19' not in Cell_band_1800_2100_list5 and '20' not in Cell_band_1800_2100_list5 and '21' not in Cell_band_1800_2100_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 22, 136, 15, 2, 2)
            if '22' in Cell_band_1800_2100_list5 and '18' not in Cell_band_1800_2100_list5 and '19' not in Cell_band_1800_2100_list5 and '20' not in Cell_band_1800_2100_list5 and '21' in Cell_band_1800_2100_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 22, 136, 15, 2, 2)
            if '23' in Cell_band_1800_2100_list5 and '18' not in Cell_band_1800_2100_list5 and '19' not in Cell_band_1800_2100_list5 and '20' not in Cell_band_1800_2100_list5 and '21' not in Cell_band_1800_2100_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 23, 136, 15, 2, 2)
            if '23' in Cell_band_1800_2100_list5 and '18' not in Cell_band_1800_2100_list5 and '19' not in Cell_band_1800_2100_list5 and '20' not in Cell_band_1800_2100_list5 and '21' in Cell_band_1800_2100_list5 and '22' in Cell_band_1800_2100_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 23, 136, 15, 2, 2)

        if CheckVar14.get() == 1 or CheckkVar14==1:
            if '4' in Cell_band_2600_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 64, 0, 4, 8, 15, 4, 4)
            if '5' in Cell_band_2600_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 65, 0, 5, 8, 15, 4, 4)
            if '6' in Cell_band_2600_list5:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 66, 0, 6, 8, 15, 4, 4)


        # 800 REC AB
        if '60' in RRU_list5_unique and '0' in Cell_band_800_list5_unique:
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 80, 0, 0, 328, 15, 4, 4)
            rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(60, 0)
            rmv_RETSUBUNIT_only(0) # in case above function does not remove retsubunit
            #mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(60, 0, 80, 0, [0,1,2, 3])
            rmv_old_SECTOR_ID(0)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 80, 0, [0, 1, 2, 3])
            mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(60, 0, 80, 0, [0, 1])
        if '61' in RRU_list5_unique and '1' in Cell_band_800_list5_unique:
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 82, 0, 1, 328, 15, 4, 4)
            rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(61, 1)
            rmv_RETSUBUNIT_only(1) # in case above function does not remove retsubunit
            #mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(61, 0, 82, 0, [0,1,2, 3])
            rmv_old_SECTOR_ID(1)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 82, 0, [0, 1, 2, 3])
            mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(61, 0, 82, 0, [0, 1])
        if '62' in RRU_list5_unique  and '2' in Cell_band_800_list5_unique:
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 84, 0, 2, 328, 15, 4, 4)
            rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(62, 2)
            rmv_RETSUBUNIT_only(2) # in case above function does not remove retsubunit
            #mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(62, 0, 84, 0, [0,1,2, 3])
            rmv_old_SECTOR_ID(2)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 0, 84, 0, [0, 1, 2, 3])
            mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(62, 0, 84, 0, [0, 1])

        # insert700 CD
        if '30' not in SECTOREQM_list5_unique and '30' in Cell_band_700_list5_unique:
            #insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            #insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(30,0,0,80,0,[2,3])
            if '0' in SECTOR_list5_unique and '0' in SECTOREQM_list5_unique and '60' not in RRU_list5_unique:
                rmv_old_SECTOR_ID(0)
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 80, 0, [0, 1, 2, 3])
            if '0' not in SECTOR_list5_unique and '0' not in Cell_band_800_list5_unique:
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 80, 0, [0, 1, 2, 3])
            if '60' not in RRU_list5_unique and '0' not in Chain_list5_unique:
                #mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(80, 0, 80, 0, [0, 1, 2, 3])
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(0, 0, Subrack_global, 4, 0)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 80, 0, 0, 328, 15, 4, 4)

        # insert800 AB
        if '0' not in SECTOREQM_list5_unique and '0' in Cell_band_800_list5_unique and '60' not in RRU_list5_unique:
            #insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            #insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(0,0,0,80,0,[0,1])
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 80, 0, [0, 1, 2, 3])
            if '30' not in Cell_band_700_list5_unique:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(0, 0, Subrack_global, 4, 0)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 80, 0, 0, 328, 15, 4, 4)

        # insert700 CD
        if '31' not in SECTOREQM_list5_unique and '31' in Cell_band_700_list5_unique:
            #insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            #insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(31,1,0,82,0,[2,3])
            if '1' in SECTOR_list5_unique and '1' in SECTOREQM_list5_unique and '61' not in RRU_list5_unique:
                rmv_old_SECTOR_ID(1)
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 82, 0, [0, 1, 2, 3])
            if '1' not in SECTOR_list5_unique and '1' not in Cell_band_800_list5_unique:
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 82, 0, [0, 1, 2, 3])
            if '61' not in RRU_list5_unique and '1' not in Chain_list5_unique:
                #mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(82, 0, 82, 0, [0, 1, 2, 3])
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(1, 0, Subrack_global, 4, 1)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 82, 0, 1, 328, 15, 4, 4)

        # insert800 AB
        if '1' not in SECTOREQM_list5_unique and '1' in Cell_band_800_list5_unique and '61' not in RRU_list5_unique:
            #insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            #insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(1,1,0,82,0,[0,1])
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 82, 0, [0, 1, 2, 3])
            if '31' not in Cell_band_700_list5_unique:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(1, 0, Subrack_global, 4, 1)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 82, 0, 1, 328, 15, 4, 4)


        # insert700 CD
        if '32' not in SECTOREQM_list5_unique and '32' in Cell_band_700_list5_unique:
            #insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            #insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(32,2,0,84,0,[2,3])
            if '2' in SECTOR_list5_unique and '2' in SECTOREQM_list5_unique and '62' not in RRU_list5_unique:
                rmv_old_SECTOR_ID(2)
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 0, 84, 0, [0, 1, 2, 3])
            if '2' not in SECTOR_list5_unique and '2' not in Cell_band_800_list5_unique:
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 0, 84, 0, [0, 1, 2, 3])
            if '62' not in RRU_list5_unique and '2' not in Chain_list5_unique:
                #mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(84, 0, 84, 0, [0, 1, 2, 3])
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(2, 0, Subrack_global, 4, 2)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 84, 0, 2, 328, 15, 4, 4)

        # insert800 AB
        if '2' not in SECTOREQM_list5_unique and '2' in Cell_band_800_list5_unique and '62' not in RRU_list5_unique:
            #insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            #insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(30, 0, 80, 0, [0, 1])
            insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(2,2,0,84,0,[0,1])
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 0, 84, 0, [0, 1, 2, 3])
            if '32' not in Cell_band_700_list5_unique:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(2, 0, Subrack_global, 4, 2)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 84, 0, 2, 328, 15, 4, 4)


        #Label(window, text='RRU5519et: LTE800 AB reconfigured  + LTE700 CD INTEGRATED ..', width=80, fg='green').grid(column=2, row=13)
        fo = open('LOG.txt', 'a')
        fo.write(f'{dt.datetime.now()} --  {NEname}  ' + 'RRU5519et: LTE800 AB reconfigured  + LTE700 CD INTEGRATED ..\n')
        text = Text(window, height=1, width=130, font=("arial", 10))
        text.insert(INSERT, 'RRU5519et: LTE800 AB reconfigured  + LTE700 CD INTEGRATED ..')
        text.grid(column=2, row=13)
    #########################################################################################################################################################################################
    # 5900: L700 INT 90-91-92 SN 5
    if CheckVar6.get() == 1 or CheckkVar6 == 1:
        BBP6 = parse_existing_MO_name(root, 'BBP')
        # BBP6=parse_BBP()
        SECTOR_var6 = parse_existing_MO_name(root, 'SECTOR')
        Chain_var6 = parse_existing_MO_name(root, 'RRUCHAIN')
        Cell_var6 = parse_existing_MO_name(root, 'Cell')
        RRU_var6 = parse_existing_MO_name(root, 'RRU')
        # print(BBP1)
        BBP_list6_unique = parse_BBP()
        SECTOR_list6_unique = []
        Chain_list6_unique = []
        Cell_band_700_list6 = []
        RRU_list6_unique = []
        # print('LIST OF BBP_______________', BBP_list6_unique)
        for i in BBP6:
            for j in i:
                for j in range(0, len(BBP6)):
                    for item in BBP6[j]:
                        for key in item.keys():
                            if key == 'SN':
                                if item[key] not in BBP_list6_unique:
                                    BBP_list6_unique.append(item[key])
        #print(BBP_list6_unique)
        # print(BBP6)
        for i in SECTOR_var6:
            for j in i:
                for j in range(0, len(SECTOR_var6)):
                    for item in SECTOR_var6[j]:
                        for key in item.keys():
                            if key == 'SECTORID':
                                if item[key] not in SECTOR_list6_unique:
                                    SECTOR_list6_unique.append(item[key])
        for i in Chain_var6:
            for j in i:
                for j in range(0, len(Chain_var6)):
                    for item in Chain_var6[j]:
                        for key in item.keys():
                            if key == 'RCN':
                                if item[key] not in Chain_list6_unique:
                                    Chain_list6_unique.append(item[key])
        for i in RRU_var6:
            for j in i:
                for j in range(0, len(RRU_var6)):
                    for item in RRU_var6[j]:
                        for key in item.keys():
                            if key == 'SRN':
                                if item[key] not in RRU_list6_unique:
                                    RRU_list6_unique.append(item[key])

        for i in Cell_var6:
            for j in i:
                for j in range(0, len(Cell_var6)):
                    for item in Cell_var6[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['30', '31', '32']:
                                if item[key] not in Cell_band_700_list6:
                                    Cell_band_700_list6.append(item[key])
        # if '5' not in BBP_list6_unique:
        if CheckVar3.get() == 0 and CheckkVar3==0:  # lte800 60-61-62 reconfig to slot 4  set to OFF
            if '5' not in BBP_list6_unique:
                insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, Subrack_global, 5, 2, 4)

        for i in range(0, len(Cell_band_700_list6)):
            if str(30 + i) not in Chain_list6_unique:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(30 + i, 0, Subrack_global, 5, i)
            if str(90 + i) not in RRU_list6_unique:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 90 + i, 0, 30 + i, 8, 15, 2, 2)
            if str(30 + i) not in SECTOR_list6_unique:
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(30 + i, 0, 90 + i, 0, [0, 1])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(30 + i, 0, 90 + i, 0, [0, 1])
        #Label(window, text='LTE700 AB INTEGRATED 90-91-92 slot 5 ..', width=80, fg='green').grid(column=2, row=14)
        fo = open('LOG.txt', 'a')
        fo.write(f'{dt.datetime.now()} --  {NEname}  ' + 'LTE700 AB INTEGRATED 90-91-92 slot 5 ..\n')
        text = Text(window, height=1, width=130, font=("arial", 10))
        text.insert(INSERT, 'LTE700 AB INTEGRATED 90-91-92 slot 5 ..')
        text.grid(column=2, row=14)
    #########################################################################################################################################################################################
    # 5900: L1800 INT ABCD
    def lte1800_single_function(Checkvar,CheckkVar, TEXT, row_to_display_status):

        if Checkvar.get() == 1 or CheckkVar == 1:
            BBP7 = parse_existing_MO_name(root, 'BBP')
            SECTOR_var7 = parse_existing_MO_name(root, 'SECTOR')
            Chain_var7 = parse_existing_MO_name(root, 'RRUCHAIN')
            Cell_var7 = parse_existing_MO_name(root, 'Cell')
            #print(Cell_var7)
            RRU_var7 = parse_existing_MO_name(root, 'RRU')
            # print(BBP1)
            BBP_list7_unique = []
            SECTOR_list7_unique = []
            Chain_list7_unique = []
            Cell_band_1800_2100_list7 = []
            RRU_list7_unique = []
            Cell_list7 = []
            Cell_1800_work_mode_list7=[]
            flag=False
            for i in BBP7:
                for j in i:
                    for j in range(0, len(BBP7)):
                        for item in BBP7[j]:
                            for key in item.keys():
                                if key == 'SN':
                                    if item[key] not in BBP_list7_unique:
                                        BBP_list7_unique.append(item[key])
            for i in Cell_var7:
                for j in i:
                    for j in range(0, len(Cell_var7)):
                        flag = False
                        for item in Cell_var7[j]:
                            for key in item.keys():
                                if key == 'LocalCellId' and item[key] in ['7', '8', '9']:
                                    flag=True
                                    if item[key] not in Cell_band_1800_2100_list7:
                                        Cell_band_1800_2100_list7.append(item[key])
                                if flag and key == 'TxRxMode':
                                    if item[key] not in Cell_1800_work_mode_list7:
                                        Cell_1800_work_mode_list7.append(item[key])
                                        #example: Cell_1800_work_mode_list=['2','2','2'] , ['4','4','4'],['3','3','3'] / 2T2R~2, 2T4R~3, 4T4R~4

            #print('CELL WORK MODE:   ',Cell_1800_work_mode_list,Cell_band_1800_2100_list7)
            #defining 1800 port_list for SECTOR/SECTOREQM depending on Cell TxRxMode
            if '4' in Cell_1800_work_mode_list7 and '2' not in Cell_1800_work_mode_list7:   #all 3 1800 cells are 4T4R
                port_list_1800=[0,1,2,3]
            if '4' in Cell_1800_work_mode_list7 and '2' in Cell_1800_work_mode_list7:   #some are 4T4R some are 2T2R - not found in the network
                port_list_1800=[0,1,2,3]
            if '2' in Cell_1800_work_mode_list7 and '4' not in Cell_1800_work_mode_list7:  #all 3 1800 cells are 2T2R
                port_list_1800=[0,1]
            if '3' in Cell_1800_work_mode_list7:
                port_list_1800 = [0, 1, 2, 3]
            if '0' in Cell_1800_work_mode_list7:
                port_list_1800=[0]

            for i in Cell_var7:
                for j in i:
                    for j in range(0, len(Cell_var7)):
                        for item in Cell_var7[j]:
                            for key in item.keys():
                                if key == 'LocalCellId':
                                    if item[key] not in Cell_list7:
                                        Cell_list7.append(item[key])
            for i in SECTOR_var7:
                for j in i:
                    for j in range(0, len(SECTOR_var7)):
                        for item in SECTOR_var7[j]:
                            for key in item.keys():
                                if key == 'SECTORID':
                                    if item[key] not in SECTOR_list7_unique:
                                        SECTOR_list7_unique.append(item[key])
            if '3' not in BBP_list7_unique:
                insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, Subrack_global, 3, 2, 4)
            # if '7' in Cell_band_1800_2100_list7 and '7' not in SECTOR_list7_unique:
            if '7' not in SECTOR_list7_unique:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(7, 0, Subrack_global, 3, 0)
                # are added by LTE800+LTE700 if button is checked for this option
                if CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0:
                    insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 67, 0, 7, 8, 15, 4, 4)
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 67, 0, port_list_1800)
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 67, 0, port_list_1800)
            # if '8' in Cell_band_1800_2100_list7 and '8' not in SECTOR_list7_unique:
            if '8' not in SECTOR_list7_unique:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(8, 0, Subrack_global, 3, 1)
                # are added by LTE800+LTE700 if button is checked for this option
                if CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0:
                    insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 68, 0, 8, 8, 15, 4, 4)
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 68, 0, port_list_1800)
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 68, 0, port_list_1800)
            # if '9' in Cell_band_1800_2100_list7 and '9' not in SECTOR_list7_unique:
            if '9' not in SECTOR_list7_unique:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(9, 0, Subrack_global, 3, 2)
                # are added by LTE800+LTE700 if button is checked for this option
                if  CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0:
                    insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 69, 0, 9, 8, 15, 4, 4)
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 69, 0, port_list_1800)
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 69, 0, port_list_1800)

            if '2' in BBP_list7_unique and ('4' in Cell_list7 or '5' in Cell_list7 or '6' in Cell_list7 or '12' in Cell_list7 or '13' in Cell_list7 or '14' in Cell_list7):
                pass
            else:
                rmv_old_bbp_from_slotnumber(2)
            #Label(window, text=TEXT, width=80, fg='green').grid(column=2, row=row_to_display_status)
            text = Text(window, height=1, width=130, font=("arial", 10))
            text.insert(INSERT, TEXT)
            text.grid(column=2, row=row_to_display_status)
            fo = open('LOG.txt', 'a')
            fo.write(f'{dt.datetime.now()} --  {NEname}  ' + TEXT + '\n')


    lte1800_single_function(CheckVar7,CheckkVar7, 'LTE1800 ABCD 67-68-69 singleband RRU integrated', 15)
    # corelated with 800 options on slot 3

    #########################################################################################################################################################################################
    # 5900: L1800 INT AC
    #lte1800_single_function(CheckVar80,CheckkVar80, [0, 2], 'LTE1800 AC 67-68-69 INTEGRATED', 16)
    # corelated with 800 options on slot 3
    # fo = open('LOG.txt', 'a')
    # fo.write(f'{dt.datetime.now()} --  {NEname}  '+'LTE1800 AC 67-68-69 INTEGRATED..\n')
    #########################################################################################################################################################################################
    # 5900: L1800 INT AB
    #lte1800_single_function(CheckVar8,CheckkVar8, [0, 1], 'LTE1800 AB 67-68-69 INTEGRATED', 17)
    # corelated with 800 options on slot 3
    # fo = open('LOG.txt', 'a')
    # fo.write(f'{dt.datetime.now()} --  {NEname}  '+'LTE1800 AB 67-68-69 INTEGRATED..\n')
    #########################################################################################################################################################################################
    # 5900: L2100 INT AB
    if CheckVar9.get() == 1 or CheckkVar9==1:
        BBP9 = parse_existing_MO_name(root, 'BBP')
        SECTOR_var9 = parse_existing_MO_name(root, 'SECTOR')
        Chain_var9 = parse_existing_MO_name(root, 'RRUCHAIN')
        Cell_var9 = parse_existing_MO_name(root, 'Cell')
        RRU_var9 = parse_existing_MO_name(root, 'RRU')
        # print(BBP9)
        BBP_list9_unique = []
        SECTOR_list9_unique = []
        Chain_list9_unique = []
        Cell_band_1800_2100_list9 = []
        Cell_band_2100_list9 = []
        RRU_list9_unique = []
        Cell_list9 = []
        for i in BBP9:
            for j in i:
                for j in range(0, len(BBP9)):
                    for item in BBP9[j]:
                        for key in item.keys():
                            if key == 'SN':
                                if item[key] not in BBP_list9_unique:
                                    BBP_list9_unique.append(item[key])
        for i in Cell_var9:
            for j in i:
                for j in range(0, len(Cell_var9)):
                    for item in Cell_var9[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['7', '8', '9']:
                                if item[key] not in Cell_band_1800_2100_list9:
                                    Cell_band_1800_2100_list9.append(item[key])
        for i in Cell_var9:
            for j in i:
                for j in range(0, len(Cell_var9)):
                    for item in Cell_var9[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['18', '19', '20','21','22','23']:
                                if item[key] not in Cell_band_2100_list9:
                                    Cell_band_2100_list9.append(item[key])
        for i in Cell_var9:
            for j in i:
                for j in range(0, len(Cell_var9)):
                    for item in Cell_var9[j]:
                        for key in item.keys():
                            if key == 'LocalCellId':
                                if item[key] not in Cell_list9:
                                    Cell_list9.append(item[key])
        for i in SECTOR_var9:
            for j in i:
                for j in range(0, len(SECTOR_var9)):
                    for item in SECTOR_var9[j]:
                        for key in item.keys():
                            if key == 'SECTORID':
                                if item[key] not in SECTOR_list9_unique:
                                    SECTOR_list9_unique.append(item[key])
        if '3' not in BBP_list9_unique:
            #if CheckVar7.get() == 0 and CheckVar80.get() == 0 and CheckVar8.get() == 0:
            if CheckVar7.get() == 0:
                insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, Subrack_global, 3, 2, 4)

        # if '18' in Cell_band_1800_2100_list7 and '18' not in SECTOR_list9_unique:
        if '18' in Cell_band_2100_list9 and '18' not in SECTOR_list9_unique:
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(18, 0, Subrack_global, 3, 3)
            # are added by LTE800+LTE700 if button is checked for this option
            if CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 18, 136, 15, 2, 2)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(18, 0, 100, 0, [0, 1])
            insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(18, 0, 100, 0, [0, 1])

        # if '19' in Cell_band_1800_2100_list9 and '19' not in SECTOR_list9_unique:
        if '19' in Cell_band_2100_list9 and '19' not in SECTOR_list9_unique:
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(19, 0, Subrack_global, 3, 4)
            # are added by LTE800+LTE700 if button is checked for this option
            if CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 19, 136, 15, 2, 2)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 0, 101, 0, [0, 1])
            insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 0, 101, 0, [0, 1])

        # if '20' in Cell_band_1800_2100_list7 and '20' not in SECTOR_list9_unique:
        if '20' in Cell_band_2100_list9 and '20' not in SECTOR_list9_unique and '18' in Cell_band_2100_list9 and '19' in Cell_band_2100_list9:
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(20, 0, Subrack_global, 3, 5)
            # are added by LTE800+LTE700 if button is checked for this option
            if CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 20, 136, 15, 2, 2)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 0, 102, 0, [0, 1])
            insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 0, 102, 0, [0, 1])
        if '20' in Cell_band_2100_list9 and '20' not in SECTOR_list9_unique and '18' not in Cell_band_2100_list9 and '19' not in Cell_band_2100_list9:
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(20, 0, Subrack_global, 3, 3)
            # are added by LTE800+LTE700 if button is checked for this option
            if CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 20, 136, 15, 2, 2)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 0, 100, 0, [0, 1])
            insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 0, 100, 0, [0, 1])

        if '21' in Cell_band_2100_list9 and '21' not in SECTOR_list9_unique and '18' not in Cell_band_2100_list9 and '19' not in Cell_band_2100_list9 and '20'  in Cell_band_2100_list9:
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(21, 0, Subrack_global, 3, 4)
            # are added by LTE800+LTE700 if button is checked for this option
            if CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 21, 136, 15, 2, 2)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(21, 0, 101, 0, [0, 1])
            insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(21, 0, 101, 0, [0, 1])
        if '21' in Cell_band_2100_list9 and '21' not in SECTOR_list9_unique and '18' not in Cell_band_2100_list9 and '19' not in Cell_band_2100_list9 and '20' not in Cell_band_2100_list9:
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(21, 0, Subrack_global, 3, 3)
            # are added by LTE800+LTE700 if button is checked for this option
            if CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 21, 136, 15, 2, 2)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(21, 0, 100, 0, [0, 1])
            insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(21, 0, 100, 0, [0, 1])

        if '22' in Cell_band_2100_list9 and '22' not in SECTOR_list9_unique and '18' not in Cell_band_2100_list9 and '19' not in Cell_band_2100_list9 and '20' not in Cell_band_2100_list9 and '21' not in Cell_band_2100_list9:
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(22, 0, Subrack_global, 3, 3)
            # are added by LTE800+LTE700 if button is checked for this option
            if CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 22, 136, 15, 2, 2)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(22, 0, 100, 0, [0, 1])
            insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(22, 0, 100, 0, [0, 1])
        if '22' in Cell_band_2100_list9 and '22' not in SECTOR_list9_unique and '18' not in Cell_band_2100_list9 and '19' not in Cell_band_2100_list9 and '20' not in Cell_band_2100_list9 and '21'  in Cell_band_2100_list9:
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(22, 0, Subrack_global, 3, 4)
            # are added by LTE800+LTE700 if button is checked for this option
            if CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 22, 136, 15, 2, 2)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(22, 0, 101, 0, [0, 1])
            insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(22, 0, 101, 0, [0, 1])

        if '23' in Cell_band_2100_list9 and '23' not in SECTOR_list9_unique and '18' not in Cell_band_2100_list9 and '19' not in Cell_band_2100_list9 and '20' not in Cell_band_2100_list9 and '21' not in Cell_band_2100_list9:
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(23, 0, Subrack_global, 3, 4)
            # are added by LTE800+LTE700 if button is checked for this option
            if CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 23, 136, 15, 2, 2)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(23, 0, 101, 0, [0, 1])
            insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(23, 0, 101, 0, [0, 1])
        if '23' in Cell_band_2100_list9 and '23' not in SECTOR_list9_unique and '18' not in Cell_band_2100_list9 and '19' not in Cell_band_2100_list9 and '20' not in Cell_band_2100_list9 and '21'  in Cell_band_2100_list9 and '22'  in Cell_band_2100_list9:
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(23, 0, Subrack_global, 3, 5)
            # are added by LTE800+LTE700 if button is checked for this option
            if CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 23, 136, 15, 2, 2)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(23, 0, 102, 0, [0, 1])
            insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(23, 0, 102, 0, [0, 1])


        if '2' in BBP_list9_unique and ('4' in Cell_list9 or '5' in Cell_list9 or '6' in Cell_list9 or '12' in Cell_list9 or '13' in Cell_list9 or '14' in Cell_list9):
            pass
        else:
            rmv_old_bbp_from_slotnumber(2)
        #Label(window, text='5900: L2100 INT AB INTEGRATED..', width=80, fg='green').grid(column=2, row=18)
        text = Text(window, height=1, width=130, font=("arial", 10))
        text.insert(INSERT, '5900: L2100  AB singleband RRU integrated..')
        text.grid(column=2, row=18)
        fo = open('LOG.txt', 'a')
        fo.write(f'{dt.datetime.now()} --  {NEname}  ' + '5900: L2100  AB singleband RRU integrated..\n')
    #########################################################################################################################################################################################
    # 5900: L2100 INT AC INT +L1800 INT ABCD or BD existing shared RRU (moved to 100,101,102)
    if CheckVar10.get() == 1 or CheckkVar10 == 1:
        BBP10 = parse_existing_MO_name(root, 'BBP')
        SECTOR_var10 = parse_existing_MO_name(root, 'SECTOR')
        SECTOREQM_var10 = parse_existing_MO_name(root, 'SECTOREQM')
        Chain_var10 = parse_existing_MO_name(root, 'RRUCHAIN')
        Cell_var10 = parse_existing_MO_name(root, 'Cell')
        RRU_var10 = parse_existing_MO_name(root, 'RRU')
        # print(BBP9)
        BBP_list10_unique = []
        SECTOR_list10_unique = []
        SECTOREQM_list10_unique=[]
        Chain_list10_unique = []
        Cell_band_1800_list10 = []
        Cell_band_2100_list10 = []
        RRU_list10_unique = []
        Cell_list10 = []
        Cell_band_1800_2100_list10=[]
        Cell_1800_work_mode_list10=[]
        flag_10 = False
        for i in BBP10:
            for j in i:
                for j in range(0, len(BBP10)):
                    for item in BBP10[j]:
                        for key in item.keys():
                            if key == 'SN':
                                if item[key] not in BBP_list10_unique:
                                    BBP_list10_unique.append(item[key])
        for i in Cell_var10:
            for j in i:
                for j in range(0, len(Cell_var10)):
                    for item in Cell_var10[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['7', '8', '9']:
                                if item[key] not in Cell_band_1800_list10:
                                    Cell_band_1800_list10.append(item[key])
                            if key == 'LocalCellId' and item[key] in ['18', '19', '20', '21']:
                                if item[key] not in Cell_band_2100_list10:
                                    Cell_band_2100_list10.append(item[key])
        for i in Cell_var10:
            for j in i:
                for j in range(0, len(Cell_var10)):
                    for item in Cell_var10[j]:
                        for key in item.keys():
                            if key == 'LocalCellId':
                                if item[key] not in Cell_list10:
                                    Cell_list10.append(item[key])
        for i in SECTOR_var10:
            for j in i:
                for j in range(0, len(SECTOR_var10)):
                    for item in SECTOR_var10[j]:
                        for key in item.keys():
                            if key == 'SECTORID':
                                if item[key] not in SECTOR_list10_unique:
                                    SECTOR_list10_unique.append(item[key])
        for i in SECTOREQM_var10:
            for j in i:
                for j in range(0, len(SECTOREQM_var10)):
                    for item in SECTOREQM_var10[j]:
                        for key in item.keys():
                            if key == 'SECTOREQMID':
                                if item[key] not in SECTOREQM_list10_unique:
                                    SECTOREQM_list10_unique.append(item[key])
        for i in RRU_var10:
            for j in i:
                for j in range(0, len(RRU_var10)):
                    for item in RRU_var10[j]:
                        for key in item.keys():
                            if key == 'SRN':
                                if item[key] not in RRU_list10_unique:
                                    RRU_list10_unique.append(item[key])
        for i in Cell_var10:
            for j in i:
                for j in range(0, len(Cell_var10)):
                    flag_10 = False
                    for item in Cell_var10[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['7', '8', '9']:
                                flag_10 = True
                                if item[key] not in Cell_band_1800_2100_list10:
                                    Cell_band_1800_2100_list10.append(item[key])
                            if flag_10 and key == 'TxRxMode':
                                if item[key] not in Cell_1800_work_mode_list10:
                                    Cell_1800_work_mode_list10.append(item[key])
                                    # example: Cell_1800_work_mode_list=['2','2','2'] , ['4','4','4'],['3','3','3'] / 2T2R~2, 2T4R~3, 4T4R~4

        # print('CELL WORK MODE:   ',Cell_1800_work_mode_list,Cell_band_1800_2100_list7)
        # defining port_list for SECTOR/SECTOREQM depending on Cell TxRxMode
        if len(Cell_1800_work_mode_list10)==0:
            sector_port_list_1800='0T0R'
        if '4' in Cell_1800_work_mode_list10 and '2' not in Cell_1800_work_mode_list10:  # all 3 1800 cells are 4T4R
            sector_port_list_1800 = '4T4R'
        if '4' in Cell_1800_work_mode_list10 and '2' in Cell_1800_work_mode_list10:  # some are 4T4R some are 2T2R - not found in the network
            sector_port_list_1800 = '4T4R'
        if '2' in Cell_1800_work_mode_list10 and '4' not in Cell_1800_work_mode_list10:  # all 3 1800 cells are 2T2R
            sector_port_list_1800 = '2T2R'
        if '3' in Cell_1800_work_mode_list10:
            sector_port_list_1800 = '4T4R'

        if '3' not in BBP_list10_unique:
            #if CheckVar7.get() == 0 and CheckVar80.get() == 0 and CheckVar8.get() == 0:
            if CheckVar7.get() == 0 and CheckkVar7==0:
                insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, Subrack_global, 3, 2, 4)

        if '18' in Cell_band_2100_list10 and '18' not in SECTOREQM_list10_unique:
            #rmv_old_rruchain_from_slotnumber_portnumber(3,3)
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(18, 0, Subrack_global, 3, 0)
            if CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0 and '100' not in RRU_list10_unique:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 18, 136, 15, 4, 4)
            if sector_port_list_1800=='2T2R':
                #insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(18, 0, 100, 0, [0, 2])
                #insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(18, 0, 100, 0, [0, 2])
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(18,7,0,100,0,[0,2])
            if sector_port_list_1800=='4T4R' and '7' in SECTOR_list10_unique:
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(18,7,0,100,0,[0,2])
            if sector_port_list_1800=='4T4R' and '7' not in SECTOR_list10_unique and '7' in Cell_band_1800_list10:
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(18, 7, 0, 100, 0, [0, 2])
            if sector_port_list_1800 == '0T0R': #no 1800
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(18, 0, 100, 0, [0, 2])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(18, 0, 100, 0, [0, 2])

        if '7' in Cell_band_1800_list10 and '7' in SECTOR_list10_unique and '67' in RRU_list10_unique:
            if sector_port_list_1800 == '2T2R':
                #mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(67, 0, 100, 0, [1, 3])
                mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(67,0,100,0,[0,1,2,3])
                mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(67, 0, 100, 0, [1, 3])
            if sector_port_list_1800 == '4T4R':
                mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(67, 0, 100, 0, [0,1,2, 3])
                mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(67, 0, 100, 0, [0,1,2, 3])
            if '18' not in Cell_band_2100_list10:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(18, 0, Subrack_global, 3, 0)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 18, 136, 15, 4, 4)
        if '7' in Cell_band_1800_list10 and '7' not in SECTOR_list10_unique and '67' not in RRU_list10_unique:
            if sector_port_list_1800 == '2T2R':
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0,1,2, 3])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [1, 3])
            if sector_port_list_1800 == '4T4R':
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0,1,2, 3])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0,1,2, 3])
            if '18' not in Cell_band_2100_list10:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(18, 0, Subrack_global, 3, 0)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 18, 136, 15, 4, 4)

        if '19' in Cell_band_2100_list10 and '19' not in SECTOREQM_list10_unique:
            #rmv_old_rruchain_from_slotnumber_portnumber(3, 4)
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(19, 0, Subrack_global, 3, 1)
            if CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0 and '101' not in RRU_list10_unique:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 19, 136, 15, 4, 4)
            if sector_port_list_1800 == '2T2R':
                # insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 0, 101, 0, [0, 2])
                # insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 0, 101, 0, [0, 2])
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 8, 0, 101, 0, [0, 2])
            if sector_port_list_1800 == '4T4R' and '8' in SECTOR_list10_unique:
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 8, 0, 101, 0, [0, 2])
            if sector_port_list_1800=='4T4R' and '8' not in SECTOR_list10_unique and '8' in Cell_band_1800_list10:
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 8, 0, 101, 0, [0, 2])
            if sector_port_list_1800 == '0T0R':  # no 1800
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 0, 101, 0, [0, 2])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 0, 101, 0, [0, 2])

        if '8' in Cell_band_1800_list10 and '8' in SECTOR_list10_unique and '68' in RRU_list10_unique:
            if sector_port_list_1800 == '2T2R':
                #mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(68, 0, 101, 0, [1, 3])
                mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(68, 0, 101, 0, [0, 1, 2, 3])
                mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(68, 0, 101, 0, [1, 3])
            if sector_port_list_1800 == '4T4R':
                mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(68, 0, 101, 0, [0, 1, 2, 3])
                mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(68, 0, 101, 0, [0, 1, 2, 3])
            if '19' not in Cell_band_2100_list10:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(19, 0, Subrack_global, 3, 1)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 19, 136, 15, 4, 4)
        if '8' in Cell_band_1800_list10 and '8' not in SECTOR_list10_unique and '68' not in RRU_list10_unique:
            if sector_port_list_1800 == '2T2R':
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0,1,2, 3])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [1, 3])
            if sector_port_list_1800 == '4T4R':
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 1, 2, 3])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 1, 2, 3])
            if '19' not in Cell_band_2100_list10:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(19, 0, Subrack_global, 3, 1)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 19, 136, 15, 4, 4)

        if '20' in Cell_band_2100_list10 and '20' not in SECTOREQM_list10_unique:
            #rmv_old_rruchain_from_slotnumber_portnumber(3, 5)
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(20, 0, Subrack_global, 3, 2)
            if CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0 and '102' not in RRU_list10_unique:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 20, 136, 15, 4, 4)
            if sector_port_list_1800 == '2T2R':
                # insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 0, 102, 0, [0, 2])
                # insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 0, 102, 0, [0, 2])
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 9, 0, 102, 0, [0, 2])
            if sector_port_list_1800 == '4T4R' and '9' in SECTOR_list10_unique:
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 9, 0, 102, 0, [0, 2])
            if sector_port_list_1800=='4T4R' and '9' not in SECTOR_list10_unique and '9' in Cell_band_1800_list10:
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 9, 0, 102, 0, [0, 2])
            if sector_port_list_1800 == '0T0R':  # no 1800
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 0, 102, 0, [0, 2])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 0, 102, 0, [0, 2])

        if '9' in Cell_band_1800_list10 and '9' in SECTOR_list10_unique and '69' in RRU_list10_unique:
            if sector_port_list_1800 == '2T2R':
                #mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(69, 0, 102, 0, [1, 3])
                mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(69, 0, 102, 0, [0, 1, 2, 3])
                mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(69, 0, 102, 0, [1, 3])
            if sector_port_list_1800 == '4T4R':
                mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(69, 0, 102, 0, [0, 1, 2, 3])
                mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(69, 0, 102, 0, [0, 1, 2, 3])
            if '20' not in Cell_band_2100_list10:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(20, 0, Subrack_global, 3, 2)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 20, 136, 15, 4, 4)
        if '9' in Cell_band_1800_list10 and '9' not in SECTOR_list10_unique and '69' not in RRU_list10_unique:
            if sector_port_list_1800 == '2T2R':
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0,1,2, 3])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [1, 3])
            if sector_port_list_1800 == '4T4R':
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 1, 2, 3])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 1, 2, 3])
            if '20' not in Cell_band_2100_list10:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(20, 0, Subrack_global, 3, 2)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 20, 136, 15, 4, 4)

        # rmv_old_rruchain_from_slotnumber_portnumber(3, 0)
        # rmv_old_rruchain_from_slotnumber_portnumber(3, 1)
        # rmv_old_rruchain_from_slotnumber_portnumber(3, 2)
        rmv_old_rruchain_from_slotnumber_portnumber(3, 3)
        rmv_old_rruchain_from_slotnumber_portnumber(3, 4)
        rmv_old_rruchain_from_slotnumber_portnumber(3, 5)
        rmv_old_rruchain_bychainID(7)
        rmv_old_rruchain_bychainID(8)
        rmv_old_rruchain_bychainID(9)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(67, 7)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(68, 8)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(69, 9)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(67, 17)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(68, 18)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(69, 19)
        rmv_RET_only(100)
        rmv_RET_only(101)
        rmv_RET_only(102)
        rmv_RET_only(4) # delete RETs controlled by RFU on SRN 4 1800/2100
        rmv_RETSUBUNIT_only(17)
        rmv_RETSUBUNIT_only(18)
        rmv_RETSUBUNIT_only(19)
        rmv_RETSUBUNIT_only(7)
        rmv_RETSUBUNIT_only(8)
        rmv_RETSUBUNIT_only(9)
        rmv_RETSUBUNIT_only(20)
        rmv_RETSUBUNIT_only(21)
        rmv_RETSUBUNIT_only(22)

        if '2' in BBP_list10_unique and ('4' in Cell_list10 or '5' in Cell_list10 or '6' in Cell_list10 or '12' in Cell_list10 or '13' in Cell_list10 or '14' in Cell_list10):
            pass
        else:
            rmv_old_bbp_from_slotnumber(2)

        #Label(window, text='5900: L2100 AC INT + L1800 BD REC/INT PN_3.4.5', width=80, fg='green').grid(column=2,row=19)
        text = Text(window, height=1, width=130, font=("arial", 10))
        text.insert(INSERT, '5900: L2100 AC INT + L1800 BD/ABCD REC/INT PN_0.1.2')
        text.grid(column=2, row=19)
        fo = open('LOG.txt', 'a')
        fo.write(f'{dt.datetime.now()} --  {NEname}  ' + '5900: L2100 AC INT + L1800 BD/ABCD REC/INT PN_0.1.2\n')
    #########################################################################################################################################################################################
    # 5900: L2100 INT BD INT +L1800 ABCD or AC INT or existing shared RRU (moved to 100,101,102)
    if CheckVar11.get() == 1 or CheckkVar11 == 1:
        BBP11 = parse_existing_MO_name(root, 'BBP')
        SECTOR_var11 = parse_existing_MO_name(root, 'SECTOR')
        SECTOREQM_var11 = parse_existing_MO_name(root, 'SECTOREQM')
        Chain_var11 = parse_existing_MO_name(root, 'RRUCHAIN')
        Cell_var11 = parse_existing_MO_name(root, 'Cell')
        RRU_var11 = parse_existing_MO_name(root, 'RRU')
        # print(BBP11)
        BBP_list11_unique = []
        SECTOR_list11_unique = []
        SECTOREQM_list11_unique=[]
        Chain_list11_unique = []
        Cell_band_1800_list11 = []
        Cell_band_2100_list11 = []
        RRU_list11_unique = []
        Cell_list11 = []
        Cell_band_1800_2100_list11 = []
        Cell_1800_work_mode_list11 = []
        flag_11=False
        for i in BBP11:
            for j in i:
                for j in range(0, len(BBP11)):
                    for item in BBP11[j]:
                        for key in item.keys():
                            if key == 'SN':
                                if item[key] not in BBP_list11_unique:
                                    BBP_list11_unique.append(item[key])
        for i in Cell_var11:
            for j in i:
                for j in range(0, len(Cell_var11)):
                    for item in Cell_var11[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['7', '8', '9']:
                                if item[key] not in Cell_band_1800_list11:
                                    Cell_band_1800_list11.append(item[key])
                            if key == 'LocalCellId' and item[key] in ['18', '19', '20', '21']:
                                if item[key] not in Cell_band_2100_list11:
                                    Cell_band_2100_list11.append(item[key])
        for i in Cell_var11:
            for j in i:
                for j in range(0, len(Cell_var11)):
                    for item in Cell_var11[j]:
                        for key in item.keys():
                            if key == 'LocalCellId':
                                if item[key] not in Cell_list11:
                                    Cell_list11.append(item[key])
        for i in SECTOR_var11:
            for j in i:
                for j in range(0, len(SECTOR_var11)):
                    for item in SECTOR_var11[j]:
                        for key in item.keys():
                            if key == 'SECTORID':
                                if item[key] not in SECTOR_list11_unique:
                                    SECTOR_list11_unique.append(item[key])
        for i in SECTOREQM_var11:
            for j in i:
                for j in range(0, len(SECTOREQM_var11)):
                    for item in SECTOREQM_var11[j]:
                        for key in item.keys():
                            if key == 'SECTOREQMID':
                                if item[key] not in SECTOREQM_list11_unique:
                                    SECTOREQM_list11_unique.append(item[key])
        for i in RRU_var11:
            for j in i:
                for j in range(0, len(RRU_var11)):
                    for item in RRU_var11[j]:
                        for key in item.keys():
                            if key == 'SRN':
                                if item[key] not in RRU_list11_unique:
                                    RRU_list11_unique.append(item[key])
        for i in Cell_var11:
            for j in i:
                for j in range(0, len(Cell_var11)):
                    flag_11 = False
                    for item in Cell_var11[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['7', '8', '9']:
                                flag_11 = True
                                if item[key] not in Cell_band_1800_2100_list11:
                                    Cell_band_1800_2100_list11.append(item[key])
                            if flag_11 and key == 'TxRxMode':
                                if item[key] not in Cell_1800_work_mode_list11:
                                    Cell_1800_work_mode_list11.append(item[key])
                                    # example: Cell_1800_work_mode_list=['2','2','2'] , ['4','4','4'],['3','3','3'] / 2T2R~2, 2T4R~3, 4T4R~4

        # print('CELL WORK MODE:   ',Cell_1800_work_mode_list,Cell_band_1800_2100_list7)
        # defining port_list for SECTOR/SECTOREQM depending on Cell TxRxMode
        if len(Cell_1800_work_mode_list11)==0:
            sector_port_list_1800='0T0R'
        if '4' in Cell_1800_work_mode_list11 and '2' not in Cell_1800_work_mode_list11:  # all 3 1800 cells are 4T4R
            sector_port_list_1800 = '4T4R'
        if '4' in Cell_1800_work_mode_list11 and '2' in Cell_1800_work_mode_list11:  # some are 4T4R some are 2T2R - not found in the network
            sector_port_list_1800 = '4T4R'
        if '2' in Cell_1800_work_mode_list11 and '4' not in Cell_1800_work_mode_list11:  # all 3 1800 cells are 2T2R
            sector_port_list_1800 = '2T2R'
        if '3' in Cell_1800_work_mode_list11:
            sector_port_list_1800 = '4T4R'

        if '3' not in BBP_list11_unique:
            #if CheckVar7.get() == 0 and CheckVar80.get() == 0 and CheckVar8.get() == 0:
            if CheckVar7.get() == 0 and CheckkVar7==0:
                insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, Subrack_global, 3, 2, 4)

        if '18' in Cell_band_2100_list11 and '18' not in SECTOREQM_list11_unique:
            #rmv_old_rruchain_from_slotnumber_portnumber(3,3)
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(18, 0, Subrack_global, 3, 0)
            if CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0 and '100' not in RRU_list11_unique:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 18, 136, 15, 4, 4)
            if sector_port_list_1800=='2T2R':
                #insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(18, 0, 100, 0, [1, 3])
                #insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(18, 0, 100, 0, [1, 3])
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(18,7,0,100,0,[1,3])
            if sector_port_list_1800=='4T4R' and '7' in SECTOR_list11_unique:
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(18,7,0,100,0,[1,3])
            if sector_port_list_1800=='4T4R' and '7' not in SECTOR_list11_unique and '7' in Cell_band_1800_list11:
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(18, 7, 0, 100, 0, [1, 3])
            if sector_port_list_1800 == '0T0R': #no 1800
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(18, 0, 100, 0, [1, 3])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(18, 0, 100, 0, [1, 3])

        if '7' in Cell_band_1800_list11 and '7' in SECTOR_list11_unique and '67' in RRU_list11_unique:
            if sector_port_list_1800 == '2T2R':
                #mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(67, 0, 100, 0, [1, 3])
                mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(67,0,100,0,[0,1,2,3])
                mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(67, 0, 100, 0, [0, 2])
            if sector_port_list_1800 == '4T4R':
                mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(67, 0, 100, 0, [0,1,2, 3])
                mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(67, 0, 100, 0, [0,1,2, 3])
            if '18' not in Cell_band_2100_list11:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(18, 0, Subrack_global, 3, 0)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 18, 136, 15, 4, 4)
        if '7' in Cell_band_1800_list11 and '7' not in SECTOR_list11_unique and '67' not in RRU_list11_unique:
            if sector_port_list_1800 == '2T2R':
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0,1,2, 3])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0, 2])
            if sector_port_list_1800 == '4T4R':
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0,1,2, 3])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0,1,2, 3])
            if '18' not in Cell_band_2100_list11:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(18, 0, Subrack_global, 3, 0)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 18, 136, 15, 4, 4)

        if '19' in Cell_band_2100_list11 and '19' not in SECTOREQM_list11_unique:
            #rmv_old_rruchain_from_slotnumber_portnumber(3, 4)
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(19, 0, Subrack_global, 3, 1)
            if CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0 and '101' not in RRU_list11_unique:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 19, 136, 15, 4, 4)
            if sector_port_list_1800 == '2T2R':
                # insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 0, 101, 0, [1, 3])
                # insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 0, 101, 0, [1, 3])
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 8, 0, 101, 0, [1, 3])
            if sector_port_list_1800 == '4T4R' and '8' in SECTOR_list11_unique:
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 8, 0, 101, 0, [1, 3])
            if sector_port_list_1800=='4T4R' and '8' not in SECTOR_list11_unique and '8' in Cell_band_1800_list11:
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 8, 0, 101, 0, [1, 3])
            if sector_port_list_1800 == '0T0R':  # no 1800
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 0, 101, 0, [1, 3])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 0, 101, 0, [1, 3])

        if '8' in Cell_band_1800_list11 and '8' in SECTOR_list11_unique and '68' in RRU_list11_unique:
            if sector_port_list_1800 == '2T2R':
                #mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(68, 0, 101, 0, [1, 3])
                mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(68, 0, 101, 0, [0, 1, 2, 3])
                mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(68, 0, 101, 0, [0, 2])
            if sector_port_list_1800 == '4T4R':
                mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(68, 0, 101, 0, [0, 1, 2, 3])
                mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(68, 0, 101, 0, [0, 1, 2, 3])
            if '19' not in Cell_band_2100_list11:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(19, 0, Subrack_global, 3, 1)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 19, 136, 15, 4, 4)
        if '8' in Cell_band_1800_list11 and '8' not in SECTOR_list11_unique and '68' not in RRU_list11_unique:
            if sector_port_list_1800 == '2T2R':
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0,1,2, 3])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 2])
            if sector_port_list_1800 == '4T4R':
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 1, 2, 3])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 1, 2, 3])
            if '19' not in Cell_band_2100_list11:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(19, 0, Subrack_global, 3, 1)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 19, 136, 15, 4, 4)

        if '20' in Cell_band_2100_list11 and '20' not in SECTOREQM_list11_unique:
            #rmv_old_rruchain_from_slotnumber_portnumber(3, 5)
            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(20, 0, Subrack_global, 3, 2)
            if CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 == 0 and CheckkVar4 == 0 and CheckkVar5 == 0 and '102' not in RRU_list11_unique:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 20, 136, 15, 4, 4)
            if sector_port_list_1800 == '2T2R':
                # insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 0, 102, 0, [1, 3])
                # insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 0, 102, 0, [1, 3])
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 9, 0, 102, 0, [1, 3])
            if sector_port_list_1800 == '4T4R' and '9' in SECTOR_list11_unique:
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 9, 0, 102, 0, [1, 3])
            if sector_port_list_1800=='4T4R' and '9' not in SECTOR_list11_unique and '9' in Cell_band_1800_list11:
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 9, 0, 102, 0, [1, 3])
            if sector_port_list_1800 == '0T0R':  # no 1800
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 0, 102, 0, [1, 3])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 0, 102, 0, [1, 3])

        if '9' in Cell_band_1800_list11 and '9' in SECTOR_list11_unique and '69' in RRU_list11_unique:
            if sector_port_list_1800 == '2T2R':
                #mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(69, 0, 102, 0, [1, 3])
                mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(69, 0, 102, 0, [0, 1, 2, 3])
                mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(69, 0, 102, 0, [0, 2])
            if sector_port_list_1800 == '4T4R':
                mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(69, 0, 102, 0, [0, 1, 2, 3])
                mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(69, 0, 102, 0, [0, 1, 2, 3])
            if '20' not in Cell_band_2100_list11:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(20, 0, Subrack_global, 3, 2)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 20, 136, 15, 4, 4)
        if '9' in Cell_band_1800_list11 and '9' not in SECTOR_list11_unique and '69' not in RRU_list11_unique:
            if sector_port_list_1800 == '2T2R':
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0,1,2, 3])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 2])
            if sector_port_list_1800 == '4T4R':
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 1, 2, 3])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 1, 2, 3])
            if '20' not in Cell_band_2100_list11:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(20, 0, Subrack_global, 3, 2)
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 20, 136, 15, 4, 4)

        # rmv_old_rruchain_from_slotnumber_portnumber(3, 0)
        # rmv_old_rruchain_from_slotnumber_portnumber(3, 1)
        # rmv_old_rruchain_from_slotnumber_portnumber(3, 2)
        rmv_old_rruchain_from_slotnumber_portnumber(3, 3)
        rmv_old_rruchain_from_slotnumber_portnumber(3, 4)
        rmv_old_rruchain_from_slotnumber_portnumber(3, 5)
        rmv_old_rruchain_bychainID(7)
        rmv_old_rruchain_bychainID(8)
        rmv_old_rruchain_bychainID(9)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(67, 7)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(68, 8)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(69, 9)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(67, 17)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(68, 18)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(69, 19)
        rmv_RET_only(100)
        rmv_RET_only(101)
        rmv_RET_only(102)
        rmv_RET_only(4) # delete RETs controlled by RFU on SRN 4 1800/2100
        rmv_RETSUBUNIT_only(17)
        rmv_RETSUBUNIT_only(18)
        rmv_RETSUBUNIT_only(19)
        rmv_RETSUBUNIT_only(7)
        rmv_RETSUBUNIT_only(8)
        rmv_RETSUBUNIT_only(9)
        rmv_RETSUBUNIT_only(20)
        rmv_RETSUBUNIT_only(21)
        rmv_RETSUBUNIT_only(22)

        if '2' in BBP_list11_unique and ('4' in Cell_list11 or '5' in Cell_list11 or '6' in Cell_list11 or '12' in Cell_list11 or '13' in Cell_list11 or '14' in Cell_list11):
            pass
        else:
            rmv_old_bbp_from_slotnumber(2)

        #Label(window, text='5900: L2100 BD INT + L1800 AC REC/INT PN_3.4.5', width=80, fg='green').grid(column=2,row=20)
        text = Text(window, height=1, width=130, font=("arial", 10))
        text.insert(INSERT, '5900: L2100 BD INT + L1800 AC/ABCD REC/INT PN_0.1.2')
        text.grid(column=2, row=20)
        fo = open('LOG.txt', 'a')
        fo.write(f'{dt.datetime.now()} --  {NEname}  ' + '5900: L2100 BD INT + L1800 AC/ABCD REC/INT PN_0.1.2\n')
    #########################################################################################################################################################################################
    # 5900: L1800 AC INT + OnAir L2100 BD _PN_3.4.5
    if CheckVar12.get() == 1 or CheckkVar12 == 1:
        BBP12 = parse_existing_MO_name(root, 'BBP')
        SECTOR_var12 = parse_existing_MO_name(root, 'SECTOR')
        SECTOREQM_var12=parse_existing_MO_name(root, 'SECTOREQM')
        Chain_var12 = parse_existing_MO_name(root, 'RRUCHAIN')
        Cell_var12 = parse_existing_MO_name(root, 'Cell')
        RRU_var12 = parse_existing_MO_name(root, 'RRU')
        # print(BBP12)
        BBP_list12_unique = []
        SECTOR_list12_unique = []
        SECTOREQM_list12_unique=[]
        Chain_list12_2100_unique = []
        Chain_list12_1800_unique = []
        Cell_band_1800_list12 = []
        Cell_band_2100_list12 = []
        Cell_band_1800_2100_list12=[]
        Cell_1800_work_mode_list12=[]
        RRU_list12_unique = []
        Cell_list12 = []
        flag_12=False

        for i in BBP12:
            for j in i:
                for j in range(0, len(BBP12)):
                    for item in BBP12[j]:
                        for key in item.keys():
                            if key == 'SN':
                                if item[key] not in BBP_list12_unique:
                                    BBP_list12_unique.append(item[key])
        for i in Chain_var12:
            for j in i:
                for j in range(0, len(Chain_var12)):
                    for item in Chain_var12[j]:
                        for key in item.keys():
                            if key == 'RCN' and item[key] in ['18', '19', '20', '21']:
                                if item[key] not in Chain_list12_2100_unique:
                                    Chain_list12_2100_unique.append(item[key])
                            if key == 'RCN' and item[key] in ['7', '8', '9']:
                                if item[key] not in Chain_list12_1800_unique:
                                    Chain_list12_1800_unique.append(item[key])
        for i in Cell_var12:
            for j in i:
                for j in range(0, len(Cell_var12)):
                    for item in Cell_var12[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['7', '8', '9']:
                                if item[key] not in Cell_band_1800_list12:
                                    Cell_band_1800_list12.append(item[key])
                            if key == 'LocalCellId' and item[key] in ['18', '19', '20', '21']:
                                if item[key] not in Cell_band_2100_list12:
                                    Cell_band_2100_list12.append(item[key])
        for i in Cell_var12:
            for j in i:
                for j in range(0, len(Cell_var12)):
                    for item in Cell_var12[j]:
                        for key in item.keys():
                            if key == 'LocalCellId':
                                if item[key] not in Cell_list12:
                                    Cell_list12.append(item[key])
        for i in SECTOR_var12:
            for j in i:
                for j in range(0, len(SECTOR_var12)):
                    for item in SECTOR_var12[j]:
                        for key in item.keys():
                            if key == 'SECTORID':
                                if item[key] not in SECTOR_list12_unique:
                                    SECTOR_list12_unique.append(item[key])
        for i in SECTOREQM_var12:
            for j in i:
                for j in range(0, len(SECTOREQM_var12)):
                    for item in SECTOREQM_var12[j]:
                        for key in item.keys():
                            if key == 'SECTOREQMID':
                                if item[key] not in SECTOREQM_list12_unique:
                                    SECTOREQM_list12_unique.append(item[key])
        for i in RRU_var12:
            for j in i:
                for j in range(0, len(RRU_var12)):
                    for item in RRU_var12[j]:
                        for key in item.keys():
                            if key == 'SRN':
                                if item[key] not in RRU_list12_unique:
                                    RRU_list12_unique.append(item[key])


        for i in Cell_var12:
            for j in i:
                for j in range(0, len(Cell_var12)):
                    flag_12 = False
                    for item in Cell_var12[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['7', '8', '9']:
                                flag_12 = True
                                if item[key] not in Cell_band_1800_2100_list12:
                                    Cell_band_1800_2100_list12.append(item[key])
                            if flag_12 and key == 'TxRxMode':
                                if item[key] not in Cell_1800_work_mode_list12:
                                    Cell_1800_work_mode_list12.append(item[key])
                                    # example: Cell_1800_work_mode_list=['2','2','2'] , ['4','4','4'],['3','3','3'] / 2T2R~2, 2T4R~3, 4T4R~4

        # print('CELL WORK MODE:   ',Cell_1800_work_mode_list,Cell_band_1800_2100_list7)
        # defining port_list for SECTOR/SECTOREQM depending on Cell TxRxMode
        if len(Cell_1800_work_mode_list12)==0:
            sector_port_list_1800='0T0R'
        if '4' in Cell_1800_work_mode_list12 and '2' not in Cell_1800_work_mode_list12:  # all 3 1800 cells are 4T4R
            sector_port_list_1800 = '4T4R'
        if '4' in Cell_1800_work_mode_list12 and '2' in Cell_1800_work_mode_list12:  # some are 4T4R some are 2T2R - not found in the network
            sector_port_list_1800 = '4T4R'
        if '2' in Cell_1800_work_mode_list12 and '4' not in Cell_1800_work_mode_list12:  # all 3 1800 cells are 2T2R
            sector_port_list_1800 = '2T2R'
        if '3' in Cell_1800_work_mode_list12:
            sector_port_list_1800 = '4T4R'

        if '3' not in BBP_list12_unique:
            insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, Subrack_global, 3, 2, 4)

        if '18' in Cell_band_2100_list12 and '18' in SECTOREQM_list12_unique and '7' not in Cell_band_1800_list12:
            mod_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(100, 0, 100, 0, 4, 4, 136, 15)
            # leave the RRUCHAIN 18,19,20 on ports 3.4.5
            mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(100, 0, 100, 0, [0,1,2, 3])
            mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(100, 0, 100, 0, [1, 3])

        if '19' in Cell_band_2100_list12 and '19' in SECTOREQM_list12_unique and '8' not in Cell_band_1800_list12:
            mod_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(101, 0, 101, 0, 4, 4, 136, 15)
            # leave the RRUCHAIN 18,19,20 on ports 3.4.5
            mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(101, 0, 101, 0, [0, 1, 2, 3])
            mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(101, 0, 101, 0, [1, 3])

        if '20' in Cell_band_2100_list12 and '20' in SECTOREQM_list12_unique and '9' not in Cell_band_1800_list12:
            mod_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(102, 0, 102, 0, 4, 4, 136, 15)
            # leave the RRUCHAIN 18,19,20 on ports 3.4.5
            mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(102, 0, 102, 0, [0, 1, 2, 3])
            mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(102, 0, 102, 0, [1, 3])

        if '21' in Cell_band_2100_list12 and '21' in SECTOREQM_list12_unique and  len(Cell_band_1800_list12)==0:
            mod_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(103, 0, 103, 0, 4, 4, 136, 15)
            # leave the RRUCHAIN 18,19,20,21 on ports 3.4.5
            mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(103, 0, 103, 0, [0, 1, 2, 3])
            mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(103, 0, 103, 0, [1, 3])

        if '7' in Cell_band_1800_list12 and '7' not in SECTOREQM_list12_unique:
            if '18' not in Cell_band_2100_list12  and '7' not in Chain_list12_1800_unique:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(7, 0, Subrack_global, 3, 0)
                rmv_old_rruchain_bychainID(18)
            if '18' not in Cell_band_2100_list12 and '100' not in RRU_list12_unique:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 7, 136, 15, 4, 4)
            if '18' not in Cell_band_2100_list12 and '18' not in SECTOREQM_list12_unique and '18' not in SECTOR_list12_unique:
                if sector_port_list_1800 =='2T2R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0,1, 2,3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0, 2])
                if sector_port_list_1800 == '4T4R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0,1, 2,3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0,1, 2,3])
            if '18' in Cell_band_2100_list12 and '18' in SECTOREQM_list12_unique:
                if sector_port_list_1800 == '2T2R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(7,0,100,0,[0,1,2,3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7,0,100,0,[0,2])
                    #insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7,18,0,100,0,[0,2])
                if sector_port_list_1800 == '4T4R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0, 1, 2, 3])
                    #insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 7, 0, 100, 0, [0,1, 2,3])
                mod_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(100, 0, 100, 0, 4, 4, 136, 15)
                mod_RRU_onlyRRUCHAINno(100,18,7) # new function in version 1.6 used for 2100 RRU moved on new chain value
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(7, 0, Subrack_global, 3, 0)
                rmv_old_rruchain_bychainID(18)
                rmv_old_SECTOR_ID(18)
                rmv_old_SECTOREQM_ID(18)
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(18,7,0,100,0,[1,3])

        if '8' in Cell_band_1800_list12 and '8' not in SECTOREQM_list12_unique:
            if '19' not in Cell_band_2100_list12 and '8' not in Chain_list12_1800_unique:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(8, 0, Subrack_global, 3, 1)
                rmv_old_rruchain_bychainID(19)
            if '19' not in Cell_band_2100_list12 and '101' not in RRU_list12_unique:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 8, 136, 15, 4, 4)
            if '19' not in Cell_band_2100_list12 and '19' not in SECTOREQM_list12_unique and '19' not in SECTOR_list12_unique:
                if sector_port_list_1800 == '2T2R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 2])
                if sector_port_list_1800 == '4T4R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 1, 2, 3])
            if '19' in Cell_band_2100_list12 and '19' in SECTOREQM_list12_unique:
                if sector_port_list_1800 == '2T2R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 2])
                    # insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7,18,0,100,0,[0,2])
                if sector_port_list_1800 == '4T4R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 1, 2, 3])
                    # insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 7, 0, 100, 0, [0,1, 2,3])
                mod_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(101, 0, 101, 0, 4, 4, 136, 15)
                mod_RRU_onlyRRUCHAINno(101, 19,8)  # new function in version 1.6 used for 2100 RRU moved on new chain value
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(8, 0, Subrack_global, 3, 1)
                rmv_old_rruchain_bychainID(19)
                rmv_old_SECTOR_ID(19)
                rmv_old_SECTOREQM_ID(19)
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 8, 0, 101, 0, [1, 3])

        if '9' in Cell_band_1800_list12 and '9' not in SECTOREQM_list12_unique:
            if '20' not in Cell_band_2100_list12 and '9' not in Chain_list12_1800_unique:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(9, 0, Subrack_global, 3, 2)
                rmv_old_rruchain_bychainID(20)
            if '20' not in Cell_band_2100_list12 and '102' not in RRU_list12_unique:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102,0, 9, 136, 15, 4, 4)
            if '20' not in Cell_band_2100_list12 and '20' not in SECTOREQM_list12_unique and '20' not in SECTOR_list12_unique:
                if sector_port_list_1800 == '2T2R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 2])
                if sector_port_list_1800 == '4T4R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 1, 2, 3])
            if '20' in Cell_band_2100_list12 and '20' in SECTOREQM_list12_unique:
                if sector_port_list_1800 == '2T2R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 2])
                    # insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7,18,0,100,0,[0,2])
                if sector_port_list_1800 == '4T4R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 1, 2, 3])
                    # insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 7, 0, 100, 0, [0,1, 2,3])
                mod_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(102, 0, 102, 0, 4, 4, 136, 15)
                mod_RRU_onlyRRUCHAINno(102, 20,9)  # new function in version 1.6 used for 2100 RRU moved on new chain value
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(9, 0, Subrack_global, 3, 2)
                rmv_old_rruchain_bychainID(20)
                rmv_old_SECTOR_ID(20)
                rmv_old_SECTOREQM_ID(20)
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 9, 0, 102, 0, [1, 3])

        if '2' in BBP_list12_unique and ('4' in Cell_list12 or '5' in Cell_list12 or '6' in Cell_list12 or '12' in Cell_list12 or '13' in Cell_list12 or '14' in Cell_list12):
            pass
        else:
            rmv_old_bbp_from_slotnumber(2)

        #Label(window, text='L1800 AC INT + OnAir L2100 BD _PN_3.4.5', width=80, fg='green').grid(column=2, row=21)
        text = Text(window, height=1, width=130, font=("arial", 10))
        text.insert(INSERT, 'L1800 AC/ABCD INT + OnAir L2100 BD _PN_0.1.2')
        text.grid(column=2, row=21)
        fo = open('LOG.txt', 'a')
        fo.write(f'{dt.datetime.now()} --  {NEname}  ' + 'L1800 AC or ABCD INT + OnAir L2100 BD _PN_0.1.2\n')
    #########################################################################################################################################################################################
    # 5900: L1800 BD INT + OnAir L2100 AC _PN_3.4.5
    if CheckVar13.get() == 1 or CheckkVar13 == 1:
        BBP13 = parse_existing_MO_name(root, 'BBP')
        SECTOR_var13 = parse_existing_MO_name(root, 'SECTOR')
        SECTOREQM_var13 = parse_existing_MO_name(root, 'SECTOREQM')
        Chain_var13 = parse_existing_MO_name(root, 'RRUCHAIN')
        Cell_var13 = parse_existing_MO_name(root, 'Cell')
        RRU_var13 = parse_existing_MO_name(root, 'RRU')
        # print(BBP13)
        BBP_list13_unique = []
        SECTOR_list13_unique = []
        SECTOREQM_list13_unique = []
        Chain_list13_2100_unique = []
        Chain_list13_1800_unique=[]
        Cell_band_1800_list13 = []
        Cell_band_2100_list13 = []
        Cell_band_1800_2100_list13 = []
        Cell_1800_work_mode_list13 = []
        RRU_list13_unique = []
        Cell_list13 = []
        flag_13=False

        for i in BBP13:
            for j in i:
                for j in range(0, len(BBP13)):
                    for item in BBP13[j]:
                        for key in item.keys():
                            if key == 'SN':
                                if item[key] not in BBP_list13_unique:
                                    BBP_list13_unique.append(item[key])
        for i in Chain_var13:
            for j in i:
                for j in range(0, len(Chain_var13)):
                    for item in Chain_var13[j]:
                        for key in item.keys():
                            if key == 'RCN' and item[key] in ['18', '19', '20', '21']:
                                if item[key] not in Chain_list13_2100_unique:
                                    Chain_list13_2100_unique.append(item[key])
                            if key == 'RCN' and item[key] in ['7', '8', '9']:
                                if item[key] not in Chain_list13_1800_unique:
                                    Chain_list13_1800_unique.append(item[key])
        for i in Cell_var13:
            for j in i:
                for j in range(0, len(Cell_var13)):
                    for item in Cell_var13[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['7', '8', '9']:
                                if item[key] not in Cell_band_1800_list13:
                                    Cell_band_1800_list13.append(item[key])
                            if key == 'LocalCellId' and item[key] in ['18', '19', '20', '21']:
                                if item[key] not in Cell_band_2100_list13:
                                    Cell_band_2100_list13.append(item[key])
        for i in Cell_var13:
            for j in i:
                for j in range(0, len(Cell_var13)):
                    for item in Cell_var13[j]:
                        for key in item.keys():
                            if key == 'LocalCellId':
                                if item[key] not in Cell_list13:
                                    Cell_list13.append(item[key])
        for i in SECTOR_var13:
            for j in i:
                for j in range(0, len(SECTOR_var13)):
                    for item in SECTOR_var13[j]:
                        for key in item.keys():
                            if key == 'SECTORID':
                                if item[key] not in SECTOR_list13_unique:
                                    SECTOR_list13_unique.append(item[key])
        for i in SECTOREQM_var13:
            for j in i:
                for j in range(0, len(SECTOREQM_var13)):
                    for item in SECTOREQM_var13[j]:
                        for key in item.keys():
                            if key == 'SECTOREQMID':
                                if item[key] not in SECTOREQM_list13_unique:
                                    SECTOREQM_list13_unique.append(item[key])
        for i in RRU_var13:
            for j in i:
                for j in range(0, len(RRU_var13)):
                    for item in RRU_var13[j]:
                        for key in item.keys():
                            if key == 'SRN':
                                if item[key] not in RRU_list13_unique:
                                    RRU_list13_unique.append(item[key])
        for i in Cell_var13:
            for j in i:
                for j in range(0, len(Cell_var13)):
                    flag_13 = False
                    for item in Cell_var13[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['7', '8', '9']:
                                flag_13 = True
                                if item[key] not in Cell_band_1800_2100_list13:
                                    Cell_band_1800_2100_list13.append(item[key])
                            if flag_13 and key == 'TxRxMode':
                                if item[key] not in Cell_1800_work_mode_list13:
                                    Cell_1800_work_mode_list13.append(item[key])
                                    # example: Cell_1800_work_mode_list=['2','2','2'] , ['4','4','4'],['3','3','3'] / 2T2R~2, 2T4R~3, 4T4R~4

        # print('CELL WORK MODE:   ',Cell_1800_work_mode_list,Cell_band_1800_2100_list7)
        # defining port_list for SECTOR/SECTOREQM depending on Cell TxRxMode
        if len(Cell_1800_work_mode_list13)==0:
            sector_port_list_1800='0T0R'
        if '4' in Cell_1800_work_mode_list13 and '2' not in Cell_1800_work_mode_list13:  # all 3 1800 cells are 4T4R
            sector_port_list_1800 = '4T4R'
        if '4' in Cell_1800_work_mode_list13 and '2' in Cell_1800_work_mode_list13:  # some are 4T4R some are 2T2R - not found in the network
            sector_port_list_1800 = '4T4R'
        if '2' in Cell_1800_work_mode_list13 and '4' not in Cell_1800_work_mode_list13:  # all 3 1800 cells are 2T2R
            sector_port_list_1800 = '2T2R'
        if '3' in Cell_1800_work_mode_list13:
            sector_port_list_1800 = '4T4R'

        if '3' not in BBP_list13_unique:
            insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, Subrack_global, 3, 2, 4)

        if '18' in Cell_band_2100_list13 and '18' in SECTOREQM_list13_unique and '7' not in Cell_band_1800_list13:
            mod_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(100, 0, 100, 0, 4, 4, 136, 15)
            # leave the RRUCHAIN 18,19,20 on ports 3.4.5
            mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(100, 0, 100, 0, [0, 1, 2, 3])
            mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(100, 0, 100, 0, [0, 2])

        if '19' in Cell_band_2100_list13 and '19' in SECTOREQM_list13_unique and '8' not in Cell_band_1800_list13:
            mod_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(101, 0, 101, 0, 4, 4, 136, 15)
            # leave the RRUCHAIN 18,19,20 on ports 3.4.5
            mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(101, 0, 101, 0, [0, 1, 2, 3])
            mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(101, 0, 101, 0, [0, 2])

        if '20' in Cell_band_2100_list13 and '20' in SECTOREQM_list13_unique and '9' not in Cell_band_1800_list13:
            mod_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(102, 0, 102, 0, 4, 4, 136, 15)
            # leave the RRUCHAIN 18,19,20 on ports 3.4.5
            mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(102, 0, 102, 0, [0, 1, 2, 3])
            mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(102, 0, 102, 0, [0, 2])

        if '21' in Cell_band_2100_list13 and '21' in SECTOREQM_list13_unique and len(Cell_band_1800_list13) == 0:
            mod_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(103, 0, 103, 0, 4, 4, 136, 15)
            # leave the RRUCHAIN 18,19,20,21 on ports 3.4.5
            mod_SECTOR_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(103, 0, 103, 0, [0, 1, 2, 3])
            mod_SECTOREQM_oldSRN_newCN_newSRN_newSN_newANTENNAPORTlist(103, 0, 103, 0, [0, 2])

        if '7' in Cell_band_1800_list13 and '7' not in SECTOREQM_list13_unique:
            if '18' not in Cell_band_2100_list13 and '7' not in Chain_list13_1800_unique:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(7, 0, Subrack_global, 3, 0)
                rmv_old_rruchain_bychainID(18)
            if '18' not in Cell_band_2100_list13 and '100' not in RRU_list13_unique:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 7, 136, 15, 4, 4)
            if '18' not in Cell_band_2100_list13 and '18' not in SECTOREQM_list13_unique and '18' not in SECTOR_list13_unique:
                if sector_port_list_1800 == '2T2R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [1, 3])
                if sector_port_list_1800 == '4T4R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0, 1, 2, 3])
            if '18' in Cell_band_2100_list13 and '18' in SECTOREQM_list13_unique:
                if sector_port_list_1800 == '2T2R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [1, 3])
                    # insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7,18,0,100,0,[0,2])
                if sector_port_list_1800 == '4T4R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 0, 100, 0, [0, 1, 2, 3])
                    # insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 7, 0, 100, 0, [0,1, 2,3])
                mod_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(100, 0, 100, 0, 4, 4, 136, 15)
                mod_RRU_onlyRRUCHAINno(100, 18,7)  # new function in version 1.6 used for 2100 RRU moved on new chain value
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(7, 0, Subrack_global, 3, 0)
                rmv_old_rruchain_bychainID(18)
                rmv_old_SECTOR_ID(18)
                rmv_old_SECTOREQM_ID(18)
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(18, 7, 0, 100, 0, [0, 2])

        if '8' in Cell_band_1800_list13 and '8' not in SECTOREQM_list13_unique:
            if '19' not in Cell_band_2100_list13 and '8' not in Chain_list13_1800_unique:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(8, 0, Subrack_global, 3, 1)
                rmv_old_rruchain_bychainID(19)
            if '19' not in Cell_band_2100_list13 and '101' not in RRU_list13_unique:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 8, 136, 15, 4, 4)
            if '19' not in Cell_band_2100_list13 and '19' not in SECTOREQM_list13_unique and '19' not in SECTOR_list13_unique:
                if sector_port_list_1800 == '2T2R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [1, 3])
                if sector_port_list_1800 == '4T4R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 1, 2, 3])
            if '19' in Cell_band_2100_list13 and '19' in SECTOREQM_list13_unique:
                if sector_port_list_1800 == '2T2R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [1, 3])
                    # insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7,18,0,100,0,[0,2])
                if sector_port_list_1800 == '4T4R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(8, 0, 101, 0, [0, 1, 2, 3])
                    # insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 7, 0, 100, 0, [0,1, 2,3])
                mod_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(101, 0, 101, 0, 4, 4, 136, 15)
                mod_RRU_onlyRRUCHAINno(101, 19,8)  # new function in version 1.6 used for 2100 RRU moved on new chain value
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(8, 0, Subrack_global, 3, 1)
                rmv_old_rruchain_bychainID(19)
                rmv_old_SECTOR_ID(19)
                rmv_old_SECTOREQM_ID(19)
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(19, 8, 0, 101, 0, [0, 2])

        if '9' in Cell_band_1800_list13 and '9' not in SECTOREQM_list13_unique:
            if '20' not in Cell_band_2100_list13 and '9' not in Chain_list13_1800_unique:
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(9, 0, Subrack_global, 3, 2)
                rmv_old_rruchain_bychainID(20)
            if '20' not in Cell_band_2100_list13 and '102' not in RRU_list13_unique:
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102,0, 9, 136, 15, 4, 4)
            if '20' not in Cell_band_2100_list13 and '20' not in SECTOREQM_list13_unique and '20' not in SECTOR_list13_unique:
                if sector_port_list_1800 == '2T2R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [1, 3])
                if sector_port_list_1800 == '4T4R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 1, 2, 3])
            if '20' in Cell_band_2100_list13 and '20' in SECTOREQM_list13_unique:
                if sector_port_list_1800 == '2T2R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [1, 3])
                    # insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7,18,0,100,0,[0,2])
                if sector_port_list_1800 == '4T4R':
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 1, 2, 3])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(9, 0, 102, 0, [0, 1, 2, 3])
                    # insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(7, 7, 0, 100, 0, [0,1, 2,3])
                mod_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_args_oldSRN_newCN_newSRN_newSN__newTXNUM_newRXNUM_newWM_newRRUType(102, 0, 102, 0, 4, 4, 136, 15)
                mod_RRU_onlyRRUCHAINno(102, 20,9)  # new function in version 1.6 used for 2100 RRU moved on new chain value
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(9, 0, Subrack_global, 3, 2)
                rmv_old_rruchain_bychainID(20)
                rmv_old_SECTOR_ID(20)
                rmv_old_SECTOREQM_ID(20)
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(20, 9, 0, 102, 0, [0, 2])

        if '2' in BBP_list13_unique and ('4' in Cell_list13 or '5' in Cell_list13 or '6' in Cell_list13 or '12' in Cell_list13 or '13' in Cell_list13 or '14' in Cell_list13):
            pass
        else:
            rmv_old_bbp_from_slotnumber(2)

        #Label(window, text='L1800 BD INT + OnAir L2100 AC _PN_3.4.5', width=80, fg='green').grid(column=2, row=22)
        text = Text(window, height=1, width=130, font=("arial", 10))
        text.insert(INSERT, 'L1800 BD/ABCD INT + OnAir L2100 AC _PN_0.1.2')
        text.grid(column=2, row=22)
        fo = open('LOG.txt', 'a')
        fo.write(f'{dt.datetime.now()} --  {NEname}  ' + 'L1800 BD or ABCD INT + OnAir L2100 AC _PN_0.1.2\n')
    #########################################################################################################################################################################################
    # 5900: LTE2600 INT AB/ABCD
    if CheckVar14.get() == 1 or CheckkVar14 == 1:
        BBP14 = parse_existing_MO_name(root, 'BBP')
        SECTOR_var14 = parse_existing_MO_name(root, 'SECTOR')
        Chain_var14 = parse_existing_MO_name(root, 'RRUCHAIN')
        Cell_var14 = parse_existing_MO_name(root, 'Cell')
        RRU_var14 = parse_existing_MO_name(root, 'RRU')
        # print(BBP14)
        BBP_list14_unique = []
        SECTOR_list14_unique = []
        Chain_list14_2100_unique = []
        Cell_band_1800_list14 = []
        Cell_band_2100_list14 = []
        RRU_list14_unique = []
        Cell_band_2600_list14 = []
        Cell_band_2600_list14_1=[]
        Cell_2600_work_mode_list14=[]
        flag_14 = False

        for i in BBP14:
            for j in i:
                for j in range(0, len(BBP14)):
                    for item in BBP14[j]:
                        for key in item.keys():
                            if key == 'SN':
                                if item[key] not in BBP_list14_unique:
                                    BBP_list14_unique.append(item[key])

        for i in Cell_var14:
            for j in i:
                for j in range(0, len(Cell_var14)):
                    for item in Cell_var14[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['4', '5', '6']:
                                if item[key] not in Cell_band_2600_list14:
                                    Cell_band_2600_list14.append(item[key])
        for i in SECTOR_var14:
            for j in i:
                for j in range(0, len(SECTOR_var14)):
                    for item in SECTOR_var14[j]:
                        for key in item.keys():
                            if key == 'SECTORID':
                                if item[key] not in SECTOR_list14_unique:
                                    SECTOR_list14_unique.append(item[key])
        for i in RRU_var14:
            for j in i:
                for j in range(0, len(RRU_var14)):
                    for item in RRU_var14[j]:
                        for key in item.keys():
                            if key == 'SRN':
                                if item[key] not in RRU_list14_unique:
                                    RRU_list14_unique.append(item[key])


        for i in Cell_var14:
            for j in i:
                for j in range(0, len(Cell_var14)):
                    flag_14 = False
                    for item in Cell_var14[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['4', '5', '6']:
                                flag_14 = True
                                if item[key] not in Cell_band_2600_list14_1:
                                    Cell_band_2600_list14_1.append(item[key])
                            if flag_14 and key == 'TxRxMode':
                                if item[key] not in Cell_2600_work_mode_list14:
                                    Cell_2600_work_mode_list14.append(item[key])
                                    # example: Cell_1800_work_mode_list=['2','2','2'] , ['4','4','4'],['3','3','3'] / 2T2R~2, 2T4R~3, 4T4R~4

        # print('CELL WORK MODE:   ',Cell_1800_work_mode_list,Cell_band_1800_2100_list7)
        # defining port_list for SECTOR/SECTOREQM depending on Cell TxRxMode
        if len(Cell_2600_work_mode_list14)==0:
            sector_port_list_2600='0T0R'
        if '4' in Cell_2600_work_mode_list14 and '2' not in Cell_2600_work_mode_list14:  # all 3 1800 cells are 4T4R
            sector_port_list_2600 = '4T4R'
        if '4' in Cell_2600_work_mode_list14 and '2' in Cell_2600_work_mode_list14:  # some are 4T4R some are 2T2R - not found in the network
            sector_port_list_2600 = '4T4R'
        if '2' in Cell_2600_work_mode_list14 and '4' not in Cell_2600_work_mode_list14:  # all 3 1800 cells are 2T2R
            sector_port_list_2600 = '2T2R'
        if '3' in Cell_2600_work_mode_list14:
            sector_port_list_2600 = '4T4R'


        if '2' in BBP_list14_unique:
            mod_BBP_TYPE_from_slot_to_slot(2, 2, 12290, 2)

        # if '2' not in BBP_list14_unique:
        #     insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, Subrack_global, 2, 2, 4)

        if '5' not in BBP_list14_unique and CheckVar3.get() ==0 and CheckVar4.get() ==0 and CheckVar5.get() ==0 and CheckVar6.get() ==0 and CheckkVar3==0 and CheckkVar4==0 and CheckkVar5==0 and CheckkVar6==0:
            # board slot 5 is not added yet at this point by any previous option or from the beggining.
            insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0, Subrack_global, 5, 2, 4)

        Cell_band_2600_list14_int = [int(i) for i in Cell_band_2600_list14]
        for i in Cell_band_2600_list14_int:

            insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(i, 0, Subrack_global, 5, i - 1)
            if CheckVar30.get() == 0 and CheckVar3.get() == 0 and CheckVar4.get() == 0 and CheckVar5.get() == 0 and CheckkVar3 ==0 and CheckkVar4 ==0 and CheckkVar5 ==0 and CheckkVar30==0 :
                insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 60 + i, 0, i, 8, 15, 4, 4)
            if sector_port_list_2600=='2T2R':
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(i, 0, 60 + i, 0, [0, 1])  #LTE2600 AB
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(i, 0, 60 + i, 0, [0, 1])
            if sector_port_list_2600=='4T4R':
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(i, 0, 60 + i, 0, [0, 1,2,3])  #LTE2600 ABCD
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(i, 0, 60 + i, 0, [0, 1,2,3])

        #Label(window, text='5900: LTE2600 INTEGRATION slot 5/3.4.5 Executed.. ', width=80, fg='green').grid(column=2,row=23)
        text = Text(window, height=1, width=130, font=("arial", 10))
        text.insert(INSERT, '5900: LTE2600 INTEGRATION slot 5/3.4.5 Executed..')
        text.grid(column=2, row=23)
        fo = open('LOG.txt', 'a')
        fo.write(f'{dt.datetime.now()} --  {NEname}  ' + '5900: LTE2600 INTEGRATION slot 5/3.4.5 Executed..\n')

    #########################################################################################################################################################################################
    # LTE900  INTEGRATION
    if CheckVar15.get() == 1 or CheckkVar15 == 1:
        BBP15 = parse_existing_MO_name(root, 'BBP')
        SECTOR_var15 = parse_existing_MO_name(root, 'SECTOR')
        Chain_var15 = parse_existing_MO_name(root, 'RRUCHAIN')
        Cell_var15 = parse_existing_MO_name(root, 'Cell')
        RRU_var15 = parse_existing_MO_name(root, 'RRU')
        #print(RRU_var15)
        BBP_list15_unique = []
        SECTOR_list15_unique = []
        Chain_list15__unique = []
        RRU_list15_unique = []
        Cell_list15 = []
        Cell_band_900_list15 = []
        for i in BBP15:
            for j in i:
                for j in range(0, len(BBP15)):
                    for item in BBP15[j]:
                        for key in item.keys():
                            if key == 'SN':
                                if item[key] not in BBP_list15_unique:
                                    BBP_list15_unique.append(item[key])

        for i in SECTOR_var15:
            for j in i:
                for j in range(0, len(SECTOR_var15)):
                    for item in SECTOR_var15[j]:
                        for key in item.keys():
                            if key == 'SECTORID':
                                if item[key] not in SECTOR_list15_unique:
                                    SECTOR_list15_unique.append(item[key])
        for i in RRU_var15:
            for j in i:
                for j in range(0, len(RRU_var15)):
                    for item in RRU_var15[j]:
                        for key in item.keys():
                            if key == 'SRN':
                                if item[key] not in RRU_list15_unique:
                                    RRU_list15_unique.append(item[key])
        #print(RRU_list15_unique)

        for i in Cell_var15:
            for j in i:
                for j in range(0, len(Cell_var15)):
                    for item in Cell_var15[j]:
                        for key in item.keys():
                            if key == 'LocalCellId':
                                if item[key] not in Cell_list15:
                                    Cell_list15.append(item[key])
        for i in Cell_var15:
            for j in i:
                for j in range(0, len(Cell_var15)):
                    for item in Cell_var15[j]:
                        for key in item.keys():
                            if key == 'LocalCellId' and item[key] in ['24', '25', '26']:
                                if item[key] not in Cell_band_900_list15:
                                    Cell_band_900_list15.append(item[key])

        # L800 AB INT 80-82-84, no L700 5519et"
        if (CheckVar300.get() == 1 or CheckkVar300 ==1):
            for i in range(0, len(Cell_band_900_list15)):
                if '30' not in Cell_list15 and '31' not in Cell_list15 and '32' not in Cell_list15:
                    #insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(24 + i, 0, 80 + 2 * i, 0, [0, 1])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(24 + i, 0, 80 + 2 * i, 0, [0, 1])

            #Label(window, text='5900: LTE900 INTEGRATED , sharing RRU with LTE800, no LTE700 cells on site , [RRU5519et]', width=80,fg='green').grid(column=2, row=24)
            text = Text(window, height=1, width=130, font=("arial", 10))
            text.insert(INSERT, '5900: LTE900 INTEGRATED , sharing RRU with LTE800, no LTE700 cells on site , [RRU5519et]')
            text.grid(column=2, row=24)
            fo = open('LOG.txt', 'a')
            fo.write(f'{dt.datetime.now()} --  {NEname}  ' + '5900: LTE900 INTEGRATED , sharing RRU with LTE800, no LTE700 on site..\n')


        # L800 CD INT 80-82-84, no L700 5509t"
        if (CheckVar30.get() == 1 or CheckkVar30 ==1):
            for i in range(0, len(Cell_band_900_list15)):
                if '30' not in Cell_list15 and '31' not in Cell_list15 and '32' not in Cell_list15:
                    #insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(24 + i, 0, 80 + 2 * i, 0, [0, 1])
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(24 + i, 0, 80 + 2 * i, 0, [0, 1])

            #Label(window, text='5900: LTE900 INTEGRATED , sharing RRU with LTE800, no LTE700 cells on site, [RRU5509t]', width=80,fg='green').grid(column=2, row=24)
            text = Text(window, height=1, width=130, font=("arial", 10))
            text.insert(INSERT,'5900: LTE900 INTEGRATED , sharing RRU with LTE800, no LTE700 cells on site, [RRU5509t]')
            text.grid(column=2, row=24)
            fo = open('LOG.txt', 'a')
            fo.write(f'{dt.datetime.now()} --  {NEname}  ' + '5900: LTE900 INTEGRATED , sharing RRU with LTE800, no LTE700 on site..\n')

        # L800 CD REC + L700 AB INT 80-82-84 5509t
        if ((CheckVar4.get() == 1 or CheckkVar4 == 1) and '30' in Cell_list15) or ((CheckVar4.get() == 1 or CheckkVar4 == 1) and '31' in Cell_list15) or ((CheckVar4.get() == 1 or CheckkVar4 == 1) and '32' in Cell_list15):
            for i in range(0, len(Cell_band_900_list15)):
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(24 + i, 0 + i, 0, 80 + 2 * i, 0, [0, 1])
            #Label(window, text='5900:  LTE900 INTEGRATED , sharing RRU with LTE800 CD + LTE700 AB on site.. ', width=80,fg='green').grid(column=2, row=24)
            text = Text(window, height=1, width=130, font=("arial", 10))
            text.insert(INSERT,'5900:  LTE900 INTEGRATED , sharing RRU with LTE800 CD + LTE700 AB on site..')
            text.grid(column=2, row=24)
            fo = open('LOG.txt', 'a')
            fo.write(f'{dt.datetime.now()} --  {NEname}  ' + '5900: LTE900 INTEGRATED , sharing RRU with LTE800 CD + LTE700 AB on site..\n')

        # L800 CD REC + L700 AB INT 80-82-84 5509t
        if (CheckVar4.get() == 1 or CheckkVar4 ==1) and ('30' not in Cell_list15 and '31' not in Cell_list15 and '32' not in Cell_list15):
            for i in range(0, len(Cell_band_900_list15)):
                # insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST((24 + i, 0, 80 + 2 * i, 0, [0, 1]))
                # insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(24 + i, 0, 80 + 2 * i, 0, [0, 1])
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(24 + i, 0 + i, 0, 80 + 2 * i, 0, [0, 1])
            #Label(window, text='5900: LTE900 INTEGRATED , sharing RRU with LTE800 CD + no LTE700 AB on site ', width=80,fg='green').grid(column=2, row=24)
            text = Text(window, height=1, width=130, font=("arial", 10))
            text.insert(INSERT, '5900: LTE900 INTEGRATED , sharing RRU with LTE800 CD + no LTE700 AB on site ')
            text.grid(column=2, row=24)
            fo = open('LOG.txt', 'a')
            fo.write(f'{dt.datetime.now()} --  {NEname}  ' + '5900:LTE900 INTEGRATED , sharing RRU with LTE800 CD + no LTE700 AB on site..\n')

        # L800 AB REC + L700 CD INT 80-82-84 5519et
        if ((CheckVar5.get() == 1 or CheckkVar5 == 1) and '30' in Cell_list15) or ((CheckVar5.get() == 1 or CheckkVar5 == 1) and '31' in Cell_list15) or ((CheckVar5.get() == 1 or CheckkVar5 == 1) and '32' in Cell_list15):
            for i in range(0, len(Cell_band_900_list15)):
                if str(0+i) in SECTOR_list15_unique:
                    insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(24 + i, 0 + i, 0, 80 + 2 * i, 0, [0, 1])
                if str(0+i) not in SECTOR_list15_unique and str(0+i) not in Cell_list15:
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST((24 + i, 0, 80 + 2 * i, 0, [0, 1]))
                    insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(24 + i, 0, 80 + 2 * i, 0, [0, 1])
            #Label(window, text='5900:  LTE900 INTEGRATED , sharing RRU with LTE800 AB + LTE700 CD on site.. ', width=80,fg='green').grid(column=2, row=24)
            text = Text(window, height=1, width=130, font=("arial", 10))
            text.insert(INSERT, '5900:  LTE900 INTEGRATED , sharing RRU with LTE800 AB + LTE700 CD on site.. ')
            text.grid(column=2, row=24)
            fo = open('LOG.txt', 'a')
            fo.write(f'{dt.datetime.now()} --  {NEname}  ' + '5900: LTE900 INTEGRATED , sharing RRU with LTE800 AB + LTE700 CD on site..\n')

        # L800 AB REC + L700 CD INT 80-82-84 5519et
        if (CheckVar5.get() == 1 or CheckkVar5 == 1) and ('30' not in Cell_list15 and '31' not in Cell_list15 and '32' not in Cell_list15):
            for i in range(0, len(Cell_band_900_list15)):
                # insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST((24 + i, 0, 80 + i, 0, [0, 1]))
                insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(24 + i, 0 + i, 0, 80 + 2 * i, 0, [0, 1])
            #Label(window, text='5900:  LTE900 INTEGRATED , sharing RRU with LTE800 AB, no LTE700 CD on site.. ',width=80, fg='green').grid(column=2, row=24)
            text = Text(window, height=1, width=130, font=("arial", 10))
            text.insert(INSERT, '5900:  LTE900 INTEGRATED , sharing RRU with LTE800 AB, no LTE700 CD on site.. ')
            text.grid(column=2, row=24)
            fo = open('LOG.txt', 'a')
            fo.write(f'{dt.datetime.now()} --  {NEname}  ' + '5900: LTE900 INTEGRATED , sharing RRU with LTE800 AB, no LTE700 CD on site..\n')

        # all above ifs  are auto excluding each other! (if one is happening, the others cannot happen)

        # No selection for 800 or 700 will lead either to 900 in single band RRUs, slot 4/3.4.5 or alongside L800 with existing 80-82-84 on ports 4/0.1.2
        if CheckVar3.get() == 0 and CheckkVar3 == 0 and CheckVar5.get()==0 and CheckkVar5==0 and CheckVar4.get()==0 and CheckkVar4==0 and CheckVar30.get()==0 and CheckkVar30==0 and CheckVar300.get()==0 and CheckkVar300==0:
            # no multiband 80-82-84 on site
            if '80' not in RRU_list15_unique and '82' not in RRU_list15_unique and '84' not in RRU_list15_unique: # we assume that L800 is not integrated on 80-82-84 / it could be on 60-61-62
                for i in range(0, len(Cell_band_900_list15)):
                    insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(24 + i, 0, Subrack_global, 4, 3 + i)
                    insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 80 + 2 * i, 0, 24 + i, 328, 15, 2, 2)
                    insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(24 + i, 0, 80 + 2 * i, 0, [0, 1])
                    insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(24+i,24+i,0,80+2*i,0,[0,1])
                    #insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(24 + i, 0, 80 + 2 * i, 0, [0, 1])
                #Label(window, text='5900:  LTE900 INTEGRATED , slot 4', width=80,fg='green').grid(column=2, row=24)
                text = Text(window, height=1, width=130, font=("arial", 10))
                text.insert(INSERT, '5900:  LTE900 INTEGRATED , slot 4/3.4.5, no multiband RRUs on site')
                text.grid(column=2, row=24)
                fo = open('LOG.txt', 'a')
                fo.write(f'{dt.datetime.now()} --  {NEname}  ' + '5900: LTE900 INTEGRATED , slot 4 , \n')

            #already existing multiband rrus 80-82-84
            if '80' in RRU_list15_unique and '82' in RRU_list15_unique and '84' in RRU_list15_unique: # we assume L800 is already integrated with 80-82-84 RRus
                for i in range(0, len(Cell_band_900_list15)):
                    insert_new_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(24 + i, i, 0, 80 + 2 * i, 0, [0, 1])
                #Label(window, text='5900:  LTE900 INTEGRATED , slot 4, on L800 multiband RRUs', width=80,fg='green').grid(column=2, row=24)
                text = Text(window, height=1, width=130, font=("arial", 10))
                text.insert(INSERT, '5900:  LTE900 INTEGRATED , slot 4, on L800 multiband RRUs')
                text.grid(column=2, row=24)
                fo = open('LOG.txt', 'a')
                fo.write(f'{dt.datetime.now()} --  {NEname}  ' + '5900:  LTE900 INTEGRATED , slot 4, on L800 multiband RRUs \n')

        # L800 AB REC/INT 60-61-62 SN 2,3 -> 4
        if (CheckVar3.get() == 1 or CheckkVar3 ==1):
            #if '80' not in RRU_list15_unique and '82' not in RRU_list15_unique and '84' not in RRU_list15_unique:
            for i in range(0, len(Cell_band_900_list15)):
                insert_new_RRUCHAIN_rruchainno_HCN_HSRN_HSN_HPN(24 + i, 0, Subrack_global, 4, 3 + i)
                #RRU inserted by CheckVar3/CheckkVar3 which is enabled
                insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(24 + i, 0, 80 + 2 * i, 0, [0, 1])
                insert_new_SECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(24 + i, 0, 80 + 2 * i, 0, [0, 1])
            #Label(window, text='5900:  LTE900 INTEGRATED , slot 4 ports 3.4.5 not shared RRU', width=80,fg='green').grid(column=2, row=24)
            text = Text(window, height=1, width=130, font=("arial", 10))
            text.insert(INSERT, '5900:  LTE900 INTEGRATED , slot 4 ports 3.4.5 not shared RRU')
            text.grid(column=2, row=24)
            fo = open('LOG.txt', 'a')
            fo.write(f'{dt.datetime.now()} --  {NEname}  ' + '5900: LTE900 INTEGRATED , slot 4 ports 3.4.5 not shared RRU , \n')
    #########################################################################################################################################################################################
    # Subrack ID change 1 to 0
    if CheckVar16.get() == 1 or CheckkVar16==1:
        #Label(window, text='BBU subrack ID change 1 ->0 executed..', width=80, fg='green').grid(column=2, row=8)
        change_bbu_subrack_id10()
        for child0 in root:
            for child1 in child0:
                for child2 in child1[::-1]:
                    if child2.tag == 'CTRLLNK':
                        for child3 in child2:
                            for child4 in child3:
                                if child4.tag == 'LN' and child4.text == str(0):
                                    child1.remove(child2)
                                    # ET.dump(child1)
                                    break
        for child0 in root:
            for child1 in child0:
                for child2 in child1[::-1]:
                    if child2.tag == 'TUNNEL':
                        child1.remove(child2)
                        # ET.dump(child1)
                        break

        tree.write(OUTPUT)
        fo = open('LOG.txt', 'a')
        fo.write(f'{dt.datetime.now()} --  {NEname}  ' + 'BBU subrack ID change  1 ->0 executed..\n')
        text = Text(window, height=1, width=130, font=("arial", 10))
        text.insert(INSERT, 'BBU subrack ID change 1 ->0 executed..')
        text.grid(column=2, row=8)
    #########################################################################################################################################################################################
    if CheckVar17.get() == 1:# or CheckkVar16 == 1:
        product_type_change_5900LTEto5900()
        #Label(window, text='Ne Type Reconstruction from the BTS5900 LTE to BTS5900 done..', width=80, fg='green').grid(column=2, row=25)
        text = Text(window, height=1, width=130, font=("arial", 10))
        text.insert(INSERT, 'Ne Type Reconstruction from the BTS5900 LTE to BTS5900 done..')
        text.grid(column=2, row=25)
    #########################################################################################################################################################################################
    # 3G  UBBP  refarming
    def UBBP_function(Checkvar,CheckkVar, TEXT, row):
        if Checkvar.get() == 1 or CheckkVar == 1:
            BBP18 = parse_existing_MO_name(root, 'BBP')
            SECTOR_var18 = parse_existing_MO_name(root, 'SECTOR')
            Chain_var18 = parse_existing_MO_name(root, 'RRUCHAIN')
            ULOCELL_var18 = parse_existing_MO_name(root, 'ULOCELL')
            RRU_var18 = parse_existing_MO_name(root, 'RRU')
            # print(BBP14)
            BBP_list18_unique = []
            SECTOR_list18_unique = []
            BBP_type_list18_unique = []
            Chain_list18__unique = []
            RRU_list18_unique = []
            Cell_list18 = []
            RRU_list18_unique = []
            # insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0,0,3,2,2)

            for i in RRU_var18:
                for j in i:
                    for j in range(0, len(RRU_var18)):
                        for item in RRU_var18[j]:
                            for key in item.keys():
                                if key == 'SRN':
                                    if item[key] not in RRU_list18_unique:
                                        RRU_list18_unique.append(item[key])
            for child0 in root:
                for child1 in child0:
                    for child2 in child1:
                        if child2.tag == '' + 'BBP':
                            for child3 in child2:
                                for child4 in child3:
                                    if child4.tag == 'TYPE' and child4.text not in BBP_type_list18_unique:
                                        BBP_type_list18_unique.append(child4.text)
            # print('BBP list content',BBP_type_list18_unique)
            for child0 in root:
                for child1 in child0:
                    for child2 in child1:
                        if child2.tag == '' + 'ULOCELL':
                            for child3 in child2:
                                for child4 in child3:
                                    if child4.tag == 'ULBASEBANDEQMID' and child4.text == str(
                                            1):
                                        child4.text = str(0)
            for child0 in root:
                for child1 in child0:
                    for child2 in child1:
                        if child2.tag == '' + 'RRU':
                            for child3 in child2:
                                for child4 in child3:
                                    if child4.tag == 'RS':
                                        child4.text = str(136)

            if '2' not in BBP_type_list18_unique:
                for child0 in root:
                    for child1 in child0:
                        for child2 in child1:
                            if child2.tag == '' + 'BASEBANDEQM':
                                sector_subelement0 = ET.SubElement(child1, 'BASEBANDEQM')
                                sector_subelement1 = ET.SubElement(sector_subelement0, 'attributes')
                                sector_subelement2 = ET.SubElement(sector_subelement1, 'BASEBANDEQMID')
                                sector_subelement2.text = str(0)
                                sector_subelement2 = ET.SubElement(sector_subelement1, 'BASEBANDEQMTYPE')
                                sector_subelement2.text = str(2)
                                sector_subelement2 = ET.SubElement(sector_subelement1, 'UMTSDEMMODE')
                                sector_subelement2.text = str(4)
                                sector_subelement2 = ET.SubElement(sector_subelement1, 'BASEBANDEQMBOARD')
                                sector_subelement3 = ET.SubElement(sector_subelement2, 'element')
                                sector_subelement4 = ET.SubElement(sector_subelement3, 'CN')
                                sector_subelement4.text = str(0)
                                sector_subelement4 = ET.SubElement(sector_subelement3, 'SRN')
                                sector_subelement4.text = str(Subrack_global)
                                sector_subelement4 = ET.SubElement(sector_subelement3, 'SN')
                                sector_subelement4.text = str(3)
                                # ET.dump(child1)
                                break
                for child0 in root:
                    for child1 in child0:
                        for child2 in child1[::-1]:
                            if child2.tag == 'BASEBANDEQM':
                                for child3 in child2:
                                    for child4 in child3:
                                        if child4.tag == 'BASEBANDEQMTYPE' and child4.text != str(
                                                2):
                                            child1.remove(child2)
                                            # ET.dump(child1)
            mod_BBP_wbbptoubbp_TYPE_from_slot_to_slot(3, 3, 8194, 2)
            rmv_old_bbp_from_slotnumber(0)
            rmv_old_bbp_from_slotnumber(1)
            rmv_old_bbp_from_slotnumber(2)

            UMPT_flag = 1
            # UMPT~1, WMPT~8193, LMPT~12289
            for child0 in root:
                for child1 in child0:
                    for child2 in child1:
                        if child2.tag == '' + 'MPT':
                            for child3 in child2:
                                for child4 in child3:
                                    if child4.tag == 'TYPE' and child4.text == '8193':
                                        UMPT_flag = 0
                                        child4.text = '1'
                                        ET.dump(child1)
                                        break
            ethport_taglist = []
            for child0 in root:
                for child1 in child0:
                    for child2 in child1:
                        if child2.tag == '' + 'ETHPORT':
                            for child3 in child2:
                                for child4 in child3:
                                    if child4.tag not in ethport_taglist:
                                        ethport_taglist.append(child4.tag)
            if UMPT_flag != 1 or UMPT_flag == 1:
                # print('ETHTAGPORTLIST',ethport_taglist)
                flag0 = False
                for child0 in root:
                    flag0 = False
                    for child1 in child0:
                        for child2 in child1:
                            if child2.tag == '' + 'ETHPORT':
                                for child3 in child2:
                                    # if 'RXBCPKTALMOCRTHD' not in ethport_taglist:
                                    #     sector_subelement2 = ET.SubElement(child3, 'RXBCPKTALMOCRTHD')
                                    #     sector_subelement2.text = str(320)
                                    # if 'RXBCPKTALMCLRTHD' not in ethport_taglist:
                                    #     sector_subelement2 = ET.SubElement(child3, 'RXBCPKTALMCLRTHD')
                                    #     sector_subelement2.text = str(288)
                                    flag0 = False
                                    for child4 in child3:
                                        if child4.tag == 'PN' and child4.text == str(
                                                0):
                                            flag0 = True
                                        if flag0 and child4.tag == 'PA':
                                            child4.text = str(0)  # COPPER~0, FIBER~1, AUTO~2, UNCONFIG~255
                                        if flag0 and child4.tag == 'SPEED':
                                            child4.text = str(3)
                                        if flag0 and child4.tag == 'DUPLEX':
                                            child4.text = str(2)
                                        if flag0 and child4.tag == 'RXBCPKTALMOCRTHD':
                                            child4.text = str(
                                                320)  # bug corrected, when Thds are valid for one port and null for the other,ethport_taglist has the tag
                                        if flag0 and child4.tag == 'RXBCPKTALMCLRTHD':
                                            child4.text = str(
                                                288)  # bug corrected, when Thds are valid for one port and null for the other,ethport_taglist has the tag
                                            ET.dump(child1)
                                            break

                flag1 = False
                for child0 in root:
                    flag1 = False
                    for child1 in child0:
                        for child2 in child1:
                            if child2.tag == '' + 'ETHPORT':
                                for child3 in child2:
                                    flag1 = False
                                    for child4 in child3:
                                        if child4.tag == 'PN' and child4.text == str(
                                                1):
                                            flag1 = True
                                        if flag1 and child4.tag == 'PA':
                                            child4.text = str(1)  # COPPER~0, FIBER~1, AUTO~2, UNCONFIG~255
                                        if flag1 and child4.tag == 'SPEED':
                                            child4.text = str(1)
                                        if flag1 and child4.tag == 'DUPLEX':
                                            child4.text = str(1)
                                        if flag1 and child4.tag == 'RXBCPKTALMOCRTHD':
                                            child4.text = str(
                                                320)  # bug corrected, when Thds are valid for one port and null for the other,ethport_taglist has the tag
                                        if flag1 and child4.tag == 'RXBCPKTALMCLRTHD':
                                            child4.text = str(
                                                288)  # bug corrected, when Thds are valid for one port and null for the other,ethport_taglist has the tag
                                            ET.dump(child1)
                                            break
            ETHPORT_adjust_thd()
            tree.write(OUTPUT)
            # Label(window, text='3G Refarming: changed to UBBP slot 3', width=80, fg='green').grid(column=2, row=26)
            Label(window, text=TEXT, width=80, fg='green').grid(column=2, row=row)
            fo = open('LOG.txt', 'a')
            # fo.write(f'{dt.datetime.now()} --  {NEname}  ' + '3G Refarming: changed to UBBP slot 3 , \n')
            fo.write(f'{dt.datetime.now()} --  {NEname}  ' + TEXT)

    UBBP_function(CheckVar18,CheckkVar18, '3G Refarming: changed to UBBP slot 3\n', 26)

    #########################################################################################################################################################################################
    # 3G  WBBP  refarming
    def WBBP_function(Checkvar,CheckkVar, TEXT, row):
        if Checkvar.get() == 1 or CheckkVar == 1:
            BBP19 = parse_existing_MO_name(root, 'BBP')
            SECTOR_var18 = parse_existing_MO_name(root, 'SECTOR')
            Chain_var18 = parse_existing_MO_name(root, 'RRUCHAIN')
            ULOCELL_var18 = parse_existing_MO_name(root, 'ULOCELL')
            RRU_var19 = parse_existing_MO_name(root, 'RRU')
            RRU_list19_unique = []
            # insert_new_BBP_CN_SRN_SN_TYPE_BBWS(0,0,3,2,2)
            for i in RRU_var19:
                for j in i:
                    for j in range(0, len(RRU_var19)):
                        for item in RRU_var19[j]:
                            for key in item.keys():
                                if key == 'SRN':
                                    if item[key] not in RRU_list19_unique:
                                        RRU_list19_unique.append(item[key])
            for child0 in root:
                for child1 in child0:
                    for child2 in child1:
                        if child2.tag == '' + 'RRU':
                            for child3 in child2:
                                for child4 in child3:
                                    if child4.tag == 'RS':
                                        child4.text = str(136)

            UMPT_flag = 1
            # UMPT~1, WMPT~8193, LMPT~12289
            for child0 in root:
                for child1 in child0:
                    for child2 in child1:
                        if child2.tag == '' + 'MPT':
                            for child3 in child2:
                                for child4 in child3:
                                    if child4.tag == 'TYPE' and child4.text == '8193':
                                        UMPT_flag = 0
                                        child4.text = '1'
                                        ET.dump(child1)
                                        break
            ethport_taglist = []
            for child0 in root:
                for child1 in child0:
                    for child2 in child1:
                        if child2.tag == '' + 'ETHPORT':
                            for child3 in child2:
                                for child4 in child3:
                                    if child4.tag not in ethport_taglist:
                                        ethport_taglist.append(child4.tag)

            if UMPT_flag != 1 or UMPT_flag == 1:
                # print('ETHTAGPORTLIST',ethport_taglist)
                flag0 = False
                for child0 in root:
                    flag0 = False
                    for child1 in child0:
                        for child2 in child1:
                            if child2.tag == '' + 'ETHPORT':
                                for child3 in child2:
                                    # if 'RXBCPKTALMOCRTHD' not in ethport_taglist:
                                    #     sector_subelement2 = ET.SubElement(child3, 'RXBCPKTALMOCRTHD')
                                    #     sector_subelement2.text = str(320)
                                    # if 'RXBCPKTALMCLRTHD' not in ethport_taglist:
                                    #     sector_subelement2 = ET.SubElement(child3, 'RXBCPKTALMCLRTHD')
                                    #     sector_subelement2.text = str(288)
                                    flag0 = False
                                    for child4 in child3:
                                        if child4.tag == 'PN' and child4.text == str(
                                                0):
                                            flag0 = True
                                        if flag0 and child4.tag == 'PA':
                                            child4.text = str(0)  # COPPER~0, FIBER~1, AUTO~2, UNCONFIG~255
                                        if flag0 and child4.tag == 'SPEED':
                                            child4.text = str(3)
                                        if flag0 and child4.tag == 'DUPLEX':
                                            child4.text = str(2)
                                            ET.dump(child1)
                                        if flag0 and child4.tag == 'RXBCPKTALMOCRTHD':
                                            child4.text = str(
                                                320)  # bug corrected, when Thds are valid for one port and null for the other,ethport_taglist has the tag
                                        if flag0 and child4.tag == 'RXBCPKTALMCLRTHD':
                                            child4.text = str(
                                                288)  # bug corrected, when Thds are valid for one port and null for the other,ethport_taglist has the tag
                                            ET.dump(child1)
                                            break

                flag1 = False
                for child0 in root:
                    flag1 = False
                    for child1 in child0:
                        for child2 in child1:
                            if child2.tag == '' + 'ETHPORT':
                                for child3 in child2:
                                    flag1 = False
                                    for child4 in child3:
                                        if child4.tag == 'PN' and child4.text == str(
                                                1):
                                            flag1 = True
                                        if flag1 and child4.tag == 'PA':
                                            child4.text = str(1)  # COPPER~0, FIBER~1, AUTO~2, UNCONFIG~255
                                        if flag1 and child4.tag == 'SPEED':
                                            child4.text = str(1)
                                        if flag1 and child4.tag == 'DUPLEX':
                                            child4.text = str(1)
                                        if flag1 and child4.tag == 'RXBCPKTALMOCRTHD':
                                            child4.text = str(
                                                320)  # bug corrected, when Thds are valid for one port and null for the other,ethport_taglist has the tag
                                        if flag1 and child4.tag == 'RXBCPKTALMCLRTHD':
                                            child4.text = str(
                                                288)  # bug corrected, when Thds are valid for one port and null for the other,ethport_taglist has the tag
                                            ET.dump(child1)
                                            break

            ETHPORT_adjust_thd()
            tree.write(OUTPUT)
            # Label(window, text='3G Refarming: KEEP WBBP boards', width=80, fg='green').grid(column=2, row=27)
            Label(window, text=TEXT, width=80, fg='green').grid(column=2, row=row)
            fo = open('LOG.txt', 'a')
            # fo.write(f'{dt.datetime.now()} --  {NEname}  ' + '3G Refarming: KEEP WBBP boards , \n')
            fo.write(f'{dt.datetime.now()} --  {NEname}  ' + TEXT)

    WBBP_function(CheckVar19,CheckkVar19, '3G Refarming: KEEP WBBP boards, keep Antenna Configuration', 27)

    #########################################################################################################################################################################################
    # 3G UBBP AC
    UBBP_function(CheckVar20,CheckkVar20, '3G Refarming: UBBP + Antennaports AC', 28)
    ULOCELLSECTOREQM_change()
    ulocelllist = []
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'ULOCELL':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'ULOCELLID' and str(
                                    child4.text[-1]) in ['1', '2', '3', '5', '6', '7', '8', '9', '0']:
                                if str(child4.text[-1]) not in ulocelllist:
                                    ulocelllist.append(str(child4.text[-1]))

    if CheckVar20.get() == 1 or CheckkVar20 == 1:
        mod_RRU_SECTOR_SECTOREQM_SRN(100, 200)
        mod_RRU_SECTOR_SECTOREQM_SRN(101, 201)
        mod_RRU_SECTOR_SECTOREQM_SRN(102, 202)
        mod_SECTOR_ID(0, 200)
        mod_SECTOR_ID(1, 201)
        mod_SECTOR_ID(2, 202)
        mod_SECTOREQM_ID(0, 300)
        mod_SECTOREQM_ID(1, 301)
        mod_SECTOREQM_ID(2, 302)
        mod_SECTOREQM_ID(3, 303)
        mod_SECTOREQM_ID(4, 304)
        mod_SECTOREQM_ID(5, 305)

        if '1' in ulocelllist:
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 0, 136, 15, 4, 4)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 100, 0, [0, 2])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 0, 100, 0, [[0, 3], [2, 1]])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 0, 100, 0, [[0, 1], [2, 3]])

        if '2' in ulocelllist:
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 1, 136, 15, 4, 4)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 101, 0, [0, 2])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 1, 0, 101, 0, [[0, 3], [2, 1]])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(3, 1, 0, 101, 0, [[0, 1], [2, 3]])

        if '3' in ulocelllist:
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 2, 136, 15, 4, 4)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 0, 102, 0, [0, 2])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(4, 2, 0, 102, 0, [[0, 3], [2, 1]])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(5, 2, 0, 102, 0, [[0, 1], [2, 3]])

        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(200, 100)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(201, 101)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(202, 102)

        rmv_old_SECTOR_ID(200)
        rmv_old_SECTOR_ID(201)
        rmv_old_SECTOR_ID(202)
        rmv_old_SECTOREQM_ID(300)
        rmv_old_SECTOREQM_ID(301)
        rmv_old_SECTOREQM_ID(302)
        rmv_old_SECTOREQM_ID(303)
        rmv_old_SECTOREQM_ID(304)
        rmv_old_SECTOREQM_ID(305)

        rmv_RET_only(100)
        rmv_RET_only(101)
        rmv_RET_only(102)
        rmv_3G_all_ret_retsubunit_retdevicedata()
    #########################################################################################################################################################################################
    # 3G UBBP BD
    UBBP_function(CheckVar21,CheckkVar21, '3G Refarming: UBBP + Antennaports BD', 29)
    ULOCELLSECTOREQM_change()
    ulocelllist = []
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'ULOCELL':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'ULOCELLID' and str(
                                    child4.text[-1]) in ['1', '2', '3', '5', '6', '7', '8', '9', '0']:
                                if str(child4.text[-1]) not in ulocelllist:
                                    ulocelllist.append(str(child4.text[-1]))
    if CheckVar21.get() == 1 or CheckkVar21 == 1:
        mod_RRU_SECTOR_SECTOREQM_SRN(100, 200)
        mod_RRU_SECTOR_SECTOREQM_SRN(101, 201)
        mod_RRU_SECTOR_SECTOREQM_SRN(102, 202)
        mod_SECTOR_ID(0, 200)
        mod_SECTOR_ID(1, 201)
        mod_SECTOR_ID(2, 202)
        mod_SECTOREQM_ID(0, 300)
        mod_SECTOREQM_ID(1, 301)
        mod_SECTOREQM_ID(2, 302)
        mod_SECTOREQM_ID(3, 303)
        mod_SECTOREQM_ID(4, 304)
        mod_SECTOREQM_ID(5, 305)

        if '1' in ulocelllist:
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 0, 136, 15, 4, 4)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 100, 0, [1, 3])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 0, 100, 0, [[1, 3], [3, 1]])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 0, 100, 0, [[1, 1], [3, 3]])

        if '2' in ulocelllist:
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 1, 136, 15, 4, 4)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 101, 0, [1, 3])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 1, 0, 101, 0, [[1, 3], [3, 1]])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(3, 1, 0, 101, 0, [[1, 1], [3, 3]])

        if '3' in ulocelllist:
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 2, 136, 15, 4, 4)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 0, 102, 0, [1, 3])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(4, 2, 0, 102, 0, [[1, 3], [3, 1]])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(5, 2, 0, 102, 0, [[1, 1], [3, 3]])

        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(200, 100)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(201, 101)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(202, 102)

        rmv_old_SECTOR_ID(200)
        rmv_old_SECTOR_ID(201)
        rmv_old_SECTOR_ID(202)
        rmv_old_SECTOREQM_ID(300)
        rmv_old_SECTOREQM_ID(301)
        rmv_old_SECTOREQM_ID(302)
        rmv_old_SECTOREQM_ID(303)
        rmv_old_SECTOREQM_ID(304)
        rmv_old_SECTOREQM_ID(305)
        rmv_RET_only(100)
        rmv_RET_only(101)
        rmv_RET_only(102)
        rmv_3G_all_ret_retsubunit_retdevicedata()
    #########################################################################################################################################################################################
    # 3G WBBP AC
    WBBP_function(CheckVar22,CheckkVar22, '3G Refarming: WBBP + Antennaports AC', 30)
    ULOCELLSECTOREQM_change()
    ulocelllist = []
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'ULOCELL':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'ULOCELLID' and str(
                                    child4.text[-1]) in ['1', '2', '3', '5', '6', '7', '8', '9', '0']:
                                if str(child4.text[-1]) not in ulocelllist:
                                    ulocelllist.append(str(child4.text[-1]))
    if CheckVar22.get() == 1 or CheckkVar22 == 1:

        mod_RRU_SECTOR_SECTOREQM_SRN(100, 200)
        mod_RRU_SECTOR_SECTOREQM_SRN(101, 201)
        mod_RRU_SECTOR_SECTOREQM_SRN(102, 202)
        mod_SECTOR_ID(0, 200)
        mod_SECTOR_ID(1, 201)
        mod_SECTOR_ID(2, 202)
        mod_SECTOREQM_ID(0, 300)
        mod_SECTOREQM_ID(1, 301)
        mod_SECTOREQM_ID(2, 302)
        mod_SECTOREQM_ID(3, 303)
        mod_SECTOREQM_ID(4, 304)
        mod_SECTOREQM_ID(5, 305)

        if '1' in ulocelllist:
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 0, 136, 15, 4, 4)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 100, 0, [0, 2])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 0, 100, 0, [[0, 3], [2, 1]])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 0, 100, 0, [[0, 1], [2, 3]])

        if '2' in ulocelllist:
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 1, 136, 15, 4, 4)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 101, 0, [0, 2])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 1, 0, 101, 0, [[0, 3], [2, 1]])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(3, 1, 0, 101, 0, [[0, 1], [2, 3]])

        if '3' in ulocelllist:
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 2, 136, 15, 4, 4)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 0, 102, 0, [0, 2])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(4, 2, 0, 102, 0, [[0, 3], [2, 1]])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(5, 2, 0, 102, 0, [[0, 1], [2, 3]])

        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(200, 100)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(201, 101)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(202, 102)

        rmv_old_SECTOR_ID(200)
        rmv_old_SECTOR_ID(201)
        rmv_old_SECTOR_ID(202)
        rmv_old_SECTOREQM_ID(300)
        rmv_old_SECTOREQM_ID(301)
        rmv_old_SECTOREQM_ID(302)
        rmv_old_SECTOREQM_ID(303)
        rmv_old_SECTOREQM_ID(304)
        rmv_old_SECTOREQM_ID(305)
        rmv_RET_only(100)
        rmv_RET_only(101)
        rmv_RET_only(102)
        rmv_3G_all_ret_retsubunit_retdevicedata()
    #########################################################################################################################################################################################
    # 3G WBBP BD
    WBBP_function(CheckVar23,CheckkVar23, '3G Refarming: WBBP + Antennaports BD', 31)
    ULOCELLSECTOREQM_change()
    ulocelllist = []
    for child0 in root:
        for child1 in child0:
            for child2 in child1:
                if child2.tag == '' + 'ULOCELL':
                    for child3 in child2:
                        for child4 in child3:
                            if child4.tag == 'ULOCELLID' and str(
                                    child4.text[-1]) in ['1', '2', '3', '5', '6', '7', '8', '9', '0']:
                                if str(child4.text[-1]) not in ulocelllist:
                                    ulocelllist.append(str(child4.text[-1]))
    if CheckVar23.get() == 1 or CheckkVar23 == 1:
        mod_RRU_SECTOR_SECTOREQM_SRN(100, 200)
        mod_RRU_SECTOR_SECTOREQM_SRN(101, 201)
        mod_RRU_SECTOR_SECTOREQM_SRN(102, 202)
        mod_SECTOR_ID(0, 200)
        mod_SECTOR_ID(1, 201)
        mod_SECTOR_ID(2, 202)
        mod_SECTOREQM_ID(0, 300)
        mod_SECTOREQM_ID(1, 301)
        mod_SECTOREQM_ID(2, 302)
        mod_SECTOREQM_ID(3, 303)
        mod_SECTOREQM_ID(4, 304)
        mod_SECTOREQM_ID(5, 305)

        if '1' in ulocelllist:
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 100, 0, 0, 136, 15, 4, 4)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 100, 0, [1, 3])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(0, 0, 0, 100, 0, [[1, 3], [3, 1]])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 0, 100, 0, [[1, 1], [3, 3]])

        if '2' in ulocelllist:
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 101, 0, 1, 136, 15, 4, 4)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(1, 0, 101, 0, [1, 3])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 1, 0, 101, 0, [[1, 3], [3, 1]])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(3, 1, 0, 101, 0, [[1, 1], [3, 3]])

        if '3' in ulocelllist:
            insert_new_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_with_args_CN_SRN_SN_RRUCHAINID_WM_RRUtype_RXnum_TXnum(0, 102, 0, 2, 136, 15, 4, 4)
            insert_new_SECTOR_ID_CN_SRN_SN_ANTENNAPORTLIST(2, 0, 102, 0, [1, 3])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(4, 2, 0, 102, 0, [[1, 3], [3, 1]])
            insert_new_and_superior_multiSECTOREQM_ID_CN_SRN_SN_ANTENNAPORTLIST(5, 2, 0, 102, 0, [[1, 1], [3, 3]])

        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(200, 100)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(201, 101)
        rmv_old_RRU_ALMPORT_ANTENNAPORT_RETPORT_RXB_TXB_CPRIPORT_SFP_from_RRU_subrackno(202, 102)

        rmv_old_SECTOR_ID(200)
        rmv_old_SECTOR_ID(201)
        rmv_old_SECTOR_ID(202)
        rmv_old_SECTOREQM_ID(300)
        rmv_old_SECTOREQM_ID(301)
        rmv_old_SECTOREQM_ID(302)
        rmv_old_SECTOREQM_ID(303)
        rmv_old_SECTOREQM_ID(304)
        rmv_old_SECTOREQM_ID(305)
        rmv_RET_only(100)
        rmv_RET_only(101)
        rmv_RET_only(102)
        rmv_3G_all_ret_retsubunit_retdevicedata()
    #########################################################################################################################################################################################
    # end of executor function
    fo = open('LOG.txt', 'a')
    end = time.time()
    #Label(window, text=f'FINISHED: XML file preparation time = {round(float(end - start), 2)} s', width=80, fg='green').grid(column=2, row=32)
    text = Text(window, height=1, width=130, font=("arial", 10))
    text.insert(INSERT, f'FINISHED: FullExport file preparation time = {round(float(end - start), 2)} s')
    text.grid(column=2, row=32)
    fo.write(f'{dt.datetime.now()} --  {NEname}  {round(float(end - start), 2)} s ' + 'Changes finished!\n')
    return (end - start)
    root.clear
    #########################################################################################################################################################################################


def bulk_step2():
    f1() #clear initially grid and selected options previously to BULK
    bulk_folder_information=bulk_step1()
    start0=time.time() #indicates start moment of bulk execution
    #global CheckVar1, CheckVar2, CheckVar300, CheckVar30, CheckVar3, CheckVar4, CheckVar5, CheckVar6, CheckVar7, CheckVar80, CheckVar8, CheckVar9, CheckVar10, CheckVar11, CheckVar12, CheckVar13, CheckVar14, CheckVar15, CheckVar16, CheckVar17, CheckVar18, CheckVar19, CheckVar20, CheckVar21, CheckVar22, CheckVar23
    global tree, root,OUTPUT
    global CheckkVar1, CheckkVar2, CheckkVar300, CheckkVar30, CheckkVar3, CheckkVar4, CheckkVar5, CheckkVar6, CheckkVar7, CheckkVar9, CheckkVar10, CheckkVar11, CheckkVar12, CheckkVar13, CheckkVar14, CheckkVar15, CheckkVar16, CheckkVar17, CheckkVar18, CheckkVar19, CheckkVar20, CheckkVar21, CheckkVar22, CheckkVar23

    for site in bulk_folder_information[1]:
        for file_name in bulk_folder_information[0]:
            if site in file_name:
                OUTPUT=file_name
                tree = ET.parse(OUTPUT)

                root = tree.getroot()

                #print(tree)
                #print(root)
                #print(OUTPUT)
                if bulk_folder_information[2][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar1=1
                if bulk_folder_information[3][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar2=1
                if bulk_folder_information[4][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar16=1
                if bulk_folder_information[5][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar300=1
                if bulk_folder_information[6][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar30=1
                if bulk_folder_information[7][str(site)] == 'x': #L800 AB REC/INT 60-61-62 SN 2,3 -> 4
                    print('connected to excel file ok')
                    CheckkVar3=1
                if bulk_folder_information[8][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar4=1
                if bulk_folder_information[9][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar5=1
                if bulk_folder_information[10][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar6=1
                if bulk_folder_information[11][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar7=1
                # if bulk_folder_information[12][str(site)] == 'x':
                #     print('connected to excell file ok')
                #     CheckkVar80=1
                # if bulk_folder_information[13][str(site)] == 'x':
                #     print('connected to excell file ok')
                #     CheckkVar8=1
                if bulk_folder_information[12][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar9=1
                if bulk_folder_information[13][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar10=1
                if bulk_folder_information[14][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar11=1
                if bulk_folder_information[15][str(site)] == 'x':
                    print('connected to excell file ok')
                    CheckkVar12=1
                if bulk_folder_information[16][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar13=1
                if bulk_folder_information[17][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar14=1
                if bulk_folder_information[18][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar15=1
                #3G options bellow
                if bulk_folder_information[19][str(site)] == 'x':
                    print('connected to excell file ok')
                    CheckkVar18=1
                if bulk_folder_information[20][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar19=1
                if bulk_folder_information[21][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar20=1
                if bulk_folder_information[22][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar21=1
                if bulk_folder_information[23][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar22=1
                if bulk_folder_information[24][str(site)] == 'x':
                    print('connected to excel file ok')
                    CheckkVar23=1
                executor_function()
                print(site)
                CheckVar1, CheckVar2, CheckVar300, CheckVar30, CheckVar3, CheckVar4, CheckVar5, CheckVar6, CheckVar7, CheckVar9, CheckVar10, CheckVar11, CheckVar12, CheckVar13, CheckVar14, CheckVar15, CheckVar16, CheckVar17, CheckVar18, CheckVar19, CheckVar20, CheckVar21, CheckVar22, CheckVar23 =\
                0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0
                CheckkVar1, CheckkVar2, CheckkVar300, CheckkVar30, CheckkVar3, CheckkVar4, CheckkVar5, \
                CheckkVar6, CheckkVar7, CheckkVar9, \
                CheckkVar10, CheckkVar11, CheckkVar12, CheckkVar13, CheckkVar14, CheckkVar15, CheckkVar16, \
                CheckkVar17, CheckkVar18, CheckkVar19, CheckkVar20, CheckkVar21, CheckkVar22, CheckkVar23 = \
                    0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0
                f1() #clear grid in GUI
    end0=time.time()
    #Label(window, text=f'Bulk XML files preparation time: {round(float(end0 - start0), 2)} s - TOTAL = {len(bulk_folder_information[0])}', width=80, fg='green').grid(column=2, row=32)
    text = Text(window, height=1, width=130, font=("arial", 10))
    text.insert(INSERT, f'Bulk XML files preparation time: {round(float(end0 - start0), 2)} s - TOTAL = {len(bulk_folder_information[0])}')
    text.grid(column=2, row=32)


button_start = Button(window, text='START', fg="navy", command=executor_function, width=40)
# button_start['font'] = myFont
button_start.grid(column=1, row=32)

button_clear = Button(window, text='CLEAR', fg="red", command=f1, width=40)
# button_start['font'] = myFont
button_clear.grid(column=1, row=33)

button_logs = Button(window, text='LOG file', fg="green", command=f2, width=40)
# button_start['font'] = myFont
button_logs.grid(column=1, row=34)

button_pre_start=Button(window, text='Smart Agent  "(¬_¬)"', fg="green", command=smart_agent, width=40)
button_pre_start.grid(column=1, row=3)
button_pre_start_tip= CreateToolTip(button_pre_start,'Press Me!\nStart the Smart Agent action filtering!')

Button_BULK = Button(window, text='BULK CFG', width=9, command=bulk_step2,bg = "cornflower blue")
Button_BULK.grid(column=1, row=1,sticky="E")
Button_BULK_tip=CreateToolTip(Button_BULK,'Enter Bulk Configuration mode for multiple NEs using NE.xlsx template')

Label(window, text='© 2020-2022 Valentin Mihai', fg='DarkGoldenrod4',bg='white', font='times 10 bold').place(x=50, y=710)

fo.close()
# S = tkinter.Scrollbar(window)
# T = tkinter.Text(window, height=4, width=50).grid(column=3,row=1)

# def callback(url):
#     webbrowser.open_new(r"file://c:\test\LOGS.txt")
#
#
# link1 = Label(window, text="Instructions", fg="blue", cursor="hand2")
# link1.place(x=50, y=810)
# link1.bind("<Button-1>", lambda e: callback())

# myFont=font.Font(weight="bold")

window.mainloop()
print('Good Bye!')
