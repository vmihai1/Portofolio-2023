import openpyxl
from openpyxl.styles import PatternFill

# 0	    IUB	IP	BE
# 26	IUB	IP	AF31
# 34	IUB	IP	AF41
# 40	IUB	IP	AF42
# 46	IUB	IP	AF43
# 48	IUB	IP	EF

#export file to be selected!
EXPORT_FILE_NAME='Munchen_22_export.xlsm'

#NodeB exception list: NodeBs  that do not have 6 IPPATHs  or TRMMAP 44
#NODEB_EXCEPTION_LIST=['MXUA15','OXUE53','OXUH48','OXU3E1','OXU5I3','OXU099','OXU4E5','OXU4S2','OXUV16','OXUV15','OXU1R1','OXU5N2','OXUP42','OXU268','HXU3Z7','HXU997','HXU1P1','BXU7U4','BXU0S1']
#NODEB_EXCEPTION_LIST=['HXU3Z7', 'HXU4X6', 'HXUB03', 'HXUB04', 'HXUM72', 'HXUY72', 'HXU9T5', 'HXUE30', 'HXU2Y1', 'HXUE92', 'HXUE52', 'HXUL97']
#NODEB_EXCEPTION_LIST=['OXUZ11', 'OXUX89', 'OXU3I4', 'OXUE53', 'OXUE55', 'OXUH48', 'OXU018', 'OXUE22', 'OXUE23', 'OXUE92', 'OXUH40', 'OXU5E1', 'OXU3E1', 'OXU5I3', 'OXU099', 'OXU4E5', 'OXU6T6', 'OXUH67', 'OXU0B3', 'OXU4S2', 'OXUV16', 'OXUV15', 'OXU1R1', 'OXUZ84', 'OXUS88']
NODEB_EXCEPTION_LIST=['MXUA15']

PHB_VDF_6900 = {'BE': '0', 'AF11': '10', 'AF12': '12', 'AF13': '14', 'AF21': '18', 'AF22': '20', 'AF23': '22',
                'AF31': '26', 'AF32': '28', 'AF33': '30', 'AF41': '34', 'AF42': '40', 'AF43': '46', 'EF': '48',
                'CS1': '8', 'CS2': '16', 'CS3': '24', 'CS4': '32', 'CS5': '36', 'CS6': '38', 'CS7': '56'}
TRMMAP_44_VDF={'Common channel primary path':'EF','IMS SRB primary path':'AF42','SRB primary path':'EF','AMR voice primary path':'AF43',
               'R99 CS conversational primary path':'AF43','R99 CS streaming primary path':'AF41','R99 PS conversational primary path':'AF43',
               'R99 PS streaming primary path':'AF41','R99 PS high PRI interactive primary path':'AF41','R99 PS middle PRI interactive primary path':'AF31',
               'R99 PS low PRI interactive primary path':'BE','R99 PS background primary path':'BE','HSDPA Signal primary path':'EF',
               'HSDPA IMS Signal primary path':'AF42','HSDPA Voice primary path':'AF43','HSDPA conversational primary path':'AF43',
               'HSDPA streaming primary path':'AF41','HSDPA high PRI interactive primary path':'AF41','HSDPA middle PRI interactive primary path':'AF31',
               'HSDPA low PRI interactive primary path':'BE','HSDPA background primary path':'BE','HSUPA Signal primary path':'EF',
               'HSUPA IMS Signal primary path':'AF42','HSUPA Voice primary path':'AF43','HSUPA conversational primary path':'AF43',
               'HSUPA streaming primary path':'AF41','HSUPA high PRI interactive primary path':'AF41','HSUPA middle PRI interactive primary path':'AF31',
               'HSUPA low PRI interactive primary path':'BE','HSUPA background primary path':'BE'}

RNC_name_list=[]
Subrack_PHB_list_initial=[]
DSCP_list_PHBMAP_initial=[]
DSCP_list_PHBMAP_final=[]
NB_full_IP_list=[]
ADJNODEID_full_IP_list=[]
RNC_NB_full_IP_list=[]
RNC_list=[]
RNC_ADJNODE_exception_list=[]



def open_excel_file(file_name):
    wb1=openpyxl.load_workbook(file_name,data_only=True)

def collect_data(file_name):
    wb1=openpyxl.load_workbook(file_name,data_only=True)
    sheet1=wb1['Controller Data']
    for i in range(3, len(sheet1['A'])+1):
        if sheet1.cell(row=i,column=1).value!=None:
            RNC_name_list.append(sheet1.cell(row=i,column=1).value)
    RNC_string_for_title='--'.join(RNC_name_list)


    wb1.close()
    return RNC_string_for_title
TITLE=collect_data(EXPORT_FILE_NAME)

def change_PHBMAP(file_name):
    wb1 = openpyxl.load_workbook(file_name, data_only=True)
    sheet1=wb1['PHBMAP']

    def color_cell(X, Y, color):
        sheet1.cell(row=X, column=Y).fill = PatternFill(bgColor=color, fill_type='lightUp')

    for i in range(3, len(sheet1['A'])):
        if sheet1.cell(row=i,column=3).value =='20':
            Subrack_PHB_list_initial.append('-'.join([sheet1.cell(row=i,column=2).value,sheet1.cell(row=i,column=3).value,sheet1.cell(row=i,column=4).value]))
            DSCP_list_PHBMAP_initial.append(sheet1.cell(row=i,column=5).value)
    PHBMAP_dict_initial=dict(zip(Subrack_PHB_list_initial,DSCP_list_PHBMAP_initial))
    #print(Subrack_PHB_list_initial)
    #print(PHBMAP_dict_initial)
    for i in range(3, len(sheet1['A'])):
        if sheet1.cell(row=i,column=3).value =='20':
            if sheet1.cell(row=i, column=5).value != PHB_VDF_6900[sheet1.cell(row=i, column=4).value]:
                sheet1.cell(row=i, column=5).value=PHB_VDF_6900[sheet1.cell(row=i, column=4).value]
                color_cell(i, 5,"00FFFF00")
                DSCP_list_PHBMAP_final.append(sheet1.cell(row=i, column=5).value)
            elif sheet1.cell(row=i, column=5).value == PHB_VDF_6900[sheet1.cell(row=i, column=4).value]:
                color_cell(i, 5, "00ccf5ff")

    PHBMAP_dict_final = dict(zip(Subrack_PHB_list_initial, DSCP_list_PHBMAP_final))
    #print(PHBMAP_dict_final)
    wb1.save(TITLE+'__import.xlsx')
    return [PHBMAP_dict_initial,PHBMAP_dict_final]

def change_TRMMAP(file_name):
    wb1 = openpyxl.load_workbook(file_name, data_only=True)
    sheet1 = wb1['TRMMAP']

    def color_cell(X, Y,color):
        sheet1.cell(row=X, column=Y).fill = PatternFill(bgColor=color, fill_type='lightUp')

    TRMMAP_service_list=[ sheet1.cell(row=2,column=item).value for item in range(6,66,2)]
    #print(TRMMAP_service_list)
    for i in range(3, len(sheet1['A'])):
        if sheet1.cell(row=i,column=2).value =='44':
            PHB_list=[sheet1.cell(row=i,column=item).value for item in range(6,66,2)]
            TRMMAP_dict_initial=dict(zip(TRMMAP_service_list,PHB_list))

            if sheet1.cell(row=i, column=6).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=6).value]:
                sheet1.cell(row=i, column=6).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=6).value]
                color_cell(i, 6,"00FFFF00")
            else:
                color_cell(i, 6, "00ccf5ff")

            if sheet1.cell(row=i, column=8).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=8).value]:
                sheet1.cell(row=i, column=8).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=8).value]
                color_cell(i, 8,"00FFFF00")
            else:
                color_cell(i, 8, "00ccf5ff")

            if sheet1.cell(row=i, column=10).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=10).value]:
                sheet1.cell(row=i, column=10).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=10).value]
                color_cell(i, 10,"00FFFF00")
            else:
                color_cell(i, 10, "00ccf5ff")

            if sheet1.cell(row=i, column=12).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=12).value]:
                sheet1.cell(row=i, column=12).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=12).value]
                color_cell(i, 12,"00FFFF00")
            else:
                color_cell(i, 12, "00ccf5ff")

            if sheet1.cell(row=i, column=14).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=14).value]:
                sheet1.cell(row=i, column=14).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=14).value]
                color_cell(i, 14,"00FFFF00")
            else:
                color_cell(i, 14, "00ccf5ff")

            if sheet1.cell(row=i, column=16).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=16).value]:
                sheet1.cell(row=i, column=16).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=16).value]
                color_cell(i, 16,"00FFFF00")
            else:
                color_cell(i, 16, "00ccf5ff")

            if sheet1.cell(row=i, column=18).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=18).value]:
                sheet1.cell(row=i, column=18).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=18).value]
                color_cell(i, 18,"00FFFF00")
            else:
                color_cell(i, 18, "00ccf5ff")

            if sheet1.cell(row=i, column=20).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=20).value]:
                sheet1.cell(row=i, column=20).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=20).value]
                color_cell(i, 20,"00FFFF00")
            else:
                color_cell(i, 20, "00ccf5ff")

            if sheet1.cell(row=i, column=22).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=22).value]:
                sheet1.cell(row=i, column=22).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=22).value]
                color_cell(i, 22,"00FFFF00")
            else:
                color_cell(i, 22, "00ccf5ff")

            if sheet1.cell(row=i, column=24).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=24).value]:
                sheet1.cell(row=i, column=24).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=24).value]
                color_cell(i, 24,"00FFFF00")
            else:
                color_cell(i, 24, "00ccf5ff")

            if sheet1.cell(row=i, column=26).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=26).value]:
                sheet1.cell(row=i, column=26).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=26).value]
                color_cell(i, 26,"00FFFF00")
            else:
                color_cell(i, 26, "00ccf5ff")

            if sheet1.cell(row=i, column=28).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=28).value]:
                sheet1.cell(row=i, column=28).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=28).value]
                color_cell(i, 28,"00FFFF00")
            else:
                color_cell(i, 28, "00ccf5ff")

            if sheet1.cell(row=i, column=30).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=30).value]:
                sheet1.cell(row=i, column=30).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=30).value]
                color_cell(i, 30,"00FFFF00")
            else:
                color_cell(i, 30, "00ccf5ff")

            if sheet1.cell(row=i, column=32).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=32).value]:
                sheet1.cell(row=i, column=32).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=32).value]
                color_cell(i, 32,"00FFFF00")
            else:
                color_cell(i, 32, "00ccf5ff")

            if sheet1.cell(row=i, column=34).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=34).value]:
                sheet1.cell(row=i, column=34).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=34).value]
                color_cell(i, 34,"00FFFF00")
            else:
                color_cell(i, 34, "00ccf5ff")

            if sheet1.cell(row=i, column=36).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=36).value]:
                sheet1.cell(row=i, column=36).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=36).value]
                color_cell(i, 36,"00FFFF00")
            else:
                color_cell(i, 36, "00ccf5ff")

            if sheet1.cell(row=i, column=38).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=38).value]:
                sheet1.cell(row=i, column=38).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=38).value]
                color_cell(i, 38,"00FFFF00")
            else:
                color_cell(i, 38, "00ccf5ff")

            if sheet1.cell(row=i, column=40).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=40).value]:
                sheet1.cell(row=i, column=40).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=40).value]
                color_cell(i, 40,"00FFFF00")
            else:
                color_cell(i, 40, "00ccf5ff")

            if sheet1.cell(row=i, column=42).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=42).value]:
                sheet1.cell(row=i, column=42).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=42).value]
                color_cell(i, 42,"00FFFF00")
            else:
                color_cell(i, 42, "00ccf5ff")

            if sheet1.cell(row=i, column=44).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=44).value]:
                sheet1.cell(row=i, column=44).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=44).value]
                color_cell(i, 44,"00FFFF00")
            else:
                color_cell(i, 44, "00ccf5ff")

            if sheet1.cell(row=i, column=46).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=46).value]:
                sheet1.cell(row=i, column=46).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=46).value]
                color_cell(i, 46,"00FFFF00")
            else:
                color_cell(i, 46, "00ccf5ff")

            if sheet1.cell(row=i, column=48).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=48).value]:
                sheet1.cell(row=i, column=48).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=48).value]
                color_cell(i, 48,"00FFFF00")
            else:
                color_cell(i, 48, "00ccf5ff")

            if sheet1.cell(row=i, column=50).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=50).value]:
                sheet1.cell(row=i, column=50).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=50).value]
                color_cell(i, 50,"00FFFF00")
            else:
                color_cell(i, 50, "00ccf5ff")

            if sheet1.cell(row=i, column=52).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=52).value]:
                sheet1.cell(row=i, column=52).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=52).value]
                color_cell(i, 52,"00FFFF00")
            else:
                color_cell(i, 52, "00ccf5ff")

            if sheet1.cell(row=i, column=54).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=54).value]:
                sheet1.cell(row=i, column=54).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=54).value]
                color_cell(i, 54,"00FFFF00")
            else:
                color_cell(i, 54, "00ccf5ff")

            if sheet1.cell(row=i, column=56).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=56).value]:
                sheet1.cell(row=i, column=56).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=56).value]
                color_cell(i, 56,"00FFFF00")
            else:
                color_cell(i, 56, "00ccf5ff")

            if sheet1.cell(row=i, column=58).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=58).value]:
                sheet1.cell(row=i, column=58).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=58).value]
                color_cell(i, 58,"00FFFF00")
            else:
                color_cell(i, 58, "00ccf5ff")

            if sheet1.cell(row=i, column=60).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=60).value]:
                sheet1.cell(row=i, column=60).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=60).value]
                color_cell(i, 60,"00FFFF00")
            else:
                color_cell(i, 60, "00ccf5ff")

            if sheet1.cell(row=i, column=62).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=62).value]:
                sheet1.cell(row=i, column=62).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=62).value]
                color_cell(i, 62,"00FFFF00")
            else:
                color_cell(i, 62, "00ccf5ff")

            if sheet1.cell(row=i, column=64).value != TRMMAP_44_VDF[sheet1.cell(row=2, column=64).value]:
                sheet1.cell(row=i, column=64).value = TRMMAP_44_VDF[sheet1.cell(row=2, column=64).value]
                color_cell(i, 64,"00FFFF00")
            else:
                color_cell(i, 64, "00ccf5ff")

    #print(TRMMAP_dict_initial)
    wb1.save(TITLE+'__import.xlsx')

def NB_ADJNODEID_mapping(file_name):
    wb1 = openpyxl.load_workbook(file_name, data_only=True)
    sheet1 = wb1['ADJNODE']

    def color_cell(X, Y, color):
        sheet1.cell(row=X, column=Y).fill = PatternFill(bgColor=color, fill_type='lightUp')

    #print(len(sheet1['A']))
    for i in range(3, len(sheet1['A'])):
        if sheet1.cell(row=i,column=4).value =='IUB'and sheet1.cell(row=i,column=6).value =='IP' and sheet1.cell(row=i,column=3).value not in NODEB_EXCEPTION_LIST:
            RNC_list.append(sheet1.cell(row=i,column=1).value)
            NB_full_IP_list.append(sheet1.cell(row=i,column=3).value)
            RNC_NB_full_IP_list.append('-'.join([sheet1.cell(row=i,column=1).value,sheet1.cell(row=i,column=3).value]))
            ADJNODEID_full_IP_list.append(sheet1.cell(row=i,column=2).value)
        if sheet1.cell(row=i,column=3).value  in NODEB_EXCEPTION_LIST:
            RNC_ADJNODE_exception_list.append('-'.join([sheet1.cell(row=i,column=1).value,sheet1.cell(row=i,column=2).value]))

    RNC_NB_ADJNODEID_dict=dict(zip(RNC_NB_full_IP_list,ADJNODEID_full_IP_list))
    #ADJNODEID_RNC_dict=dict(zip(ADJNODEID_full_IP_list,RNC_list))
    NB_ADJNODEID_dict=dict(zip(NB_full_IP_list,ADJNODEID_full_IP_list))

    #return value is tuple of 3 dictionaries
    return RNC_NB_ADJNODEID_dict,NB_ADJNODEID_dict,RNC_list,RNC_ADJNODE_exception_list


def IPPATH_change(file_name):
    NB_ADJNODE_tuple_dict = NB_ADJNODEID_mapping(file_name)
    wb1 = openpyxl.load_workbook(file_name, data_only=True)
    sheet1 = wb1['IPPATH']

    def color_cell(X, Y, color):
        sheet1.cell(row=X, column=Y).fill = PatternFill(bgColor=color, fill_type='lightUp')

    for i in range(3, len(sheet1['A'])):
        # if Interface Type==IUB and ADJNODEID of fullIP  and RNC in RNC list and IPLOGICPORT ==NULL and RNC-ADJNODEID not in exception list
        if sheet1.cell(row=i,column=4).value =='IUB' and sheet1.cell(row=i,column=2).value in NB_ADJNODE_tuple_dict[0].values() and sheet1.cell(row=i,column=1).value in NB_ADJNODE_tuple_dict[2] and sheet1.cell(row=i,column=12).value=='NULL' and '-'.join([sheet1.cell(row=i,column=1).value,sheet1.cell(row=i,column=2).value]) not in NB_ADJNODE_tuple_dict[3]:

            if sheet1.cell(row=i,column=3).value=='0':
                if sheet1.cell(row=i, column=6).value != 'BE':
                    sheet1.cell(row=i, column=6).value='BE'
                    color_cell(i, 6,"00FFFF00")
                elif sheet1.cell(row=i, column=6).value == 'BE':
                    color_cell(i, 6, "00ccf5ff")

            if sheet1.cell(row=i,column=3).value=='26':
                if sheet1.cell(row=i, column=6).value != 'AF31':
                    sheet1.cell(row=i, column=6).value = 'AF31'
                    color_cell(i, 6,"00FFFF00")
                elif sheet1.cell(row=i, column=6).value == 'AF31':
                    color_cell(i, 6, "00ccf5ff")

            if sheet1.cell(row=i,column=3).value=='34':
                if sheet1.cell(row=i, column=6).value != 'AF41':
                    sheet1.cell(row=i, column=6).value = 'AF41'
                    color_cell(i, 6,"00FFFF00")
                elif sheet1.cell(row=i, column=6).value == 'AF41':
                    color_cell(i, 6, "00ccf5ff")

            if sheet1.cell(row=i,column=3).value=='40':
                if sheet1.cell(row=i, column=6).value !='AF42':
                    sheet1.cell(row=i, column=6).value='AF42'
                    color_cell(i, 6,"00FFFF00")
                elif sheet1.cell(row=i, column=6).value =='AF42':
                    color_cell(i, 6, "00ccf5ff")

            if sheet1.cell(row=i,column=3).value=='46':
                if sheet1.cell(row=i, column=6).value != 'AF43':
                    sheet1.cell(row=i, column=6).value = 'AF43'
                    color_cell(i, 6,"00FFFF00")
                elif sheet1.cell(row=i, column=6).value == 'AF43':
                    color_cell(i, 6, "00ccf5ff")

            if sheet1.cell(row=i,column=3).value=='48':
                if sheet1.cell(row=i, column=6).value != 'EF':
                    sheet1.cell(row=i, column=6).value = 'EF'
                    color_cell(i,6,"00FFFF00")
                elif sheet1.cell(row=i, column=6).value == 'EF':
                    color_cell(i, 6, "00ccf5ff")

    wb1.save(TITLE+'__import.xlsx')




#main

# select Bulk Export RNC file

change_PHBMAP(EXPORT_FILE_NAME)
print("PHBMAP modifications finished\n")
change_TRMMAP(TITLE+'__import.xlsx')
print("TRMMAP modifications finished\n")
IPPATH_change(TITLE+'__import.xlsx')
print("IPPATH modifications finished\n")
