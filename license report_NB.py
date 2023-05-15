import openpyxl
#csv_export_name='Berlin.csv'
#csv_export_name='Munchen.csv'
#csv_export_name='Hannover.csv'
csv_export_name='Dresden.csv'

def export_to_excel(filein, fileout):
    import csv, openpyxl

    wb = openpyxl.Workbook()

    sheet = wb.active
    csvfile = open(filein, 'r')
    reader = csv.reader(csvfile)
    for line in reader:
        sheet.append(line)  # adaugare lista la final
    wb.save(fileout)


export_to_excel(csv_export_name, csv_export_name[0:len(csv_export_name)-4]+'.xlsx')


def splitf(string):
    ret = string.split('/')
    return ret


def collect_data(file_name):
    results_file = open('NodeB_license_report___'+csv_export_name[0:len(csv_export_name)-4]+'.txt', 'w')

    feature_name_list = []
    feature_available_proportion_value = []
    feature_max_value = []
    feature_available_remain_value = []

    wb1 = openpyxl.load_workbook(file_name, data_only=True)
    sheet1 = wb1['Sheet']
    for i in range(1, len(sheet1['A']) + 1):
        if 'XN' in str(sheet1.cell(row=i, column=1).value):
            results_file.write(str(sheet1.cell(row=i, column=1).value))
            results_file.write('\n')
            results_file.write("LICENSE_ITEM  / Distributed Proportion --- Max Value --- Remain Value"+'\n')
            for item in range(6, 140):
                if sheet1.cell(row=i + 1, column=item).value != None:
                    feature_name_list.append(str(sheet1.cell(row=i + 1, column=item).value))
            for j in range(i + 1, len(sheet1['A']) + 1):
                if 'Distributed Proportion' in str(sheet1.cell(row=j, column=4).value):
                    for item1 in range(6, 140):
                        if sheet1.cell(row=j, column=item1).value is not None:
                            feature_available_proportion_value.append(str(round(int(splitf(sheet1.cell(row=j, column=item1).value)[0]) / int(splitf(sheet1.cell(row=j, column=item1).value)[1]) * 100, 2)) + ' %')
                            feature_available_remain_value.append(sheet1.cell(row=j + 2, column=item1).value)
                            feature_max_value.append(sheet1.cell(row=j + 3, column=item1).value)
            feature_available_proportion_dict = dict(zip(feature_name_list, feature_available_proportion_value))
            feature_max_value_dict = dict(zip(feature_name_list, feature_max_value))
            feature_remain_value_dict = dict(zip(feature_name_list, feature_available_remain_value))
            # results_file.write('\n')

            # results_file.write(str(feature_available_proportion_dict)+'\n')
            # results_file.write(str(feature_max_value_dict)+'\n')
            # results_file.write(str(feature_remain_value_dict)+'\n')

            f1 = 'UL CE'
            f2 = 'DL CE'
            f3 = 'Local Cell'
            f4 = 'UMPT Multi Mode License (UMTS)(per UMPT)'
            f5 = 'PA[43dBm]'
            f6 = 'Power License (per 20W)'
            f7 = 'UBBP Multi-Mode license for UMTS (per UBBP)'
            f8 = 'Multi-carrier License for Multimode for 5000 Series RF Module (UMTS) (per Carrier)'
            f9 = '20W Power License for 5000 Series RF Module (UMTS) (per 20W)'
            f10 = 'Spectrum Sharing License for 5000 Series RF Module (UMTS)(Per Band per RRU)'
            f11 = 'Multimode License for 5000 Series RF Module (UMTS) (per RU)'

            try:
                results_file.write(f1+f'  {feature_available_proportion_dict[f1]} --- {feature_max_value_dict[f1]} --- {feature_remain_value_dict[f1]}'+'\n')
            except:
                pass
            try:
                results_file.write(f2+f'  {feature_available_proportion_dict[f2]} --- {feature_max_value_dict[f2]} --- {feature_remain_value_dict[f2]}'+'\n')
            except:
                pass
            try:
                results_file.write(f3+f'  {feature_available_proportion_dict[f3]} --- {feature_max_value_dict[f3]} --- {feature_remain_value_dict[f3]}'+'\n')
            except:
                pass
            try:
                results_file.write(f4+f'  {feature_available_proportion_dict[f4]} --- {feature_max_value_dict[f4]} --- {feature_remain_value_dict[f4]}'+'\n')
            except:
                pass
                #results_file.write(f5+f'  {feature_available_proportion_dict[f5]} --- {feature_max_value_dict[f5]} --- {feature_remain_value_dict[f5]}')
            try:
                results_file.write(f6+f'  {feature_available_proportion_dict[f6]} --- {feature_max_value_dict[f6]} --- {feature_remain_value_dict[f6]}'+'\n')
            except:
                pass
            try:
                results_file.write(f7+f'  {feature_available_proportion_dict[f7]} --- {feature_max_value_dict[f7]} --- {feature_remain_value_dict[f7]}'+'\n')
            except:
                pass
            try:
                results_file.write(f8+f'  {feature_available_proportion_dict[f8]} --- {feature_max_value_dict[f8]} --- {feature_remain_value_dict[f8]}'+'\n')
            except:
                pass
            try:
                results_file.write(f9+f'  {feature_available_proportion_dict[f9]} --- {feature_max_value_dict[f9]} --- {feature_remain_value_dict[f9]}'+'\n')
            except:
                pass
            try:
                results_file.write(f10+f'  {feature_available_proportion_dict[f10]} --- {feature_max_value_dict[f10]} --- {feature_remain_value_dict[f10]}'+'\n')
            except:
                pass
            try:
                results_file.write(f11+f'  {feature_available_proportion_dict[f11]} --- {feature_max_value_dict[f11]} --- {feature_remain_value_dict[f11]}'+'\n')
            except:
                pass
            try:
                results_file.write('\n')
            except:
                pass


        feature_name_list.clear()
        feature_available_proportion_value.clear()
        feature_max_value.clear()
        feature_available_remain_value.clear()
    wb1.close()



collect_data(csv_export_name[0:len(csv_export_name)-4]+'.xlsx')