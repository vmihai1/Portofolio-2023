# add to program directory all csv files to compare, the 2nd(reference) must have also '_r.csv';

import csv
from openpyxl import Workbook
from os import listdir

def find_csv_filenames( path_to_dir, suffix=".csv" ):
    filenames = listdir(path_to_dir)
    return [filename for filename in filenames if filename.endswith(suffix)]


def compare_csv(csv1,csv2,delimiter=','):
    Param_name_xl1=[]
    Param_Value_xl1=[]
    Param_name_xl2=[]
    Param_Value_xl2=[]

    Csv1= open(csv1,'r')
    wb_xl1 = Workbook()
    ws_xl1 = wb_xl1.active
    for row in csv.reader(Csv1):
        ws_xl1.append(row)

    wb_xl1.save(f'{csv1[0:len(csv1)-4]}'+'.xlsx')
    content_lst_csv1=Csv1.read().splitlines()
    #print(content_lst_csv1)
    # print('\n')
    # Param_name_csv1=content_lst_csv1[0].split(delimiter)
    # Param_Value_csv1=content_lst_csv1[1].split(delimiter)
    # print('\n')
    # print(Param_name_csv1)
    # print(Param_Value_csv1)

    Csv2=open(csv2,'r')
    wb_xl2 = Workbook()
    ws_xl2 = wb_xl2.active
    for row in csv.reader(Csv2):
        ws_xl2.append(row)

    wb_xl2.save(f'{csv2[0:len(csv2)-4]}'+'_ref.xlsx')
    ws_xl1_col=ws_xl1.max_column
    ws_xl2_col = ws_xl2.max_column
    # print(ws_xl1_col)
    # print(ws_xl2_col)

    for i in range(1,ws_xl1_col+1):
        Param_name_xl1.append(ws_xl1.cell(row=1,column=i).value)
    print(Param_name_xl1)
    for i in range(1, ws_xl1_col + 1):
        if ws_xl1.cell(row=2, column=i).value is None:
            Param_Value_xl1.append('EMPTY CELL')
        else:
            Param_Value_xl1.append(ws_xl1.cell(row=2, column=i).value)

    print(Param_Value_xl1)
    for i in range(1, ws_xl2_col + 1):
        Param_name_xl2.append(ws_xl2.cell(row=1, column=i).value)
    print(Param_name_xl2)
    for i in range(1, ws_xl2_col + 1):
        if ws_xl2.cell(row=2, column=i).value is None:
            Param_Value_xl2.append('EMPTY CELL')
        else:

            Param_Value_xl2.append(ws_xl2.cell(row=2, column=i).value)
    print(Param_Value_xl2)

    print(len(Param_name_xl1))
    print(len(Param_Value_xl1))
    print(len(Param_name_xl2))
    print(len(Param_Value_xl2))
    # tuplu_list=[]
    # for n in range(0,len(Param_name_xl2)):
    #     tuplu_list.append((Param_name_xl2[n],Param_Value_xl2[n]))
    # print(tuplu_list)

    file_output = open(f'COMPARE result between target MO -- {csv1[0:len(csv1) - 4]} -- and reference MO -- {csv2[0:len(csv2) - 6]} --.txt', 'w')
    for i in range(0, len(Param_name_xl1)):
        if Param_name_xl1[i] in Param_name_xl2:
            CHECK_RESULT = f'Validation in parameter name "{Param_name_xl1[i]}" between TARGET MO and REFERENCE MO'
            file_output.write("---------------------------------------------------------------------------------------"+ '\n')
            file_output.write("PARAMETER in target csv: '" + Param_name_xl1[i] + "'  VALUE in target_csv: '" + Param_Value_xl1[i] +"'"+ '\n')
            for j in range(0, len(Param_name_xl2)):
                if Param_name_xl2[j] == Param_name_xl1[i]:
                    A = Param_Value_xl2[j]
                    C = Param_name_xl2[j]
            file_output.write("PARAMETER in reference csv: '" + str(C) + "'  VALUE in reference_csv: '" + str(A) +"'"+ '\n')
            file_output.write(CHECK_RESULT + '\n')
            file_output.write("---------------------------------------------------------------------------------------" + '\n')
            file_output.write('\n')
        elif Param_name_xl1[i] not in Param_name_xl2:
            CHECK_RESULT = f'Parameter  "{Param_name_xl1[i]}"  found only in TARGET MO, not in REFERENCE MO'
            file_output.write("---------------------------------------------------------------------------------------" + '\n')
            file_output.write("PARAMETER in target csv: '" + Param_name_xl1[i] + "'  VALUE in target_csv: '" + Param_Value_xl1[i]+"'" + '\n')
            file_output.write(CHECK_RESULT + '\n')
            file_output.write("---------------------------------------------------------------------------------------" + '\n')
            file_output.write('\n')

    content_lst_csv2=Csv2.read().splitlines()



#all csv files to be compared
#print(find_csv_filenames('C:/1_valentin/Singtel/Ericsson/4G_CM/output/',".csv"),'\n')
#all csv files used as reference
#print(find_csv_filenames('C:/1_valentin/Singtel/Ericsson/4G_CM_reference_ENM/output/',".csv"),'\n')

#checking the csv file names


#


# compare_csv('RncFunction.csv','RncFunction_r.csv',delimiter=',')
#compare_csv('IubLink.csv','IubLink_r.csv',delimiter=',')
#compare_csv('UtranCell.csv','UtranCell_r.csv',delimiter=',')
#compare_csv('vsDataUtranCell.csv','vsDataUtranCell_r.csv',delimiter=',')
#compare_csv('vsDataHsdsch.csv','vsDataHsdsch_r.csv',delimiter=',')
#compare_csv('vsDataEul.csv','vsDataEul_r.csv',delimiter=',')
#compare_csv('vsDataMultiCarrier.csv','vsDataMultiCarrier_r.csv',delimiter=',')
#compare_csv('UtranRelation.csv','UtranRelation_r.csv',delimiter=',')
#compare_csv('vsDataUtranRelation.csv','vsDataUtranRelation_r.csv',delimiter=',')
#compare_csv('vsDataEUtranFreqRelation.csv','vsDataEUtranFreqRelation_r.csv',delimiter=',')
#compare_csv('vsDataEUtranFrequency.csv','vsDataEUtranFrequency_r.csv',delimiter=',')
#compare_csv('vsDataEUtraNetwork.csv','vsDataEUtraNetwork_r.csv',delimiter=',')
compare_csv('CaMgtCfg_1.csv','CaMgtCfg_2.csv',delimiter=',')